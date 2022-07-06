from rest_framework import mixins, status
from rest_framework import viewsets
from rest_framework import filters
from rest_framework.pagination import PageNumberPagination
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.decorators import action
from rest_framework.response import Response
from xToolkit import xstring

from bonus.models import AccountSales, AccountBonus, Accounts, MonthList, ExchangeRate, BasicInfo, Manager
from bonus.serializers import AccountSalesSerializer, AccountBonusSerializer, AccountsSerializer, MonthListSerializer, \
    ExchangeRateSerializer, BasicInfoSerializer, ManagerSerializer, AccountsSerializerNodepth, \
    AccountSalesSerializerNodepth
from casedance.settings import BASE_URL


class DefaultPagination(PageNumberPagination):
    """
    分页设置
    """
    page_size = 20
    page_size_query_param = 'page_size'
    page_query_param = 'page'
    max_page_size = 100


class AccountSalesViewSet(mixins.ListModelMixin,
                          mixins.CreateModelMixin,
                          mixins.UpdateModelMixin,
                          mixins.DestroyModelMixin,
                          mixins.RetrieveModelMixin,
                          viewsets.GenericViewSet):
    """
    list:
        账号销售,分页,过滤,搜索,排序
    create:
        账号销售新增
    retrieve:
        账号销售详情页
    update:
        账号销售修改
    destroy:
        账号销售删除
    """
    queryset = AccountSales.objects.all()
    serializer_class = AccountSalesSerializer  # 序列化
    pagination_class = DefaultPagination  # 分页

    filter_backends = (DjangoFilterBackend, filters.SearchFilter)  # 过滤,搜索,排序
    filter_fields = ('platform', 'platform_base', 'account_name', 'manager', 'month')  # 配置过滤字段
    search_fields = ('account_name',)  # 配置搜索字段

    # 重写create，让depth序列号字段能保存数据
    def create(self, request):
        serialized = AccountSalesSerializerNodepth(data=request.data)
        if serialized.is_valid():
            serialized.save()
            return Response(serialized.data)
        else:
            return Response(serialized.errors)

    # 重写update，让depth序列号字段能保存数据
    def update(self, request, pk=None):
        serialized = AccountSalesSerializerNodepth(self.get_object(), data=request.data)
        if serialized.is_valid():
            serialized.save()
            return Response(serialized.data)
        else:
            return Response(serialized.errors, status=status.HTTP_400_BAD_REQUEST)


class AccountBonusViewSet(mixins.ListModelMixin,
                          mixins.CreateModelMixin,
                          mixins.UpdateModelMixin,
                          mixins.DestroyModelMixin,
                          mixins.RetrieveModelMixin,
                          viewsets.GenericViewSet):
    """
    list:
        提成表,分页,过滤,搜索,排序
    create:
        提成表新增
    retrieve:
        提成表详情页
    update:
        提成表修改
    destroy:
        提成表删除
    """
    queryset = AccountBonus.objects.all()
    serializer_class = AccountBonusSerializer  # 序列化
    pagination_class = DefaultPagination  # 分页

    filter_backends = (DjangoFilterBackend, filters.SearchFilter)  # 过滤,搜索,排序
    filter_fields = ('platform', 'platform_base', 'account_name', 'manager', 'month')  # 配置过滤字段
    search_fields = ('account_name',)  # 配置搜索字段

    # 生成提成表
    @action(methods=['post'], detail=False, url_path='create_bonus')
    def create_bonus(self, request):
        month = request.data['month']
        account_sales = AccountSales.objects.filter(month=month)
        for i in account_sales:
            account_bonus = AccountBonus.objects.filter(month=month, account_name=i.account_name,
                                                        platform=i.platform).first()
            if not account_bonus:
                account_bonus = AccountBonus()
                account_bonus.platform = i.platform
                account_bonus.platform_base = i.platform_base
                account_bonus.account_name = i.account_name
                account_bonus.month = i.month
                account_bonus.ori_currency = i.ori_currency
                account_bonus.currency_rate = i.currency_rate
                account_bonus.sale_amount = i.sale_amount
                account_bonus.sale_income = i.sale_income
                account_bonus.receipts = i.receipts
                account_bonus.refund = i.refund
                account_bonus.FES = i.FES
                account_bonus.platform_fees = i.platform_fees
                account_bonus.platform_fees_rmb = i.platform_fees_rmb
                account_bonus.product_cost = i.product_cost
                account_bonus.shipping_cost = i.shipping_cost
                account_bonus.profit = i.profit
                account_bonus.profit_margin = i.profit_margin
                account_bonus.month_profit = i.month_profit
                account_bonus.orders = i.orders
                account_bonus.CUP = i.CUP
                account_bonus.ad_fees = i.ad_fees
                account_bonus.ad_fees_rmb = i.ad_fees_rmb
                account_bonus.ad_percent = i.ad_percent
                account_bonus.cp_cgf_fees = i.cp_cgf_fees
                account_bonus.cp_first_ship = i.cp_first_ship
                if i.manager:
                    account_bonus.manager = i.manager
                    account_bonus.bonus_rate = i.manager.bonus_rate
                    account_bonus.bonus = i.profit * i.manager.bonus_rate
                account_bonus.save()
        return Response({'msg': '操作成功'}, status=status.HTTP_200_OK)

    # 导出excel表
    @action(methods=['post'], detail=False, url_path='export_bonus')
    def export_bonus(self, request):
        import openpyxl
        from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
        alignment = Alignment(horizontal='center', vertical='center')
        title_style = Font(name='微软雅黑', sz=28, b=True)
        title_style2 = Font(name='微软雅黑', sz=8, color='ffffff')
        title_style3 = Font(name='微软雅黑', sz=8)
        title_style3_red = Font(name='微软雅黑', sz=8, color='eb3223', b=True)
        title_style3_b = Font(name='微软雅黑', sz=8, b=True)
        title_style_10b = Font(name='微软雅黑', sz=10, b=True)
        title_style_8grey = Font(name='微软雅黑', sz=8, color='7f7f7f')
        border = Border(
            left=Side(border_style='thin', color='e7e6e6'),
            right=Side(border_style='thin', color='e7e6e6'),
            top=Side(style='thin', color='e7e6e6'),
            bottom=Side(style='thin', color='e7e6e6'))

        all_months = request.data['months']
        months = []
        for a in all_months:
            months.append(a['month'])

        platforms = ['eBay', 'Coupang', 'Aliexpress']
        manager = request.data['manager']
        # 检查数据是否存在
        is_exist = False
        for m in months:
            for p in platforms:
                ab = AccountBonus.objects.filter(month=m, platform=p, manager__name=manager).count()
                if ab:
                    is_exist = True
                    break
        if not is_exist:
            return Response({'msg': '无数据'}, status=status.HTTP_406_NOT_ACCEPTABLE)

        area = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']
        area2 = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']
        wb = openpyxl.Workbook()
        for m in months:
            for p in platforms:
                if p == 'eBay' or p == 'Aliexpress':
                    # 获取当月ebay平台指定负责人的提成数据
                    account_bonus = AccountBonus.objects.filter(month=m, platform=p, manager__name=manager)
                    if account_bonus:
                        f_sheet = wb.create_sheet(m[0:4] + '年' + m[4:6] + '月' + p)
                        num = 0
                        for ab in account_bonus:
                            # 第一次循环
                            if num == 0:
                                f_sheet.merge_cells('A1:P1')
                                f_sheet['A1'] = m[0:4] + '年' + m[4:6] + '月' + p + '销售报表-' + ab.manager.name
                                # 第1行样式
                                f_sheet.row_dimensions[1].height = 50
                                f_sheet.row_dimensions[2].height = 30
                                f_sheet['A1'].alignment = alignment
                                f_sheet['A1'].font = title_style
                                f_sheet['A1'].fill = PatternFill(patternType='solid', fgColor='65d6b4')
                                for i in area:
                                    f_sheet[i + '1'].border = border
                                # 第二行样式
                                for i in area:
                                    f_sheet[i+'2'].alignment = alignment
                                    f_sheet[i+'2'].font = title_style2
                                    f_sheet[i+'2'].fill = PatternFill(patternType='solid', fgColor='df8244')
                                    f_sheet[i+'2'].border = border
                                f_sheet['A2'] = '账号'
                                f_sheet['B2'] = '站点'
                                f_sheet['C2'] = '销售收入'
                                f_sheet['D2'] = '净收款'
                                f_sheet['E2'] = '退款'
                                f_sheet['F2'] = '结汇收入'
                                f_sheet['G2'] = 'eBay费用'
                                f_sheet['H2'] = '产品成本'
                                f_sheet['I2'] = '物流成本'
                                f_sheet['J2'] = '毛利润'
                                f_sheet['K2'] = '订单数'
                                f_sheet['L2'] = '平均客单价'
                                f_sheet['M2'] = '广告费'
                                f_sheet['N2'] = '广告费占比'
                                f_sheet['O2'] = '提成(' + str(ab.bonus_rate*100) + '%)'
                                f_sheet['P2'] = '是否记入提成'

                                # 第3行样式
                                for i in area:
                                    f_sheet[i + '3'].alignment = alignment
                                    f_sheet[i + '3'].font = title_style3
                                    f_sheet[i + '3'].border = border
                                if ab.bonus < 0:
                                    f_sheet['J3'].font = title_style3_red
                                f_sheet['A3'] = ab.account_name
                                f_sheet['B3'] = ab.platform_base
                                f_sheet['C3'] = ab.ori_currency + ' ' + xstring.dispose(ab.sale_amount).humanized_amount(compel=True)
                                f_sheet['D3'] = ab.ori_currency + ' ' + xstring.dispose(ab.receipts).humanized_amount(compel=True)
                                f_sheet['E3'] = ab.ori_currency + ' ' + xstring.dispose(ab.refund).humanized_amount(compel=True)
                                f_sheet['F3'] = '￥' + xstring.dispose(ab.FES).humanized_amount(compel=True)
                                f_sheet['G3'] = ab.ori_currency + ' ' + xstring.dispose(ab.platform_fees).humanized_amount(compel=True)
                                f_sheet['H3'] = '￥' + xstring.dispose(ab.product_cost).humanized_amount(compel=True)
                                f_sheet['I3'] = '￥' + xstring.dispose(ab.shipping_cost).humanized_amount(compel=True)
                                f_sheet['J3'] = '￥' + xstring.dispose(ab.profit).humanized_amount(compel=True)
                                f_sheet['K3'] = ab.orders
                                f_sheet['L3'] = ab.ori_currency + ' ' + xstring.dispose(ab.CUP).humanized_amount(compel=True)
                                f_sheet['M3'] = ab.ori_currency + ' ' + xstring.dispose(ab.ad_fees).humanized_amount(compel=True)
                                f_sheet['N3'] = xstring.dispose(ab.ad_percent).humanized_amount(compel=True) + '%'
                                f_sheet['O3'] = '￥' + xstring.dispose(ab.bonus).humanized_amount(compel=True) if ab.bonus > 0 else 0
                                f_sheet['P3'] = '是' if ab.bonus >= 500 else '否'
                                if ab.bonus >= 500:
                                    f_sheet['P3'].font = title_style3_b
                                else:
                                    f_sheet['P3'].font = title_style3_red
                            elif num > 0:
                                # >=4行样式
                                for i in area:
                                    f_sheet[i + str(num+3)].alignment = alignment
                                    f_sheet[i + str(num+3)].font = title_style3
                                    f_sheet[i + str(num+3)].border = border
                                if ab.bonus < 0:
                                    f_sheet['J' + str(num+3)].font = title_style3_red
                                f_sheet['A'+str(num+3)] = ab.account_name
                                f_sheet['B'+str(num+3)] = ab.platform_base
                                f_sheet['C'+str(num+3)] = ab.ori_currency + ' ' + xstring.dispose(ab.sale_amount).humanized_amount(compel=True)
                                f_sheet['D'+str(num+3)] = ab.ori_currency + ' ' + xstring.dispose(ab.receipts).humanized_amount(compel=True)
                                f_sheet['E'+str(num+3)] = ab.ori_currency + ' ' + xstring.dispose(ab.refund).humanized_amount(compel=True)
                                f_sheet['F'+str(num+3)] = '￥' + xstring.dispose(ab.FES).humanized_amount(compel=True)
                                f_sheet['G'+str(num+3)] = ab.ori_currency + ' ' + xstring.dispose(ab.platform_fees).humanized_amount(compel=True)
                                f_sheet['H'+str(num+3)] = '￥' + xstring.dispose(ab.product_cost).humanized_amount(compel=True)
                                f_sheet['I'+str(num+3)] = '￥' + xstring.dispose(ab.shipping_cost).humanized_amount(compel=True)
                                f_sheet['J'+str(num+3)] = '￥' + xstring.dispose(ab.profit).humanized_amount(compel=True)
                                f_sheet['K'+str(num+3)] = ab.orders
                                f_sheet['L'+str(num+3)] = ab.ori_currency + ' ' + xstring.dispose(ab.CUP).humanized_amount(compel=True)
                                f_sheet['M'+str(num+3)] = ab.ori_currency + ' ' + xstring.dispose(ab.ad_fees).humanized_amount(compel=True)
                                f_sheet['N'+str(num+3)] = xstring.dispose(ab.ad_percent).humanized_amount(compel=True) + '%'
                                f_sheet['O'+str(num+3)] = '￥' + xstring.dispose(ab.bonus).humanized_amount(compel=True) if ab.bonus > 0 else 0
                                f_sheet['P'+str(num+3)] = '是' if ab.bonus >= 500 else '否'
                                if ab.bonus >= 500:
                                    f_sheet['P'+str(num+3)].font = title_style3_b
                                else:
                                    f_sheet['P'+str(num+3)].font = title_style3_red
                            num += 1
                        f_sheet['A' + str(num + 3)] = '备注'
                        f_sheet['A' + str(num + 3)].font = title_style_10b
                        f_sheet.merge_cells('A' + str(num + 3) + ':P' + str(num + 3))
                        f_sheet['A' + str(num + 4)] = '1. 收入汇率和支出汇率按相同计算'
                        f_sheet['A' + str(num + 4)].font = title_style_8grey
                        f_sheet.merge_cells('A' + str(num + 4) + ':P' + str(num + 4))
                        f_sheet['A' + str(num + 5)] = '2. 结汇RMB增加2.5%汇损'
                        f_sheet['A' + str(num + 5)].font = title_style_8grey
                        f_sheet.merge_cells('A' + str(num + 5) + ':P' + str(num + 5))
                        f_sheet['A' + str(num + 6)] = '3. 统计时间按报表美国东部标准时间EST进行计算'
                        f_sheet['A' + str(num + 6)].font = title_style_8grey
                        f_sheet.merge_cells('A' + str(num + 6) + ':P' + str(num + 6))
                        f_sheet['A' + str(num + 7)] = '月份基准汇率'
                        f_sheet['A' + str(num + 7)].font = title_style3
                        er = ExchangeRate.objects.filter(month=m)
                        for e in er:
                            f_sheet['A' + str(num + 8)] = e.currency
                            f_sheet['A' + str(num + 8)].font = title_style3
                            f_sheet['B' + str(num + 8)] = e.rate
                            f_sheet['B' + str(num + 8)].font = title_style3
                            num += 1
                if p == 'Coupang':
                    # 获取当月Coupang平台指定负责人的提成数据
                    account_bonus = AccountBonus.objects.filter(month=m, platform=p, manager__name=manager)
                    if account_bonus:
                        f_sheet = wb.create_sheet(m[0:4] + '年' + m[4:6] + '月' + 'Coupang')
                        num = 0
                        for ab in account_bonus:
                            # 第一次循环
                            if num == 0:
                                f_sheet.merge_cells('A1:L1')
                                f_sheet['A1'] = m[0:4] + '年' + m[4:6] + '月' + 'Coupang销售报表-' + ab.manager.name
                                # 第1行样式
                                f_sheet.row_dimensions[1].height = 50
                                f_sheet.row_dimensions[2].height = 30
                                f_sheet['A1'].alignment = alignment
                                f_sheet['A1'].font = title_style
                                f_sheet['A1'].fill = PatternFill(patternType='solid', fgColor='65d6b4')
                                for i in area2:
                                    f_sheet[i + '1'].border = border
                                # 第二行样式
                                for i in area2:
                                    f_sheet[i+'2'].alignment = alignment
                                    f_sheet[i+'2'].font = title_style2
                                    f_sheet[i+'2'].fill = PatternFill(patternType='solid', fgColor='df8244')
                                    f_sheet[i+'2'].border = border
                                f_sheet['A2'] = '账号'
                                f_sheet['B2'] = '站点'
                                f_sheet['C2'] = '销售收入'
                                f_sheet['D2'] = '销售手续费'
                                f_sheet['E2'] = 'CGF/CGF LITE Fees'
                                f_sheet['F2'] = '最终支付金额'
                                f_sheet['G2'] = '结汇收入'
                                f_sheet['H2'] = '产品成本'
                                f_sheet['I2'] = 'CGF发仓运费'
                                f_sheet['J2'] = '毛利润'
                                f_sheet['K2'] = '提成(' + str(ab.bonus_rate*100) + '%)'
                                f_sheet['L2'] = '是否记入提成'

                                # 第3行样式
                                for i in area2:
                                    f_sheet[i + '3'].alignment = alignment
                                    f_sheet[i + '3'].font = title_style3
                                    f_sheet[i + '3'].border = border
                                if ab.bonus < 0:
                                    f_sheet['J3'].font = title_style3_red
                                f_sheet['A3'] = ab.account_name
                                f_sheet['B3'] = ab.platform_base
                                f_sheet['C3'] = ab.ori_currency + ' ' + xstring.dispose(ab.sale_amount).humanized_amount(compel=True)
                                f_sheet['D3'] = ab.ori_currency + ' ' + xstring.dispose(ab.platform_fees).humanized_amount(compel=True)
                                f_sheet['E3'] = ab.ori_currency + ' ' + xstring.dispose(ab.cp_cgf_fees).humanized_amount(compel=True)
                                f_sheet['F3'] = ab.ori_currency + ' ' + xstring.dispose(ab.receipts).humanized_amount(compel=True)
                                f_sheet['G3'] = '￥' + xstring.dispose(ab.FES).humanized_amount(compel=True)
                                f_sheet['H3'] = '￥' + xstring.dispose(ab.product_cost).humanized_amount(compel=True)
                                f_sheet['I3'] = '￥' + xstring.dispose(ab.cp_first_ship).humanized_amount(compel=True)
                                f_sheet['J3'] = '￥' + xstring.dispose(ab.profit).humanized_amount(compel=True)
                                f_sheet['K3'] = '￥' + xstring.dispose(ab.bonus).humanized_amount(compel=True)
                                f_sheet['L3'] = '是'
                            elif num > 0:
                                # >=4行样式
                                for i in area2:
                                    f_sheet[i + str(num+3)].alignment = alignment
                                    f_sheet[i + str(num+3)].font = title_style3
                                    f_sheet[i + str(num+3)].border = border
                                if ab.bonus < 0:
                                    f_sheet['J3'].font = title_style3_red
                                f_sheet['A'+str(num+3)] = ab.account_name
                                f_sheet['B'+str(num+3)] = ab.platform_base
                                f_sheet['C'+str(num+3)] = ab.ori_currency + ' ' + xstring.dispose(ab.sale_amount).humanized_amount(compel=True)
                                f_sheet['D'+str(num+3)] = ab.ori_currency + ' ' + xstring.dispose(ab.platform_fees).humanized_amount(compel=True)
                                f_sheet['E'+str(num+3)] = ab.ori_currency + ' ' + xstring.dispose(ab.cp_cgf_fees).humanized_amount(compel=True)
                                f_sheet['F'+str(num+3)] = ab.ori_currency + ' ' + xstring.dispose(ab.receipts).humanized_amount(compel=True)
                                f_sheet['G'+str(num+3)] = '￥' + xstring.dispose(ab.FES).humanized_amount(compel=True)
                                f_sheet['H'+str(num+3)] = '￥' + xstring.dispose(ab.product_cost).humanized_amount(compel=True)
                                f_sheet['I'+str(num+3)] = '￥' + xstring.dispose(ab.cp_first_ship).humanized_amount(compel=True)
                                f_sheet['J'+str(num+3)] = '￥' + xstring.dispose(ab.profit).humanized_amount(compel=True)
                                f_sheet['K'+str(num+3)] = '￥' + xstring.dispose(ab.bonus).humanized_amount(compel=True)
                                f_sheet['L'+str(num+3)] = '是'
                            num += 1

        del wb['Sheet']
        wb.save('media/export/销售数据-' + manager + '.xlsx')
        url = BASE_URL + '/media/export/销售数据-' + manager + '.xlsx'
        return Response({'url': url}, status=status.HTTP_200_OK)


class AccountsViewSet(mixins.ListModelMixin,
                      mixins.CreateModelMixin,
                      mixins.UpdateModelMixin,
                      mixins.DestroyModelMixin,
                      mixins.RetrieveModelMixin,
                      viewsets.GenericViewSet):
    """
    list:
        帐号,分页,过滤,搜索,排序
    create:
        帐号新增
    retrieve:
        帐号详情页
    update:
        帐号修改
    destroy:
        帐号删除
    """
    queryset = Accounts.objects.all()
    serializer_class = AccountsSerializer  # 序列化
    pagination_class = DefaultPagination  # 分页

    filter_backends = (DjangoFilterBackend, filters.SearchFilter)  # 过滤,搜索,排序
    filter_fields = ('type',)  # 配置过滤字段
    search_fields = ('name',)  # 配置搜索字段

    # 重写create，让depth序列号字段能保存数据
    def create(self, request):
        serialized = AccountsSerializerNodepth(data=request.data)
        if serialized.is_valid():
            serialized.save()
            return Response(serialized.data)
        else:
            return Response(serialized.errors)

    # 重写update，让depth序列号字段能保存数据
    def update(self, request, pk=None):
        serialized = AccountsSerializerNodepth(self.get_object(), data=request.data)
        if serialized.is_valid():
            serialized.save()
            return Response(serialized.data)
        else:
            return Response(serialized.errors, status=status.HTTP_400_BAD_REQUEST)


class MonthListViewSet(mixins.ListModelMixin,
                       mixins.CreateModelMixin,
                       mixins.UpdateModelMixin,
                       mixins.DestroyModelMixin,
                       mixins.RetrieveModelMixin,
                       viewsets.GenericViewSet):
    """
    list:
        统计月份,分页,过滤,搜索,排序
    create:
        统计月份新增
    retrieve:
        统计月份详情页
    update:
        统计月份修改
    destroy:
        统计月份删除
    """
    queryset = MonthList.objects.all()
    serializer_class = MonthListSerializer  # 序列化
    pagination_class = DefaultPagination  # 分页

    filter_backends = (DjangoFilterBackend, filters.SearchFilter)  # 过滤,搜索,排序
    filter_fields = ('status',)  # 配置过滤字段
    search_fields = ('month',)  # 配置搜索字段

    # 生成当月月份
    @action(methods=['get'], detail=False, url_path='create_month')
    def create_month(self, request):
        from . import tasks
        tasks.create_month()
        return Response({'msg': '操作成功'}, status=status.HTTP_200_OK)


class ExchangeRateViewSet(mixins.ListModelMixin,
                          mixins.CreateModelMixin,
                          mixins.UpdateModelMixin,
                          mixins.DestroyModelMixin,
                          mixins.RetrieveModelMixin,
                          viewsets.GenericViewSet):
    """
    list:
        汇率,分页,过滤,搜索,排序
    create:
        汇率新增
    retrieve:
        汇率详情页
    update:
        汇率修改
    destroy:
        汇率删除
    """
    queryset = ExchangeRate.objects.all()
    serializer_class = ExchangeRateSerializer  # 序列化
    pagination_class = DefaultPagination  # 分页

    filter_backends = (DjangoFilterBackend, filters.SearchFilter)  # 过滤,搜索,排序
    filter_fields = ('currency', 'month')  # 配置过滤字段
    search_fields = ('month',)  # 配置搜索字段


class BasicInfoViewSet(mixins.ListModelMixin,
                       mixins.CreateModelMixin,
                       mixins.UpdateModelMixin,
                       mixins.DestroyModelMixin,
                       mixins.RetrieveModelMixin,
                       viewsets.GenericViewSet):
    """
    list:
        基础信息设置,分页,过滤,搜索,排序
    create:
        基础信息设置新增
    retrieve:
        基础信息设置详情页
    update:
        基础信息设置修改
    destroy:
        基础信息设置删除
    """
    queryset = BasicInfo.objects.all()
    serializer_class = BasicInfoSerializer  # 序列化
    pagination_class = DefaultPagination  # 分页

    filter_backends = (DjangoFilterBackend, filters.SearchFilter)  # 过滤,搜索,排序
    filter_fields = ('type', 'platform')  # 配置过滤字段
    search_fields = ('name',)  # 配置搜索字段


class ManagerViewSet(mixins.ListModelMixin,
                     mixins.CreateModelMixin,
                     mixins.UpdateModelMixin,
                     mixins.DestroyModelMixin,
                     mixins.RetrieveModelMixin,
                     viewsets.GenericViewSet):
    """
    list:
        运营负责人,分页,过滤,搜索,排序
    create:
        运营负责人新增
    retrieve:
        运营负责人详情页
    update:
        运营负责人修改
    destroy:
        运营负责人删除
    """
    queryset = Manager.objects.all()
    serializer_class = ManagerSerializer  # 序列化
