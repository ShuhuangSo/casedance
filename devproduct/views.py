from django.shortcuts import render
from rest_framework import mixins, viewsets, status
from rest_framework.response import Response
from rest_framework import filters
from rest_framework.pagination import PageNumberPagination
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.decorators import action
from django.db.models import F
from django.contrib.auth import get_user_model
import openpyxl
from datetime import datetime, timedelta
import numpy
from casedance.settings import BASE_URL

from devproduct.models import DevProduct, DevPrice, DevChannelData, DevListingChannel, DevListingAccount, DevOrder
from devproduct.serializers import DevProductSerializer, DevPriceSerializer, DevChannelDataSerializer, DevListingChannelSerializer, DevListingAccountSerializer, DevOrderSerializer
from bonus.models import Accounts, Manager
from mercado.models import MLOperateLog, FileUploadNotify
from devproduct import tasks


# Create your views here.
class DefaultPagination(PageNumberPagination):
    """
    分页设置
    """
    page_size = 20
    page_size_query_param = 'page_size'
    page_query_param = 'page'
    max_page_size = 1000


class DevProductViewSet(mixins.ListModelMixin, mixins.CreateModelMixin,
                        mixins.UpdateModelMixin, mixins.DestroyModelMixin,
                        mixins.RetrieveModelMixin, viewsets.GenericViewSet):
    """
    list:
        产品开发列表,分页,过滤,搜索,排序
    create:
        产品开发新增
    retrieve:
        产品开发详情页
    update:
        产品开发修改
    destroy:
        产品开发删除
    """
    queryset = DevProduct.objects.all()
    serializer_class = DevProductSerializer  # 序列化
    pagination_class = DefaultPagination  # 分页

    filter_backends = (DjangoFilterBackend, filters.SearchFilter,
                       filters.OrderingFilter)  # 过滤,搜索,排序
    filterset_fields = {
        'create_time': ['gte', 'lte', 'exact', 'gt', 'lt'],
        'online_time': ['gte', 'lte', 'exact', 'gt', 'lt'],
        'offline_time': ['gte', 'lte', 'exact', 'gt', 'lt'],
        'p_status': ['exact'],
        'local_category': ['exact'],
        'buy_status': ['exact', 'in'],
        'is_elec': ['exact'],
        'is_magnet': ['exact'],
        'is_water': ['exact'],
        'is_dust': ['exact'],
        'is_confirm_data': ['exact'],
        'is_stock': ['exact'],
        'user_id': ['exact', 'in'],
        'percent': ['exact', 'lt'],
        'id': ['exact', 'in'],
        'rate': ['exact', 'in'],
        'dev_listing_channel__platform': ['exact'],
        'dev_listing_channel__site': ['exact'],
        'dev_listing_channel__is_active': ['exact'],
    }
    search_fields = ('sku', 'cn_name', 'en_name', 'keywords', 'note',
                     'tag_name', 'cp_id')  # 配置搜索字段
    ordering_fields = ('create_time', 'online_time', 'offline_time', 'percent',
                       'rate', 'unit_cost')  # 配置排序字段

    # 开发产品批量上传
    @action(methods=['post'], detail=False, url_path='bulk_upload')
    def bulk_upload(self, request):
        import warnings
        warnings.filterwarnings('ignore')

        data = request.data
        wb = openpyxl.load_workbook(data['excel'])
        sheet = wb['上传模板']

        add_list = []
        if sheet.max_row <= 1:
            return Response({
                'msg': '表格不能为空',
                'status': 'error'
            },
                            status=status.HTTP_202_ACCEPTED)
        if sheet['M1'].value != '是否带电 1/0':
            return Response({
                'msg': '表格有误，请下载最新模板',
                'status': 'error'
            },
                            status=status.HTTP_202_ACCEPTED)

        for cell_row in list(sheet)[1:]:
            # 检查1，2列是否为空，空则跳过
            row_status = cell_row[0].value and cell_row[1].value
            if not row_status:
                continue

            cn_name = cell_row[0].value
            en_name = cell_row[1].value
            unit_cost = cell_row[2].value
            if not type(unit_cost) in [float, int]:
                unit_cost = 0
            p_weight = cell_row[3].value
            if not type(p_weight) in [float, int]:
                p_weight = 0
            p_length = cell_row[4].value
            if not type(p_length) in [float, int]:
                p_length = 0
            p_width = cell_row[5].value
            if not type(p_width) in [float, int]:
                p_width = 0
            p_height = cell_row[6].value
            if not type(p_height) in [float, int]:
                p_height = 0
            local_category = cell_row[7].value
            if not local_category:
                local_category = 0
            category = cell_row[8].value
            package_included = cell_row[9].value
            keywords = cell_row[10].value
            desc = cell_row[11].value
            is_elec = cell_row[12].value
            if is_elec == 1:
                is_elec = True
            else:
                is_elec = False
            is_magnet = cell_row[13].value
            if is_magnet == 1:
                is_magnet = True
            else:
                is_magnet = False
            is_dust = cell_row[14].value
            if is_dust == 1:
                is_dust = True
            else:
                is_dust = False
            is_water = cell_row[15].value
            if is_water == 1:
                is_water = True
            else:
                is_water = False
            buy_url1 = cell_row[16].value
            buy_url2 = cell_row[17].value
            refer_url1 = cell_row[18].value

            last_record = DevProduct.objects.all().order_by(-F('id')).first()
            # 生成id
            create_id = 10001
            if last_record:
                create_id = last_record.id + 10001
            sku = 'DEV{user_id}{create_id}'.format(user_id=request.user.id,
                                                   create_id=create_id)

            devp = DevProduct()
            devp.sku = sku
            devp.cn_name = cn_name
            devp.en_name = en_name
            devp.p_status = 'DRAFT'
            devp.is_elec = is_elec
            devp.is_magnet = is_magnet
            devp.is_water = is_water
            devp.is_dust = is_dust
            devp.keywords = keywords
            devp.p_weight = p_weight
            devp.p_length = p_length
            devp.p_width = p_width
            devp.p_height = p_height
            devp.unit_cost = unit_cost
            devp.package_included = package_included
            devp.desc = desc
            devp.local_category = local_category
            devp.category = category
            devp.buy_url1 = buy_url1
            devp.buy_url2 = buy_url2
            devp.refer_url1 = refer_url1
            devp.user_id = request.user.id
            devp.save()

            # 创建操作日志
            log = MLOperateLog()
            log.op_module = 'DEVP'
            log.op_type = 'CREATE'
            log.target_type = 'DEVP_P'
            log.target_id = devp.id
            log.desc = '创建产品'
            log.user = request.user
            log.save()
        return Response({
            'msg': '成功上传!',
            'status': 'success'
        },
                        status=status.HTTP_200_OK)

    # DEV产品图片上传
    @action(methods=['post'], detail=False, url_path='image_upload')
    def image_upload(self, request):
        from PIL import Image
        data = request.data
        product = DevProduct.objects.filter(id=data['id']).first()
        if not product:
            return Response({'msg': '产品不存在'}, status=status.HTTP_202_ACCEPTED)

        path = 'media/dev_product/' + product.sku + '.jpg'
        pic = data['pic']
        content = pic.chunks()
        with open(path, 'wb') as f:
            for i in content:
                f.write(i)
        product.image = 'dev_product/' + product.sku + '.jpg'
        product.save()

        pic_org = Image.open(path)
        # 修改保存图片大小
        pic_ori_new = pic_org.resize((800, 800), Image.ANTIALIAS)
        pic_ori_new.save(path)
        # 增加小图
        pic_new = pic_org.resize((100, 100), Image.ANTIALIAS)
        pic_new.save('media/dev_product/' + product.sku + '_100x100.jpg')

        # 创建操作日志
        log = MLOperateLog()
        log.op_module = 'DEVP'
        log.op_type = 'CREATE'
        log.target_type = 'DEVP_P'
        log.target_id = product.id
        log.desc = '修改产品图片'
        log.user = request.user
        log.save()

        return Response({
            'msg': '操作成功!',
            'status': 'success'
        },
                        status=status.HTTP_200_OK)

    # 发布产品
    @action(methods=['post'], detail=False, url_path='list_product')
    def list_product(self, request):
        data = request.data
        id = data['id']
        listing_channel = data['listingChannel']
        product = DevProduct.objects.get(id=id)

        for i in listing_channel:
            # 检查产品定价情况
            if i['default_active']:
                price = DevPrice.objects.filter(dev_p=product,
                                                platform=i['platform'],
                                                site=i['site']).first()
                if not price:
                    return Response(
                        {
                            'msg': '发布站点产品未定价，请先定价!',
                            'status': 'error'
                        },
                        status=status.HTTP_202_ACCEPTED)
        for i in listing_channel:
            # 创建产品发布渠道
            dlc = DevListingChannel()
            dlc.dev_p = product
            dlc.platform = i['platform']
            dlc.site = i['site']
            dlc.is_active = i['default_active']
            dlc.save()

            # 开发产品上架账号
            if dlc.is_active:
                # 列出可上架账号
                ac_list = Accounts.objects.filter(type=dlc.platform,
                                                  site=dlc.site,
                                                  is_active=True)
                for item in ac_list:
                    # 通过姓名获取用户id
                    User = get_user_model()
                    user = User.objects.filter(
                        first_name=item.manager.name).first()

                    dla = DevListingAccount()
                    dla.dev_p = product
                    dla.platform = dlc.platform
                    dla.site = dlc.site
                    dla.account_name = item.name
                    dla.user_name = item.manager.name
                    if user:
                        dla.user_id = user.id
                    dla.save()
        product.percent = 0
        product.p_status = 'ONLINE'
        product.online_time = datetime.now()
        product.save()

        # 创建操作日志
        log = MLOperateLog()
        log.op_module = 'DEVP'
        log.op_type = 'CREATE'
        log.target_type = 'DEVP_P'
        log.target_id = product.id
        log.desc = '发布产品'
        log.user = request.user
        log.save()

        return Response({
            'msg': '发布成功!',
            'status': 'success'
        },
                        status=status.HTTP_200_OK)

    # 下架产品
    @action(methods=['post'], detail=False, url_path='end_product')
    def end_product(self, request):
        id = request.data['id']
        product = DevProduct.objects.filter(id=id).first()
        # 检查账号链接在线情况
        listings = DevListingAccount.objects.filter(dev_p=product)
        for i in listings:
            if i.is_online:
                return Response({
                    'msg': '该产品还有链接在线，请先下架',
                    'status': 'error'
                },
                                status=status.HTTP_202_ACCEPTED)
        product.p_status = 'OFFLINE'
        product.offline_time = datetime.now()
        product.save()
        # 创建操作日志
        log = MLOperateLog()
        log.op_module = 'DEVP'
        log.op_type = 'EDIT'
        log.target_type = 'DEVP_P'
        log.target_id = product.id
        log.desc = '下架产品'
        log.user = request.user
        log.save()
        return Response({
            'msg': '操作成功!',
            'status': 'success'
        },
                        status=status.HTTP_200_OK)

    # 重新发布上架产品（恢复）
    @action(methods=['post'], detail=False, url_path='relist_product')
    def relist_product(self, request):
        id = request.data['id']
        product = DevProduct.objects.filter(id=id).first()
        product.p_status = 'ONLINE'
        product.online_time = datetime.now()
        product.save()
        # 更新账号创建时间
        DevListingAccount.objects.filter(dev_p=product).update(
            create_time=datetime.now())
        # 创建操作日志
        log = MLOperateLog()
        log.op_module = 'DEVP'
        log.op_type = 'EDIT'
        log.target_type = 'DEVP_P'
        log.target_id = product.id
        log.desc = '重新上架'
        log.user = request.user
        log.save()
        return Response({
            'msg': '操作成功!',
            'status': 'success'
        },
                        status=status.HTTP_200_OK)

    # 采购备货
    @action(methods=['post'], detail=False, url_path='buy_stock')
    def buy_stock(self, request):
        id = request.data['id']
        action = request.data['action']
        note = request.data['note']
        product = DevProduct.objects.filter(id=id).first()
        desc = ''  #日志描述
        # 申请备货
        if action == 'REQUEST':
            desc = '【申请备货】 {note}'.format(note=note)
        # 通过申请
        if action == 'APPROVED':
            desc = '【通过申请】 {note}'.format(note=note)
        # 拒绝申请
        if action == 'REJECTED':
            desc = '【拒绝申请】 {note}'.format(note=note)
        # 已采购
        if action == 'BUYING':
            desc = '【已采购】 {note}'.format(note=note)
        # 已到货
        if action == 'ARRIVED':
            desc = '【已收货】 {note}'.format(note=note)
            product.is_stock = True
        product.buy_status = action
        product.save()
        # 创建操作日志
        log = MLOperateLog()
        log.op_module = 'DEVP'
        log.op_type = 'CREATE'
        log.target_type = 'DEVP_BUY'
        log.target_id = id
        log.desc = desc
        log.user = request.user
        log.save()

        return Response({
            'msg': '操作成功!',
            'status': 'success'
        },
                        status=status.HTTP_200_OK)

    # 我的数据
    @action(methods=['get'], detail=False, url_path='get_my_data')
    def get_my_data(self, request):
        user_id = request.user.id
        # 未上架产品
        offline_product_qty = 0
        offline_product_ids = ''
        devp_list = DevProduct.objects.filter(
            p_status='ONLINE',
            dev_listing_account__user_id=user_id,
            dev_listing_account__is_online=False).distinct()
        for i in devp_list:
            offline_product_qty += 1
            offline_product_ids += str(i.id) + ','
        # 未上架账号
        offline_account_qty = 0
        offline_account_ids = ''
        offline_account_ids_list = []
        ac_list = DevListingAccount.objects.filter(dev_p__p_status='ONLINE',
                                                   user_id=user_id,
                                                   is_online=False)
        for i in ac_list:
            offline_account_qty += 1
            if str(i.dev_p.id) not in offline_account_ids_list:
                offline_account_ids_list.append(str(i.dev_p.id))
        offline_account_ids = ','.join(offline_account_ids_list)
        # 7天未上架产品
        day7 = datetime.now().date() - timedelta(days=7)
        day7_qty = 0
        day7_product_ids = ''
        devp_list = DevProduct.objects.filter(
            p_status='ONLINE',
            dev_listing_account__user_id=user_id,
            dev_listing_account__is_online=False,
            dev_listing_account__create_time__lte=day7).distinct()
        for i in devp_list:
            day7_qty += 1
            day7_product_ids += str(i.id) + ','
        # 调价确认
        price_notify = 0
        price_notify_ids = ''
        devp_list = DevProduct.objects.filter(
            dev_listing_account__user_id=user_id,
            dev_listing_account__notify=1).distinct()
        for i in devp_list:
            price_notify += 1
            price_notify_ids += str(i.id) + ','
        # 暂停确认
        pause_notify = 0
        pause_notify_ids = ''
        devp_list = DevProduct.objects.filter(
            dev_listing_account__user_id=user_id,
            dev_listing_account__notify=2).distinct()
        for i in devp_list:
            pause_notify += 1
            pause_notify_ids += str(i.id) + ','
        # 恢复确认
        restore_notify = 0
        restore_notify_ids = ''
        devp_list = DevProduct.objects.filter(
            dev_listing_account__user_id=user_id,
            dev_listing_account__notify=4).distinct()
        for i in devp_list:
            restore_notify += 1
            restore_notify_ids += str(i.id) + ','
        # 翻新确认
        relist_notify = 0
        relist_notify_ids = ''
        devp_list = DevProduct.objects.filter(
            dev_listing_account__user_id=user_id,
            dev_listing_account__notify=3).distinct()
        for i in devp_list:
            relist_notify += 1
            relist_notify_ids += str(i.id) + ','
        # 下架确认
        end_notify = 0
        end_notify_ids = ''
        devp_list = DevProduct.objects.filter(
            dev_listing_account__user_id=user_id,
            dev_listing_account__notify=5).distinct()
        for i in devp_list:
            end_notify += 1
            end_notify_ids += str(i.id) + ','
        return Response(
            {
                'offline_product_qty': offline_product_qty,
                'offline_product_ids': offline_product_ids,
                'offline_account_qty': offline_account_qty,
                'offline_account_ids': offline_account_ids,
                'day7_qty': day7_qty,
                'day7_product_ids': day7_product_ids,
                'price_notify': price_notify,
                'price_notify_ids': price_notify_ids,
                'pause_notify': pause_notify,
                'pause_notify_ids': pause_notify_ids,
                'restore_notify': restore_notify,
                'restore_notify_ids': restore_notify_ids,
                'relist_notify': relist_notify,
                'relist_notify_ids': relist_notify_ids,
                'end_notify': end_notify,
                'end_notify_ids': end_notify_ids,
            },
            status=status.HTTP_200_OK)

    # 所有人数据
    @action(methods=['get'], detail=False, url_path='get_all_users_data')
    def get_all_users_data(self, request):
        users_list = Manager.objects.filter(is_active=True)
        data_list = []
        for item in users_list:
            # 通过姓名获取用户id
            User = get_user_model()
            user = User.objects.filter(first_name=item.name).first()
            user_id = user.id

            # 未上架产品
            offline_product_qty = 0
            devp_list = DevProduct.objects.filter(
                p_status='ONLINE',
                dev_listing_account__user_id=user_id,
                dev_listing_account__is_online=False).distinct()
            for i in devp_list:
                offline_product_qty += 1
            # 未上架账号
            offline_account_qty = 0
            ac_list = DevListingAccount.objects.filter(
                dev_p__p_status='ONLINE', user_id=user_id, is_online=False)
            for i in ac_list:
                offline_account_qty += 1
            # 7天未上架产品
            day7 = datetime.now().date() - timedelta(days=7)
            day7_qty = 0
            devp_list = DevProduct.objects.filter(
                p_status='ONLINE',
                dev_listing_account__user_id=user_id,
                dev_listing_account__is_online=False,
                dev_listing_account__create_time__lte=day7).distinct()
            for i in devp_list:
                day7_qty += 1

            # 7天上架产品数量
            day7_online_qty = 0
            devp_list = DevProduct.objects.filter(
                p_status='ONLINE',
                dev_listing_account__user_id=user_id,
                dev_listing_account__is_online=True,
                dev_listing_account__online_time__gte=day7).distinct()
            for i in devp_list:
                day7_online_qty += 1

            # 平均上架时长
            ac_list = DevListingAccount.objects.filter(
                dev_p__p_status='ONLINE', user_id=user_id, is_online=True)
            time_list = []
            avg_days = 0
            for i in ac_list:
                # 天数
                days = (i.online_time - i.create_time).days
                micro_day = round(
                    (i.online_time - i.create_time).seconds / 3600 / 24, 3)
                time_list.append(days + micro_day)
            if len(time_list):
                avg_days = round(numpy.average(time_list), 1)

            data_list.append({
                'offline_product_qty': offline_product_qty,
                'offline_account_qty': offline_account_qty,
                'day7_qty': day7_qty,
                'day7_online_qty': day7_online_qty,
                'user_name': item.name,
                'avg_days': avg_days,
            })

        return Response(data_list, status=status.HTTP_200_OK)

    # 产品开发数据
    @action(methods=['get'], detail=False, url_path='get_dev_data')
    def get_dev_data(self, request):
        all_products = DevProduct.objects.filter(p_status='ONLINE')
        # 全部发布产品数量
        all_qty = all_products.count()
        # 7天发布数量
        day7 = datetime.now().date() - timedelta(days=7)
        day7_qty = all_products.filter(online_time__gte=day7).count()
        # 15天发布数量
        day15 = datetime.now().date() - timedelta(days=15)
        day15_qty = all_products.filter(online_time__gte=day15).count()
        # 30天发布数量
        day30 = datetime.now().date() - timedelta(days=30)
        day30_qty = all_products.filter(online_time__gte=day30).count()
        return Response(
            {
                'all_qty': all_qty,
                'day7_qty': day7_qty,
                'day15_qty': day15_qty,
                'day30_qty': day30_qty,
            },
            status=status.HTTP_200_OK)

    # 备货情况
    @action(methods=['get'], detail=False, url_path='get_purchase')
    def get_purchase(self, request):
        # 审批中数量
        checking_qty = DevProduct.objects.filter(buy_status='REQUEST').count()
        # 待采购数量
        wait_buy_qty = DevProduct.objects.filter(buy_status='APPROVED').count()
        # 待收货数量
        wait_rec_qty = DevProduct.objects.filter(buy_status='BUYING').count()
        # 审批不通过
        rejected_qty = DevProduct.objects.filter(buy_status='REJECTED').count()
        return Response(
            {
                'checking_qty': checking_qty,
                'wait_buy_qty': wait_buy_qty,
                'wait_rec_qty': wait_rec_qty,
                'rejected_qty': rejected_qty,
            },
            status=status.HTTP_200_OK)

    # 产品关联绑定
    @action(methods=['post'], detail=False, url_path='cp_products')
    def cp_products(self, request):
        ids = request.data['ids']
        if len(ids) == 1:
            return Response({
                'msg': '单个产品无法关联!',
                'status': 'error'
            },
                            status=status.HTTP_202_ACCEPTED)

        exist_cp = []
        for i in ids:
            dm = DevProduct.objects.get(id=i)
            if dm.cp_id:
                # 如果关联id不在列表中就加进去
                if dm.cp_id not in exist_cp:
                    exist_cp.append(dm.cp_id)
        if len(exist_cp) > 1:
            return Response({
                'msg': '产品关联冲突，请检查!',
                'status': 'error'
            },
                            status=status.HTTP_202_ACCEPTED)
        if len(exist_cp) == 1:
            cp_id = exist_cp[0]
            for i in ids:
                dm = DevProduct.objects.get(id=i)
                dm.cp_id = cp_id
                dm.save()

        if len(exist_cp) == 0:
            cp_id = ids[0]
            for i in ids:
                dm = DevProduct.objects.get(id=i)
                dm.cp_id = cp_id
                dm.save()
        # 创建操作日志
        log = MLOperateLog()
        log.op_module = 'DEVP'
        log.op_type = 'EDIT'
        log.target_type = 'DEVP_P'
        log.desc = '产品关联'
        log.user = request.user
        log.save()
        return Response({
            'msg': '成功关联!',
            'status': 'success'
        },
                        status=status.HTTP_200_OK)

    # 产品关联解绑
    @action(methods=['post'], detail=False, url_path='ucp_products')
    def ucp_products(self, request):
        m_id = request.data['id']
        dm = DevProduct.objects.get(id=m_id)
        delete_cp_id = dm.cp_id
        dm.cp_id = None
        dm.save()

        rest_count = DevProduct.objects.filter(cp_id=delete_cp_id).count()
        # 如果只有剩下1
        if rest_count == 1:
            dm1 = DevProduct.objects.get(cp_id=delete_cp_id)
            dm1.cp_id = None
            dm1.save()

        if rest_count > 1 and DevProduct.objects.filter(cp_id=m_id).count():
            new_cp = DevProduct.objects.filter(cp_id=m_id).first()
            new_cp_id = new_cp.id
            qs = DevProduct.objects.filter(cp_id=m_id)
            for i in qs:
                i.cp_id = new_cp_id
                i.save()
        # 创建操作日志
        log = MLOperateLog()
        log.op_module = 'DEVP'
        log.op_type = 'EDIT'
        log.target_type = 'DEVP_P'
        log.desc = '产品解除关联'
        log.user = request.user
        log.save()
        return Response({
            'msg': '成功解除!',
            'status': 'success'
        },
                        status=status.HTTP_200_OK)

    # 获取产品关联数据
    @action(methods=['post'], detail=False, url_path='get_cp_list')
    def get_cp_list(self, request):
        id = request.data['id']
        dp = DevProduct.objects.filter(id=id).first()
        p_list = []
        if dp.cp_id:
            # 查出关联数据
            cps = DevProduct.objects.filter(cp_id=dp.cp_id).exclude(id=id)
            for i in cps:
                p_list.append({
                    'id': i.id,
                    'sku': i.sku,
                    'p_status': i.p_status,
                    'cn_name': i.cn_name,
                    'en_name': i.en_name,
                    'image': BASE_URL + i.image.url if i.image else '',
                    'main': False
                })
            # 将本产品数据插入到列表最前面
            p_list.insert(
                0, {
                    'id': dp.id,
                    'sku': dp.sku,
                    'p_status': dp.p_status,
                    'cn_name': dp.cn_name,
                    'en_name': dp.en_name,
                    'image': BASE_URL + dp.image.url if dp.image else '',
                    'main': True
                })
        return Response(p_list, status=status.HTTP_200_OK)


class DevPriceViewSet(mixins.ListModelMixin, mixins.CreateModelMixin,
                      mixins.UpdateModelMixin, mixins.DestroyModelMixin,
                      mixins.RetrieveModelMixin, viewsets.GenericViewSet):
    """
    list:
        产品开发定价列表,分页,过滤,搜索,排序
    create:
        产品开发定价新增
    retrieve:
        产品开发定价详情页
    update:
        产品开发定价修改
    destroy:
        产品开发定价删除
    """
    queryset = DevPrice.objects.all()
    serializer_class = DevPriceSerializer  # 序列化
    pagination_class = DefaultPagination  # 分页

    filter_backends = (DjangoFilterBackend, filters.SearchFilter,
                       filters.OrderingFilter)  # 过滤,搜索,排序
    filter_fields = ('dev_p__id', 'platform')  # 配置过滤字段

    # 新增定价
    @action(methods=['post'], detail=False, url_path='create_price')
    def create_price(self, request):
        data = request.data
        dev_p_id = data['dev_p_id']  # 开发产品id
        platform = data['platform']  # 平台
        site = data['site']  # 站点
        price = data['price']  # 定价
        currency = data['currency']  # 币种
        ex_rate = data['ex_rate']  # 汇率
        profit = data['profit']  # 毛利润
        note = data['note']  # 备注

        product = DevProduct.objects.filter(id=dev_p_id).first()
        if product:
            dp = DevPrice()
            dp.dev_p = product
            dp.platform = platform
            dp.site = site
            dp.price = price
            dp.currency = currency
            dp.ex_rate = ex_rate
            dp.note = note
            dp.profit = profit
            dp.gross_margin = profit / (price * ex_rate)
            dp.save()
        else:
            return Response({
                'msg': '保存失败!',
                'status': 'error'
            },
                            status=status.HTTP_202_ACCEPTED)

        # 创建操作日志
        log = MLOperateLog()
        log.op_module = 'DEVP'
        log.op_type = 'CREATE'
        log.target_type = 'DEVP_P'
        log.target_id = product.id
        log.desc = '{platform} {site} 新增产品定价'.format(platform=platform,
                                                     site=site)
        log.user = request.user
        log.save()
        return Response({
            'msg': '操作成功!',
            'status': 'success'
        },
                        status=status.HTTP_200_OK)

    # 修改定价
    @action(methods=['post'], detail=False, url_path='update_price')
    def update_price(self, request):
        data = request.data
        id = data['id']  # 定价id
        price = data['price']  # 定价
        ex_rate = data['ex_rate']  # 汇率
        profit = data['profit']  # 毛利润
        note = data['note']  # 备注

        dp = DevPrice.objects.filter(id=id).first()
        if dp:
            desc = '修改定价信息'
            # 查看调价状态，发送调价通知
            is_change = False if dp.price == price else True
            if is_change:
                dla_list = DevListingAccount.objects.filter(
                    dev_p=dp.dev_p,
                    platform=dp.platform,
                    site=dp.site,
                    is_online=True)
                for i in dla_list:
                    i.notify = 1
                    i.save()
                # 调价日志描述
                desc = '{platform} {site} 调价: {old_p} ===> {new_p}'.format(
                    platform=dp.platform,
                    site=dp.site,
                    old_p=dp.price,
                    new_p=price)
            # 汇率变动日志描述
            if dp.ex_rate != ex_rate:
                desc = '{platform} {site} 定价汇率变动: {old_p} ===> {new_p}'.format(
                    platform=dp.platform,
                    site=dp.site,
                    old_p=dp.ex_rate,
                    new_p=ex_rate)
            # 毛利变动日志描述
            if dp.profit != profit:
                desc = '{platform} {site} 预估毛利润变动: {old_p} ===> {new_p}'.format(
                    platform=dp.platform,
                    site=dp.site,
                    old_p=dp.profit,
                    new_p=profit)
            # 创建操作日志
            log = MLOperateLog()
            log.op_module = 'DEVP'
            log.op_type = 'EDIT'
            log.target_type = 'DEVP_P'
            log.target_id = dp.dev_p.id
            log.desc = desc
            log.user = request.user
            log.save()

            dp.price = price
            dp.ex_rate = ex_rate
            dp.note = note
            dp.profit = profit
            dp.gross_margin = profit / (price * ex_rate)
            dp.save()
        else:
            return Response({
                'msg': '保存失败!',
                'status': 'error'
            },
                            status=status.HTTP_202_ACCEPTED)

        return Response({
            'msg': '操作成功!',
            'status': 'success'
        },
                        status=status.HTTP_200_OK)

    # 删除定价
    @action(methods=['post'], detail=False, url_path='delete_price')
    def delete_price(self, request):
        data = request.data
        id = data['id']  # 定价id
        dp = DevPrice.objects.filter(id=id).first()

        # 检查发布渠道情况
        dlc = DevListingChannel.objects.filter(dev_p=dp.dev_p,
                                               platform=dp.platform,
                                               site=dp.site).first()
        if dlc:
            if dlc.is_active:
                return Response(
                    {
                        'msg': '删除失败！删除之前请先取消该站点发布渠道！',
                        'status': 'error'
                    },
                    status=status.HTTP_202_ACCEPTED)

        # 创建操作日志
        log = MLOperateLog()
        log.op_module = 'DEVP'
        log.op_type = 'DEL'
        log.target_type = 'DEVP_P'
        log.target_id = dp.dev_p.id
        log.desc = '{platform} {site} 删除定价'.format(platform=dp.platform,
                                                   site=dp.site)
        log.user = request.user
        log.save()

        dp.delete()
        return Response({
            'msg': '操作成功!',
            'status': 'success'
        },
                        status=status.HTTP_200_OK)

    # 发送通知
    @action(methods=['post'], detail=False, url_path='send_notify')
    def send_notify(self, request):
        data = request.data
        id = data['id']  # 定价id
        notify = data['notify']
        dp = DevPrice.objects.filter(id=id).first()

        dla_list = DevListingAccount.objects.filter(dev_p=dp.dev_p,
                                                    platform=dp.platform,
                                                    site=dp.site,
                                                    is_online=True)
        # 发送暂停/翻新通知
        for i in dla_list:
            i.notify = notify
            i.save()
        # 创建操作日志
        msg = ''
        if notify == 2:
            msg = '暂停'
        if notify == 3:
            msg = '翻新'
        if notify == 4:
            msg = '恢复'
        log = MLOperateLog()
        log.op_module = 'DEVP'
        log.op_type = 'EDIT'
        log.target_type = 'DEVP_P'
        log.target_id = dp.dev_p.id
        log.desc = '{platform} {site} 发送{msg}通知'.format(platform=dp.platform,
                                                        site=dp.site,
                                                        msg=msg)
        log.user = request.user
        log.save()
        return Response({
            'msg': '操作成功!',
            'status': 'success'
        },
                        status=status.HTTP_200_OK)


class DevListingAccountViewSet(mixins.ListModelMixin, mixins.CreateModelMixin,
                               mixins.UpdateModelMixin,
                               mixins.DestroyModelMixin,
                               mixins.RetrieveModelMixin,
                               viewsets.GenericViewSet):
    """
    list:
        开发产品上架账号列表,分页,过滤,搜索,排序
    create:
        开发产品上架账号新增
    retrieve:
        开发产品上架账号详情页
    update:
        开发产品上架账号修改
    destroy:
        开发产品上架账号删除
    """
    queryset = DevListingAccount.objects.all()
    serializer_class = DevListingAccountSerializer  # 序列化
    pagination_class = DefaultPagination  # 分页

    filter_backends = (DjangoFilterBackend, filters.SearchFilter,
                       filters.OrderingFilter)  # 过滤,搜索,排序
    filter_fields = ('dev_p__id', 'dev_p__p_status', 'platform', 'site',
                     'is_online', 'user_name', 'user_id', 'is_paused',
                     'account_name')  # 配置过滤字段
    search_fields = ('dev_p__sku', 'dev_p__cn_name', 'dev_p__en_name',
                     'item_id', 'account_name')  # 配置搜索字段
    ordering_fields = ('online_time', 'offline_time', 'site', 'user_name',
                       'is_online')  # 配置排序字段

    # 账号上架
    @action(methods=['post'], detail=False, url_path='create_listing')
    def create_listing(self, request):
        data = request.data
        id = data['id']
        item_id = data['item_id']
        note = data['note']
        dla = DevListingAccount.objects.filter(id=id).first()
        # 检查itemID是否已经存在(关联产品允许同itemID)
        same_product_exist = DevListingAccount.objects.filter(
            item_id=item_id, dev_p__id=dla.dev_p.id).count()
        is_exist = False
        if dla.dev_p.cp_id:
            # 存在关联产品
            is_exist = DevListingAccount.objects.filter(
                item_id=item_id).exclude(dev_p__cp_id=dla.dev_p.cp_id).count()
        else:
            # 不存在关联产品
            is_exist = DevListingAccount.objects.filter(
                item_id=item_id).count()
        if is_exist or same_product_exist:
            return Response({
                'msg': '上架失败,itemID已存在!',
                'status': 'error'
            },
                            status=status.HTTP_202_ACCEPTED)

        if dla:
            dla.item_id = item_id
            dla.note = note
            dla.is_online = True
            dla.online_time = datetime.now()
            dla.save()
        else:
            return Response({
                'msg': '保存失败!',
                'status': 'error'
            },
                            status=status.HTTP_202_ACCEPTED)

        # 计算开发产品上架率
        calc_listing_online_rate(dla.dev_p.id)
        # 创建操作日志
        log = MLOperateLog()
        log.op_module = 'DEVP'
        log.op_type = 'EDIT'
        log.target_type = 'DEVP_P'
        log.target_id = dla.dev_p.id
        log.desc = '账号{account} 上架, itemID: {item_id}'.format(
            account=dla.account_name, item_id=item_id)
        log.user = request.user
        log.save()
        return Response({
            'msg': '操作成功!',
            'status': 'success'
        },
                        status=status.HTTP_200_OK)

    # 修改
    @action(methods=['post'], detail=False, url_path='update_listing')
    def update_listing(self, request):
        data = request.data
        id = data['id']
        item_id = data['item_id']
        note = data['note']

        dla = DevListingAccount.objects.filter(id=id).first()
        # 修改日志
        if dla.item_id != item_id:
            desc = '{account} 修改itemID: {old_p} ===> {new_p}'.format(
                account=dla.account_name, old_p=dla.item_id, new_p=item_id)
            # 创建操作日志
            log = MLOperateLog()
            log.op_module = 'DEVP'
            log.op_type = 'EDIT'
            log.target_type = 'DEVP_P'
            log.target_id = dla.dev_p.id
            log.desc = desc
            log.user = request.user
            log.save()
        if dla.note != note:
            desc = '{account} 修改i备注'.format(account=dla.account_name)
            # 创建操作日志
            log = MLOperateLog()
            log.op_module = 'DEVP'
            log.op_type = 'EDIT'
            log.target_type = 'DEVP_P'
            log.target_id = dla.dev_p.id
            log.desc = desc
            log.user = request.user
            log.save()
        if dla:
            dla.item_id = item_id
            dla.note = note
            dla.save()
        else:
            return Response({
                'msg': '保存失败!',
                'status': 'error'
            },
                            status=status.HTTP_202_ACCEPTED)

        return Response({
            'msg': '操作成功!',
            'status': 'success'
        },
                        status=status.HTTP_200_OK)

    # 账号下架
    @action(methods=['post'], detail=False, url_path='end_listing')
    def end_listing(self, request):
        data = request.data
        id = data['id']

        dla = DevListingAccount.objects.filter(id=id).first()
        # 检查是否下架的是自己的链接
        if not request.user.is_superuser and dla.user_id != request.user.id:
            return Response({
                'msg': '只能下架自己的链接!',
                'status': 'error'
            },
                            status=status.HTTP_202_ACCEPTED)
        old_item_id = dla.item_id
        if dla:
            dla.item_id = ''
            dla.is_online = False
            dla.notify = 0
            dla.offline_time = datetime.now()
            dla.online_time = None
            dla.save()
        else:
            return Response({
                'msg': '保存失败!',
                'status': 'error'
            },
                            status=status.HTTP_202_ACCEPTED)

        # 计算开发产品上架率
        calc_listing_online_rate(dla.dev_p.id)
        # 创建操作日志
        log = MLOperateLog()
        log.op_module = 'DEVP'
        log.op_type = 'EDIT'
        log.target_type = 'DEVP_P'
        log.target_id = dla.dev_p.id
        log.desc = '账号{account} 下架, itemID: {item_id}'.format(
            account=dla.account_name, item_id=old_item_id)
        log.user = request.user
        log.save()
        return Response({
            'msg': '操作成功!',
            'status': 'success'
        },
                        status=status.HTTP_200_OK)

    # 账号批量下架
    @action(methods=['post'], detail=False, url_path='bulk_end_listing')
    def bulk_end_listing(self, request):
        data = request.data
        ids = data['ids']
        # 检查是否下架的是自己的链接
        if not request.user.is_superuser:
            for id in ids:
                dla = DevListingAccount.objects.filter(id=id).first()
                if dla.user_id != request.user.id:
                    return Response(
                        {
                            'msg': '批量下架只能下架自己的链接!',
                            'status': 'error'
                        },
                        status=status.HTTP_202_ACCEPTED)

        for id in ids:
            dla = DevListingAccount.objects.filter(id=id).first()
            old_item_id = dla.item_id
            if dla:
                dla.item_id = ''
                dla.is_online = False
                dla.notify = 0
                dla.offline_time = datetime.now()
                dla.online_time = None
                dla.save()
            else:
                return Response({
                    'msg': '操作失败!',
                    'status': 'error'
                },
                                status=status.HTTP_202_ACCEPTED)

            # 计算开发产品上架率
            calc_listing_online_rate(dla.dev_p.id)
            # 创建操作日志
            log = MLOperateLog()
            log.op_module = 'DEVP'
            log.op_type = 'EDIT'
            log.target_type = 'DEVP_P'
            log.target_id = dla.dev_p.id
            log.desc = '(批量操作) 账号{account} 下架, itemID: {item_id}'.format(
                account=dla.account_name, item_id=old_item_id)
            log.user = request.user
            log.save()
        return Response({
            'msg': '操作成功!',
            'status': 'success'
        },
                        status=status.HTTP_200_OK)

    # 账号暂停/恢复
    @action(methods=['post'], detail=False, url_path='pause_listing')
    def pause_listing(self, request):
        data = request.data
        id = data['id']
        ac_status = data['status']
        dla = DevListingAccount.objects.filter(id=id).first()
        if dla:
            if dla.notify == 2:
                dla.notify = 0
            dla.is_paused = not ac_status
            dla.save()
        else:
            return Response({
                'msg': '保存失败!',
                'status': 'error'
            },
                            status=status.HTTP_202_ACCEPTED)
        # 创建操作日志
        if ac_status:
            desc = '账号{account} 恢复链接'.format(account=dla.account_name)
        else:
            desc = '账号{account} 暂停链接'.format(account=dla.account_name)
        log = MLOperateLog()
        log.op_module = 'DEVP'
        log.op_type = 'EDIT'
        log.target_type = 'DEVP_P'
        log.target_id = dla.dev_p.id
        log.desc = desc
        log.user = request.user
        log.save()
        return Response({
            'msg': '操作成功!',
            'status': 'success'
        },
                        status=status.HTTP_200_OK)

    # 确认变动通知
    @action(methods=['post'], detail=False, url_path='confirm_notify')
    def confirm_notify(self, request):
        data = request.data
        id = data['id']
        notify = data['notify']
        dla = DevListingAccount.objects.filter(id=id).first()

        # 确认调价通知
        if notify == 1:
            pass
        # 确认链接暂停通知
        if notify == 2:
            pass
        dla.notify = 0
        dla.save()

        return Response({
            'msg': '操作成功!',
            'status': 'success'
        },
                        status=status.HTTP_200_OK)

    # 同步账号信息到产品开发
    @action(methods=['post'], detail=False, url_path='sync_account')
    def sync_account(self, request):
        platform = request.data['platform']
        site = request.data['site']
        account_name = request.data['account_name']
        user_name = request.data['user_name']
        op_type = request.data['op_type']
        is_active = request.data['is_active']
        # 通过姓名获取用户id
        User = get_user_model()
        user = User.objects.filter(first_name=user_name).first()

        dla_list = DevListingAccount.objects.filter(account_name=account_name)

        # 修改负责人
        if op_type == 'USER':

            for i in dla_list:
                i.user_id = user.id
                i.user_name = user_name
                i.save()
        # 修改站点
        if op_type == 'SITE':
            for i in dla_list:
                if i.is_online:
                    return Response(
                        {
                            'msg': '同步失败！该账号还有在线链接，请先下架该账号所有链接！',
                            'status': 'error'
                        },
                        status=status.HTTP_202_ACCEPTED)
            for i in dla_list:
                is_exist = DevListingChannel.objects.filter(
                    dev_p=i.dev_p,
                    site=site,
                    platform=platform,
                    is_active=True).count()
                # 查看该产品对应站点是否发布
                if is_exist:
                    i.site = site
                    i.platform = platform
                    i.save()
                else:
                    i.delete()
        # 修改账号状态
        if op_type == 'STATUS':
            # 如果是下线账号
            if not is_active:
                for i in dla_list:
                    if i.is_online:
                        return Response(
                            {
                                'msg': '同步失败！该账号还有在线链接，请下架先该账号所有链接！',
                                'status': 'error'
                            },
                            status=status.HTTP_202_ACCEPTED)
                dla_list.delete()
            else:
                # 上线新账号
                dev_list = DevProduct.objects.filter(
                    dev_listing_channel__platform=platform,
                    dev_listing_channel__site=site,
                    dev_listing_channel__is_active=True)
                for i in dev_list:
                    dla = DevListingAccount()
                    dla.dev_p = i
                    dla.platform = platform
                    dla.site = site
                    dla.account_name = account_name
                    dla.user_id = user.id
                    dla.user_name = user_name
                    dla.save()

        return Response({
            'msg': '操作成功!',
            'status': 'success'
        },
                        status=status.HTTP_200_OK)


class DevListingChannelViewSet(mixins.ListModelMixin, mixins.CreateModelMixin,
                               mixins.UpdateModelMixin,
                               mixins.DestroyModelMixin,
                               mixins.RetrieveModelMixin,
                               viewsets.GenericViewSet):
    """
    list:
        开发产品发布渠道列表,分页,过滤,搜索,排序
    create:
        开发产品发布渠道新增
    retrieve:
        开发产品发布渠道详情页
    update:
        开发产品发布渠道修改
    destroy:
        开发产品发布渠道删除
    """
    queryset = DevListingChannel.objects.all()
    serializer_class = DevListingChannelSerializer  # 序列化
    filter_backends = (DjangoFilterBackend, filters.SearchFilter,
                       filters.OrderingFilter)  # 过滤,搜索,排序
    filter_fields = ('dev_p__id', 'platform')  # 配置过滤字段

    # 修改
    @action(methods=['post'], detail=False, url_path='update_channel')
    def update_channel(self, request):
        data = request.data
        id = data['id']
        is_active = data['is_active']
        dlc = DevListingChannel.objects.get(id=id)

        # 取消发布
        if not is_active:
            # 检查账号链接在线情况
            dla_list = DevListingAccount.objects.filter(dev_p=dlc.dev_p,
                                                        platform=dlc.platform,
                                                        site=dlc.site)
            for i in dla_list:
                if i.is_online:
                    return Response(
                        {
                            'msg': '删除失败！该站点还有在线链接，请先下架该站点所有链接！',
                            'status': 'error'
                        },
                        status=status.HTTP_202_ACCEPTED)
            # 删除对应账号
            for i in dla_list:
                i.delete()
            dlc.is_active = False
            dlc.save()
            # 创建操作日志
            log = MLOperateLog()
            log.op_module = 'DEVP'
            log.op_type = 'EDIT'
            log.target_type = 'DEVP_P'
            log.target_id = dlc.dev_p.id
            log.desc = '{platform} {site} 渠道取消发布'.format(platform=dlc.platform,
                                                         site=dlc.site)
            log.user = request.user
            log.save()
        else:
            # 发布渠道
            # 检查产品定价情况
            price = DevPrice.objects.filter(dev_p=dlc.dev_p,
                                            platform=dlc.platform,
                                            site=dlc.site).first()
            if not price:
                return Response({
                    'msg': '发布站点产品未定价，请先定价!',
                    'status': 'error'
                },
                                status=status.HTTP_202_ACCEPTED)
            # 列出可上架账号
            ac_list = Accounts.objects.filter(type=dlc.platform,
                                              site=dlc.site,
                                              is_active=True)
            for item in ac_list:
                # 通过姓名获取用户id
                User = get_user_model()
                user = User.objects.filter(
                    first_name=item.manager.name).first()

                dla = DevListingAccount()
                dla.dev_p = dlc.dev_p
                dla.platform = dlc.platform
                dla.site = dlc.site
                dla.account_name = item.name
                dla.user_name = item.manager.name
                if user:
                    dla.user_id = user.id
                dla.save()
            dlc.is_active = True
            dlc.save()
            # 创建操作日志
            log = MLOperateLog()
            log.op_module = 'DEVP'
            log.op_type = 'EDIT'
            log.target_type = 'DEVP_P'
            log.target_id = dlc.dev_p.id
            log.desc = '{platform} {site} 渠道发布'.format(platform=dlc.platform,
                                                       site=dlc.site)
            log.user = request.user
            log.save()

        # 计算开发产品上架率
        calc_listing_online_rate(dlc.dev_p.id)
        return Response({
            'msg': '操作成功!',
            'status': 'success'
        },
                        status=status.HTTP_200_OK)


class DevChannelDataViewSet(mixins.ListModelMixin, mixins.CreateModelMixin,
                            mixins.UpdateModelMixin, mixins.DestroyModelMixin,
                            mixins.RetrieveModelMixin,
                            viewsets.GenericViewSet):
    """
    list:
        开发平台渠道数据列表,分页,过滤,搜索,排序
    create:
        开发平台渠道数据新增
    retrieve:
        开发平台渠道数据详情页
    update:
        开发平台渠道数据修改
    destroy:
        开发平台渠道数据删除
    """
    queryset = DevChannelData.objects.all()
    serializer_class = DevChannelDataSerializer  # 序列化


# 计算开发产品上架率
def calc_listing_online_rate(id):
    import math
    dp = DevProduct.objects.filter(id=id).first()
    online_qty = DevListingAccount.objects.filter(dev_p=dp,
                                                  is_online=True).count()
    all_qty = DevListingAccount.objects.filter(dev_p=dp).count()
    if all_qty:
        dp.percent = math.trunc((online_qty / all_qty) * 100)
    else:
        dp.percent = 0
    dp.save()


class DevOrderViewSet(mixins.ListModelMixin, mixins.CreateModelMixin,
                      mixins.UpdateModelMixin, mixins.DestroyModelMixin,
                      mixins.RetrieveModelMixin, viewsets.GenericViewSet):
    """
    list:
        开发产品订单列表,分页,过滤,搜索,排序
    create:
        开发产品订单新增
    retrieve:
        开发产品订单详情页
    update:
        开发产品订单修改
    destroy:
        开发产品订单删除
    """
    queryset = DevOrder.objects.all()
    serializer_class = DevOrderSerializer  # 序列化
    pagination_class = DefaultPagination  # 分页

    filter_backends = (DjangoFilterBackend, filters.SearchFilter,
                       filters.OrderingFilter)  # 过滤,搜索,排序
    filterset_fields = {
        'order_time': ['gte', 'lte', 'exact', 'gt', 'lt'],
        'order_status': ['exact'],
        'platform': ['exact'],
        'site': ['exact', 'in'],
        'account_name': ['exact'],
        'item_id': ['exact'],
        'is_ad': ['exact'],
        'is_combined': ['exact'],
        'is_settled': ['exact'],
        'is_resent': ['exact'],
        'dev_p_id': ['exact'],
    }
    search_fields = ('sku', 'cn_name', 'item_id', 'order_number')  # 配置搜索字段
    ordering_fields = ('order_time', 'total_price')  # 配置排序字段

    # test
    @action(methods=['get'], detail=False, url_path='test')
    def test(self, request):
        from devproduct import tasks
        tasks.calc_product_sales.delay()

        return Response('OK,', status=status.HTTP_200_OK)

    # 订单上传
    @action(methods=['post'], detail=False, url_path='bulk_upload')
    def bulk_upload(self, request):
        import warnings
        warnings.filterwarnings('ignore')

        data = request.data
        wb = openpyxl.load_workbook(data['excel'])
        sheet = wb['Sheet1']

        if sheet.max_row <= 1:
            return Response({
                'msg': '表格不能为空',
                'status': 'error'
            },
                            status=status.HTTP_202_ACCEPTED)
        if sheet['A1'].value != '订单编号':
            return Response({
                'msg': '表格有误，请下载最新模板',
                'status': 'error'
            },
                            status=status.HTTP_202_ACCEPTED)

        c_num = 0  # 空单号标记码
        for cell_row in list(sheet)[1:]:
            order_number = cell_row[0].value
            account_name = cell_row[1].value
            platform = cell_row[2].value
            if platform == 'ebay':
                platform = 'eBay'
            site = cell_row[3].value
            if site == 'GB':
                site = 'UK'
            currency = cell_row[4].value
            sku = cell_row[5].value
            qty = cell_row[6].value
            item_price = cell_row[7].value
            postage_price = cell_row[8].value
            total_price = cell_row[9].value
            postage = cell_row[10].value
            profit = cell_row[11].value
            profit_rate = cell_row[12].value
            if not type(profit_rate) in [float, int]:
                if profit_rate:
                    profit_rate = float(profit_rate.replace('%', '')) / 100
                else:
                    profit_rate = 0
            ad_fee = cell_row[13].value
            item_id = cell_row[14].value
            order_time = cell_row[15].value
            is_resent = False if cell_row[16].value == '否' else True
            ex_rate = cell_row[17].value

            is_exist = DevOrder.objects.filter(
                order_number=order_number).count()
            # 订单号已存在情况
            if is_exist:
                # 合并订单的条数
                if order_number:
                    c_num = is_exist - 1
                if postage:
                    if is_exist == 1:
                        od = DevOrder.objects.filter(
                            order_number=order_number).first()
                        # 未结算订单
                        if not od.is_settled:
                            od.postage = postage
                            od.profit = profit
                            od.profit_rate = profit_rate
                            od.is_settled = True
                            od.save()
                    if is_exist > 1:
                        # 合并订单情况
                        od_list = DevOrder.objects.filter(
                            order_number=order_number)
                        for i in od_list:
                            # 未结算订单
                            if not i.is_settled:
                                i.postage = postage
                                i.profit = profit
                                i.profit_rate = profit_rate
                                i.is_settled = True
                                i.save()

                continue

            # 判断该无单号行是新订单还是重复订单
            if not order_number:
                if c_num:
                    c_num -= 1
                    continue
            if order_number and sku:
                dp = DevProduct.objects.filter(sku=sku).first()
                if dp:
                    # 创建订单
                    od = DevOrder()
                    od.dev_p_id = dp.id
                    od.unit_cost = dp.unit_cost
                    od.sku = dp.sku
                    od.cn_name = dp.cn_name
                    od.image = dp.image
                    od.order_number = order_number
                    od.platform = platform
                    od.site = site
                    od.account_name = account_name
                    od.currency = currency
                    od.qty = qty
                    od.item_price = item_price
                    od.item_id = item_id
                    od.postage_price = postage_price
                    od.total_price = total_price
                    od.postage = postage
                    od.is_settled = True if postage else False  # 有发货运费则视为已结算
                    od.profit = profit if postage else 0
                    od.profit_rate = profit_rate if postage else 0
                    od.ad_fee = ad_fee
                    od.is_ad = True if ad_fee else False
                    od.order_time = order_time
                    od.is_resent = is_resent
                    od.ex_rate = ex_rate
                    od.save()
                    # 账号销量记录
                    # 不计算重发订单
                    if not od.is_resent:
                        dla = DevListingAccount.objects.filter(
                            dev_p=dp, account_name=account_name).first()
                        if dla:
                            dla.total_sold += qty
                            dla.save()
            elif not order_number and sku:
                # 合并订单产品
                dp = DevProduct.objects.filter(sku=sku).first()
                if dp:
                    last_order = DevOrder.objects.order_by('id').last()
                    # 创建订单
                    od = DevOrder()
                    od.dev_p_id = dp.id
                    od.unit_cost = dp.unit_cost
                    od.sku = dp.sku
                    od.cn_name = dp.cn_name
                    od.image = dp.image
                    od.order_number = last_order.order_number
                    od.platform = last_order.platform
                    od.site = last_order.site
                    od.account_name = last_order.account_name
                    od.currency = last_order.currency
                    od.qty = qty
                    od.item_price = item_price
                    od.item_id = item_id
                    od.postage_price = last_order.postage_price
                    od.total_price = last_order.total_price
                    od.postage = last_order.postage
                    od.is_settled = last_order.is_settled
                    od.profit = last_order.profit
                    od.profit_rate = last_order.profit_rate
                    od.ad_fee = last_order.ad_fee
                    od.is_ad = last_order.is_ad
                    od.order_time = last_order.order_time
                    od.is_resent = last_order.is_resent
                    od.is_combined = True  # 合并订单
                    od.ex_rate = last_order.ex_rate
                    od.save()
                    last_order.is_combined = True
                    last_order.save()
                    # 账号销量记录
                    # 不计算重发订单
                    if not od.is_resent:
                        dla = DevListingAccount.objects.filter(
                            dev_p=dp, account_name=account_name).first()
                        if dla:
                            dla.total_sold += qty
                            dla.save()
            else:
                continue

        return Response({
            'msg': '成功上传!',
            'status': 'success'
        },
                        status=status.HTTP_200_OK)

    # 订单上传2
    @action(methods=['post'], detail=False, url_path='bulk_upload2')
    def bulk_upload2(self, request):
        data = request.data
        # 保存上传excel文件
        path = 'media/upload_file/dev_order_excel.xlsx'
        excel = data['excel']
        content = excel.chunks()
        with open(path, 'wb') as f:
            for i in content:
                f.write(i)

        # 创建上传通知
        file_upload = FileUploadNotify()
        file_upload.shop = None
        file_upload.user_id = request.user.id
        file_upload.upload_type = 'ORDER'
        file_upload.upload_status = 'LOADING'
        file_upload.desc = '订单正在上传中...'
        file_upload.save()

        # 处理订单
        tasks.upload_dev_order.delay(file_upload.id)

        # 计算产品销量
        tasks.calc_product_sales.delay()

        # 创建操作日志
        log = MLOperateLog()
        log.op_module = 'DEVP'
        log.op_type = 'CREATE'
        log.target_type = 'ORDER'
        log.desc = '开发产品销售订单导入'
        log.user = request.user
        log.save()

        return Response({
            'msg': '文件已上传，后台处理中，稍微刷新查看结果',
            'status': 'success'
        },
                        status=status.HTTP_200_OK)
