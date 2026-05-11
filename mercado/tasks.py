from __future__ import absolute_import
from celery import shared_task
import requests
import random
import time
import json
import urllib
import hashlib
import os, re
from datetime import datetime, timedelta
from django.utils import dateparse
from bs4 import BeautifulSoup
from django.db.models import Sum, Avg, Q
from django.db import transaction
from fpdf import FPDF
from io import BytesIO
from barcode import Code128
from barcode.writer import ImageWriter

from casedance.settings import MEDIA_ROOT, BASE_DIR, BASE_URL, MEDIA_URL
from mercado.models import ApiSetting, Listing, Seller, ListingTrack, Categories, TransApiSetting, SellerTrack, Shop, \
    MLOrder, ShopStock, ShopReport, TransStock, Ship, ShipDetail, MLProduct, CarrierTrack, ExRate, StockLog, \
    FileUploadNotify, ShipAttachment, ShipBox, MLOperateLog, GeneralSettings, FBMWarehouse
from setting.models import TaskLog

user_agent_list = [
    "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.1 (KHTML, like Gecko) Chrome/22.0.1207.1 Safari/537.1",
    "Mozilla/5.0 (X11; CrOS i686 2268.111.0) AppleWebKit/536.11 (KHTML, like Gecko) Chrome/20.0.1132.57 Safari/536.11",
    "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/536.6 (KHTML, like Gecko) Chrome/20.0.1092.0 Safari/536.6",
    "Mozilla/5.0 (Windows NT 6.2) AppleWebKit/536.6 (KHTML, like Gecko) Chrome/20.0.1090.0 Safari/536.6",
    "Mozilla/5.0 (Windows NT 6.2; WOW64) AppleWebKit/537.1 (KHTML, like Gecko) Chrome/19.77.34.5 Safari/537.1",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/536.5 (KHTML, like Gecko) Chrome/19.0.1084.9 Safari/536.5",
    "Mozilla/5.0 (Windows NT 6.0) AppleWebKit/536.5 (KHTML, like Gecko) Chrome/19.0.1084.36 Safari/536.5",
    "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1063.0 Safari/536.3",
    "Mozilla/5.0 (Windows NT 5.1) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1063.0 Safari/536.3",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_8_0) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1063.0 Safari/536.3",
    "Mozilla/5.0 (Windows NT 6.2) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1062.0 Safari/536.3",
    "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1062.0 Safari/536.3",
    "Mozilla/5.0 (Windows NT 6.2) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1061.1 Safari/536.3",
    "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1061.1 Safari/536.3",
    "Mozilla/5.0 (Windows NT 6.1) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1061.1 Safari/536.3",
    "Mozilla/5.0 (Windows NT 6.2) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1061.0 Safari/536.3",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/535.24 (KHTML, like Gecko) Chrome/19.0.1055.1 Safari/535.24",
    "Mozilla/5.0 (Windows NT 6.2; WOW64) AppleWebKit/535.24 (KHTML, like Gecko) Chrome/19.0.1055.1 Safari/535.24",
    "Mozilla/5.0 (iPhone; CPU iPhone OS 7_1_2 like Mac OS X) App leWebKit/537.51.2 (KHTML, like Gecko) Version/7.0 Mobile/11D257 Safari/9537.53"
]
UA = random.choice(user_agent_list)  # 获取随机的User_Agent
main_headers = {
    'Accept-Encoding': 'gzip, deflate, sdch',
    'Accept-Language': 'en-US,en;q=0.8',
    'Upgrade-Insecure-Requests': '1',
    'User-Agent': UA,
    'Accept':
    'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
    'Cache-Control': 'max-age=0',
}
# 商品api获取字段
attributes = 'id,site_id,title,thumbnail,permalink,price,currency_id,start_time,status,health,shipping,tags,seller_id'


# 添加跟踪商品
@shared_task
def create_listing(item_id):
    is_exist = Listing.objects.filter(item_id=item_id).count()
    if is_exist:
        return 'exist'
    at = ApiSetting.objects.all().first()
    token = at.access_token
    headers = {
        'Authorization': 'Bearer ' + token,
    }

    url = 'https://api.mercadolibre.com/items?ids=' + item_id + '&attributes=' + attributes
    resp = requests.get(url, headers=headers)
    if resp.status_code == 200:
        data = resp.json()[0]['body']
        listing = Listing()
        listing.item_id = data['id']
        listing.site_id = data['site_id']
        listing.title = data['title']
        listing.image = data['thumbnail']
        listing.link = data['permalink']
        listing.price = data['price']
        listing.currency = data['currency_id']
        listing.start_date = dateparse.parse_datetime(data['start_time'])
        listing.listing_status = data['status']
        listing.health = data['health']
        listing.ship_type = data['shipping']['logistic_type']
        listing.is_free_shipping = data['shipping']['free_shipping']
        listing.is_cbt = 'cbt_item' in data['tags']
        listing.seller_id = data['seller_id']
        listing.update_time = datetime.now()
        listing.save()

        ss = Seller.objects.filter(seller_id=data['seller_id'])
        if not ss:
            # 获取卖家信息
            seller_url = 'https://api.mercadolibre.com/users/' + str(
                data['seller_id'])
            resp2 = requests.get(seller_url, headers=headers)
            if resp2.status_code == 200:
                seller_data = resp2.json()
                seller = Seller()
                seller.seller_id = seller_data['id']
                seller.nickname = seller_data['nickname']
                seller.registration_date = dateparse.parse_datetime(
                    seller_data['registration_date'])
                seller.site_id = seller_data['site_id']
                seller.link = seller_data['permalink']
                seller.level_id = seller_data['seller_reputation']['level_id']
                seller.total = seller_data['seller_reputation'][
                    'transactions']['total']
                seller.canceled = seller_data['seller_reputation'][
                    'transactions']['canceled']
                seller.negative = seller_data['seller_reputation'][
                    'transactions']['ratings']['negative']
                seller.neutral = seller_data['seller_reputation'][
                    'transactions']['ratings']['neutral']
                seller.positive = seller_data['seller_reputation'][
                    'transactions']['ratings']['positive']
                seller.save()
                listing.seller_name = seller_data['nickname']
                listing.save()
        else:
            seller = Seller.objects.filter(seller_id=data['seller_id']).first()
            listing.seller_name = seller.nickname
            listing.save()

        # 抓取listing总销量
        resp3 = requests.get(data['permalink'], headers=main_headers)
        if resp3:
            resp3.encoding = "utf-8"
            page = BeautifulSoup(resp3.text, 'html.parser')
            sold = page.find('span', class_='ui-pdp-subtitle')
            import re
            sold_qty = re.findall("\d+", sold.text)
            if sold_qty:
                listing.total_sold = int(sold_qty[0])
                listing.save()

        # 获取评价信息
        review_url = 'https://api.mercadolibre.com/reviews/item/' + item_id
        resp4 = requests.get(review_url, headers=headers)
        if resp4.status_code == 200:
            review_data = resp4.json()
            listing.reviews = review_data['paging']['total']
            listing.rating_average = review_data['rating_average']
            listing.save()
    return 'OK'


# 跟踪商品
@shared_task
def track_listing():
    date = datetime.now().date()
    queryset = Listing.objects.filter(update_time__date__lt=date)
    attributes2 = 'id,title,thumbnail,price,status,health,shipping'
    api_url = 'https://api.mercadolibre.com/items?ids='
    url = ''
    n = 0
    for i in queryset:
        # 拼接item_id
        if n == 0:
            url = api_url + i.item_id
        else:
            url = url + ',' + i.item_id
        n += 1

        if n == 20:
            req_url = url + '&attributes=' + attributes2
            url = ''
            n = 0
            get_listing_info(req_url)

    if n != 0:
        req_url = url + '&attributes=' + attributes2
        get_listing_info(req_url)

    # 记录执行日志
    task_log = TaskLog()
    task_log.task_type = 9
    task_log.note = '美客多listing跟踪'
    task_log.save()

    return 'OK'


# 获取链接追踪信息
def get_listing_info(url):
    at = ApiSetting.objects.all().first()
    token = at.access_token
    headers = {
        'Authorization': 'Bearer ' + token,
    }
    resp = requests.get(url, headers=headers)
    if resp.status_code == 200:
        data = resp.json()
        add_list = []
        for i in data:
            listing = Listing.objects.filter(item_id=i['body']['id']).first()
            listing.title = i['body']['title']
            listing.image = i['body']['thumbnail']
            listing.price = i['body']['price']
            listing.listing_status = i['body']['status']
            listing.health = i['body']['health']
            listing.ship_type = i['body']['shipping']['logistic_type']
            listing.update_time = datetime.now()
            listing.save()

            # 获取评价信息
            review_url = 'https://api.mercadolibre.com/reviews/item/' + i[
                'body']['id']
            resp4 = requests.get(review_url, headers=headers)
            if resp4.status_code == 200:
                review_data = resp4.json()
                listing.reviews = review_data['paging']['total']
                listing.rating_average = review_data['rating_average']
                listing.save()

            # 爬取销量
            resp2 = requests.get(listing.link, headers=main_headers)
            if resp2:
                resp2.encoding = "utf-8"
                page = BeautifulSoup(resp2.text, 'html.parser')
                sold = page.find('span', class_='ui-pdp-subtitle')
                import re
                sold_qty = re.findall("\d+", sold.text)
                if sold_qty:
                    listing.total_sold = int(sold_qty[0])
                    listing.save()

            # 补充计算昨天当天销量
            lt = ListingTrack.objects.filter(
                listing=listing,
                create_time__date=datetime.now().date() -
                timedelta(days=1)).first()
            if lt:
                lt.today_sold = listing.total_sold - lt.total_sold
                lt.save()

            is_exits = ListingTrack.objects.filter(
                listing=listing,
                create_time__date=datetime.now().date()).count()
            if not is_exits:
                add_list.append(
                    ListingTrack(listing=listing,
                                 currency=listing.currency,
                                 price=listing.price,
                                 total_sold=listing.total_sold,
                                 reviews=listing.reviews,
                                 rating_average=listing.rating_average,
                                 health=listing.health))
            else:
                listing_track = ListingTrack.objects.filter(
                    listing=listing,
                    create_time__date=datetime.now().date()).first()
                today_sold = listing.total_sold - listing_track.total_sold
                listing_track.today_sold = today_sold
                listing_track.reviews = listing.reviews
                listing_track.price = listing.price
                listing_track.rating_average = listing.rating_average
                listing_track.health = listing.health
                listing_track.save()
            time.sleep(0.5)
        ListingTrack.objects.bulk_create(add_list)


# 更新一级类目信息
@shared_task
def update_categories(site_id):
    at = ApiSetting.objects.all().first()
    token = at.access_token
    headers = {
        'Authorization': 'Bearer ' + token,
    }

    url = 'https://api.mercadolibre.com/sites/' + site_id + '/categories'
    resp = requests.get(url, headers=headers)
    if resp.status_code == 200:
        data = resp.json()
        for i in data:
            categories = Categories.objects.filter(categ_id=i['id'],
                                                   site_id=site_id).first()
            # 如果目录存在，则检查需不需要更新
            if categories:
                if categories.name != i['name']:
                    categories.name = i['name']
                    categories.update_time = datetime.now()
                    categories.save()
            else:
                cate_g = Categories()
                cate_g.categ_id = i['id']
                cate_g.father_id = '0'
                cate_g.site_id = site_id
                cate_g.name = i['name']
                t_name = translate(i['name'])
                if t_name:
                    cate_g.t_name = t_name
                cate_g.path_from_root = i['name']
                cate_g.has_children = True
                cate_g.update_time = datetime.now()
                cate_g.save()


# 翻译
def translate(q):
    ta = TransApiSetting.objects.all().first()
    appid = ta.appid  # 填写你的appid
    secretKey = ta.secretKey  # 填写你的密钥
    myurl = 'https://fanyi-api.baidu.com/api/trans/vip/translate'
    fromLang = 'auto'  # 原文语种
    toLang = 'zh'  # 译文语种
    salt = random.randint(32768, 65536)
    sign = appid + q + str(salt) + secretKey
    sign = hashlib.md5(sign.encode()).hexdigest()
    myurl = myurl + '?appid=' + appid + '&q=' + urllib.parse.quote(
        q) + '&from=' + fromLang + '&to=' + toLang + '&salt=' + str(
            salt) + '&sign=' + sign
    resp = requests.get(myurl)
    time.sleep(0.01)
    if resp.status_code == 200:
        resp.encoding = "utf-8"
        result = json.loads(resp.text)
        return result['trans_result'][0]['dst']
    else:
        return ''


# 更新卖家详细信息
def create_or_update_seller(site_id, seller_id):
    at = ApiSetting.objects.all().first()
    token = at.access_token
    headers = {
        'Authorization': 'Bearer ' + token,
    }

    url = 'https://api.mercadolibre.com/sites/' + site_id + '/search?seller_id=' + seller_id
    resp = requests.get(url, headers=headers)
    if resp.status_code == 200:
        data = resp.json()
        seller = Seller.objects.filter(seller_id=seller_id).first()
        if seller:
            seller.level_id = data['seller']['seller_reputation']['level_id']
            seller.total = data['seller']['seller_reputation']['transactions'][
                'total']
            seller.canceled = data['seller']['seller_reputation'][
                'transactions']['canceled']
            seller.negative = data['seller']['seller_reputation'][
                'transactions']['ratings']['negative']
            seller.neutral = data['seller']['seller_reputation'][
                'transactions']['ratings']['neutral']
            seller.positive = data['seller']['seller_reputation'][
                'transactions']['ratings']['positive']
            seller.link = data['seller']['permalink']
            seller.sold_60d = data['seller']['seller_reputation']['metrics'][
                'sales']['completed']
            seller.total_items = data['paging']['total']
            seller.update_time = datetime.now()
            seller.save()
        else:
            seller = Seller()
            seller.seller_id = seller_id
            seller.site_id = site_id
            seller.nickname = data['seller']['nickname']
            seller.registration_date = dateparse.parse_datetime(
                data['seller']['registration_date'])
            seller.level_id = data['seller']['seller_reputation']['level_id']
            seller.total = data['seller']['seller_reputation']['transactions'][
                'total']
            seller.canceled = data['seller']['seller_reputation'][
                'transactions']['canceled']
            seller.negative = data['seller']['seller_reputation'][
                'transactions']['ratings']['negative']
            seller.neutral = data['seller']['seller_reputation'][
                'transactions']['ratings']['neutral']
            seller.positive = data['seller']['seller_reputation'][
                'transactions']['ratings']['positive']
            seller.link = data['seller']['permalink']
            seller.sold_60d = data['seller']['seller_reputation']['metrics'][
                'sales']['completed']
            seller.total_items = data['paging']['total']
            seller.update_time = datetime.now()
            seller.save()


# 跟踪卖家
@shared_task
def track_seller():
    today = datetime.now().date()
    # 跟踪所有卖家
    queryset = Seller.objects.all()
    at = ApiSetting.objects.all().first()
    token = at.access_token
    headers = {
        'Authorization': 'Bearer ' + token,
    }
    url = 'https://api.mercadolibre.com/sites/'
    for i in queryset:
        resp = requests.get(url + i.site_id + '/search?seller_id=' +
                            i.seller_id,
                            headers=headers)
        if resp.status_code == 200:
            data = resp.json()
            # 更新卖家信息
            i.level_id = data['seller']['seller_reputation']['level_id']
            i.total = data['seller']['seller_reputation']['transactions'][
                'total']
            i.canceled = data['seller']['seller_reputation']['transactions'][
                'canceled']
            i.negative = data['seller']['seller_reputation']['transactions'][
                'ratings']['negative']
            i.neutral = data['seller']['seller_reputation']['transactions'][
                'ratings']['neutral']
            i.positive = data['seller']['seller_reputation']['transactions'][
                'ratings']['positive']
            i.link = data['seller']['permalink']
            i.sold_60d = data['seller']['seller_reputation']['metrics'][
                'sales']['completed']
            i.total_items = data['paging']['total']
            i.update_time = datetime.now()
            i.save()

            # 补充计算昨天当天销量
            st = SellerTrack.objects.filter(seller=i,
                                            create_time__date=today -
                                            timedelta(days=1)).first()
            if st:
                st.today_sold = i.total - st.total
                st.save()

            seller_track = SellerTrack.objects.filter(
                seller=i, create_time__date=today).first()
            if seller_track:
                seller_track.today_sold = i.total - seller_track.total
                seller_track.negative = i.negative
                seller_track.neutral = i.neutral
                seller_track.positive = i.positive
                seller_track.total_items = i.total_items
                seller_track.save()
            else:
                seller_track = SellerTrack()
                seller_track.seller = i
                seller_track.total = i.total
                seller_track.negative = i.negative
                seller_track.neutral = i.neutral
                seller_track.positive = i.positive
                seller_track.total_items = i.total_items
                seller_track.save()

        time.sleep(0.5)

    # 记录执行日志
    task_log = TaskLog()
    task_log.task_type = 1
    task_log.note = '美客多seller跟踪'
    task_log.save()

    return 'OK'


# 跟踪运单物流运输-旧
@shared_task
def ship_tracking_old(num):
    url = 'https://client.morelink56.com/baiduroute/get_routeinfo'
    carrier = 'SHENGDE'

    resp = requests.post(url, {'num': num})
    if resp.status_code == 200:
        data = resp.json()
        if data['message'] == 'SUCCESS':
            if len(data['data']):
                for i in data['data']:
                    is_exist = CarrierTrack.objects.filter(
                        carrier_name=carrier,
                        carrier_number=num,
                        context=i['context']).first()
                    if not is_exist:
                        ct = CarrierTrack()
                        ct.carrier_name = carrier
                        ct.carrier_number = num
                        ct.context = i['context']
                        ct.location = i['location']
                        if i['time']:
                            ct.time = datetime.strptime(
                                i['time'], '%Y-%m-%d %H:%M:%S')
                        if i['optime']:
                            ct.optime = datetime.strptime(
                                i['optime'], '%Y-%m-%d %H:%M:%S')
                        ct.save()
                    else:
                        is_exist.create_time = datetime.now()
                        is_exist.save()
                return 'SUCCESS'
        else:
            return data['message']
    else:
        return '网络异常'


# 跟踪运单物流运输
@shared_task
def ship_tracking(num):
    # ================= 1. 读取配置文件获取 token/cookies =================
    try:
        c_path = os.path.join(BASE_DIR, "site_config.json")
        with open(c_path, "r", encoding="utf-8") as f:
            config = json.load(f)
        sd_cookies = config.get('sd_cookies', '')
    except:
        return '读取配置失败'

    # ================= 2. 新接口地址 =================
    url = 'http://api.more56.com/api/v1/edi/open_bigwaybill_order_client/routeinfo'

    # ================= 3. 请求头（带认证） =================
    headers = {'Authorization': sd_cookies, 'Content-Type': 'application/json'}

    # ================= 4. 请求参数 =================
    payload = {
        "param": num  # 跟踪号
    }

    carrier = 'SHENGDE'

    try:
        # 发送请求
        resp = requests.post(url, json=payload, headers=headers, timeout=10)

        if resp.status_code != 200:
            return '网络异常'

        # 解析返回
        result = resp.json()
        if not result.get('success', False):
            return result.get('message', '获取轨迹失败')

        # 获取轨迹列表
        data = result.get('data', {})
        track_list = data.get('list', [])

        if not track_list:
            return '暂无轨迹信息'

        # 遍历保存（只存中文 + 时间）
        for item in track_list:
            # 中文轨迹内容 + 时间
            context = item.get('route_wlzz', '')
            track_time = item.get('route_date', '')

            if not context or not track_time:
                continue

            # 去重：同一个运单 + 同一个内容不重复保存
            is_exist = CarrierTrack.objects.filter(carrier_name=carrier,
                                                   carrier_number=num,
                                                   context=context).first()

            if is_exist:
                is_exist.create_time = datetime.now()
                is_exist.save()
                continue

            # 保存新轨迹
            ct = CarrierTrack()
            ct.carrier_name = carrier
            ct.carrier_number = num
            ct.context = context  # 中文轨迹
            ct.location = ''  # 新接口没有 location，留空

            # 时间格式化
            try:
                ct.time = datetime.strptime(track_time, '%Y-%m-%d %H:%M:%S')
            except:
                ct.time = None

            ct.save()

        return 'SUCCESS'

    except requests.exceptions.Timeout:
        return '请求超时'
    except Exception as e:
        return f'解析异常：{str(e)}'


# 批量更新运单物流运输跟踪
@shared_task
def bulk_ship_tracking():
    queryset = Ship.objects.filter(
        send_from='CN',
        carrier='盛德物流').filter(Q(s_status='SHIPPED') | Q(s_status='BOOKED'))
    for i in queryset:
        if i.s_number:
            ship_tracking(i.s_number)
            time.sleep(0.2)
    # 记录执行日志
    task_log = TaskLog()
    task_log.task_type = 14
    task_log.note = '美客多物流跟踪信息更新'
    task_log.save()
    return 'SUCCESS'


# 同步物流收货的尺寸重量
def update_sd_ship_size():
    # 1. 读取认证配置
    try:
        c_path = os.path.join(BASE_DIR, "site_config.json")
        with open(c_path, "r", encoding="utf-8") as f:
            config = json.load(f)
        sd_cookies = config.get('sd_cookies', '')
    except Exception as e:
        return f"读取配置失败: {str(e)}"

    # 2. 接口地址
    url = "http://api.more56.com/api/v1/edi/open_bigwaybill_order_client/ordersize"

    headers = {'Authorization': sd_cookies, 'Content-Type': 'application/json'}

    # 3. 查询需要更新尺寸重量的运单
    shipset = Ship.objects.filter(carrier='盛德物流',
                                  carrier_order_status='ACCEPTED',
                                  carrier_rec_check='UNCHECK').filter(
                                      Q(s_status='PREPARING')
                                      | Q(s_status='SHIPPED')
                                      | Q(s_status='BOOKED'))

    if not shipset.exists():
        return "暂无需要更新重量尺寸的运单"

    # 4. 循环逐个查询（接口不支持批量）
    for ship in shipset:
        sono = ship.s_number
        if not sono:
            continue

        # 接口只支持单个 sono
        payload = {"ids": [sono]}

        try:
            resp = requests.post(url,
                                 json=payload,
                                 headers=headers,
                                 timeout=15)
            if resp.status_code != 200:
                continue

            result = resp.json()
            if not result.get('success', False):
                continue

            # 获取返回列表
            data = result.get('data', {})
            size_list = data.get('list', [])
            if not size_list:
                continue

            # 取第一条
            item = size_list[0]
            total_qty = item.get('total_qty')
            total_kg = item.get('total_kg')
            total_cbm = item.get('total_cbm')
            cost_kg = item.get('cost_kg')

            # 只有有数据才保存
            if all([total_qty, total_kg, total_cbm, cost_kg]):
                with transaction.atomic():
                    ship.carrier_GoodsNum = total_qty  # 总数量
                    ship.carrier_ckweight = total_kg  # 重量
                    ship.carrier_ckcbm = total_cbm  # 体积
                    ship.carrier_ckvolume = cost_kg  # 计费重
                    # ============= ✅ 开始比对逻辑 =============
                    # 本地数量
                    local_qty = ship.total_box or 0
                    # 物流返回数量
                    carrier_qty = ship.carrier_GoodsNum or 0

                    # 本地重量
                    local_weight = round(float(ship.weight or 0), 2)  # 保留2位

                    # 物流返回计费重
                    carrier_weight = round(float(ship.carrier_ckvolume or 0),
                                           2)  # 保留2位

                    # 数量必须完全相等 | 重量差异 ≤ 1
                    qty_ok = (local_qty == carrier_qty)
                    weight_ok = (abs(local_weight - carrier_weight)
                                 <= 1.0)  # 安全判断

                    # 标记状态
                    if qty_ok and weight_ok:
                        ship.carrier_rec_check = 'CHECKED'
                    else:
                        ship.carrier_rec_check = 'ERROR'
                    ship.save()

                    # 记录日志
                    log = MLOperateLog()
                    log.op_module = 'SHIP'
                    log.op_type = 'EDIT'
                    log.target_type = 'SHIP'
                    log.target_id = ship.id
                    log.desc = f'同步收货尺寸/重量 | 状态：{ship.carrier_rec_check}'
                    log.save()

        except Exception as e:
            continue

    return "SUCCESS"


# sd物流自动同步/跟踪
@shared_task
def sd_auto_sync():
    # 1.自动检查备货中已打包运单
    shipset = Ship.objects.filter(carrier='盛德物流', s_status='PREPARING').filter(
        Q(carrier_order_status='')
        | Q(carrier_order_status=None))
    for ship in shipset:
        sd_set = ShipDetail.objects.filter(ship=ship)
        is_packed = True  # 产品打包状态
        for i in sd_set:
            if not i.box_number:
                is_packed = False
        if is_packed:
            ship.carrier_order_status = 'PACKED'
            ship.save()
            # 记录日志
            log = MLOperateLog()
            log.op_module = 'SHIP'
            log.op_type = 'EDIT'
            log.target_type = 'SHIP'
            log.target_id = ship.id
            log.desc = '检测到运单已装箱'
            log.save()

    # 2.检查资料是否齐全
    shipset = Ship.objects.filter(carrier='盛德物流').filter(
        carrier_order_status='PACKED').filter(s_status='PREPARING')
    for ship in shipset:
        sd_set = ShipDetail.objects.filter(ship=ship)
        is_declare = True  # 产品申报状态
        is_file = False  # 箱唛附件
        for i in sd_set:
            if not i.custom_code or not i.en_name or not i.brand or not i.declared_value or not i.cn_material:
                is_declare = False

        sa_set = ShipAttachment.objects.filter(ship=ship, a_type='BOX_LABEL')
        for i in sa_set:
            extension = i.name.split(".")[-1]
            if extension == 'pdf':
                is_file = True

        if is_declare and is_file:
            ship.carrier_order_status = 'UNSUBMIT'
            ship.save()
            # 记录日志
            log = MLOperateLog()
            log.op_module = 'SHIP'
            log.op_type = 'EDIT'
            log.target_type = 'SHIP'
            log.target_id = ship.id
            log.desc = '检测到运单资料已上传'
            log.save()

    # 3，自动预报交运
    shipset = Ship.objects.filter(carrier='盛德物流').filter(
        carrier_order_status='UNSUBMIT').filter(s_status='PREPARING')
    today_time = datetime.now().strftime('%Y-%m-%d 18:00')  # 默认今天日期
    for ship in shipset:
        fbm = FBMWarehouse.objects.filter(w_code=ship.fbm_warehouse).first()
        shop = Shop.objects.filter(name=ship.shop).first()
        ship_data = {
            'd_code': ship.fbm_warehouse,  # fbm仓库编号
            'sh_addr': fbm.address,  # fbm仓库地址
            'sh_zip_code': fbm.zip if fbm.zip else ' ',  # fbm仓库邮编
            'delivery_time': today_time,  # 预约日期
            'transport_type': '空运',  # 货运方式
            'ck_name': '东莞凤岗仓',  # 送货仓库名称
            'refid':
            'Local' if shop.shop_type == 'LOCAL' else shop.seller_id,  #店铺账号id
            'fbx_no': ship.envio_number,  # 运单编号
            'product': '手机壳',  # 品名
            'ProductNature': '普货',  # 产品性质
        }

        info = sd_place_order_api(ship.id, ship_data)
        # 记录日志
        log = MLOperateLog()
        log.op_module = 'SHIP'
        log.op_type = 'EDIT'
        log.target_type = 'SHIP'
        log.target_id = ship.id
        if info['status'] == 'success':
            log.desc = '自动交运成功'
        else:
            log.desc = '自动交运失败:' + info['msg']
        log.save()

    # 4. 查询受理情况
    query_sd_order_status()
    # 5. 同步收货重量/尺寸
    update_sd_ship_size()
    # 批量更新跟踪
    bulk_ship_tracking()

    # 记录执行日志
    task_log = TaskLog()
    task_log.task_type = 4
    task_log.note = '自动检查sd物流交运/信息同步'
    task_log.save()

    return 'SUCCESS'


# 计算mercado产品销量
@shared_task
def calc_product_sales():
    shops = Shop.objects.filter(warehouse_type='FBM', is_active=True)
    for s in shops:
        is_exist = MLOrder.objects.filter(shop=s).count()
        if not is_exist:
            continue
        shop_stocks = ShopStock.objects.filter(shop=s, is_active=True)
        for st in shop_stocks:
            # 7天销量
            day7 = datetime.now().date() - timedelta(days=7)
            sum_day7 = MLOrder.objects.filter(shop=s,
                                              sku=st.sku,
                                              item_id=st.item_id,
                                              order_time__gte=day7).aggregate(
                                                  Sum('qty'))
            day7_sold = sum_day7['qty__sum']

            # 15天销量
            day15 = datetime.now().date() - timedelta(days=15)
            sum_day15 = MLOrder.objects.filter(
                shop=s, sku=st.sku, item_id=st.item_id,
                order_time__gte=day15).aggregate(Sum('qty'))
            day15_sold = sum_day15['qty__sum']

            # 30天销量
            day30 = datetime.now().date() - timedelta(days=30)
            sum_day30 = MLOrder.objects.filter(
                shop=s, sku=st.sku, item_id=st.item_id,
                order_time__gte=day30).aggregate(Sum('qty'))
            day30_sold = sum_day30['qty__sum']

            # 累计销量
            sum_total = MLOrder.objects.filter(
                shop=s,
                sku=st.sku,
                item_id=st.item_id,
            ).aggregate(Sum('qty'))
            total_sold = sum_total['qty__sum']

            # 累计利润
            sum_profit = MLOrder.objects.filter(
                shop=s,
                sku=st.sku,
                item_id=st.item_id,
            ).aggregate(Sum('profit'))
            total_profit = sum_profit['profit__sum']

            # 平均毛利润
            avg_profit = MLOrder.objects.filter(
                shop=s,
                sku=st.sku,
                item_id=st.item_id,
            ).aggregate(Avg('profit'))
            avg_profit = avg_profit['profit__avg']

            # 平均毛利率
            a_profit_rate = MLOrder.objects.filter(
                shop=s,
                sku=st.sku,
                item_id=st.item_id,
            ).aggregate(Avg('profit_rate'))
            avg_profit_rate = a_profit_rate['profit_rate__avg']

            from django.db.models import Q
            # 退款率
            refund_count = MLOrder.objects.filter(
                shop=s,
                sku=st.sku,
                item_id=st.item_id,
            ).filter(Q(order_status='RETURN')
                     | Q(order_status='CASE')).count()
            total_count = MLOrder.objects.filter(
                shop=s,
                sku=st.sku,
                item_id=st.item_id,
            ).count()
            if total_count:
                refund_rate = refund_count / total_count
            else:
                refund_rate = 0

            st.day15_sold = day15_sold if day15_sold else 0
            st.day7_sold = day7_sold if day7_sold else 0
            st.day30_sold = day30_sold if day30_sold else 0
            st.total_sold = total_sold if total_sold else 0
            st.total_profit = total_profit if sum_profit else 0
            st.avg_profit = avg_profit if avg_profit else 0
            st.avg_profit_rate = avg_profit_rate if avg_profit_rate else 0
            st.refund_rate = refund_rate
            st.save()

    # 记录执行日志
    task_log = TaskLog()
    task_log.task_type = 11
    task_log.note = 'FBM销量计算'
    task_log.save()

    return 'OK'


# 计算店铺30天每天累积总销量
@shared_task()
def calc_shop_sale():
    # 删除订单状态过滤
    # q = Q()
    # q.connector = 'OR'
    # q.children.append(('order_status', 'FINISHED'))
    # q.children.append(('order_status', 'RETURN'))
    # q.children.append(('order_status', 'CASE'))

    shops = Shop.objects.filter(warehouse_type='FBM', is_active=True)
    for s in shops:
        add_list = []
        for i in range(30):
            date = datetime.now().date() - timedelta(days=i)
            # 删除订单状态过滤（弃用）
            # order_set = MLOrder.objects.filter(order_time__date=date, shop=s).filter(q)
            order_set = MLOrder.objects.filter(order_time__date=date, shop=s)
            qty = 0
            amount = 0.0
            profit = 0.0

            for item in order_set:
                qty += item.qty
                amount += item.price
                profit += item.profit

            sr = ShopReport.objects.filter(calc_date=date, shop=s).first()
            if sr:
                sr.qty = qty
                sr.amount = amount
                sr.profit = profit
                sr.save()
            else:
                add_list.append(
                    ShopReport(qty=qty,
                               amount=amount,
                               profit=profit,
                               shop=s,
                               calc_date=date))
        if len(add_list):
            ShopReport.objects.bulk_create(add_list)
    # 记录执行日志
    task_log = TaskLog()
    task_log.task_type = 13
    task_log.note = 'fbm店铺销量计算'
    task_log.save()

    return 'OK'


# 获取店铺额度
@shared_task()
def get_shop_quota(shop_id):
    shop = Shop.objects.filter(id=shop_id).first()

    # FBM库存统计
    queryset = ShopStock.objects.filter(
        is_active=True, qty__gt=0,
        shop__id=shop_id).exclude(p_status='OFFLINE')
    total_amount = 0
    for i in queryset:
        total_amount += (i.unit_cost + i.first_ship_cost) * i.qty

    # 中转仓库存统计
    ts = TransStock.objects.filter(listing_shop=shop.name, is_out=False)
    trans_amount = 0
    for i in ts:
        trans_amount += (i.unit_cost + i.first_ship_cost) * i.qty

    # 在途运单统计,含备货中
    ships = Ship.objects.filter(shop=shop.name).filter(
        Q(s_status='SHIPPED') | Q(s_status='BOOKED') | Q(s_status='PREPARING'))
    onway_amount = 0
    for i in ships:
        onway_amount += i.shipping_fee
        onway_amount += i.extra_fee
        onway_amount += i.products_cost

    # 在途中转运单统计,含备货中
    ship_detail = ShipDetail.objects.filter(
        Q(ship__s_status='SHIPPED') | Q(ship__s_status='BOOKED')
        | Q(ship__s_status='PREPARING')).filter(ship__target='TRANSIT')
    for i in ship_detail:
        is_shop_product = MLProduct.objects.filter(shop=shop.name,
                                                   sku=i.sku).first()
        # 如果是该店铺下的产品
        if is_shop_product:
            onway_amount += (i.unit_cost + i.avg_ship_fee) * i.qty

    used_quota = total_amount + trans_amount + onway_amount
    return used_quota


# 上传美客多订单
@shared_task()
def upload_mercado_order(shop_id, notify_id, mel_row):
    import warnings
    import re
    import openpyxl

    warnings.filterwarnings('ignore')

    try:
        data = MEDIA_ROOT + '/upload_file/order_excel_' + shop_id + '.xlsx'
        wb = openpyxl.load_workbook(data)
        sheet = wb.active

        month_dict = {
            'enero': '01',
            'febrero': '02',
            'marzo': '03',
            'abril': '04',
            'mayo': '05',
            'junio': '06',
            'julio': '07',
            'agosto': '08',
            'septiembre': '09',
            'octubre': '10',
            'noviembre': '11',
            'diciembre': '12'
        }
        # 定义所需表头
        required_headers = [
            '# de venta', 'Fecha de venta', 'Estado', 'Unidades',
            'Cargo por venta e impuestos (MXN)', 'Costos de envío (MXN)',
            'Total (MXN)', 'Venta por publicidad', 'SKU', '# de publicación',
            'Precio unitario de venta de la publicación (MXN)', 'Comprador',
            'Domicilio', 'Municipio/Alcaldía', 'Estado', 'Código postal',
            'País'
        ]

        shop = Shop.objects.filter(id=shop_id).first()
        er = ExRate.objects.filter(currency=shop.currency).first()
        ex_rate = er.value  # 获取汇率
        sp = GeneralSettings.objects.filter(
            item_name='mel_sp').first()  # 获取服务商费用

        # 模板格式检查
        # 记录表头对应的列索引
        header_index = {}
        header_row = sheet[mel_row]
        if header_row is None:
            file_upload = FileUploadNotify.objects.filter(id=notify_id).first()
            file_upload.upload_status = 'ERROR'
            file_upload.desc = '未找到表头行'
            file_upload.save()
            return 'ERROR'
        for col_index, cell in enumerate(header_row, start=0):
            header = cell.value
            if header in required_headers and header not in header_index:
                header_index[header] = col_index

        # 检查是否所有必需的表头都存在
        if not all(header in header_index for header in required_headers):
            # 修改上传通知
            file_upload = FileUploadNotify.objects.filter(id=notify_id).first()
            file_upload.upload_status = 'ERROR'
            file_upload.desc = '模板格式有误，缺少必要的表头，请检查!'
            file_upload.save()
            return 'ERROR'

        add_list = []
        for cell_row in list(sheet)[int(mel_row):]:
            if cell_row is None:
                continue
            qty = cell_row[header_index['Unidades']].value
            if not qty:
                continue
            sku = cell_row[header_index['SKU']].value
            item_id = cell_row[header_index['# de publicación']].value[3:]

            # 如果不在fmb库存中，或者所在店铺不对应，则跳出
            shop_stock = ShopStock.objects.filter(sku=sku,
                                                  item_id=item_id,
                                                  shop=shop).first()
            if not shop_stock:
                continue
            first_ship_cost = shop_stock.first_ship_cost
            if not first_ship_cost:
                first_ship_cost = 0

            order_number = cell_row[header_index['# de venta']].value

            t = cell_row[header_index['Fecha de venta']].value
            de_locate = [m.start() for m in re.finditer('de', t)]
            day = t[:de_locate[0] - 1]
            if int(day) < 10:
                day = '0' + day
            month = t[de_locate[0] + 3:de_locate[1] - 1]
            year = t[de_locate[1] + 3:de_locate[1] + 7]
            hour = t[de_locate[1] + 8:de_locate[1] + 10]
            min = t[de_locate[1] + 11:de_locate[1] + 13]
            order_time = '%s-%s-%s %s:%s:00' % (year, month_dict[month], day,
                                                hour, min)

            bj = datetime.strptime(order_time,
                                   '%Y-%m-%d %H:%M:%S') + timedelta(hours=14)
            order_time_bj = bj.strftime('%Y-%m-%d %H:%M:%S')

            price = cell_row[header_index[
                'Precio unitario de venta de la publicación (MXN)']].value if cell_row[
                    header_index[
                        'Precio unitario de venta de la publicación (MXN)']].value else 0
            postage = cell_row[
                header_index['Costos de envío (MXN)']].value if cell_row[
                    header_index['Costos de envío (MXN)']].value else 0
            fees = cell_row[header_index[
                'Cargo por venta e impuestos (MXN)']].value if cell_row[
                    header_index[
                        'Cargo por venta e impuestos (MXN)']].value else 0
            receive_fund = cell_row[
                header_index['Total (MXN)']].value if cell_row[
                    header_index['Total (MXN)']].value else 0
            is_ad = True if cell_row[
                header_index['Venta por publicidad']].value == 'Sí' else False

            buyer_name = cell_row[header_index['Comprador']].value
            buyer_address = cell_row[header_index['Domicilio']].value
            buyer_city = cell_row[header_index['Municipio/Alcaldía']].value
            buyer_state = ''  # 与订单状态表头冲突，不使用
            buyer_postcode = cell_row[header_index['Código postal']].value
            buyer_country = cell_row[header_index['País']].value

            # 计算服务商费用
            sp_fee = 0
            sp_fee_rate = 0
            if sp:
                sp_fee_rate = sp.value1
                sp_fee = price * sp_fee_rate
            # 利润计算
            profit = (
                (float(receive_fund) - sp_fee) * 0.99
            ) * ex_rate - shop_stock.unit_cost * qty - shop_stock.first_ship_cost * qty
            profit_rate = profit / (price * ex_rate)
            if profit_rate < 0:
                profit_rate = 0

            order_status = 'FINISHED'

            if cell_row[
                    header_index['Estado']].value == 'Procesando en la bodega':
                order_status = 'PROCESS'
            if cell_row[header_index[
                    'Estado']].value == 'Cancelada por el comprador':
                order_status = 'CANCEL'
            if cell_row[header_index[
                    'Estado']].value == 'Paquete cancelado por Mercado Libre':
                order_status = 'CANCEL'
            if cell_row[
                    header_index['Estado']].value == 'Devolución en camino':
                order_status = 'RETURN'
            if cell_row[header_index[
                    'Estado']].value == 'Reclamo cerrado con reembolso al comprador':
                order_status = 'CASE'
            if cell_row[header_index['Estado']].value[:8] == 'Devuelto':
                order_status = 'RETURN'
            if cell_row[header_index['Estado']].value == 'En devolución':
                order_status = 'RETURN'
            if cell_row[header_index[
                    'Estado']].value == 'Devolución finalizada con reembolso al comprador':
                order_status = 'RETURN'

            # 检查同一店铺订单编号是否存在
            ml_order = MLOrder.objects.filter(order_number=order_number,
                                              shop=shop).first()
            if not ml_order:
                add_list.append(
                    MLOrder(
                        shop=shop,
                        platform='MERCADO',
                        order_number=order_number,
                        order_status=order_status,
                        order_time=order_time,
                        order_time_bj=order_time_bj,
                        qty=qty,
                        currency=shop.currency,
                        ex_rate=ex_rate,
                        price=price,
                        fees=fees,
                        postage=postage,
                        receive_fund=receive_fund,
                        is_ad=is_ad,
                        sku=sku,
                        p_name=shop_stock.p_name,
                        item_id=item_id,
                        image=shop_stock.image,
                        unit_cost=shop_stock.unit_cost * qty,
                        first_ship_cost=first_ship_cost * qty,
                        profit=profit,
                        profit_rate=profit_rate,
                        buyer_name=buyer_name,
                        buyer_address=buyer_address,
                        buyer_city=buyer_city,
                        buyer_state=buyer_state,
                        buyer_postcode=buyer_postcode,
                        buyer_country=buyer_country,
                        sp_fee=sp_fee,
                        sp_fee_rate=sp_fee_rate,
                    ))
                shop_stock.qty -= qty
                shop_stock.save()

                # 创建库存日志
                stock_log = StockLog()
                stock_log.shop_stock = shop_stock
                stock_log.current_stock = shop_stock.qty
                stock_log.qty = qty
                stock_log.in_out = 'OUT'
                stock_log.action = 'SALE'
                stock_log.desc = '销售订单号: ' + order_number
                stock_log.user_id = 0
                stock_log.save()
                stock_log.create_time = order_time  # order_time
                stock_log.save()
            else:
                # 如果费用更新
                if ml_order.fees != fees:
                    ml_order.receive_fund = receive_fund
                    ml_order.fees = fees
                    ml_order.profit = profit
                    ml_order.profit_rate = profit_rate
                    ml_order.save()
                # 如果订单状态更新
                if ml_order.order_status != order_status:
                    ml_order.order_status = order_status
                    ml_order.receive_fund = receive_fund
                    ml_order.fees = fees
                    ml_order.profit = profit
                    ml_order.profit_rate = profit_rate
                    ml_order.save()

                    # 如果订单是取消状态，库存增加回来
                    if order_status == 'CANCEL':
                        shop_stock.qty += qty
                        shop_stock.save()

                        # 创建库存日志
                        stock_log = StockLog()
                        stock_log.shop_stock = shop_stock
                        stock_log.current_stock = shop_stock.qty
                        stock_log.qty = qty
                        stock_log.in_out = 'IN'
                        stock_log.action = 'CANCEL'
                        stock_log.desc = '取消订单入库, 订单号: ' + order_number
                        stock_log.user_id = 0
                        stock_log.save()
        if len(add_list):
            MLOrder.objects.bulk_create(add_list)

        # 修改上传通知
        file_upload = FileUploadNotify.objects.filter(id=notify_id).first()
        file_upload.upload_status = 'SUCCESS'
        file_upload.desc = '上传成功!'
        file_upload.save()

        return 'SUCESS'
    except Exception as e:
        # 处理异常并更新上传通知
        file_upload = FileUploadNotify.objects.filter(id=notify_id).first()
        file_upload.upload_status = 'ERROR'
        file_upload.desc = f'上传过程中出现错误: {str(e)}'
        file_upload.save()
        return 'ERROR'


# 上传美客多订单(跨境号)
@shared_task()
def upload_mercado_kj_order(shop_id, notify_id, mel_row):
    import warnings
    import re
    import openpyxl

    warnings.filterwarnings('ignore')

    try:
        data = MEDIA_ROOT + '/upload_file/order_excel_' + shop_id + '.xlsx'
        wb = openpyxl.load_workbook(data)
        sheet = wb.active

        # 定义所需表头
        required_headers = [
            'Order #', 'Order date', 'Status', 'Units', 'Selling fee (USD)',
            'Shipping cost based on the declared weight (USD)', 'Total (USD)',
            'Income per products (USD)', 'Sale from advertising', 'SKU',
            'Taxes (USD)', '# of listing', 'Listing sale unit price (USD)',
            'Buyer', 'Address', 'City', 'States', 'Zip Code', 'Country'
        ]

        shop = Shop.objects.filter(id=shop_id).first()
        er = ExRate.objects.filter(currency=shop.currency).first()
        ex_rate = er.value

        # 模板格式检查
        # 记录表头对应的列索引
        header_index = {}
        header_row = sheet[mel_row]
        if header_row is None:
            file_upload = FileUploadNotify.objects.filter(id=notify_id).first()
            file_upload.upload_status = 'ERROR'
            file_upload.desc = '未找到表头行'
            file_upload.save()
            return 'ERROR'
        for col_index, cell in enumerate(header_row, start=0):
            header = cell.value
            if header in required_headers and header not in header_index:
                header_index[header] = col_index

        # 检查是否所有必需的表头都存在
        if not all(header in header_index for header in required_headers):
            # 修改上传通知
            file_upload = FileUploadNotify.objects.filter(id=notify_id).first()
            file_upload.upload_status = 'ERROR'
            file_upload.desc = '模板格式有误，缺少必要的表头，请检查!'
            file_upload.save()
            return 'ERROR'

        add_list = []
        for cell_row in list(sheet)[int(mel_row):]:
            if cell_row is None:
                continue
            qty = cell_row[header_index['Units']].value
            if not qty:
                continue
            sku = cell_row[header_index['SKU']].value
            item_id = cell_row[header_index['# of listing']].value[3:]

            # 如果不在fmb库存中，或者所在店铺不对应，则跳出
            shop_stock = ShopStock.objects.filter(sku=sku,
                                                  item_id=item_id,
                                                  shop=shop).first()
            if not shop_stock:
                continue
            first_ship_cost = shop_stock.first_ship_cost
            if not first_ship_cost:
                first_ship_cost = 0

            order_number = cell_row[header_index['Order #']].value

            # 格式化时间
            # 首先去除GMT时区部分
            t = cell_row[header_index['Order date']].value
            date_str = t.split('GMT')[0].strip()
            # 解析原始日期字符串
            parsed_date = datetime.strptime(date_str, "%B %d, %Y %I:%M %p")
            # 格式化为"年-月-日 时-分-00"（增加固定的00秒）
            order_time = parsed_date.strftime("%Y-%m-%d %H:%M:00")

            bj = datetime.strptime(order_time,
                                   '%Y-%m-%d %H:%M:%S') + timedelta(hours=14)
            order_time_bj = bj.strftime('%Y-%m-%d %H:%M:%S')

            price = cell_row[
                header_index['Income per products (USD)']].value if cell_row[
                    header_index['Income per products (USD)']].value else 0
            postage = cell_row[header_index[
                'Shipping cost based on the declared weight (USD)']].value if cell_row[
                    header_index[
                        'Shipping cost based on the declared weight (USD)']].value else 0
            fees = cell_row[
                header_index['Selling fee (USD)']].value if cell_row[
                    header_index['Selling fee (USD)']].value else 0
            receive_fund = cell_row[
                header_index['Total (USD)']].value if cell_row[
                    header_index['Total (USD)']].value else 0
            VAT = cell_row[header_index['Taxes (USD)']].value if cell_row[
                header_index['Taxes (USD)']].value else 0
            is_ad = True if cell_row[header_index[
                'Sale from advertising']].value == 'Yes' else False

            buyer_name = cell_row[header_index['Buyer']].value
            buyer_address = cell_row[header_index['Address']].value
            buyer_city = cell_row[header_index['City']].value
            buyer_state = ''  # 与订单状态表头冲突，不使用
            buyer_postcode = cell_row[header_index['Zip Code']].value
            buyer_country = cell_row[header_index['Country']].value

            profit = (
                float(receive_fund) * 0.99
            ) * ex_rate - shop_stock.unit_cost * qty - shop_stock.first_ship_cost * qty
            # 合并订单，导出的表中没有价格显示
            if price:
                profit_rate = profit / (price * ex_rate)
            else:
                profit_rate = 0
            if profit_rate < 0:
                profit_rate = 0

            order_status = 'FINISHED'

            if cell_row[
                    header_index['Status']].value == 'Procesando en la bodega':
                order_status = 'PROCESS'
            if cell_row[header_index[
                    'Status']].value == 'Cancelada por el comprador':
                order_status = 'CANCEL'
            if cell_row[header_index[
                    'Status']].value == 'Paquete cancelado por Mercado Libre':
                order_status = 'CANCEL'
            if cell_row[
                    header_index['Status']].value == 'Devolución en camino':
                order_status = 'RETURN'
            if cell_row[header_index[
                    'Status']].value == 'Reclamo cerrado con reembolso al comprador':
                order_status = 'CASE'
            if cell_row[header_index['Status']].value[:8] == 'Devuelto':
                order_status = 'RETURN'
            if cell_row[header_index['Status']].value == 'En devolución':
                order_status = 'RETURN'
            if cell_row[header_index[
                    'Status']].value == 'Devolución finalizada con reembolso al comprador':
                order_status = 'RETURN'

            # 检查同一店铺订单编号是否存在
            ml_order = MLOrder.objects.filter(order_number=order_number,
                                              shop=shop).first()
            if not ml_order:
                add_list.append(
                    MLOrder(
                        shop=shop,
                        platform='MERCADO',
                        order_number=order_number,
                        order_status=order_status,
                        order_time=order_time,
                        order_time_bj=order_time_bj,
                        qty=qty,
                        currency=shop.currency,
                        ex_rate=ex_rate,
                        price=price,
                        fees=fees,
                        postage=postage,
                        receive_fund=receive_fund,
                        VAT=VAT,
                        is_ad=is_ad,
                        sku=sku,
                        p_name=shop_stock.p_name,
                        item_id=item_id,
                        image=shop_stock.image,
                        unit_cost=shop_stock.unit_cost * qty,
                        first_ship_cost=first_ship_cost * qty,
                        profit=profit,
                        profit_rate=profit_rate,
                        buyer_name=buyer_name,
                        buyer_address=buyer_address,
                        buyer_city=buyer_city,
                        buyer_state=buyer_state,
                        buyer_postcode=buyer_postcode,
                        buyer_country=buyer_country,
                    ))
                shop_stock.qty -= qty
                shop_stock.save()

                # 创建库存日志
                stock_log = StockLog()
                stock_log.shop_stock = shop_stock
                stock_log.current_stock = shop_stock.qty
                stock_log.qty = qty
                stock_log.in_out = 'OUT'
                stock_log.action = 'SALE'
                stock_log.desc = '销售订单号: ' + order_number
                stock_log.user_id = 0
                stock_log.save()
                stock_log.create_time = order_time  # order_time
                stock_log.save()
            else:
                # 如果费用更新
                if ml_order.fees != fees:
                    ml_order.receive_fund = receive_fund
                    ml_order.fees = fees
                    ml_order.profit = profit
                    ml_order.profit_rate = profit_rate
                    ml_order.save()
                # 如果订单状态更新
                if ml_order.order_status != order_status:
                    ml_order.order_status = order_status
                    ml_order.receive_fund = receive_fund
                    ml_order.fees = fees
                    ml_order.profit = profit
                    ml_order.profit_rate = profit_rate
                    ml_order.save()

                    # 如果订单是取消状态，库存增加回来
                    if order_status == 'CANCEL':
                        shop_stock.qty += qty
                        shop_stock.save()

                        # 创建库存日志
                        stock_log = StockLog()
                        stock_log.shop_stock = shop_stock
                        stock_log.current_stock = shop_stock.qty
                        stock_log.qty = qty
                        stock_log.in_out = 'IN'
                        stock_log.action = 'CANCEL'
                        stock_log.desc = '取消订单入库, 订单号: ' + order_number
                        stock_log.user_id = 0
                        stock_log.save()
        if len(add_list):
            MLOrder.objects.bulk_create(add_list)

        # 修改上传通知
        file_upload = FileUploadNotify.objects.filter(id=notify_id).first()
        file_upload.upload_status = 'SUCCESS'
        file_upload.desc = '上传成功!'
        file_upload.save()

        return 'SUCESS'
    except Exception as e:
        import sys
        exc_type, exc_value, exc_traceback = sys.exc_info()

        print("异常类型：", exc_type.__name__)
        print("异常描述：", exc_value)
        print("错误文件：", exc_traceback.tb_frame.f_code.co_filename)  # 出错文件
        print("错误行号：", exc_traceback.tb_lineno)  # 出错行号
        print("错误函数：", exc_traceback.tb_frame.f_code.co_name)  # 出错函数
        # 处理异常并更新上传通知
        file_upload = FileUploadNotify.objects.filter(id=notify_id).first()
        file_upload.upload_status = 'ERROR'
        file_upload.desc = f'上传过程中出现错误: {str(e)}'
        file_upload.save()
        return 'ERROR'


# 上传Emag订单
@shared_task()
def upload_emag_order(shop_id, notify_id, mel_row):
    import warnings
    import re
    import openpyxl

    warnings.filterwarnings('ignore')

    try:
        data = MEDIA_ROOT + '/upload_file/order_excel_' + shop_id + '.xlsx'
        wb = openpyxl.load_workbook(data)
        sheet = wb.active

        # 定义所需表头
        required_headers = [
            '订单编号', '订单日期', '产品代码', 'PNK', '数量', '不含增值税价格/件', '含增值税价格', '货币',
            '增值税', '订单状态', '客户姓名', '电话号码', '送货地址'
        ]

        shop = Shop.objects.filter(id=shop_id).first()

        # 模板格式检查
        # 记录表头对应的列索引
        header_index = {}
        header_row = sheet[mel_row]
        if header_row is None:
            file_upload = FileUploadNotify.objects.filter(id=notify_id).first()
            file_upload.upload_status = 'ERROR'
            file_upload.desc = '未找到表头行'
            file_upload.save()
            return 'ERROR'
        for col_index, cell in enumerate(header_row, start=0):
            header = cell.value
            if header in required_headers and header not in header_index:
                header_index[header] = col_index

        # 检查是否所有必需的表头都存在
        if not all(header in header_index for header in required_headers):
            # 修改上传通知
            file_upload = FileUploadNotify.objects.filter(id=notify_id).first()
            file_upload.upload_status = 'ERROR'
            file_upload.desc = '模板格式有误，缺少必要的表头，请检查!'
            file_upload.save()
            return 'ERROR'

        add_list = []
        for cell_row in list(sheet)[int(mel_row):]:
            if cell_row is None:
                continue
            qty = float(cell_row[header_index['数量']].value)
            if not qty:
                continue
            sku = cell_row[header_index['产品代码']].value
            item_id = cell_row[header_index['PNK']].value

            # 如果不在fmb库存中，或者所在店铺不对应，则跳出
            shop_stock = ShopStock.objects.filter(sku=sku,
                                                  item_id=item_id,
                                                  shop=shop).first()
            if not shop_stock:
                continue
            first_ship_cost = shop_stock.first_ship_cost
            if not first_ship_cost:
                first_ship_cost = 0

            order_number = cell_row[header_index['订单编号']].value

            order_time = cell_row[header_index['订单日期']].value
            currency = cell_row[header_index['货币']].value
            VAT = cell_row[header_index['增值税']].value

            price = float(cell_row[header_index['不含增值税价格/件']].value
                          if cell_row[header_index['不含增值税价格/件']].value else 0)
            invoice_price = cell_row[header_index['含增值税价格']].value if cell_row[
                header_index['含增值税价格']].value else 0

            buyer_name = cell_row[header_index['客户姓名']].value
            buyer_address = cell_row[header_index['送货地址']].value

            # 汇率 订单有不同货币，汇率不同
            er = ExRate.objects.filter(currency=currency).first()
            ex_rate = er.value

            postage = 8.00  # 尾程运费按100g重计算
            fees = round(price * 0.23, 2)  # 佣金率固定按23%计算

            profit = (
                price - fees
            ) * ex_rate * qty - postage * qty - shop_stock.unit_cost * qty - shop_stock.first_ship_cost * qty
            profit_rate = profit / (price * ex_rate)
            if profit_rate < 0:
                profit_rate = 0

            order_status = 'FINISHED'

            if cell_row[header_index['订单状态']].value == 'Storno':
                order_status = 'CANCEL'

            # 检查同一店铺订单编号是否存在
            ml_order = MLOrder.objects.filter(order_number=order_number,
                                              shop=shop).first()
            if not ml_order:
                add_list.append(
                    MLOrder(
                        shop=shop,
                        platform='EMAG',
                        order_number=order_number,
                        order_status=order_status,
                        order_time=order_time,
                        qty=qty,
                        currency=currency,
                        ex_rate=ex_rate,
                        price=price,
                        fees=fees,
                        invoice_price=invoice_price,
                        postage=postage,
                        receive_fund=round(
                            (price - fees) * ex_rate * qty - postage * qty, 2),
                        sku=sku,
                        p_name=shop_stock.p_name,
                        item_id=item_id,
                        image=shop_stock.image,
                        unit_cost=shop_stock.unit_cost * qty,
                        first_ship_cost=first_ship_cost * qty,
                        profit=profit,
                        profit_rate=profit_rate,
                        buyer_name=buyer_name,
                        buyer_address=buyer_address,
                    ))
                shop_stock.qty -= qty
                shop_stock.save()

                # 创建库存日志
                stock_log = StockLog()
                stock_log.shop_stock = shop_stock
                stock_log.current_stock = shop_stock.qty
                stock_log.qty = qty
                stock_log.in_out = 'OUT'
                stock_log.action = 'SALE'
                stock_log.desc = '销售订单号: ' + order_number
                stock_log.user_id = 0
                stock_log.save()
                stock_log.create_time = order_time  # order_time
                stock_log.save()
            else:
                # 如果订单状态更新
                if ml_order.order_status != order_status:
                    ml_order.order_status = order_status
                    ml_order.save()

                    # 如果订单是取消状态，库存增加回来
                    if order_status == 'CANCEL':
                        shop_stock.qty += qty
                        shop_stock.save()

                        # 创建库存日志
                        stock_log = StockLog()
                        stock_log.shop_stock = shop_stock
                        stock_log.current_stock = shop_stock.qty
                        stock_log.qty = qty
                        stock_log.in_out = 'IN'
                        stock_log.action = 'CANCEL'
                        stock_log.desc = '取消订单入库, 订单号: ' + order_number
                        stock_log.user_id = 0
                        stock_log.save()
        if len(add_list):
            MLOrder.objects.bulk_create(add_list)

        # 修改上传通知
        file_upload = FileUploadNotify.objects.filter(id=notify_id).first()
        file_upload.upload_status = 'SUCCESS'
        file_upload.desc = '上传成功!'
        file_upload.save()

        return 'SUCESS'
    except Exception as e:
        print("堆栈信息:")
        import traceback
        traceback.print_exc()  # 直接打印到控制台
        # 处理异常并更新上传通知
        file_upload = FileUploadNotify.objects.filter(id=notify_id).first()
        file_upload.upload_status = 'ERROR'
        file_upload.desc = f'上传过程中出现错误: {str(e)}'
        file_upload.save()
        return 'ERROR'


# 上传NOON订单
@shared_task()
def upload_noon_order(shop_id, notify_id):
    import warnings
    import openpyxl
    warnings.filterwarnings('ignore')

    data = MEDIA_ROOT + '/upload_file/order_excel_' + shop_id + '.xlsx'
    wb = openpyxl.load_workbook(data)
    sheet = wb.active

    shop = Shop.objects.filter(id=shop_id).first()
    er = ExRate.objects.filter(currency=shop.currency).first()
    ex_rate = er.value

    # 如果是invoice表
    if sheet['A1'].value == 'Invoice Nr':
        # 模板格式检查
        format_checked = True
        if sheet['A1'].value != 'Invoice Nr':
            format_checked = False
        if sheet['C1'].value != 'Invoice Date':
            format_checked = False
        if sheet['G1'].value != 'Source Document Type':
            format_checked = False
        if sheet['K1'].value != 'Vat Amount':
            format_checked = False
        if sheet['L1'].value != 'Price incl tax':
            format_checked = False
        if sheet['N1'].value != 'Item Nr':
            format_checked = False
        if sheet['O1'].value != 'SKU':
            format_checked = False
        if sheet['P1'].value != 'Partner SKU':
            format_checked = False
        if sheet['S1'].value != 'Customer ID':
            format_checked = False
        if sheet['T1'].value != 'Customer Name':
            format_checked = False
        if sheet['U1'].value != 'Customer City':
            format_checked = False
        if sheet['V1'].value != 'Customer Country':
            format_checked = False
        if not format_checked:
            # 修改上传通知
            file_upload = FileUploadNotify.objects.filter(id=notify_id).first()
            file_upload.upload_status = 'ERROR'
            file_upload.desc = '模板格式有误，请检查!'
            file_upload.save()
            return 'ERROR'

        add_list = []
        for cell_row in list(sheet)[1:]:
            row_type = cell_row[6].value
            # 检查行类型
            if row_type != 'Customer order':
                continue
            if not row_type:
                break

            sku = cell_row[15].value
            item_id = cell_row[14].value[:-2]

            # 如果不在fmb库存中，或者所在店铺不对应，则跳出
            shop_stock = ShopStock.objects.filter(sku=sku,
                                                  item_id=item_id,
                                                  shop=shop).first()
            if not shop_stock:
                continue
            first_ship_cost = shop_stock.first_ship_cost
            if not first_ship_cost:
                first_ship_cost = 0

            order_number = cell_row[13].value
            order_time = cell_row[2].value
            # 该字段可能有datetime和str 2种类型, 需要进行判断
            if type(cell_row[2].value) == 'str':
                order_time = cell_row[2].value + ' 00:00:00'

            buyer_name = cell_row[19].value
            buyer_id = cell_row[18].value
            buyer_city = cell_row[20].value
            buyer_country = cell_row[21].value
            VAT = -abs(cell_row[10].value) if cell_row[10].value else 0
            invoice_price = cell_row[11].value
            order_status = 'UNCHECK'

            # 检查同一店铺订单编号是否存在
            ml_order = MLOrder.objects.filter(order_number=order_number,
                                              shop=shop).first()
            if not ml_order:
                add_list.append(
                    MLOrder(
                        shop=shop,
                        platform='NOON',
                        order_number=order_number,
                        order_status=order_status,
                        order_time=order_time,
                        qty=1,
                        currency=shop.currency,
                        ex_rate=ex_rate,
                        VAT=VAT,
                        invoice_price=invoice_price,
                        sku=sku,
                        p_name=shop_stock.p_name,
                        item_id=item_id,
                        image=shop_stock.image,
                        unit_cost=shop_stock.unit_cost * 1,
                        first_ship_cost=first_ship_cost * 1,
                        buyer_name=buyer_name,
                        buyer_id=buyer_id,
                        buyer_city=buyer_city,
                        buyer_country=buyer_country,
                    ))
                shop_stock.qty -= 1
                shop_stock.save()

                # 创建库存日志
                stock_log = StockLog()
                stock_log.shop_stock = shop_stock
                stock_log.current_stock = shop_stock.qty
                stock_log.qty = 1
                stock_log.in_out = 'OUT'
                stock_log.action = 'SALE'
                stock_log.desc = '销售订单号: ' + order_number
                stock_log.user_id = 0
                stock_log.save()
                stock_log.create_time = order_time  # order_time
                stock_log.save()
        if len(add_list):
            MLOrder.objects.bulk_create(add_list)

    elif sheet['D1'].value == 'item_nr':
        # 模板格式检查
        format_checked = True
        if sheet['D1'].value != 'item_nr':
            format_checked = False
        if sheet['H1'].value != 'partner_sku':
            format_checked = False
        if sheet['G1'].value != 'sku':
            format_checked = False
        if sheet['O1'].value != 'country_code':
            format_checked = False
        if sheet['N1'].value != 'item_status':
            format_checked = False
        if sheet['Q1'].value != 'ordered_date':
            format_checked = False
        if sheet['U1'].value != 'currency_code':
            format_checked = False
        if sheet['AA1'].value != 'offer_price':
            format_checked = False
        if sheet['AF1'].value != 'fee_referral':
            format_checked = False
        if sheet['AH1'].value != 'fee_outbound_fbn':
            format_checked = False
        if sheet['AV1'].value != 'total_payment':
            format_checked = False
        if not format_checked:
            return 'ERROR'
        for cell_row in list(sheet)[1:]:
            order_number = cell_row[3].value
            if not order_number:
                break

            # 检查同一店铺订单编号是否存在
            ml_order = MLOrder.objects.filter(order_number=order_number,
                                              shop=shop).first()
            if ml_order:
                order_status = cell_row[13].value
                price = cell_row[26].value if cell_row[26].value else 0
                promo_coupon = cell_row[27].value if cell_row[27].value else 0
                fees = cell_row[31].value if cell_row[31].value else 0
                postage = cell_row[32].value if cell_row[32].value else 0
                payment_due = cell_row[46].value if cell_row[46].value else 0
                receive_fund = round(payment_due + ml_order.VAT, 2)
                shipped_date = ''
                delivered_date = ''
                if cell_row[16].value:
                    shipped_date = cell_row[16].value
                    # 该字段可能有datetime和str 2种类型, 需要进行判断
                    if type(cell_row[16].value) == 'str':
                        shipped_date = cell_row[16].value + ' 00:00:00'
                if cell_row[17].value:
                    delivered_date = cell_row[17].value
                    if type(cell_row[17].value) == 'str':
                        delivered_date = cell_row[17].value + ' 00:00:00'

                # 如果不在fmb库存中，或者所在店铺不对应，则跳出
                shop_stock = ShopStock.objects.filter(sku=ml_order.sku,
                                                      item_id=ml_order.item_id,
                                                      shop=shop).first()
                if not shop_stock:
                    continue
                first_ship_cost = shop_stock.first_ship_cost
                if not first_ship_cost:
                    first_ship_cost = 0
                # 如果没数据，跳出
                if price == 0:
                    continue

                if order_status == 'delivered':
                    # 订单送达
                    if ml_order.order_status != 'FINISHED':
                        ml_order.order_status = 'FINISHED'
                        ml_order.price = price
                        ml_order.promo_coupon = promo_coupon
                        ml_order.postage = postage
                        ml_order.fees = fees
                        ml_order.receive_fund = receive_fund
                        if shipped_date:
                            ml_order.shipped_date = shipped_date
                        if delivered_date:
                            ml_order.delivered_date = delivered_date

                        profit = (
                            float(receive_fund) * 0.99
                        ) * ex_rate - shop_stock.unit_cost * ml_order.qty - first_ship_cost * ml_order.qty
                        profit_rate = profit / (price * ex_rate)
                        if profit_rate < 0:
                            profit_rate = 0
                        ml_order.profit = profit
                        ml_order.profit_rate = profit_rate
                        ml_order.save()
                elif order_status == 'returned':
                    # 订单退货
                    if ml_order.order_status != 'RETURN':
                        ml_order.order_status = 'RETURN'
                        ml_order.VAT = 0
                        ml_order.price = price
                        ml_order.promo_coupon = promo_coupon
                        ml_order.postage = postage
                        ml_order.fees = fees
                        ml_order.receive_fund = payment_due
                        if shipped_date:
                            ml_order.shipped_date = shipped_date
                        if delivered_date:
                            ml_order.delivered_date = delivered_date

                        profit = (
                            float(payment_due) * 0.99
                        ) * ex_rate - shop_stock.unit_cost * ml_order.qty - first_ship_cost * ml_order.qty
                        profit_rate = profit / (price * ex_rate)
                        if profit_rate < 0:
                            profit_rate = 0
                        ml_order.profit = profit
                        ml_order.profit_rate = profit_rate
                        ml_order.save()

                else:
                    continue
            else:
                continue
    else:
        # 修改上传通知
        file_upload = FileUploadNotify.objects.filter(id=notify_id).first()
        file_upload.upload_status = 'ERROR'
        file_upload.desc = '模板格式有误，请检查!'
        file_upload.save()
        return 'ERROR'
    # 修改上传通知
    file_upload = FileUploadNotify.objects.filter(id=notify_id).first()
    file_upload.upload_status = 'SUCCESS'
    file_upload.desc = '上传成功!'
    file_upload.save()
    return 'SUCESS'


# 上传OZON订单
@shared_task()
def upload_ozon_order(shop_id, notify_id):
    import warnings
    import openpyxl
    warnings.filterwarnings('ignore')

    data = MEDIA_ROOT + '/upload_file/order_excel_' + shop_id + '.xlsx'
    wb = openpyxl.load_workbook(data)
    sheet = wb.active

    shop = Shop.objects.filter(id=shop_id).first()
    er = ExRate.objects.filter(currency=shop.currency).first()
    ex_rate = er.value

    oz_row = 2  # 默认标题行
    title_group_a = ['Номер заказа', '订单号']
    title_group_b = ['Номер отправления', '发货号码']
    title_group_c = ['Принят в обработку', '正在处理中']
    title_group_e = ['Статус', '状态']
    title_group_g = ['Фактическая дата передачи в доставку', '实际转移配送日期']
    title_group_h = ['Сумма отправления', '发货的金额']
    title_group_i = ['Код валюты отправления', '货件的货币代码']
    title_group_k = ['OZON id', 'Ozon ID', 'SKU']
    title_group_l = ['Артикул', '货号']
    title_group_q = ['Количество', '数量']
    # 如果是订单表
    if sheet['A' + str(oz_row)].value in title_group_a and sheet[
            'B' + str(oz_row)].value in title_group_b:
        # 模板格式检查
        format_checked = True
        if sheet['C' + str(oz_row)].value not in title_group_c:
            format_checked = False
        if sheet['E' + str(oz_row)].value not in title_group_e:
            format_checked = False
        if sheet['G' + str(oz_row)].value not in title_group_g:
            format_checked = False
        if sheet['H' + str(oz_row)].value not in title_group_h:
            format_checked = False
        if sheet['I' + str(oz_row)].value not in title_group_i:
            format_checked = False
        if sheet['K' + str(oz_row)].value not in title_group_k:
            format_checked = False
        if sheet['L' + str(oz_row)].value not in title_group_l:
            format_checked = False
        if sheet['Q' + str(oz_row)].value not in title_group_q:
            format_checked = False
        if not format_checked:
            # 修改上传通知
            file_upload = FileUploadNotify.objects.filter(id=notify_id).first()
            file_upload.upload_status = 'ERROR'
            file_upload.desc = '模板格式有误，请检查!'
            file_upload.save()
            return 'ERROR'

        add_list = []
        for cell_row in list(sheet)[oz_row:]:
            row_type = cell_row[0].value
            # 检查行是否有数据
            if not row_type:
                break

            sku = cell_row[11].value
            item_id = cell_row[10].value

            # 如果不在fmb库存中，或者所在店铺不对应，则跳出
            shop_stock = ShopStock.objects.filter(sku=sku,
                                                  item_id=item_id,
                                                  shop=shop).first()
            if not shop_stock:
                continue
            first_ship_cost = shop_stock.first_ship_cost
            if not first_ship_cost:
                first_ship_cost = 0

            order_number = cell_row[0].value
            dispatch_number = cell_row[1].value
            order_time = cell_row[2].value
            qty = cell_row[16].value
            price = cell_row[7].value

            oz_status = cell_row[4].value
            order_status = 'OTHERS'
            status_group_ass = ['Ожидает сборки', '待备货']
            status_group_ship = ['Ожидает отгрузки', '等待发运']
            status_group_deliver = ['Доставляется', '运输中']
            status_group_finish = ['Доставлен', '已签收']
            status_group_cancel = ['Отменен', '已取消']
            if oz_status in status_group_ass:
                order_status = 'WAIT_ASS'
            if oz_status in status_group_ship:
                order_status = 'WAIT_SHIP'
            if oz_status in status_group_deliver:
                order_status = 'DELIVERED'
            if oz_status in status_group_finish:
                order_status = 'FINISHED'
            if oz_status in status_group_cancel:
                order_status = 'CANCEL'

            # 检查同一店铺订单编号是否存在
            ml_order = MLOrder.objects.filter(order_number=order_number,
                                              dispatch_number=dispatch_number,
                                              sku=sku,
                                              shop=shop).first()
            if not ml_order:
                add_list.append(
                    MLOrder(
                        shop=shop,
                        platform='OZON',
                        order_number=order_number,
                        dispatch_number=dispatch_number,
                        order_status=order_status,
                        order_time=order_time,
                        qty=qty,
                        currency=shop.currency,
                        ex_rate=ex_rate,
                        price=price,
                        sku=sku,
                        p_name=shop_stock.p_name,
                        item_id=item_id,
                        image=shop_stock.image,
                        unit_cost=shop_stock.unit_cost * qty,
                        first_ship_cost=first_ship_cost * qty,
                    ))
                # 取消订单不扣库存
                if order_status != 'CANCEL':
                    shop_stock.qty -= qty
                    shop_stock.save()

                # 创建库存日志
                stock_log = StockLog()
                stock_log.shop_stock = shop_stock
                stock_log.current_stock = shop_stock.qty
                stock_log.qty = qty
                stock_log.in_out = 'OUT'
                stock_log.action = 'SALE'
                stock_log.desc = '销售订单号: ' + order_number
                stock_log.user_id = 0
                stock_log.save()
                stock_log.create_time = order_time  # order_time
                stock_log.save()
            else:
                # 如果订单状态更新
                if ml_order.order_status != order_status:
                    ml_order.order_status = order_status
                    ml_order.save()

                    # 如果订单是取消状态，库存增加回来
                    if order_status == 'CANCEL':
                        shop_stock.qty += qty
                        shop_stock.save()

                        # 创建库存日志
                        stock_log = StockLog()
                        stock_log.shop_stock = shop_stock
                        stock_log.current_stock = shop_stock.qty
                        stock_log.qty = qty
                        stock_log.in_out = 'IN'
                        stock_log.action = 'CANCEL'
                        stock_log.desc = '取消订单入库, 订单号: ' + order_number
                        stock_log.user_id = 0
                        stock_log.save()
        if len(add_list):
            MLOrder.objects.bulk_create(add_list)

    # 如果是费用更新表
    elif sheet['A2'].value == 'ID начисления' and sheet[
            'D2'].value == 'Тип начисления':
        # 模板格式检查
        format_checked = True
        if sheet['E2'].value != 'Артикул':
            format_checked = False
        if sheet['M2'].value != 'Вознаграждение Ozon, %':
            format_checked = False
        if sheet['P2'].value != 'Сумма итого, руб.':
            format_checked = False
        if not format_checked:
            # 修改上传通知
            file_upload = FileUploadNotify.objects.filter(id=notify_id).first()
            file_upload.upload_status = 'ERROR'
            file_upload.desc = '模板格式有误，请检查!'
            file_upload.save()
            return 'ERROR'

        # 费用类型
        fees_type_group = [
            'Последняя миля', 'Логистика', 'Вознаграждение за продажу',
            'Выручка'
        ]
        for cell_row in list(sheet)[2:]:
            item_number = cell_row[0].value
            order_type = cell_row[3].value
            sku = cell_row[4].value
            fee_rate = cell_row[12].value
            fees = cell_row[15].value  # 各种费用类型

            if not order_type:
                break

            # 检查账单项目
            if order_type == 'Эквайринг':
                # 收单费用
                # 检查同一店铺订单编号是否存在
                ml_order = MLOrder.objects.filter(order_number=item_number,
                                                  sku=sku,
                                                  shop=shop).first()
                if ml_order:
                    # 保存收单费用
                    if not ml_order.payment_fee:
                        ml_order.payment_fee = fees
                        ml_order.save()
            elif order_type in fees_type_group:
                # 最后一公里物流费用
                ml_order = MLOrder.objects.filter(dispatch_number=item_number,
                                                  sku=sku,
                                                  shop=shop).first()
                if not ml_order:
                    continue
                if ml_order.finance_check1:
                    continue

                # 保存数据
                if order_type == 'Выдача товара':
                    ml_order.last_mile_fee = fees  #最后一公里
                if order_type == 'Логистика':
                    ml_order.fbo_fee = fees  #fbo物流费
                if order_type == 'Вознаграждение за продажу':
                    ml_order.fees = fees  #平台佣金
                    ml_order.fee_rate = fee_rate  #平台佣金率
                if order_type == 'Выручка':  #收入资金
                    ml_order.receive_fund = fees
                ml_order.save()
            else:
                continue
        # 计算费用项
        orders = MLOrder.objects.filter(shop=shop, finance_check1=False)
        sp = GeneralSettings.objects.filter(item_name='ozon_sp').first()
        for i in orders:
            if i.fbo_fee and i.fees and i.receive_fund:
                # 如果不在fmb库存中，或者所在店铺不对应，则跳出
                shop_stock = ShopStock.objects.filter(sku=i.sku,
                                                      item_id=i.item_id,
                                                      shop=shop).first()
                if not shop_stock:
                    continue
                first_ship_cost = shop_stock.first_ship_cost
                if not first_ship_cost:
                    first_ship_cost = 0
                postage = round(i.fbo_fee + i.last_mile_fee, 2)  # 平台物流总费用
                # 计算收入资金
                # receive_fund = round(
                #     i.price - abs(i.fees) - abs(postage) - abs(i.payment_fee),
                #     2)
                receive_fund = i.receive_fund
                # 计算服务商费用
                sp_fee = 0
                sp_fee_rate = 0
                if sp:
                    sp_fee_rate = sp.value1
                    sp_fee = receive_fund * sp_fee_rate
                # 计算利润
                profit = (
                    receive_fund - sp_fee
                ) * ex_rate - shop_stock.unit_cost * i.qty - first_ship_cost * i.qty
                profit_rate = profit / (i.price * ex_rate)
                # 保存数据
                i.postage = postage
                i.receive_fund = receive_fund
                i.sp_fee = sp_fee
                i.sp_fee_rate = sp_fee_rate
                i.profit = profit
                i.profit_rate = profit_rate
                if i.last_mile_fee:
                    i.finance_check1 = True
                i.save()
    else:
        # 修改上传通知
        file_upload = FileUploadNotify.objects.filter(id=notify_id).first()
        file_upload.upload_status = 'ERROR'
        file_upload.desc = '模板格式有误，请检查!'
        file_upload.save()
        return 'ERROR'
    # 修改上传通知
    file_upload = FileUploadNotify.objects.filter(id=notify_id).first()
    file_upload.upload_status = 'SUCCESS'
    file_upload.desc = '上传成功!'
    file_upload.save()
    return 'SUCESS'


# 盛德交运运单-旧
@shared_task()
def sd_place_order(ship_id, data):
    ship = Ship.objects.filter(id=ship_id).first()
    if not ship:
        return {'status': 'error', 'msg': '运单不存在！'}
    if ship.carrier != '盛德物流':
        return {'status': 'error', 'msg': '仅支持盛德物流交运！'}
    if not ship.envio_number:
        return {'status': 'error', 'msg': 'envio号不能为空！'}

    # 交运数据
    payload = {
        'txtconsigneephoto': 'undefined',
        'operNo': '',
        'hkremarks': '',
        'country_code': '墨西哥',
        'd_code': data['d_code'],  # fbm仓库代码
        'fbano': '',
        'address1': data['address1'],  # fbm仓库地址
        'zip_code': data['zip_code'],  # fbm仓库邮编
        'channelcode': '',
        'delivertype': '卡派',
        'warehousename': '东莞塘厦仓',
        'channeltype2': '墨西哥极速达',
        'jfunit': 'kg',
        'postcode': 'mx',
        'reserveid': data['reserveid'],  # 预约id
        'apptdate': data['apptdate'],  # 预约日期
        'ckyyservice': 1,
        'rsum': 1,
        'invoicelink': '',
        'cargoagency': '',
        'EntrustType': data['EntrustType'],  # 空运
        'tihuotype': '自送货',
        'nid': '03E85014-136E-4E5B-B2F9-751E21FF5D88',  # 公司id,固定
        'khchannel_id': '2498F847-EA54-44EE-9EFF-797BF44AAEF6',  # 不知道，固定
        'warehouseid': data['warehouseid'],  # 仓库id
        'DeliveryTime': data['DeliveryTime'],  # 预约送仓时间
        'LoadingDriver': '',
        'channelway': '',
        'kimsvolume': '',
        'ConsignorContacts': '钟生',
        'ConsignorPhone': '13823289200',
        'ConsignorAddress': '深圳市 福田区 赛格科技园 4栋西  703E',
        'extendoperno': '',
        'khremarks': '',
        'isbx': '否',
        'insure_currency': '',
        'xmai': '',
        'isOtherImporter': 0,
        'importername': '',
        'destination': '',
        'eori': '',
        'address': '',
        'importerid': '',
        'isfba': 'FBM',
        'sellerid': data['sellerid'],  # Seller ID
        'envio': data['envio'],  # fbm入仓号
        'intowarehouseweeks': '',
        'shop': '',
        'cwremark': '',
        'fbl': '',
        'buildtime': '',
        'shippingdate': '',
        'companyname': '',
        'consignee': '',
        'telephone': '',
        'province': '',
        'city': '',
        'yyno': '',
        'po': '',
        'consigneephoto': '',
        'ftype': '箱唛',  # 上传文件类型
        'product': data['product'],  # 品名
        'ProductNature': data['ProductNature'],  # 产品性质，可能多个
        'vat_clear_customs': '无',
        'GoodsNum': ship.total_box,  # 预计箱数
        'yjweight': ship.weight,  # 预计重量
        'yjvolume': round(ship.cbm, 4),  # 预计体积
        'CustomsDeclaration': '无单证',
        'fruit': 1,
        'quantity': ship.total_qty,  # 产品总数量
        'rcid': 'EBCF9B79-2504-49D7-AD07-BE325670DE22',  # 待查看
        'bxjine': 0.00,
        'email': '',
        'valtype': 1,
        'somrequest': '',
        'vat': '',
        'vat_company': '',
        'vat_address': '',
        'vat_contact': '',
        'vat_companyphone': '',
        'eroi': '',
        'eoricompany': '',
        'eoriaddr': '',
        'iskims': '否',
    }
    # 上传附件
    files = {}
    f_id = time.strftime('%Y%m%d%H%M%S')  # 箱唛附件编号

    f_name = ''  #  pdf文件名
    sa_set = ShipAttachment.objects.filter(ship=ship, a_type='BOX_LABEL')
    for i in sa_set:
        extension = i.name.split(".")[-1]
        if extension == 'pdf':
            f_name = i.name
    path = '{batch}_{id}/{name}'.format(batch=ship.batch,
                                        id=ship.id,
                                        name=f_name)
    book_file = MEDIA_ROOT + '/ml_ships/' + path
    files['txtupload' + f_id] = open(book_file, 'rb')  # 上传箱唛附件

    # 箱唛附件描述信息
    payload['filegrid'] = json.dumps([{
        "id": f_id,
        "filename": f_name,
        "filetype": "箱唛",
        "size": "2.805KB",
        "fileurl": ""
    }],
                                     ensure_ascii=False)

    p_list = []  # 运单产品列表
    boxes = ShipBox.objects.filter(ship=ship)
    box_num = 0
    num = 0
    for b in boxes:
        sd_set = ShipDetail.objects.filter(ship=ship, box_number=b.box_number)
        box_tag = 1
        for i in sd_set:
            if i.image:
                # 获取完整文件名（含路径，如 "uploads/images/2025/07/01/test.jpg"）
                full_name = i.image.name
                # 提取仅文件名（不含路径和扩展名，如 "test"）
                file_base_name = os.path.basename(full_name)  # 结果："test.jpg"
                img_name = os.path.splitext(file_base_name)[0]  # 结果："test"
                product_pic = MEDIA_ROOT + '/ml_product/' + img_name + '_100x100.jpg'
                # 上传产品图片
                files['txtimg' + str(num + 1)] = open(product_pic, 'rb')
            else:
                return {'status': 'error', 'msg': '产品图片缺失，请补充'}

            p_list.append({
                'ContainerNo':
                i.ship.batch + '/' + str(box_num + 1),  # 货箱编号
                'ContaineNum':
                box_tag,  # 件数
                'ContainerWeight':
                '0.00',
                'ContainerLength':
                '0.00',
                'ContainerWidth':
                '0.00',
                'ContainerHeight':
                '0.00',
                'GoodsSKU':
                i.brand,  # 品牌
                'EnglishProduct':
                i.en_name,  # 英文品名
                'ChineseProduct':
                i.cn_name,  # 中文品名
                'DeclaredValue':
                i.declared_value,  # 单个产品申报价值(usd)
                'DeclaredNum':
                i.qty,  # 单箱申报数量
                'Material':
                i.cn_material,  # 材质
                'Purpose':
                i.use,  # 用途
                'CustomsCode':
                i.custom_code,  # 海关编码
                'SalesWebsite':
                'https://articulo.mercadolibre.com.mx/MLM-' +
                i.item_id,  # 销售网址
                'SellingPice':
                '0.00',
                "PicturesLink":
                "1",
                "ProductWeight":
                "0.00",
                "ProductSize":
                "",
                "ASIN":
                "无",
                "FNSKU":
                "无",
                "model":
                "",
                "netweight":
                "0.00",
                "roughweight":
                "0.00",
                "english_material":
                i.cn_material,  # 英文材质
                "id":
                num + 1,
                "isdd":
                "",
                "isdc":
                "",
                "GoodsSKUtype":
                "",
                "custom1":
                "",
                "custom2":
                "",
                "custom3":
                "",
                "custom4":
                "",
                "custom5":
                ""
            })
            box_tag = 0
            num += 1
        box_num += 1
    payload['tb_GoodsInfo2'] = json.dumps(p_list, ensure_ascii=False)  # 产品列表

    c_path = os.path.join(BASE_DIR, "site_config.json")
    with open(c_path, "r", encoding="utf-8") as f:
        data = json.load(f)  # 加载配置数据
    header = {'Cookie': data['sd_cookies']}
    url = 'http://client.sanstar.net.cn/console/customer_order/newadd?somrequest='

    resp = requests.post(url, files=files, data=payload, headers=header)

    if 'success' in resp.json():
        if resp.json()['success']:
            # 交运成功
            ship.s_number = resp.json()['sono']
            ship.carrier_order_status = 'WAIT'
            ship.carrier_rec_check = 'UNCHECK'
            ship.carrier_order_time = datetime.now()
            ship.save()

            # 创建操作日志
            log = MLOperateLog()
            log.op_module = 'SHIP'
            log.op_type = 'EDIT'
            log.target_type = 'SHIP'
            log.target_id = ship.id
            log.desc = '盛德交运成功，获取运单号-{track_num}'.format(
                track_num=ship.s_number)
            log.save()
            return {'status': 'success', 'msg': resp.json()['msg']}
        else:
            return {'status': 'error', 'msg': resp.json()['msg']}

    return 'SUCESS'


# 盛德交运运单-新api
@shared_task()
def sd_place_order_api(ship_id, data):
    ship = Ship.objects.filter(id=ship_id).first()
    if not ship:
        return {'status': 'error', 'msg': '运单不存在！'}
    if ship.carrier != '盛德物流':
        return {'status': 'error', 'msg': '仅支持盛德物流交运！'}
    if not ship.envio_number:
        return {'status': 'error', 'msg': 'envio号不能为空！'}

    # 交运数据
    payload = {
        "order": {
            "sono": "",
            "transport_type": "空运",
            "is_tax": True,
            "start_city": "东莞市",
            "country_name": "墨西哥",
            "country_code": "MX",
            "delivery_mode": 2,
            "delivery_time": data['delivery_time'],  # 预约日期
            "fh_contacts": "",
            "fh_phone": "",
            "fh_province": "",
            "fh_city": "",
            "fh_area": "",
            "fh_addr": "",
            "fh_longitude": "",
            "fh_latitude": "",
            "ckid": "695128929226885",
            "ck_name": "东莞凤岗仓",
            "is_insure": False,
            "tb_jine": "",
            "tb_currency": "",
            "insured": "",
            "kh_chid": "650521541155397",
            "kh_channel_name": "墨西哥极速达",
            "fbx_name": "美客多MELI",
            "addr_type": 0,
            "fbx_code": "FBL",
            "d_code": data['d_code'],  # fbm仓库代码
            "fbx_no": data['fbx_no'],  # fbm入仓号
            "sh_province": "",
            "sh_city": "",
            "sh_addr": data['sh_addr'],  # fbm仓库地址
            "sh_email": "",
            "sh_zip_code": data['sh_zip_code'],  # fbm仓库邮编
            "sh_company": "",
            "final_delivery": "",
            "sh_contacts": "",
            "sh_phone": "",
            "product": data['product'],  # 品名
            "product_nature": data['ProductNature'],  # 产品性质，可能多个
            "packing": "箱",
            "yj_qty": ship.total_box,  # 预计箱数
            "yj_kg": ship.weight,  # 预计重量
            "yj_cbm": round(ship.cbm, 4),  # 预计体积
            "bg_mode": "无单证",
            "qg_mode": "无",
            "have_no": "",
            "ve_vat_tax": "",
            "ve_eori_tax": "",
            "ve_company": "",
            "ve_addr": "",
            "ve_contacts": "",
            "ve_tel": "",
            "billing_unit": "KG",
            "offer_price": "",
            "kh_remark": "",
            "order_state": 0,
            "infoid": "",
            "coupon_name": "",
            "vid": "",
            "v_title": ""
        }
    }

    boxgauge_list = []  # 箱列表
    boxes = ShipBox.objects.filter(ship=ship)
    box_num = 0
    for b in boxes:
        sd_set = ShipDetail.objects.filter(ship=ship, box_number=b.box_number)
        one_box = {
            "bgid": "",
            'container_no':
            b.ship.envio_number + '/' + str(box_num + 1),  # 货箱编号
            'tracking_no': data['refid'],  #店铺账号id
            "length": b.length,
            "width": b.width,
            "height": b.heigth,
            "build_date": "",
            "singlebox_kg": 0,
            "refid": data['refid'],  #店铺账号id
            "ps_timeslot": "",
            # 预留字段，后面赋值
            "product_list": [],
            "qty": 0
        }
        p_list = []  # 运单产品列表
        product_qty_in_box = 0  #箱内产品数量
        for i in sd_set:
            product_qty_in_box += i.qty
            if i.image:
                # 获取完整文件名（含路径，如 "uploads/images/2025/07/01/test.jpg"）
                full_name = i.image.name
                # 提取仅文件名（不含路径和扩展名，如 "test"）
                file_base_name = os.path.basename(full_name)  # 结果："test.jpg"
                img_name = os.path.splitext(file_base_name)[0]  # 结果："test"
                product_pic_url = BASE_URL + MEDIA_URL + '/ml_product/' + img_name + '_100x100.jpg'  # 产品图片链接
            else:
                return {'status': 'error', 'msg': '产品图片缺失，请补充'}

            p_list.append({
                "cn_product_name": i.cn_name,  # 中文品名
                "us_product_name": i.en_name,  # 英文品名
                "brand": i.brand,  # 品牌
                "model": i.sku,  #型号
                "purpose": i.use,  # 用途
                "material": i.cn_material,  # 材质
                "hs_code": i.custom_code,  # 海关编码
                "single_cost": i.declared_value,  # 单个产品申报价值(usd)
                "currency": "USD",
                "total_declared": "",
                "asin": "",
                "product_nature": "",
                "sales_links": "",
                "back_map": "",
                "sales_records_map": "",
                "product_map": product_pic_url,  # 图片地址
                "sales_price": "",
                "attestation": "",
                "product_pcs": i.qty,  #数量
                "sku": i.sku,  #型号
                "pt_gross_weight": "",
                "pt_net_weight": "",
                "tax_rate": "",
                "bgid": "",
                "product_dx_pcs": 0
            })
        one_box['product_list'] = p_list  #该箱产品列表
        one_box['qty'] = product_qty_in_box  #该箱产品数量
        boxgauge_list.append(one_box)  #  把这个箱子加入总列表
        box_num += 1
    payload['boxgauge_list'] = boxgauge_list  # 箱列表

    c_path = os.path.join(BASE_DIR, "site_config.json")
    with open(c_path, "r", encoding="utf-8") as f:
        config = json.load(f)  # 加载配置数据
    header = {
        'Authorization': config['sd_cookies'],
        'Content-Type': 'application/json'
    }
    url = 'http://api.more56.com/api/v1/edi/open_bigwaybill_order_client/create'

    try:
        resp = requests.post(url, json=payload, headers=header, timeout=20)
        print(f"接口状态码：{resp.status_code}")
        print(f"接口返回原始内容：{resp.text}")

        # 1. 先判断 HTTP 状态码
        if resp.status_code != 200:
            return {
                "status": "error",
                "msg": f"接口异常，HTTP状态码：{resp.status_code}"
            }

        # 2. 只解析一次 json（避免多次解析报错）
        result = resp.json()

        # 3. 业务状态判断
        if not result.get("success", False):
            msg = result.get("message", "接口返回失败，无错误信息")
            return {"status": "error", "msg": f"交运失败：{msg}"}

        # 4. 成功 → 拿 sono
        data_info = result.get("data", {})
        sono = data_info.get("sono", "")
        if not sono:
            return {"status": "error", "msg": "接口返回成功，但未获取到 sono"}

        # ====================== 开始上传箱唛文件 ======================
        def upload_box_label_files(ship, sono, config):
            """上传箱唛附件到盛德"""
            sa_set = ShipAttachment.objects.filter(ship=ship,
                                                   a_type='BOX_LABEL')
            if not sa_set.exists():
                return True  # 无附件，跳过

            upload_url = "http://api.more56.com/api/v1/fss/sys_files_client/filesadd"

            form_data = {}
            files = {}

            # 遍历多个附件，自动生成 ifs[0], ifs[1], ifs[2]...
            for idx, att in enumerate(sa_set):
                # 拼接文件路径
                path = f"{ship.batch}_{ship.id}/{att.name}"
                file_path = os.path.join(MEDIA_ROOT, 'ml_ships', path)

                if not os.path.exists(file_path):
                    continue

                with open(file_path, 'rb') as f_obj:
                    form_data[f"ifs[{idx}].file_no"] = sono
                    form_data[f"ifs[{idx}].file_classify"] = "箱唛"
                    form_data[f"ifs[{idx}].fileposition"] = "customerfiles"
                    form_data[f"ifs[{idx}].fields"] = "file"
                    files[f"ifs[{idx}].files"] = (att.name, f_obj.read(),
                                                  'application/pdf')

                # 文件
                files[f"ifs[{idx}].files"] = (att.name, open(file_path, 'rb'),
                                              'application/pdf')

            if not files:
                return True

            # 上传请求头（不要 Content-Type，requests 自动生成 form-data）
            upload_headers = {'Authorization': config['sd_cookies']}

            try:
                res = requests.post(upload_url,
                                    data=form_data,
                                    files=files,
                                    headers=upload_headers,
                                    timeout=30)
                print(f"上传附件返回：{res.status_code} | {res.text}")
                return res.status_code == 200
            except Exception as e:
                print(f"上传附件失败：{str(e)}")
                return False

        # 执行上传
        upload_ok = upload_box_label_files(ship, sono, config)

        # ====================== 保存数据（事务保证安全） ======================
        with transaction.atomic():

            ship.s_number = sono
            ship.carrier_order_status = "WAIT"
            ship.carrier_rec_check = "UNCHECK"
            ship.carrier_order_time = datetime.now()
            ship.save()

            # 日志
            log = MLOperateLog()
            log.op_module = "SHIP"
            log.op_type = "EDIT"
            log.target_type = "SHIP"
            log.target_id = ship.id
            log.desc = f"盛德交运成功，运单号：{sono}"
            log.save()

        return {"status": "success", "msg": f"交运成功，运单号：{sono}"}

    except requests.exceptions.Timeout:
        return {"status": "error", "msg": "请求盛德接口超时"}
    except ValueError:
        return {"status": "error", "msg": "接口返回非JSON格式数据"}
    except Exception as e:
        return {"status": "error", "msg": f"系统异常：{str(e)}"}

    return 'SUCESS'


# 查询盛德运单受理状态
@shared_task()
def query_sd_order_status():
    # ===================== 1. 读取配置获取认证 =====================
    try:
        c_path = os.path.join(BASE_DIR, "site_config.json")
        with open(c_path, "r", encoding="utf-8") as f:
            config = json.load(f)
        sd_cookies = config.get('sd_cookies', '')
    except:
        return '读取配置失败'

    # ===================== 2. 新接口地址 =====================
    url = 'http://api.more56.com/api/v1/edi/open_bigwaybill_order_client/orderinfo'

    # ===================== 3. 请求头（认证） =====================
    headers = {'Authorization': sd_cookies, 'Content-Type': 'application/json'}

    # ===================== 4. 自动计算日期（30天前 ~ 今天） =====================
    today = datetime.now().date()
    s_date = today - timedelta(days=30)  # 30天前
    e_date = today  # 今天

    # ===================== 5. 请求参数（ids 留空，按时间批量查询） =====================
    payload = {
        "sel_type": "系统SO号",  # 固定
        "ids": [],  # 留空 = 按时间查询
        "pagesize": 50,  # 固定
        "pageindex": 1,  # 固定
        "s_date": s_date.strftime('%Y-%m-%d'),  # 30天前
        "e_date": e_date.strftime('%Y-%m-%d'),  # 今天
        "order_state": "2"  # 固定
    }

    try:
        # ===================== 6. 请求接口 =====================
        resp = requests.post(url, json=payload, headers=headers, timeout=20)
        if resp.status_code != 200:
            return f'接口异常 {resp.status_code}'

        result = resp.json()
        if not result.get('success', False):
            return result.get('message', '查询失败')

        # ===================== 7. 获取订单列表 =====================
        data = result.get('data', {})
        order_list = data.get('list', [])
        if not order_list:
            return '暂无订单数据'

        # ===================== 8. 取出所有返回的 sono 列表（加速查询） =====================
        sono_list = [
            item.get('sono') for item in order_list if item.get('sono')
        ]

        # ===================== 9. 匹配并更新运单状态 =====================
        # 只更新：待受理 + 盛德物流 的运单
        ship_set = Ship.objects.filter(
            carrier='盛德物流',
            s_status='PREPARING',
            carrier_order_status='WAIT',
            s_number__in=sono_list  # 只查接口返回的运单，速度极快
        )

        for ship in ship_set:
            # 状态改为：已受理
            ship.carrier_order_status = 'ACCEPTED'
            ship.save()

            # 记录日志
            log = MLOperateLog()
            log.op_module = 'SHIP'
            log.op_type = 'EDIT'
            log.target_type = 'SHIP'
            log.target_id = ship.id
            log.desc = '盛德运单已受理'
            log.save()

        return 'SUCCESS'

    except requests.exceptions.Timeout:
        return '请求超时'
    except Exception as e:
        return f'系统异常：{str(e)}'


# ==============================
# 生成 Code128 条码
# ==============================
# ==========================
# 修复 ZPL 十六进制编码（如 _C3_A9 → é）
# ==========================
def fix_zpl_chars(text):
    if not text:
        return ""
    import re
    pattern = re.compile(r'_C3_([0-9A-F]{2})')

    def replace_match(match):
        hex_byte = match.group(1)
        return bytes([0xC3, int(hex_byte, 16)]).decode('utf-8')

    return pattern.sub(replace_match, text).strip()


# ==========================
# 生成清晰条码
# ==========================
def generate_barcode(code):
    try:
        rv = BytesIO()
        Code128(code, writer=ImageWriter()).write(rv,
                                                  options={
                                                      "module_width": 0.25,
                                                      "module_height": 8,
                                                      "quiet_zone": 1,
                                                      "text_distance": 0,
                                                      "font_size": 0,
                                                  })
        rv.seek(0)
        return rv
    except:
        return None


# ==========================
# 完整解析 ZPL（含颜色/型号）
# ==========================
def parse_zpl(zpl_text):
    labels = []
    blocks = zpl_text.strip().split("^XA")
    for b in blocks:
        if not b.strip():
            continue

        # 1. 追踪码
        code = re.search(r"\^FD([A-Z0-9]{9})\^FS", b)
        code = code.group(1) if code else ""

        # 2. 产品名称
        desc = re.search(r"\^FO40,185.*?\^FD(.*?)\^FS", b, re.DOTALL)
        desc = desc.group(1) if desc else ""
        desc = fix_zpl_chars(desc)

        # 3. 颜色 / 型号（第二行描述）
        color = ""
        color1 = re.search(r"\^FO40,240.*?\^FD(.*?)\^FS", b, re.DOTALL)
        color2 = re.search(r"\^FO39,240.*?\^FD(.*?)\^FS", b, re.DOTALL)
        if color1:
            color = color1.group(1)
        if color2 and not color:
            color = color2.group(1)
        color = fix_zpl_chars(color)

        # 4. SKU
        sku = re.search(r"SKU:\s*(MD\d+)\^FS", b)
        sku = sku.group(1) if sku else ""

        # 5. 数量
        qty = re.search(r"\^PQ(\d+),", b)
        qty = int(qty.group(1)) if qty else 1

        if code:
            labels.append({
                "code": code,
                "desc": desc,
                "color": color,
                "sku": sku,
                "qty": qty
            })
    return labels


# ==========================
# 【纯渲染·可复用】生成ml平台标签PDF
# 入参：
#   label_list: 数组，固定结构
#   output_path: 输出路径
# 结构示例：
# [
#     {"code":"追踪码", "desc":"产品标题", "color":"型号/颜色", "sku":"MDxxxx", "qty":2},
# ]
# ==========================
def generate_ml_label_pdf(label_list, output_path=""):
    pdf = FPDF(unit="mm", format=[60, 40])
    pdf.set_auto_page_break(False)

    for lab in label_list:
        # 根据数量循环生成多页标签
        for _ in range(lab.get("qty", 1)):
            pdf.add_page()

            # 条码居中 45mm
            bc = generate_barcode(lab.get("code", ""))
            if bc:
                pdf.image(bc, x=7.5, y=2, w=45)

            # 顶部追踪码
            pdf.set_font("Arial", "B", 12)
            pdf.set_xy(0, 15)
            pdf.cell(60, 5, lab.get("code", ""), align="C")

            # 产品描述
            pdf.set_font("Arial", "", 8)
            pdf.set_xy(4, 21)
            pdf.multi_cell(52, 3, lab.get("desc", ""), align='L')

            # 颜色/型号选项
            if lab.get("color"):
                pdf.set_font("Arial", "B", 8)
                pdf.set_xy(4, pdf.get_y())
                pdf.multi_cell(52, 3, lab["color"], align='L')

            # 底部SKU
            pdf.set_font("Arial", "B", 9)
            pdf.set_xy(4, 34)
            pdf.cell(50, 4, f"SKU: {lab.get('sku', '')}")

    pdf.output(output_path)
    return output_path


def process_sku_label_file(file_path):
    """
    处理ml平台上传的SKU_LABEL标签TXT文件
    解析ZPL -> 提取SKU、追踪码、标题、选项 -> 自动填充MLProduct
    """
    try:
        # 1. 读取文件内容
        with open(file_path, 'r', encoding='utf-8') as f:
            zpl_content = f.read()

        # 2. 分割每个标签 ^XA ... ^XZ
        label_blocks = re.split(r'\^XA', zpl_content)
        product_list = []

        for block in label_blocks:
            if not block.strip():
                continue

            # ====================== 解析字段 ======================
            # 追踪码（如 PMCM72911）
            code_match = re.search(r'\^FD([A-Z0-9]{9})\^FS', block)
            label_code = code_match.group(1) if code_match else ""

            # 产品描述 label_title
            desc_match = re.search(r'\^FO40,185.*?\^FD(.*?)\^FS', block,
                                   re.DOTALL)
            label_title = desc_match.group(1).strip() if desc_match else ""

            # 颜色/型号 label_option
            option_match1 = re.search(r'\^FO40,240.*?\^FD(.*?)\^FS', block,
                                      re.DOTALL)
            option_match2 = re.search(r'\^FO39,240.*?\^FD(.*?)\^FS', block,
                                      re.DOTALL)
            label_option = ""
            if option_match1:
                label_option = option_match1.group(1).strip()
            if option_match2 and not label_option:
                label_option = option_match2.group(1).strip()

            # SKU
            sku_match = re.search(r'SKU:\s*(MD\d+)\^FS', block)
            sku = sku_match.group(1) if sku_match else ""

            # 西语字符解码
            def fix_zpl_text(s):
                if not s:
                    return ""
                pattern = re.compile(r'_C3_([0-9A-F]{2})')

                def rep(m):
                    return bytes([0xC3, int(m.group(1), 16)]).decode('utf-8')

                return pattern.sub(rep, s).strip()

            label_title = fix_zpl_text(label_title)
            label_option = fix_zpl_text(label_option)
            # ======================================================

            if sku:
                product_list.append({
                    "sku": sku,
                    "label_code": label_code,
                    "label_title": label_title,
                    "label_option": label_option
                })

        # 3. 批量更新产品（为空才更新）
        update_count = 0

        for item in product_list:
            sku = item["sku"]
            product = MLProduct.objects.filter(sku=sku).first()
            if not product:
                continue

            need_save = False

            # 只要其中一个为空，就赋值
            if not product.label_code or not product.label_title:
                product.label_code = item["label_code"]
                product.label_title = item["label_title"]
                need_save = True
            # 颜色/型号 label_option
            product.label_option = item["label_option"]

            if need_save:
                product.save(update_fields=[
                    'label_code', 'label_title', 'label_option'
                ])
                update_count += 1

        return {
            "status": "ok",
            "parsed_count": len(product_list),
            "updated_count": update_count
        }

    except Exception as e:
        return {"status": "error", "msg": str(e)}
