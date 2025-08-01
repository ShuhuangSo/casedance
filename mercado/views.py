from rest_framework import mixins, status
from rest_framework import viewsets
from rest_framework.response import Response
from rest_framework import filters
from rest_framework.pagination import PageNumberPagination
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.decorators import action
import requests
import openpyxl
import time, os
import random
import json
import urllib
import hashlib
from datetime import datetime, timedelta
from django.db.models import Sum
from django.db.models import Q

from casedance.settings import BASE_URL, MEDIA_ROOT, BASE_DIR
from mercado.models import Listing, ListingTrack, Categories, ApiSetting, TransApiSetting, Keywords, Seller, \
    SellerTrack, MLProduct, Shop, ShopStock, Ship, ShipDetail, ShipBox, Carrier, TransStock, MLSite, FBMWarehouse, \
    MLOrder, ExRate, Finance, Packing, MLOperateLog, ShopReport, PurchaseManage, ShipItemRemove, ShipAttachment, UPC, \
    RefillRecommend, RefillSettings, CarrierTrack, StockLog, FileUploadNotify, PlatformCategoryRate, ShippingPrice, ShippingAreaCode
from mercado.serializers import ListingSerializer, ListingTrackSerializer, CategoriesSerializer, SellerSerializer, \
    SellerTrackSerializer, MLProductSerializer, ShopSerializer, ShopStockSerializer, ShipSerializer, \
    ShipDetailSerializer, ShipBoxSerializer, CarrierSerializer, TransStockSerializer, MLSiteSerializer, \
    FBMWarehouseSerializer, MLOrderSerializer, FinanceSerializer, PackingSerializer, MLOperateLogSerializer, \
    ShopReportSerializer, PurchaseManageSerializer, ShipItemRemoveSerializer, ShipAttachmentSerializer, UPCSerializer, \
    RefillRecommendSerializer, RefillSettingsSerializer, CarrierTrackSerializer, StockLogSerializer, \
    FileUploadNotifySerializer, PlatformCategoryRateSerializer
from mercado import tasks
from report.models import ProductReport


class DefaultPagination(PageNumberPagination):
    """
    分页设置
    """
    page_size = 20
    page_size_query_param = 'page_size'
    page_query_param = 'page'
    max_page_size = 1000


class ListingViewSet(mixins.ListModelMixin, mixins.CreateModelMixin,
                     mixins.UpdateModelMixin, mixins.DestroyModelMixin,
                     mixins.RetrieveModelMixin, viewsets.GenericViewSet):
    """
    list:
        在线产品列表,分页,过滤,搜索,排序
    create:
        在线产品新增
    retrieve:
        在线产品详情页
    update:
        在线产品修改
    destroy:
        在线产品删除
    """
    queryset = Listing.objects.all()
    serializer_class = ListingSerializer  # 序列化
    pagination_class = DefaultPagination  # 分页

    filter_backends = (DjangoFilterBackend, filters.SearchFilter,
                       filters.OrderingFilter)  # 过滤,搜索,排序
    filter_fields = ('site_id', 'listing_status', 'is_cbt', 'seller_id',
                     'seller_name', 'brand', 'collection')  # 配置过滤字段
    search_fields = ('item_id', 'title', 'seller_id', 'seller_name', 'brand'
                     )  # 配置搜索字段
    ordering_fields = ('create_time', 'total_sold', 'reviews', 'stock_num',
                       'profit', 'update_time')  # 配置排序字段

    # test
    @action(methods=['get'], detail=False, url_path='test')
    def test(self, request):

        set = FBMWarehouse.objects.all()
        for i in set:
            if i.w_code == 'MXRC02':
                i.zip = '54616'
                i.save()
            if i.w_code == 'MXRCO1 RC':
                i.zip = '54803'
                i.save()
        fbm = FBMWarehouse.objects.filter(
            w_code='Traslada la etiqueta').first()
        if not fbm:
            wh = FBMWarehouse()
            wh.country = '墨西哥MX'
            wh.w_code = 'Traslada la etiqueta'
            wh.address = 'SANSTAR'
            wh.platform = 'MERCADO'
            wh.name = '盛德中转仓(墨西哥)'
            wh.save()
        return Response({'batch_list': 'ok'}, status=status.HTTP_200_OK)

    # 添加商品链接
    @action(methods=['post'], detail=False, url_path='create_listing')
    def create_listing(self, request):
        item_id = request.data['item_id']
        is_exist = Listing.objects.filter(item_id=item_id).count()
        if is_exist:
            return Response({'msg': '商品已存在'},
                            status=status.HTTP_406_NOT_ACCEPTABLE)
        tasks.create_listing(item_id)

        return Response({'msg': '操作成功'}, status=status.HTTP_200_OK)

    # 计算利润
    @action(methods=['post'], detail=False, url_path='calc_profit')
    def calc_profit(self, request):
        id = request.data['id']
        cost = request.data['cost']
        listing = Listing.objects.filter(id=id).first()
        if not listing:
            return Response({'msg': '商品不存在'},
                            status=status.HTTP_406_NOT_ACCEPTABLE)

        mercado_fee = listing.price * 0.305 if listing.price >= 299 else listing.price * 0.305 + 25
        shipping_fee = 57 if listing.price >= 299 else 79.8
        if not listing.is_free_shipping:
            shipping_fee = 0
        profit = (
            (listing.price - mercado_fee - shipping_fee) / 3.1 - cost) * 0.99

        from xToolkit import xstring
        listing.profit = xstring.dispose(profit).humanized_amount(compel=True)
        listing.cost = cost
        listing.save()

        return Response({'msg': '操作成功'}, status=status.HTTP_200_OK)


class ListingTrackViewSet(mixins.ListModelMixin, mixins.CreateModelMixin,
                          mixins.UpdateModelMixin, mixins.DestroyModelMixin,
                          mixins.RetrieveModelMixin, viewsets.GenericViewSet):
    """
    list:
        商品跟踪列表,分页,过滤,搜索,排序
    create:
        商品跟踪新增
    retrieve:
        商品跟踪详情页
    update:
        商品跟踪修改
    destroy:
        商品跟踪删除
    """
    queryset = ListingTrack.objects.all()
    serializer_class = ListingTrackSerializer  # 序列化

    filter_backends = (DjangoFilterBackend, filters.SearchFilter,
                       filters.OrderingFilter)  # 过滤,搜索,排序
    # filter_fields = ('listing__id', 'health')  # 配置过滤字段
    filterset_fields = {
        'create_time': ['gte', 'lte', 'exact', 'gt', 'lt'],
        'listing__id': ['exact']
    }
    ordering_fields = ('create_time', )  # 配置排序字段


class CategoriesViewSet(mixins.ListModelMixin, mixins.CreateModelMixin,
                        mixins.UpdateModelMixin, mixins.DestroyModelMixin,
                        mixins.RetrieveModelMixin, viewsets.GenericViewSet):
    """
    list:
        站点类目列表,分页,过滤,搜索,排序
    create:
        站点类目新增
    retrieve:
        站点类目详情页
    update:
        站点类目修改
    destroy:
        站点类目删除
    """
    queryset = Categories.objects.all()
    serializer_class = CategoriesSerializer  # 序列化

    filter_backends = (DjangoFilterBackend, filters.SearchFilter,
                       filters.OrderingFilter)  # 过滤,搜索,排序
    filter_fields = ('father_id', 'site_id', 'collection', 'has_children'
                     )  # 配置过滤字段
    search_fields = ('categ_id', 'name', 't_name')  # 配置搜索字段
    ordering_fields = ('name', 'update_time')  # 配置排序字段

    # 获取类目
    @action(methods=['get'], detail=False, url_path='get_categories')
    def get_categories(self, request):
        father_id = self.request.query_params.get('father_id')
        site_id = self.request.query_params.get('site_id')
        if father_id != '0':
            is_exist = Categories.objects.filter(father_id=father_id,
                                                 site_id=site_id).count()
            if not is_exist:
                # 如果类目不在数据库里，则发起请求查询并保存
                at = ApiSetting.objects.all().first()
                token = at.access_token
                headers = {
                    'Authorization': 'Bearer ' + token,
                }
                c_url = 'https://api.mercadolibre.com/categories/' + father_id
                resp = requests.get(c_url, headers=headers)
                if resp.status_code == 200:
                    data = resp.json()
                    add_list = []
                    path_from_root = ''
                    for p in data['path_from_root']:
                        path_from_root += p['name'] + ' > '

                    if len(data['children_categories']) == 0:
                        cate = Categories.objects.filter(
                            categ_id=father_id, site_id=site_id).first()
                        cate.has_children = False
                        cate.save()
                    for item in data['children_categories']:
                        result = tasks.translate(item['name'])
                        add_list.append(
                            Categories(categ_id=item['id'],
                                       father_id=father_id,
                                       site_id=site_id,
                                       name=item['name'],
                                       t_name=result if result else None,
                                       path_from_root=path_from_root +
                                       item['name'],
                                       total_items=item[
                                           'total_items_in_this_category'],
                                       update_time=datetime.now()))
                    Categories.objects.bulk_create(add_list)
        else:
            # 如果一级目录不存在，则请求数据
            is_exist = Categories.objects.filter(father_id=father_id,
                                                 site_id=site_id).count()
            if not is_exist:
                tasks.update_categories(site_id)

        queryset = Categories.objects.filter(father_id=father_id,
                                             site_id=site_id)
        data_list = []
        for i in queryset:
            data_list.append({
                'id': i.id,
                'categ_id': i.categ_id,
                'father_id': i.father_id,
                'site_id': i.site_id,
                'name': i.name,
                't_name': i.t_name,
                'path_from_root': i.path_from_root,
                'total_items': i.total_items,
                'has_children': i.has_children,
                'collection': i.collection,
                'update_time': i.update_time,
            })
        return Response(data_list, status=status.HTTP_200_OK)

    # 获取热搜关键词
    @action(methods=['get'], detail=False, url_path='get_trends')
    def get_trends(self, request):
        categ_id = self.request.query_params.get('categ_id')
        site_id = self.request.query_params.get('site_id')
        # 如果当天有数据
        today = datetime.now().date()
        queryset = Keywords.objects.filter(update_time__date=today,
                                           categ_id=categ_id)
        if queryset:
            query_list = []
            for i in queryset:
                query_list.append({
                    'id': i.id,
                    'categ_id': i.categ_id,
                    'keyword': i.keyword,
                    't_keyword': i.t_keyword,
                    'url': i.url,
                    'rank': i.rank,
                    'status': i.status,
                    'rank_changed': i.rank_changed,
                    'update_time': i.update_time,
                })
            return Response(query_list, status=status.HTTP_200_OK)

        at = ApiSetting.objects.all().first()
        token = at.access_token
        headers = {
            'Authorization': 'Bearer ' + token,
        }
        url = 'https://api.mercadolibre.com/trends/' + site_id + '/' + categ_id
        # 如果目录id是整站, 则请求整站数据
        if categ_id in ['MLM', 'MLC', 'MLB']:
            url = 'https://api.mercadolibre.com/trends/' + site_id
        resp = requests.get(url, headers=headers)
        query_list = []
        if resp.status_code == 200:
            data = resp.json()
            n = 1
            if data:
                add_list = []
                for i in data:
                    kw = Keywords.objects.filter(keyword=i['keyword'],
                                                 categ_id=categ_id).first()
                    rank_status = 'NEW'
                    rank_changed = 0
                    if kw:
                        if n == kw.rank:
                            rank_status = 'EQ'
                        elif n < kw.rank:
                            rank_status = 'UP'
                            rank_changed = kw.rank - n
                        elif n > kw.rank:
                            rank_status = 'DOWN'
                            rank_changed = n - kw.rank
                        else:
                            rank_status = None
                        t_keyword = kw.t_keyword
                    else:
                        t_keyword = tasks.translate(i['keyword'])
                    add_list.append(
                        Keywords(categ_id=categ_id,
                                 keyword=i['keyword'],
                                 t_keyword=t_keyword,
                                 url=i['url'],
                                 rank=n,
                                 status=rank_status,
                                 rank_changed=rank_changed))
                    n += 1
                Keywords.objects.filter(categ_id=categ_id).delete()
                Keywords.objects.bulk_create(add_list)

                keywords_data = Keywords.objects.filter(categ_id=categ_id)
                if keywords_data:
                    for i in keywords_data:
                        query_list.append({
                            'id': i.id,
                            'categ_id': i.categ_id,
                            'keyword': i.keyword,
                            't_keyword': i.t_keyword,
                            'url': i.url,
                            'rank': i.rank,
                            'status': i.status,
                            'rank_changed': i.rank_changed,
                            'update_time': i.update_time,
                        })

        return Response(query_list, status=status.HTTP_200_OK)

    # 翻译类目
    @action(methods=['get'], detail=False, url_path='translate')
    def translate(self, request):
        queryset = Categories.objects.filter(t_name=None)
        for i in queryset:
            result = tasks.translate(i.name)
            i.t_name = result
            i.save()

        return Response({'data': 'OK'}, status=status.HTTP_200_OK)


class SellerViewSet(mixins.ListModelMixin, mixins.CreateModelMixin,
                    mixins.UpdateModelMixin, mixins.DestroyModelMixin,
                    mixins.RetrieveModelMixin, viewsets.GenericViewSet):
    """
    list:
        卖家列表,分页,过滤,搜索,排序
    create:
        卖家新增
    retrieve:
        卖家详情页
    update:
        卖家修改
    destroy:
        卖家删除
    """
    queryset = Seller.objects.all()
    serializer_class = SellerSerializer  # 序列化
    pagination_class = DefaultPagination  # 分页

    filter_backends = (DjangoFilterBackend, filters.SearchFilter,
                       filters.OrderingFilter)  # 过滤,搜索,排序
    filter_fields = ('site_id', 'level_id', 'collection')  # 配置过滤字段
    search_fields = ('nickname', 'seller_id')  # 配置搜索字段
    ordering_fields = ('registration_date', 'total')  # 配置排序字段

    # 更新卖家详细信息
    @action(methods=['get'], detail=False, url_path='update_seller')
    def update_seller(self, request):
        seller_id = self.request.query_params.get('seller_id')

        seller = Seller.objects.filter(seller_id=seller_id).first()
        tasks.create_or_update_seller(seller.site_id, seller_id)
        return Response({'msg': '操作成功'}, status=status.HTTP_200_OK)

    # 新增卖家，按昵称
    @action(methods=['get'], detail=False, url_path='create_seller')
    def create_seller(self, request):
        seller_info = self.request.query_params.get('seller_info')
        site_id = self.request.query_params.get('site_id')

        at = ApiSetting.objects.all().first()
        token = at.access_token
        headers = {
            'Authorization': 'Bearer ' + token,
        }

        url = 'https://api.mercadolibre.com/sites/' + site_id + '/search?nickname=' + seller_info
        resp = requests.get(url, headers=headers)
        if resp.status_code == 200:
            data = resp.json()
            if 'seller' in data:
                seller_id = str(data['seller']['id'])
                tasks.create_or_update_seller(site_id, seller_id)
                return Response({'msg': '操作成功'}, status=status.HTTP_200_OK)
            else:
                return Response({'msg': '卖家不存在，请检查卖家名称'},
                                status=status.HTTP_406_NOT_ACCEPTABLE)
        return Response({'msg': 'api操作不成功'},
                        status=status.HTTP_406_NOT_ACCEPTABLE)


class SellerTrackViewSet(mixins.ListModelMixin, mixins.CreateModelMixin,
                         mixins.UpdateModelMixin, mixins.DestroyModelMixin,
                         mixins.RetrieveModelMixin, viewsets.GenericViewSet):
    """
    list:
        卖家跟踪列表,分页,过滤,搜索,排序
    create:
        卖家跟踪新增
    retrieve:
        卖家跟踪详情页
    update:
        卖家跟踪修改
    destroy:
        卖家跟踪删除
    """
    queryset = SellerTrack.objects.all()
    serializer_class = SellerTrackSerializer  # 序列化

    filter_backends = (DjangoFilterBackend, filters.SearchFilter,
                       filters.OrderingFilter)  # 过滤,搜索,排序
    # filter_fields = ('listing__id', 'health')  # 配置过滤字段
    filterset_fields = {
        'create_time': ['gte', 'lte', 'exact', 'gt', 'lt'],
        'seller__id': ['exact']
    }
    ordering_fields = ('create_time', )  # 配置排序字段


class MLProductViewSet(mixins.ListModelMixin, mixins.CreateModelMixin,
                       mixins.UpdateModelMixin, mixins.DestroyModelMixin,
                       mixins.RetrieveModelMixin, viewsets.GenericViewSet):
    """
    list:
        ML产品列表,分页,过滤,搜索,排序
    create:
        ML产品新增
    retrieve:
        ML产品详情页
    update:
        ML产品修改
    destroy:
        ML产品删除
    """
    queryset = MLProduct.objects.all()
    serializer_class = MLProductSerializer  # 序列化
    pagination_class = DefaultPagination  # 分页

    filter_backends = (DjangoFilterBackend, filters.SearchFilter,
                       filters.OrderingFilter)  # 过滤,搜索,排序
    filter_fields = ('p_status', 'site', 'shop', 'is_checked', 'user_id'
                     )  # 配置过滤字段
    search_fields = ('sku', 'p_name', 'label_code', 'upc', 'item_id')  # 配置搜索字段
    ordering_fields = ('create_time', 'sku', 'item_id')  # 配置排序字段

    # ML产品批量上传
    @action(methods=['post'], detail=False, url_path='bulk_upload')
    def bulk_upload(self, request):
        import warnings
        warnings.filterwarnings('ignore')

        data = request.data
        wb = openpyxl.load_workbook(data['excel'])
        sheet = wb['上传模板']

        add_list = []
        if sheet.max_row <= 1:
            return Response({
                'msg': '表格不能为空',
                'status': 'error'
            },
                            status=status.HTTP_202_ACCEPTED)
        if sheet['V1'].value != '是否带电 1/0':
            return Response({
                'msg': '表格有误，请下载最新模板',
                'status': 'error'
            },
                            status=status.HTTP_202_ACCEPTED)

        for cell_row in list(sheet)[1:]:
            # 检查1，2列是否为空，空则跳过
            row_status = cell_row[0].value and cell_row[1].value
            if not row_status:
                continue

            # 检查型号是否已存在
            is_exist = MLProduct.objects.filter(
                sku=cell_row[0].value.strip()).count()
            if is_exist:
                continue
            # 检查店铺是否存在
            sp = Shop.objects.filter(name=cell_row[6].value.strip()).first()
            if not sp:
                continue
            platform = sp.platform
            # 检查对应平台upc列是否强制
            if platform == 'MERCADO':
                if not cell_row[2].value:
                    continue
            sku = cell_row[0].value
            p_name = cell_row[1].value
            upc = cell_row[2].value
            # 自动申请upc
            if upc == 'AUTO_CREATE':
                upc_num = UPC.objects.filter(is_used=False).first()
                if upc_num:
                    upc_num.is_used = True
                    upc_num.user = request.user
                    upc_num.use_time = datetime.now()
                    upc_num.save()

                    upc = upc_num.number
            item_id = cell_row[3].value
            label_code = cell_row[4].value
            site = cell_row[5].value
            shop = cell_row[6].value
            unit_cost = cell_row[7].value
            if not type(unit_cost) in [float, int]:
                unit_cost = 0
            weight = cell_row[8].value
            if not type(weight) in [float, int]:
                weight = 0
            length = cell_row[9].value
            if not type(length) in [float, int]:
                length = 0
            width = cell_row[10].value
            if not type(width) in [float, int]:
                width = 0
            heigth = cell_row[11].value
            if not type(heigth) in [float, int]:
                heigth = 0
            first_ship_cost = cell_row[12].value
            if not type(first_ship_cost) in [float, int]:
                first_ship_cost = 0
            custom_code = cell_row[13].value
            cn_name = cell_row[14].value
            en_name = cell_row[15].value
            brand = cell_row[16].value
            declared_value = cell_row[17].value
            if not type(declared_value) in [float, int]:
                declared_value = 0
            cn_material = cell_row[18].value
            en_material = cell_row[19].value
            use = cell_row[20].value
            is_elec = cell_row[21].value
            if is_elec == 1:
                is_elec = True
            else:
                is_elec = False
            is_water = cell_row[22].value
            if is_water == 1:
                is_water = True
            else:
                is_water = False
            buy_url = cell_row[23].value
            sale_url = cell_row[24].value
            refer_url = cell_row[25].value

            add_list.append(
                MLProduct(
                    sku=sku,
                    p_name=p_name,
                    platform=platform,
                    upc=upc,
                    item_id=item_id,
                    label_code=label_code,
                    site=site,
                    shop=shop,
                    unit_cost=unit_cost,
                    weight=weight,
                    length=length,
                    width=width,
                    heigth=heigth,
                    first_ship_cost=first_ship_cost,
                    custom_code=custom_code,
                    cn_name=cn_name,
                    en_name=en_name,
                    brand=brand,
                    declared_value=declared_value,
                    cn_material=cn_material,
                    en_material=en_material,
                    use=use,
                    is_elec=is_elec,
                    is_water=is_water,
                    buy_url=buy_url,
                    sale_url=sale_url,
                    refer_url=refer_url,
                    user_id=request.user.id,
                ))
        MLProduct.objects.bulk_create(add_list)

        # 创建操作日志
        log = MLOperateLog()
        log.op_module = 'PRODUCT'
        log.op_type = 'CREATE'
        log.desc = '导入新产品 {name}个'.format(name=len(add_list))
        log.user = request.user
        log.save()

        return Response({'msg': '成功上传'}, status=status.HTTP_200_OK)

    # ML产品图片上传
    @action(methods=['post'], detail=False, url_path='image_upload')
    def image_upload(self, request):
        from PIL import Image
        data = request.data
        product = MLProduct.objects.filter(id=data['id']).first()
        if not product:
            return Response({'msg': '产品不存在'}, status=status.HTTP_202_ACCEPTED)

        path = 'media/ml_product/' + product.sku + '.jpg'
        pic = data['pic']
        content = pic.chunks()
        with open(path, 'wb') as f:
            for i in content:
                f.write(i)
        product.image = 'ml_product/' + product.sku + '.jpg'
        product.save()

        pic_org = Image.open(path)
        # 修改保存图片大小
        pic_ori_new = pic_org.resize((800, 800), Image.ANTIALIAS)
        pic_ori_new.save(path)
        # 增加小图
        pic_new = pic_org.resize((100, 100), Image.ANTIALIAS)
        pic_new.save('media/ml_product/' + product.sku + '_100x100.jpg')

        # 创建操作日志
        log = MLOperateLog()
        log.op_module = 'PRODUCT'
        log.op_type = 'EDIT'
        log.target_type = 'PRODUCT'
        log.target_id = product.id
        log.desc = '上传修改图片'
        log.user = request.user
        log.save()

        return Response({'msg': '成功上传'}, status=status.HTTP_200_OK)

    # 查询ML产品id
    @action(methods=['post'], detail=False, url_path='get_product_id')
    def get_product_id(self, request):
        sku = request.data['sku']
        product = MLProduct.objects.filter(sku=sku).first()
        pid = 0
        if product:
            pid = product.id

        return Response({'id': pid}, status=status.HTTP_200_OK)

    # 复制产品
    @action(methods=['post'], detail=False, url_path='cp_product')
    def cp_product(self, request):
        old_id = request.data['old_id']
        new_sku = request.data['new_sku']
        is_exist = MLProduct.objects.filter(sku=new_sku).first()
        if is_exist:
            return Response({'msg': 'SKU已存在'}, status=status.HTTP_202_ACCEPTED)

        product = MLProduct.objects.filter(id=old_id).first()
        if product:
            new_product = MLProduct(
                sku=new_sku,
                p_name=product.p_name,
                image=product.image,
                platform=product.platform,
                upc=product.upc,
                label_code=product.label_code,
                site=product.site,
                shop=product.shop,
                unit_cost=product.unit_cost,
                weight=product.weight,
                length=product.length,
                width=product.width,
                heigth=product.heigth,
                first_ship_cost=product.first_ship_cost,
                custom_code=product.custom_code,
                cn_name=product.cn_name,
                en_name=product.en_name,
                brand=product.brand,
                declared_value=product.declared_value,
                cn_material=product.cn_material,
                en_material=product.en_material,
                use=product.use,
                is_elec=product.is_elec,
                is_water=product.is_water,
                buy_url=product.buy_url,
                sale_url=product.sale_url,
                refer_url=product.refer_url,
                user_id=request.user.id,
            )
            new_product.save()
            # 创建操作日志
            log = MLOperateLog()
            log.op_module = 'PRODUCT'
            log.op_type = 'EDIT'
            log.target_type = 'PRODUCT'
            log.target_id = new_product.id
            log.desc = '复制产品创建'
            log.user = request.user
            log.save()
        else:
            return Response({'msg': '源产品不存在'}, status=status.HTTP_202_ACCEPTED)

        return Response({
            'msg': '产品复制成功',
            'status': 'success'
        },
                        status=status.HTTP_200_OK)

    # 产品标签生成(upc 标签)
    @action(methods=['post'], detail=False, url_path='create_label')
    def create_label(self, request):
        from fpdf import FPDF
        import barcode
        data = request.data['products']

        # 检查是否有upc
        for item in data:
            sku = item['sku']
            product = MLProduct.objects.filter(sku=sku).first()
            if not product.upc:
                return Response({
                    'msg': '产品没有upc码, 无法打印标签',
                    'status': 'error'
                },
                                status=status.HTTP_202_ACCEPTED)

        path = 'media/label/'
        # pdf文件信息
        pdf = FPDF('P', 'mm', (60, 40))
        pdf.set_font('Helvetica', size=12)
        pdf.set_margin(0.5)

        # 条码图片信息
        options = {
            'module_height': 4,  # 默认值15.0，条码高度，单位为毫米
            'module_width': 0.13,  # 默认值0.2，每个条码宽度（？），单位为毫米
            'quiet_zone': 2,  # 默认值6.5，两端空白宽度，单位为毫米
            'font_size': 3,  # 默认值10，文本字体大小，单位为磅
            'text_distance': 1,  # 默认值5.0，文本和条码之间的距离，单位为毫米
            'dpi': 600,  # 图片分辨率
        }
        if data:
            file_name = 'product_label'
            for item in data:
                sku = item['sku']
                qty = item['qty']
                product = MLProduct.objects.filter(sku=sku).first()
                if len(data) == 1:
                    file_name = sku

                # 生成条码图片png
                barcode.generate(
                    'code128',
                    product.upc,
                    writer=barcode.writer.ImageWriter(),
                    output=path + product.sku,
                    writer_options=options,
                )
                for i in range(qty):
                    pdf.add_page()
                    # pdf.image("media/sys/logo.png", x=17, y=2, w=25)
                    pdf.cell(0, 20, align='C', txt=product.sku)
                    pdf.image(path + product.sku + '.png', x=-5, y=13, w=70)
            output_name = path + file_name + '.pdf'
            pdf.output(output_name)
            url = BASE_URL + '/' + output_name
        return Response({'url': url}, status=status.HTTP_200_OK)

    #  重写产品删除
    def destroy(self, request, *args, **kwargs):
        product = self.get_object()

        is_ship_exist = ShipDetail.objects.filter(sku=product.sku).count()
        is_stock_exist = ShopStock.objects.filter(sku=product.sku).count()
        is_purchase_exist = PurchaseManage.objects.filter(
            sku=product.sku).count()
        if is_ship_exist or is_stock_exist or is_purchase_exist:
            return Response({
                'msg': '产品已被引用，无法删除',
                'status': 'error'
            },
                            status=status.HTTP_202_ACCEPTED)

        # 创建操作日志
        log = MLOperateLog()
        log.op_module = 'PRODUCT'
        log.op_type = 'DEL'
        log.target_type = 'PRODUCT'
        log.target_id = product.id
        log.desc = '删除产品 {sku} {p_name}'.format(sku=product.sku,
                                                p_name=product.p_name)
        log.user = request.user
        log.save()

        product.delete()

        return Response({
            'msg': '操作成功!',
            'status': 'success'
        },
                        status=status.HTTP_200_OK)


class PackingViewSet(mixins.ListModelMixin, mixins.CreateModelMixin,
                     mixins.UpdateModelMixin, mixins.DestroyModelMixin,
                     mixins.RetrieveModelMixin, viewsets.GenericViewSet):
    """
    list:
        包材管理列表,分页,过滤,搜索,排序
    create:
        包材管理新增
    retrieve:
        包材管理详情页
    update:
        包材管理修改
    destroy:
        包材管理删除
    """
    queryset = Packing.objects.all()
    serializer_class = PackingSerializer  # 序列化
    pagination_class = DefaultPagination  # 分页

    filter_backends = (DjangoFilterBackend, filters.SearchFilter,
                       filters.OrderingFilter)  # 过滤,搜索,排序
    filter_fields = ('name', 'size')  # 配置过滤字段
    search_fields = ('name', )  # 配置搜索字段
    ordering_fields = ('create_time', )  # 配置排序字段

    # 重写
    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)

        # 创建操作日志
        log = MLOperateLog()
        log.op_module = 'PRODUCT'
        log.op_type = 'CREATE'
        log.target_type = 'PACKING'
        log.desc = '新增包材 {name}'.format(name=request.data['name'])
        log.user = request.user
        log.save()
        return Response(serializer.data,
                        status=status.HTTP_201_CREATED,
                        headers=headers)

    # 重写
    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance,
                                         data=request.data,
                                         partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)

        if getattr(instance, '_prefetched_objects_cache', None):
            # If 'prefetch_related' has been applied to a queryset, we need to
            # forcibly invalidate the prefetch cache on the instance.
            instance._prefetched_objects_cache = {}

        # 创建操作日志
        log = MLOperateLog()
        log.op_module = 'PRODUCT'
        log.op_type = 'EDIT'
        log.target_type = 'PACKING'
        log.target_id = serializer.data['id']
        log.desc = '修改包材 {name}'.format(name=serializer.data['name'])
        log.user = request.user
        log.save()
        return Response(serializer.data)

    # 重写
    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()

        # 创建操作日志
        log = MLOperateLog()
        log.op_module = 'PRODUCT'
        log.op_type = 'DEL'
        log.target_type = 'PACKING'
        log.target_id = instance.id
        log.desc = '删除包材 {name}'.format(name=instance.name)
        log.user = request.user
        log.save()

        self.perform_destroy(instance)
        return Response(status=status.HTTP_204_NO_CONTENT)


class ShopViewSet(mixins.ListModelMixin, mixins.CreateModelMixin,
                  mixins.UpdateModelMixin, mixins.DestroyModelMixin,
                  mixins.RetrieveModelMixin, viewsets.GenericViewSet):
    """
    list:
        FBM店铺列表,分页,过滤,搜索,排序
    create:
        FBM店铺新增
    retrieve:
        FBM店铺详情页
    update:
        FBM店铺修改
    destroy:
        FBM店铺删除
    """
    queryset = Shop.objects.all()
    serializer_class = ShopSerializer  # 序列化
    pagination_class = DefaultPagination  # 分页

    filter_backends = (DjangoFilterBackend, filters.SearchFilter,
                       filters.OrderingFilter)  # 过滤,搜索,排序
    filter_fields = ('warehouse_type', 'platform', 'shop_type', 'site',
                     'is_active', 'user')  # 配置过滤字段
    search_fields = ('name', 'seller_id', 'nickname')  # 配置搜索字段
    ordering_fields = ('create_time', 'total_profit', 'total_weight')  # 配置排序字段

    def get_queryset(self):
        # 返回已启用的店铺
        return Shop.objects.filter(is_active=True)

    # 代办事项
    @action(methods=['get'], detail=False, url_path='get_daiban')
    def get_daiban(self, request):
        # 运单管理
        overtime_ship = 0  # 入仓核对数量
        need_book = 0  # 需预约运单数量

        income_confirm = 0  # 提现待确认数量
        if request.user.is_superuser:
            pre_qty = Ship.objects.filter(s_status='PREPARING').count()
            shipped_qty = Ship.objects.filter(s_status='SHIPPED').count()
            booked_qty = Ship.objects.filter(s_status='BOOKED').count()

            ships = Ship.objects.filter(s_status='BOOKED')
            need_book = Ship.objects.filter(s_status='SHIPPED',
                                            target='FBM').count()

            income_confirm = Finance.objects.filter(f_type='WD',
                                                    is_received=False).count()
        else:
            pre_qty = Ship.objects.filter(s_status='PREPARING').filter(
                Q(user_id=request.user.id) | Q(user_id=0)).count()
            shipped_qty = Ship.objects.filter(s_status='SHIPPED').filter(
                Q(user_id=request.user.id) | Q(user_id=0)).count()
            booked_qty = Ship.objects.filter(s_status='BOOKED').filter(
                Q(user_id=request.user.id) | Q(user_id=0)).count()

            ships = Ship.objects.filter(s_status='BOOKED',
                                        user_id=request.user.id)
            need_book = Ship.objects.filter(s_status='SHIPPED',
                                            target='FBM',
                                            user_id=request.user.id).count()

            income_confirm = Finance.objects.filter(
                f_type='WD', is_received=False,
                shop__user_id=request.user.id).count()
        for i in ships:
            if i.book_date:
                ad = str(i.book_date)
                dd = datetime.strptime(ad, '%Y-%m-%d')
                delta = datetime.now() - dd
                if delta.days > 0:
                    overtime_ship += 1

        # 采购管理
        wait_buy_num = PurchaseManage.objects.filter(p_status='WAITBUY',
                                                     buy_qty__gt=0).count()
        purchased_num = PurchaseManage.objects.filter(
            p_status='PURCHASED').count()
        rec_num = PurchaseManage.objects.filter(p_status='RECEIVED').count()
        pack_num = PurchaseManage.objects.filter(p_status='PACKED').count()

        # 检查产品完整状态
        all_product_incomplete = False
        products = MLProduct.objects.filter(user_id=request.user.id).exclude(
            p_status='OFFLINE')
        for i in products:

            if not (i.site and i.unit_cost and i.image and i.shop
                    and i.item_id):
                all_product_incomplete = True
                break
            if not (i.custom_code and i.cn_name and i.en_name and i.brand
                    and i.declared_value and i.cn_material and i.en_material):
                all_product_incomplete = True
                break
            if not (i.use and i.weight and i.length and i.width and i.heigth
                    and i.first_ship_cost):
                all_product_incomplete = True
                break

        # 变动清单待处理产品
        shops_set = Shop.objects.filter(user__id=request.user.id)
        remove_items_count = 0
        if request.user.is_superuser:
            remove_items_count = ShipItemRemove.objects.filter(
                handle=0).count()
        else:
            for i in shops_set:
                remove_items_count += ShipItemRemove.objects.filter(
                    handle=0).filter(belong_shop=i.name).count()

        return Response(
            {
                'pre_qty': pre_qty,
                'shipped_qty': shipped_qty,
                'booked_qty': booked_qty,
                'wait_buy_num': wait_buy_num,
                'purchased_num': purchased_num,
                'rec_num': rec_num,
                'pack_num': pack_num,
                'overtime_ship': overtime_ship,
                'need_book': need_book,
                'income_confirm': income_confirm,
                'all_product_incomplete': all_product_incomplete,
                'remove_items_count': remove_items_count
            },
            status=status.HTTP_200_OK)

    # 店铺信息
    @action(methods=['post'], detail=False, url_path='shop_info')
    def shop_info(self, request):
        shop_id = request.data['id']

        shop = Shop.objects.filter(id=shop_id).first()
        manager = shop.user.first_name if shop.user else '管理员'  # 负责人

        ss = ShopStock.objects.filter(shop=shop)
        total_sku = 0  # 在售产品
        total_amount = 0  # 库存金额
        total_qty = 0  # 库存数量
        out_stock = 0  # 缺货产品
        for i in ss:
            if i.p_status != 'OFFLINE' and i.qty > 0:
                total_sku += 1
                total_qty += i.qty
                total_amount += (i.unit_cost + i.first_ship_cost) * i.qty
            if i.p_status != 'OFFLINE' and i.qty <= 0:
                out_stock += 1

        used_quota = tasks.get_shop_quota(shop_id)  # 获取店铺已用额度
        return Response(
            {
                'manager': manager,
                'total_sku': total_sku,
                'total_qty': total_qty,
                'total_amount': total_amount,
                'quota': shop.quota,
                'out_stock': out_stock,
                'used_quota': used_quota
            },
            status=status.HTTP_200_OK)

    # 店铺收支管理
    @action(methods=['post'], detail=False, url_path='shop_finance')
    def shop_finance(self, request):
        start_date = request.data['start_date']
        end_date = request.data['end_date']

        shops = Shop.objects.filter(warehouse_type='FBM', is_active=True)
        data = []
        for s in shops:
            # 直发目标店铺的支出统计(不统计中转入仓运单)
            total_pay = 0  # 总支出
            ships = Ship.objects.filter(shop=s.name, send_from='CN').filter(
                Q(s_status='SHIPPED') | Q(s_status='BOOKED')
                | Q(s_status='FINISHED')).filter(
                    sent_time__date__gte=start_date,
                    sent_time__date__lte=end_date)

            for i in ships:
                total_pay += i.shipping_fee
                total_pay += i.extra_fee
                total_pay += i.products_cost
            # 中转运单的支出统计
            ships = Ship.objects.filter(target='TRANSIT').filter(
                Q(s_status='SHIPPED') | Q(s_status='BOOKED')
                | Q(s_status='FINISHED')).filter(
                    sent_time__date__gte=start_date,
                    sent_time__date__lte=end_date)
            for i in ships:
                sd = ShipDetail.objects.filter(ship=i)
                for item in sd:
                    # 统计中转运单中该店铺产品部分
                    if item.target_FBM == s.name:
                        total_pay += (item.unit_cost +
                                      item.avg_ship_fee) * item.qty
            # 店铺费用支出统计
            shop_fees = Finance.objects.filter(shop=s, f_type='FEE').filter(
                wd_date__gte=start_date, wd_date__lte=end_date)
            for i in shop_fees:
                total_pay += i.income

            finance = Finance.objects.filter(shop=s, f_type='EXC').filter(
                exc_date__gte=start_date, exc_date__lte=end_date)
            total_rec = 0  # 总收入
            for i in finance:
                total_rec += i.income_rmb
            data.append({
                'shop_id': s.id,
                'shop_name': s.name,
                'total_pay': total_pay,
                'total_rec': total_rec,
                'profit': total_rec - total_pay
            })

        return Response({'data': data}, status=status.HTTP_200_OK)

    # 导出店铺收支明细
    @action(methods=['post'], detail=False, url_path='export_shop_finance')
    def export_shop_finance(self, request):
        import openpyxl
        from decimal import Decimal
        start_date = request.data['start_date']
        end_date = request.data['end_date']
        shop_id = request.data['shop_id']
        shop = Shop.objects.filter(id=shop_id).first()

        wb = openpyxl.Workbook()

        total_pay = 0  # 总支出

        f_sheet = wb.create_sheet('支出明细')
        f_sheet['A1'] = '发货日期'
        f_sheet['B1'] = '批次号'
        f_sheet['C1'] = '目标店铺'
        f_sheet['D1'] = '类型'
        f_sheet['E1'] = '货品成本'
        f_sheet['F1'] = '头程运费'
        f_sheet['G1'] = '杂费'
        f_sheet['H1'] = '小计'

        ships = Ship.objects.filter(send_from='CN').filter(
            Q(s_status='SHIPPED') | Q(s_status='BOOKED')
            | Q(s_status='FINISHED')).filter(sent_time__date__gte=start_date,
                                             sent_time__date__lte=end_date)
        num = 2
        for i in ships:
            # 直发目标店铺的支出统计(不含中转入仓运单)
            if i.shop == shop.name:
                total_pay += i.shipping_fee
                total_pay += i.extra_fee
                total_pay += i.products_cost
                subtotal = i.shipping_fee + i.extra_fee + i.products_cost

                f_sheet['A' + str(num)] = i.sent_time.strftime('%Y-%m-%d')
                f_sheet['B' + str(num)] = i.batch
                f_sheet['C' + str(num)] = shop.name
                f_sheet['D' + str(num)] = 'FBM直发'
                f_sheet['E' + str(num)] = Decimal(i.products_cost).quantize(
                    Decimal("0.00"))
                if i.shipping_fee:
                    f_sheet['F' + str(num)] = Decimal(i.shipping_fee).quantize(
                        Decimal("0.00"))
                if i.extra_fee:
                    f_sheet['G' + str(num)] = Decimal(i.extra_fee).quantize(
                        Decimal("0.00"))
                f_sheet['H' + str(num)] = subtotal
                num += 1

            # 中转运单的支出统计
            if i.target == 'TRANSIT':
                product_cost = 0  # 货品成本
                shipping_fee = 0  # 运费（含杂费）
                sd = ShipDetail.objects.filter(ship=i)
                if sd:
                    for item in sd:
                        # 统计中转运单中该店铺产品部分
                        if item.target_FBM == shop.name:
                            product_cost += item.unit_cost * item.qty
                            shipping_fee += item.avg_ship_fee * item.qty
                            total_pay += (item.unit_cost +
                                          item.avg_ship_fee) * item.qty
                    f_sheet['A' + str(num)] = i.sent_time.strftime('%Y-%m-%d')
                    f_sheet['B' + str(num)] = i.batch
                    f_sheet['C' + str(num)] = shop.name
                    f_sheet['D' + str(num)] = '中转'
                    f_sheet['E' + str(num)] = Decimal(product_cost).quantize(
                        Decimal("0.00"))
                    if shipping_fee:
                        f_sheet['F' +
                                str(num)] = Decimal(shipping_fee).quantize(
                                    Decimal("0.00"))
                    if product_cost + shipping_fee:
                        f_sheet['H' +
                                str(num)] = Decimal(product_cost +
                                                    shipping_fee).quantize(
                                                        Decimal("0.00"))
                    num += 1

        f_sheet = wb.create_sheet('其它费用')
        f_sheet['A1'] = '产生日期'
        f_sheet['B1'] = '店铺名称'
        f_sheet['C1'] = '费用金额(RMB)'
        f_sheet['D1'] = '说明'
        shop_fees = Finance.objects.filter(shop=shop, f_type='FEE').filter(
            wd_date__gte=start_date, wd_date__lte=end_date)
        num = 2
        fee_note = ''
        for i in shop_fees:
            total_pay += i.income

            f_sheet['A' + str(num)] = i.wd_date
            f_sheet['B' + str(num)] = shop.name
            f_sheet['C' + str(num)] = Decimal(i.income).quantize(
                Decimal("0.00"))
            f_sheet['D' + str(num)] = i.note
            num += 1

        f_sheet = wb.create_sheet('收入明细')
        f_sheet['A1'] = '结汇日期'
        f_sheet['B1'] = '店铺名称'
        f_sheet['C1'] = '外汇金额({currency})'.format(currency=shop.currency)
        f_sheet['D1'] = '到账金额(RMB)'
        finance = Finance.objects.filter(shop=shop, f_type='EXC').filter(
            exc_date__gte=start_date, exc_date__lte=end_date)
        total_rec = 0  # 总收入
        num = 2
        for i in finance:
            total_rec += i.income_rmb

            f_sheet['A' + str(num)] = i.exc_date
            f_sheet['B' + str(num)] = shop.name
            f_sheet['C' + str(num)] = Decimal(i.exchange).quantize(
                Decimal("0.00"))
            f_sheet['D' + str(num)] = Decimal(i.income_rmb).quantize(
                Decimal("0.00"))
            num += 1

        f_sheet = wb.create_sheet('汇总')
        f_sheet['A1'] = '店铺名称'
        f_sheet['B1'] = '负责人'
        f_sheet['C1'] = '统计日期'
        f_sheet['D1'] = '总支出'
        f_sheet['E1'] = '总收入'
        f_sheet['F1'] = '盈亏'
        f_sheet['A2'] = shop.name
        f_sheet['B2'] = shop.user.first_name if shop.user else ''
        f_sheet['C2'] = start_date + '至' + end_date
        f_sheet['D2'] = Decimal(total_pay).quantize(Decimal("0.00"))
        f_sheet['E2'] = Decimal(total_rec).quantize(Decimal("0.00"))
        f_sheet['F2'] = Decimal(total_rec - total_pay).quantize(
            Decimal("0.00"))

        del wb['Sheet']
        wb.save('media/export/shop_finance/收支明细报表-' + shop.name + '.xlsx')
        url = BASE_URL + '/media/export/shop_finance/收支明细报表-' + shop.name + '.xlsx'
        return Response({'url': url}, status=status.HTTP_200_OK)


class ShopStockViewSet(mixins.ListModelMixin, mixins.CreateModelMixin,
                       mixins.UpdateModelMixin, mixins.DestroyModelMixin,
                       mixins.RetrieveModelMixin, viewsets.GenericViewSet):
    """
    list:
        店铺库存列表,分页,过滤,搜索,排序
    create:
        店铺库存新增
    retrieve:
        店铺库存详情页
    update:
        店铺库存修改
    destroy:
        店铺库存删除
    """
    queryset = ShopStock.objects.all()
    serializer_class = ShopStockSerializer  # 序列化
    pagination_class = DefaultPagination  # 分页

    filter_backends = (DjangoFilterBackend, filters.SearchFilter,
                       filters.OrderingFilter)  # 过滤,搜索,排序
    # filter_fields = ('shop', 'p_status', 'is_active', 'is_collect')  # 配置过滤字段
    filterset_fields = {
        'qty': ['gte', 'lte', 'exact', 'gt', 'lt'],
        'onway_qty': ['gte', 'lte', 'exact', 'gt', 'lt'],
        'trans_qty': ['gte', 'lte', 'exact', 'gt', 'lt'],
        'day15_sold': ['gte', 'lte', 'exact', 'gt', 'lt'],
        'day30_sold': ['gte', 'lte', 'exact', 'gt', 'lt'],
        'total_sold': ['gte', 'lte', 'exact', 'gt', 'lt'],
        'total_profit': ['gte', 'lte', 'exact', 'gt', 'lt'],
        'total_weight': ['gte', 'lte', 'exact', 'gt', 'lt'],
        'total_cbm': ['gte', 'lte', 'exact', 'gt', 'lt'],
        'stock_value': ['gte', 'lte', 'exact', 'gt', 'lt'],
        'refund_rate': ['gte', 'lte', 'exact', 'gt', 'lt'],
        'avg_profit': ['gte', 'lte', 'exact', 'gt', 'lt'],
        'avg_profit_rate': ['gte', 'lte', 'exact', 'gt', 'lt'],
        'shop': ['exact'],
        'p_status': ['exact', 'in'],
        'is_active': ['exact'],
        'is_collect': ['exact'],
    }
    search_fields = ('sku', 'p_name', 'label_code', 'upc', 'item_id')  # 配置搜索字段
    ordering_fields = ('create_time', 'item_id', 'qty', 'day15_sold',
                       'day30_sold', 'total_sold', 'total_profit',
                       'total_weight', 'total_cbm', 'stock_value', 'onway_qty',
                       'refund_rate', 'avg_profit', 'avg_profit_rate'
                       )  # 配置排序字段

    # FBM库存上传
    @action(methods=['post'], detail=False, url_path='fbm_upload')
    def fbm_upload(self, request):
        import warnings
        warnings.filterwarnings('ignore')

        data = request.data
        shop_id = data['id']
        wb = openpyxl.load_workbook(data['excel'])
        sheet = wb['Reporte general de stock']
        shop = Shop.objects.filter(id=shop_id).first()

        for cell_row in list(sheet)[5:]:
            sku = cell_row[1].value
            item_id = cell_row[3].value
            qty = cell_row[16].value

            shop_stock = ShopStock.objects.filter(sku=sku,
                                                  item_id=item_id).first()
            if shop_stock:
                shop_stock.qty = qty
                shop_stock.save()
            else:
                ml_product = MLProduct.objects.filter(sku=sku,
                                                      item_id=item_id).first()
                url = ''
                if shop.platform == 'MERCADO':
                    url = 'https://articulo.mercadolibre.com.mx/' + shop.site + '-' + ml_product.item_id
                if shop.platform == 'NOON':
                    url = 'https://www.noon.com/product/{item_id}/p/?o={item_id}-1'.format(
                        item_id=ml_product.item_id)
                if ml_product:
                    shop_stock = ShopStock()
                    shop_stock.shop = shop
                    shop_stock.sku = ml_product.sku
                    shop_stock.p_name = ml_product.p_name
                    shop_stock.label_code = ml_product.label_code
                    shop_stock.upc = ml_product.upc
                    shop_stock.item_id = ml_product.item_id
                    shop_stock.image = ml_product.image
                    shop_stock.qty = qty
                    shop_stock.weight = ml_product.weight
                    shop_stock.length = ml_product.length
                    shop_stock.width = ml_product.width
                    shop_stock.heigth = ml_product.heigth
                    shop_stock.unit_cost = ml_product.unit_cost
                    shop_stock.first_ship_cost = ml_product.first_ship_cost
                    shop_stock.sale_url = url
                    shop_stock.save()

        if shop:
            # 创建操作日志
            log = MLOperateLog()
            log.op_module = 'FBM'
            log.op_type = 'CREATE'
            log.target_type = 'FBM'
            log.desc = 'FBM库存导入 店铺: {name}'.format(name=shop.name)
            log.user = request.user
            log.save()

        return Response({'msg': '成功上传'}, status=status.HTTP_200_OK)

    # 统计店铺数据
    @action(methods=['post'], detail=False, url_path='calc_stock')
    def calc_stock(self, request):
        shop_id = request.data['id']
        shop = Shop.objects.filter(id=shop_id).first()
        ex = ExRate.objects.filter(currency=shop.currency).first()
        ex_rate = ex.value if ex else 0

        # FBM库存统计
        queryset = ShopStock.objects.filter(is_active=True,
                                            qty__gt=0,
                                            shop__id=shop_id)
        total_amount = 0
        total_qty = 0
        for i in queryset:
            total_qty += i.qty
            total_amount += (i.unit_cost + i.first_ship_cost) * i.qty

        # 中转仓库存统计
        ts = TransStock.objects.filter(listing_shop=shop.name, is_out=False)
        trans_amount = 0
        for i in ts:
            trans_amount += (i.unit_cost + i.first_ship_cost) * i.qty

        # 在途运单统计
        ships = Ship.objects.filter(shop=shop.name).filter(
            Q(s_status='SHIPPED') | Q(s_status='BOOKED'))
        onway_amount = 0
        for i in ships:
            onway_amount += i.shipping_fee
            onway_amount += i.extra_fee
            onway_amount += i.products_cost

        date = datetime.now().date() - timedelta(days=30)
        sum_qty = MLOrder.objects.filter(shop__id=shop_id,
                                         order_time__gte=date).aggregate(
                                             Sum('qty'))
        sold_qty = sum_qty['qty__sum']
        sum_amount = MLOrder.objects.filter(shop__id=shop_id,
                                            order_time__gte=date).aggregate(
                                                Sum('price'))
        sold_amount = sum_amount['price__sum']
        sum_profit = MLOrder.objects.filter(shop__id=shop_id,
                                            order_time__gte=date).aggregate(
                                                Sum('profit'))
        sold_profit = sum_profit['profit__sum']

        sum_unit_cost = MLOrder.objects.filter(shop__id=shop_id,
                                               order_time__gte=date).aggregate(
                                                   Sum('unit_cost'))
        total_unit_cost = sum_unit_cost['unit_cost__sum']
        if not total_unit_cost:
            total_unit_cost = 0

        sum_first_ship_cost = MLOrder.objects.filter(
            shop__id=shop_id,
            order_time__gte=date).aggregate(Sum('first_ship_cost'))
        total_first_ship_cost = sum_first_ship_cost['first_ship_cost__sum']
        if not total_first_ship_cost:
            total_first_ship_cost = 0
        total_cost = total_unit_cost + total_first_ship_cost

        # 结汇外汇
        sum_exchange = Finance.objects.filter(shop__id=shop_id,
                                              f_type='EXC').aggregate(
                                                  Sum('exchange'))
        exchange_fund = sum_exchange['exchange__sum']
        if not exchange_fund:
            exchange_fund = 0

        # 店铺提现外汇
        sum_income_fund = Finance.objects.filter(shop__id=shop_id,
                                                 is_received=True,
                                                 f_type='WD').aggregate(
                                                     Sum('income'))
        income_fund = sum_income_fund['income__sum']
        if not income_fund:
            income_fund = 0

        # 店铺剩余外汇
        rest_income = income_fund - exchange_fund

        # 店铺结汇资金
        sum_income_rmb = Finance.objects.filter(shop__id=shop_id,
                                                f_type='EXC',
                                                exc_date__gte=date).aggregate(
                                                    Sum('income_rmb'))
        income_rmb = sum_income_rmb['income_rmb__sum']
        if not income_rmb:
            income_rmb = 0
        total_fund = rest_income * ex_rate + income_rmb

        real_profit = total_fund - total_cost

        return Response(
            {
                'todayStockQty': total_qty,
                'todayStockAmount': total_amount,
                'sold_qty': sold_qty,
                'sold_amount': sold_amount,
                'sold_profit': sold_profit,
                'real_profit': real_profit,
                'onway_amount': onway_amount,
                'trans_amount': trans_amount
            },
            status=status.HTTP_200_OK)

    # 查询库存在途情况
    @action(methods=['post'], detail=False, url_path='get_stock_detail')
    def get_stock_detail(self, request):
        sku = request.data['sku']
        op_type = request.data['op_type']

        data = []
        if op_type == 'FBM_ONWAY':
            querySet = ShipDetail.objects.filter(
                sku=sku, ship__target='FBM').filter(
                    Q(ship__s_status='SHIPPED') | Q(ship__s_status='BOOKED'))
            if querySet:
                for i in querySet:
                    data.append({
                        'qty': i.qty,
                        "s_status": i.ship.s_status,
                        "book_date": i.ship.book_date,
                        "tag_name": i.ship.tag_name,
                        "tag_color": i.ship.tag_color,
                        "batch": i.ship.batch,
                    })
        if op_type == 'TRANS_ONWAY':
            querySet = ShipDetail.objects.filter(
                sku=sku, ship__target='TRANSIT').filter(
                    Q(ship__s_status='SHIPPED') | Q(ship__s_status='BOOKED'))
            if querySet:
                for i in querySet:
                    data.append({
                        'qty': i.qty,
                        "s_status": i.ship.s_status,
                        "book_date": i.ship.book_date,
                        "tag_name": i.ship.tag_name,
                        "tag_color": i.ship.tag_color,
                        "batch": i.ship.batch,
                    })
        if op_type == 'FINISH':
            date = datetime.now().date() - timedelta(days=30)
            querySet = ShipDetail.objects.filter(
                sku=sku, ship__s_status='FINISHED',
                ship__target='FBM').filter(ship__book_date__gte=date)
            if querySet:
                for i in querySet:
                    data.append({
                        'qty': i.qty,
                        "s_status": i.ship.s_status,
                        "book_date": i.ship.book_date,
                        "tag_name": i.ship.tag_name,
                        "tag_color": i.ship.tag_color,
                        "batch": i.ship.batch,
                    })
        if op_type == 'TRANS':
            ts = TransStock.objects.filter(sku=sku, is_out=False)
            if ts:
                for i in ts:
                    data.append({
                        'qty': i.qty,
                        'arrived_date': i.arrived_date,
                        'batch': i.batch,
                        'warehouse': i.shop.name,
                    })

        return Response(data, status=status.HTTP_200_OK)

    # 盘点fbm库存
    @action(methods=['post'], detail=False, url_path='change_stock')
    def change_stock(self, request):
        sid = request.data['id']
        new_qty = request.data['qty']
        reason = request.data['reason']

        shop_stock = ShopStock.objects.filter(id=sid).first()
        old_qty = shop_stock.qty
        shop_stock.qty = new_qty
        shop_stock.save()

        changed_qty = new_qty - old_qty
        # 创建库存日志
        stock_log = StockLog()
        stock_log.shop_stock = shop_stock
        stock_log.current_stock = shop_stock.qty
        stock_log.qty = abs(changed_qty)
        stock_log.in_out = 'IN' if changed_qty > 0 else 'OUT'
        stock_log.action = 'TAKING'
        stock_log.desc = '盘点理由: ' + reason
        stock_log.user_id = request.user.id
        stock_log.save()

        # 创建操作日志
        log = MLOperateLog()
        log.op_module = 'FBM'
        log.op_type = 'EDIT'
        log.target_id = sid
        log.target_type = 'FBM'
        log.desc = '库存盘点: {sku}数量 {old_qty} ===>> {new_qty}, 理由：{reason}'.format(
            sku=shop_stock.sku,
            old_qty=old_qty,
            new_qty=new_qty,
            reason=reason)
        log.user = request.user
        log.save()

        return Response({'msg': '操作成功'}, status=status.HTTP_200_OK)

    # 修改产品状态
    @action(methods=['post'], detail=False, url_path='change_status')
    def change_status(self, request):
        sid = request.data['id']
        new_status = request.data['status']

        shop_stock = ShopStock.objects.filter(id=sid).first()
        if shop_stock.p_status == new_status:
            return Response({'msg': '操作成功'}, status=status.HTTP_200_OK)
        old_status = shop_stock.p_status
        shop_stock.p_status = new_status
        shop_stock.save()

        # 创建操作日志
        log = MLOperateLog()
        log.op_module = 'FBM'
        log.op_type = 'EDIT'
        log.target_id = sid
        log.target_type = 'FBM'
        log.desc = '修改状态: {sku}状态 {old_status} ===>> {new_status}'.format(
            sku=shop_stock.sku, old_status=old_status, new_status=new_status)
        log.user = request.user
        log.save()

        return Response({'msg': '操作成功'}, status=status.HTTP_200_OK)

    # 同步中转库存数量
    @action(methods=['post'], detail=False, url_path='sycn_trans_stock')
    def sycn_trans_stock(self, request):
        shop_id = request.data['id']
        queryset = ShopStock.objects.filter(shop__id=shop_id)
        for i in queryset:
            qty = 0
            ts = TransStock.objects.filter(sku=i.sku, is_out=False)
            for t in ts:
                qty += t.qty
            i.trans_qty = qty
            i.save()

        result = True
        return Response({'result': result}, status=status.HTTP_200_OK)

    @action(methods=['get'], detail=False, url_path='test')
    def test(self, request):
        message = 'runing'
        tasks.bulk_ship_tracking.delay()

        return Response({'message': message}, status=status.HTTP_200_OK)


class ShipViewSet(mixins.ListModelMixin, mixins.CreateModelMixin,
                  mixins.UpdateModelMixin, mixins.DestroyModelMixin,
                  mixins.RetrieveModelMixin, viewsets.GenericViewSet):
    """
    list:
        头程运单列表,分页,过滤,搜索,排序
    create:
        头程运单新增
    retrieve:
        头程运单详情页
    update:
        头程运单修改
    destroy:
        头程运单删除
    """
    queryset = Ship.objects.all()
    serializer_class = ShipSerializer  # 序列化
    pagination_class = DefaultPagination  # 分页

    filter_backends = (DjangoFilterBackend, filters.SearchFilter,
                       filters.OrderingFilter)  # 过滤,搜索,排序
    # filter_fields = ('s_status', 'shop', 'target', 'ship_type', 'carrier', 'user_id')  # 配置过滤字段
    filterset_fields = {
        'book_date': ['gte', 'lte', 'exact', 'gt', 'lt'],
        'create_time': ['gte', 'lte', 'exact', 'gt', 'lt'],
        'total_box': ['gte', 'lte', 'exact', 'gt', 'lt'],
        'shop': ['exact'],
        'platform': ['exact'],
        's_status': ['exact', 'in'],
        'target': ['exact'],
        'batch': ['exact'],
        'ship_type': ['exact'],
        'carrier': ['exact'],
        'user_id': ['exact', 'in'],
    }
    search_fields = ('s_number', 'batch', 'envio_number', 'note',
                     'ship_shipDetail__sku', 'ship_shipDetail__item_id',
                     'ship_shipDetail__p_name', 'shop')  # 配置搜索字段
    ordering_fields = ('create_time', 'book_date', 'shop')  # 配置排序字段

    # 创建运单
    @action(methods=['post'], detail=False, url_path='create_ship')
    def create_ship(self, request):
        shop = request.data['shop']
        shop_obj = Shop.objects.filter(name=shop).first()
        shop_id = shop_obj.id
        platform = request.data['platform']
        target = request.data['target']
        ship_type = request.data['ship_type']
        carrier = request.data['carrier']
        end_date = request.data['end_date']
        ship_date = request.data['ship_date']
        note = request.data['note']
        all_see = request.data['all_see']
        ship_detail = request.data['ship_detail']

        # 检查店铺额度(仅检查直发运单额度)
        if target == 'FBM':
            products_cost = 0  # 总货品成本
            for i in ship_detail:
                product = MLProduct.objects.filter(sku=i['sku']).first()
                products_cost += product.unit_cost * i['qty']
            used_quota = tasks.get_shop_quota(shop_id)  # 获取店铺已用额度
            if (products_cost + used_quota) > shop_obj.quota:
                return Response({
                    'msg': '店铺额度不足,请减少发货数量!',
                    'status': 'error'
                },
                                status=status.HTTP_202_ACCEPTED)

        batch = 'P{time_str}'.format(time_str=time.strftime('%m%d'))
        if end_date:
            b_num = end_date.replace('-', '')[2:]
            batch = 'P{time_str}'.format(time_str=b_num)
        if not end_date:
            end_date = None
        if not ship_date:
            ship_date = None
        user_id = request.user.id
        # 如果全员可见，id=0
        if all_see:
            user_id = 0
        ship = Ship(s_status='PREPARING',
                    send_from='CN',
                    platform=platform,
                    batch=batch,
                    shop=shop,
                    target=target,
                    carrier=carrier,
                    ship_type=ship_type,
                    end_date=end_date,
                    ship_date=ship_date,
                    note=note,
                    user_id=user_id)
        ship.save()

        total_qty = 0  # 总数量
        products_cost = 0  # 总货品成本

        # 创建运单详情
        for i in ship_detail:
            product = MLProduct.objects.filter(sku=i['sku']).first()
            if product:
                sd = ShipDetail()
                sd.ship = ship
                sd.s_type = i['s_type']
                sd.qty = i['qty']
                sd.plan_qty = i['qty']
                sd.note = i['note']
                sd.sku = i['sku']
                sd.target_FBM = product.shop
                sd.p_name = product.p_name
                sd.label_code = product.label_code
                sd.upc = product.upc
                sd.item_id = product.item_id
                sd.custom_code = product.custom_code
                sd.cn_name = product.cn_name
                sd.en_name = product.en_name
                sd.brand = product.brand
                sd.declared_value = product.declared_value
                sd.cn_material = product.cn_material
                sd.en_material = product.en_material
                sd.use = product.use
                sd.is_elec = product.is_elec
                sd.is_magnet = product.is_magnet
                sd.is_water = product.is_water
                sd.image = product.image
                sd.unit_cost = product.unit_cost
                sd.weight = product.weight
                sd.length = product.length
                sd.width = product.width
                sd.heigth = product.heigth

                packing = Packing.objects.filter(id=product.packing_id).first()
                if packing:
                    sd.packing_name = packing.name
                    sd.packing_size = packing.size

                # 如果店铺库存中没有，发货在途没有，自动标为新品
                is_exist = ShopStock.objects.filter(sku=sd.sku).count()
                is_ship = ShipDetail.objects.filter(sku=sd.sku).filter(
                    Q(ship__s_status='SHIPPED')
                    | Q(ship__s_status='BOOKED')).count()
                if not is_exist and not is_ship:
                    sd.s_type = 'NEW'

                sd.save()

                total_qty += i['qty']
                products_cost += product.unit_cost * i['qty']

        ship.total_qty = total_qty
        ship.products_cost = products_cost
        ship.save()

        # 创建操作日志
        log = MLOperateLog()
        log.op_module = 'SHIP'
        log.op_type = 'CREATE'
        log.target_type = 'SHIP'
        log.target_id = ship.id
        log.desc = '创建运单'
        log.user = request.user
        log.save()

        return Response({
            'msg': '成功创建运单',
            'status': 'success'
        },
                        status=status.HTTP_200_OK)

    # 编辑运单
    @action(methods=['post'], detail=False, url_path='edit_ship')
    def edit_ship(self, request):
        id = request.data['id']
        shop = request.data['shop']
        shop_obj = Shop.objects.filter(name=shop).first()
        shop_id = shop_obj.id
        target = request.data['target']
        fbm_warehouse = request.data['fbm_warehouse']
        ship_type = request.data['ship_type']
        carrier = request.data['carrier']
        end_date = request.data['end_date']
        ship_date = request.data['ship_date']
        note = request.data['note']
        all_see = request.data['all_see']
        s_number = request.data['s_number']
        batch = request.data['batch']
        envio_number = request.data['envio_number']
        ship_detail = request.data['ship_shipDetail']

        # 检查店铺额度
        if target == 'FBM':
            products_cost = 0  # 新总货品成本
            for i in ship_detail:
                product = MLProduct.objects.filter(sku=i['sku']).first()
                products_cost += product.unit_cost * i['qty']

            ori_products_cost = 0  # 原货品成本
            sd_set = ShipDetail.objects.filter(ship__id=id)
            for i in sd_set:
                ori_products_cost += i.unit_cost * i.qty

            used_quota = tasks.get_shop_quota(shop_id)  # 获取店铺已用额度
            if (products_cost - ori_products_cost +
                    used_quota) > shop_obj.quota:
                return Response({
                    'msg': '店铺额度不足,请减少发货数量!',
                    'status': 'error'
                },
                                status=status.HTTP_202_ACCEPTED)

        ship = Ship.objects.filter(id=id).first()
        if not ship:
            return Response({'msg': '数据错误'}, status=status.HTTP_202_ACCEPTED)
        ship.shop = shop
        ship.target = target
        ship.fbm_warehouse = fbm_warehouse
        ship.ship_type = ship_type
        ship.carrier = carrier

        # 修改批次号
        if ship.batch != batch:
            # 判断是否有文件夹需要修改
            path = 'media/ml_ships/'
            head_path = '{path}{batch}_{id}'.format(path=path,
                                                    batch=ship.batch,
                                                    id=ship.id)
            # 如果存在文件夹，则修改文件夹名称
            if os.path.exists(head_path):
                new_head_path = '{path}{batch}_{id}'.format(path=path,
                                                            batch=batch,
                                                            id=ship.id)
                os.rename(head_path, new_head_path)
            ship.batch = batch

        if end_date:
            ship.end_date = end_date
        if ship_date:
            ship.ship_date = ship_date
        ship.note = note
        ship.s_number = s_number
        ship.envio_number = envio_number

        # 如果中转运单关闭全员可见，修改用户id
        if target == 'TRANSIT':
            if not all_see:
                ship.user_id = request.user.id
            else:
                ship.user_id = 0

        # 需要更新的sku列表
        sku_list = []
        for i in ship_detail:
            sku_list.append(i['sku'])

        # 移除不需要的产品
        queryset = ShipDetail.objects.filter(ship=ship)
        for i in queryset:
            if i.sku not in sku_list:
                # 创建操作日志
                log = MLOperateLog()
                log.op_module = 'SHIP'
                log.op_type = 'DEL'
                log.target_type = 'SHIP'
                log.target_id = ship.id
                log.desc = '移除产品 {sku} {p_name}'.format(sku=i.sku,
                                                        p_name=i.p_name)
                log.user = request.user
                log.save()

                i.delete()

        total_qty = 0  # 总数量
        products_cost = 0  # 总货品成本

        # 更新运单详情
        for i in ship_detail:
            product = MLProduct.objects.filter(sku=i['sku']).first()
            if product:
                if 'id' in i.keys():
                    sd = ShipDetail.objects.filter(id=i['id']).first()
                else:
                    sd = ShipDetail()

                    # 创建操作日志
                    log = MLOperateLog()
                    log.op_module = 'SHIP'
                    log.op_type = 'CREATE'
                    log.target_type = 'SHIP'
                    log.target_id = ship.id
                    log.desc = '新增产品 {sku} {p_name} {qty}个'.format(
                        sku=product.sku, p_name=product.p_name, qty=i['qty'])
                    log.user = request.user
                    log.save()

                sd.ship = ship
                sd.s_type = i['s_type']
                sd.qty = i['qty']
                sd.plan_qty = i['plan_qty']
                sd.note = i['note']
                sd.sku = i['sku']
                sd.target_FBM = product.shop
                sd.p_name = product.p_name
                sd.label_code = product.label_code
                sd.upc = product.upc
                sd.item_id = product.item_id
                sd.custom_code = product.custom_code
                sd.cn_name = product.cn_name
                sd.en_name = product.en_name
                sd.brand = product.brand
                sd.declared_value = product.declared_value
                sd.cn_material = product.cn_material
                sd.en_material = product.en_material
                sd.use = product.use
                sd.is_elec = product.is_elec
                sd.is_magnet = product.is_magnet
                sd.is_water = product.is_water
                sd.image = product.image
                sd.unit_cost = product.unit_cost
                sd.weight = product.weight
                sd.length = product.length
                sd.width = product.width
                sd.heigth = product.heigth

                packing = Packing.objects.filter(id=product.packing_id).first()
                if packing:
                    sd.packing_name = packing.name
                    sd.packing_size = packing.size

                # 如果店铺库存中没有，自动标为新品
                if 'id' not in i.keys():
                    # 如果店铺库存中没有，发货在途没有，自动标为新品
                    is_exist = ShopStock.objects.filter(sku=sd.sku).count()
                    is_ship = ShipDetail.objects.filter(sku=sd.sku).filter(
                        Q(ship__s_status='SHIPPED')
                        | Q(ship__s_status='BOOKED')).count()
                    if not is_exist and not is_ship:
                        sd.s_type = 'NEW'

                sd.save()

                total_qty += i['qty']
                products_cost += product.unit_cost * i['qty']

        ship.total_qty = total_qty
        ship.products_cost = products_cost
        ship.save()
        return Response({
            'msg': '成功更新运单',
            'status': 'success'
        },
                        status=status.HTTP_200_OK)

    # 运单发货/保存
    @action(methods=['post'], detail=False, url_path='send_ship')
    def send_ship(self, request):
        ship_id = request.data['id']
        ship_note = request.data['note']
        ship_detail = request.data['ship_shipDetail']
        ship_action = request.data['action']

        # 运单发货前检查
        if ship_action == 'SHIPPED':
            # 检查是否重复发货
            is_exist = Ship.objects.filter(id=ship_id,
                                           s_status='PREPARING').count()
            if not is_exist:
                return Response({
                    'msg': '重复发货!',
                    'status': 'error'
                },
                                status=status.HTTP_202_ACCEPTED)
            # 检查库存是否足够
            for i in ship_detail:
                pm = PurchaseManage.objects.filter(p_status='PACKED',
                                                   sku=i['sku']).first()
                if pm:
                    if pm.pack_qty >= i['qty']:
                        continue
                return Response({
                    'msg': '已打包货品库存不足!',
                    'status': 'error'
                },
                                status=status.HTTP_202_ACCEPTED)

        products_cost = 0  # 总货品成本
        for i in ship_detail:
            if not i['qty']:
                # 发货为0的加入遗弃清单
                sd = ShipDetail.objects.filter(id=i['id']).first()

                ship_ir = ShipItemRemove.objects.filter(ship=sd.ship,
                                                        sku=sd.sku).first()
                if not ship_ir:
                    ship_ir = ShipItemRemove()
                    ship_ir.ship = sd.ship
                    ship_ir.belong_shop = sd.target_FBM
                    ship_ir.sku = sd.sku
                    ship_ir.p_name = sd.p_name
                    ship_ir.image = sd.image
                    ship_ir.item_id = sd.item_id
                    ship_ir.plan_qty = sd.qty
                    ship_ir.send_qty = 0
                    ship_ir.item_type = 'REMOVE'
                    ship_ir.save()
                else:
                    if ship_ir.item_type == 'REDUCE':
                        ship_ir.send_qty = 0
                        ship_ir.item_type = 'REMOVE'
                        ship_ir.save()

                # 发货数量为0的删除
                sd.delete()
                continue

            sd = ShipDetail.objects.filter(id=i['id']).first()

            # 发货数量少于计划数量的加入遗弃清单
            if i['qty'] < sd.qty:
                ship_ir = ShipItemRemove.objects.filter(ship=sd.ship,
                                                        sku=sd.sku).first()
                if not ship_ir:
                    ship_ir = ShipItemRemove()
                    ship_ir.ship = sd.ship
                    ship_ir.belong_shop = sd.target_FBM
                    ship_ir.sku = sd.sku
                    ship_ir.p_name = sd.p_name
                    ship_ir.image = sd.image
                    ship_ir.item_id = sd.item_id
                    ship_ir.plan_qty = sd.qty
                    ship_ir.send_qty = i['qty']
                    ship_ir.item_type = 'REDUCE'
                    ship_ir.save()
                else:
                    ship_ir.send_qty = i['qty']
                    ship_ir.item_type = 'REDUCE'
                    ship_ir.save()

            sd.qty = i['qty']
            sd.box_number = i['box_number']
            sd.note = i['note']
            sd.save()
            products_cost += sd.unit_cost * sd.qty

            # 发货后减去打包库存数量(旧代码)
            # if ship_action == 'SHIPPED':
            #     pm = PurchaseManage.objects.filter(sku=sd.sku, p_status='PACKED').first()
            #
            #     if pm:
            #         # 创建操作日志
            #         log = MLOperateLog()
            #         log.op_module = 'PURCHASE'
            #         log.op_type = 'CREATE'
            #         log.target_type = 'PURCHASE'
            #         log.desc = '库存扣除 {qty}个 {sku} {p_name}'.format(sku=pm.sku, p_name=pm.p_name, qty=sd.qty)
            #         log.save()
            #         purchase_manage = PurchaseManage(
            #             p_status='USED',
            #             s_type=pm.s_type,
            #             create_type=pm.create_type,
            #             sku=pm.sku,
            #             p_name=pm.p_name,
            #             item_id=pm.item_id,
            #             label_code=pm.label_code,
            #             image=pm.image,
            #             unit_cost=pm.unit_cost,
            #             weight=pm.weight,
            #             length=pm.length,
            #             width=pm.width,
            #             heigth=pm.heigth,
            #             used_qty=sd.qty,
            #             used_batch=sd.ship.batch,
            #             note=pm.note,
            #             shop=pm.shop,
            #             shop_color=pm.shop_color,
            #             packing_size=pm.packing_size,
            #             packing_name=pm.packing_name,
            #             used_time=datetime.now()
            #         )
            #         purchase_manage.save()
            #         if pm.pack_qty > sd.qty:
            #             pm.pack_qty = pm.pack_qty - sd.qty
            #             pm.save()
            #         else:
            #             pm.delete()

        ship = Ship.objects.filter(id=ship_id).first()
        if ship_note:
            ship.note = ship_note
        ship.s_status = ship_action
        ship.products_cost = products_cost
        # 总箱数
        box_qty = ShipBox.objects.filter(ship=ship).count()
        ship.total_box = box_qty

        sum_weight = ShipBox.objects.filter(ship=ship).aggregate(Sum('weight'))
        ship.weight = sum_weight['weight__sum']

        # 总数量
        result = ShipDetail.objects.filter(ship=ship).aggregate(Sum('qty'))
        ship.total_qty = result['qty__sum']

        # 总体积cbm
        sum_cbm = ShipBox.objects.filter(ship=ship).aggregate(Sum('cbm'))
        ship.cbm = sum_cbm['cbm__sum']

        # 添加发货时间
        if ship_action == 'SHIPPED':
            ship.sent_time = datetime.now()

        ship.save()

        # 发货后减去打包库存数量
        if ship_action == 'SHIPPED':
            sd_set = ShipDetail.objects.filter(ship=ship)
            for sd in sd_set:
                pm = PurchaseManage.objects.filter(sku=sd.sku,
                                                   p_status='PACKED').first()

                if pm:
                    # 创建操作日志
                    log = MLOperateLog()
                    log.op_module = 'PURCHASE'
                    log.op_type = 'CREATE'
                    log.target_type = 'PURCHASE'
                    log.desc = '库存扣除 {qty}个 {sku} {p_name}'.format(
                        sku=pm.sku, p_name=pm.p_name, qty=sd.qty)
                    log.save()
                    purchase_manage = PurchaseManage(
                        p_status='USED',
                        s_type=pm.s_type,
                        create_type=pm.create_type,
                        sku=pm.sku,
                        p_name=pm.p_name,
                        item_id=pm.item_id,
                        label_code=pm.label_code,
                        image=pm.image,
                        unit_cost=pm.unit_cost,
                        weight=pm.weight,
                        length=pm.length,
                        width=pm.width,
                        heigth=pm.heigth,
                        used_qty=sd.qty,
                        used_batch=sd.ship.batch,
                        note=pm.note,
                        shop=pm.shop,
                        shop_color=pm.shop_color,
                        packing_size=pm.packing_size,
                        packing_name=pm.packing_name,
                        used_time=datetime.now())
                    purchase_manage.save()
                    if pm.pack_qty > sd.qty:
                        pm.pack_qty = pm.pack_qty - sd.qty
                        pm.save()
                    else:
                        pm.delete()

        # 添加fbm库存在途数量
        if ship.target == 'FBM' and ship_action == 'SHIPPED':
            queryset = ShipDetail.objects.filter(ship=ship)
            for i in queryset:
                shop_stock = ShopStock.objects.filter(
                    sku=i.sku, shop__name=ship.shop).first()
                if shop_stock:
                    shop_stock.onway_qty += i.qty
                    shop_stock.save()

        if ship_action == 'SHIPPED':
            msg = '成功发货!'

            # 创建操作日志
            log = MLOperateLog()
            log.op_module = 'SHIP'
            log.op_type = 'EDIT'
            log.target_type = 'SHIP'
            log.target_id = ship.id
            log.desc = '运单发货!'
            log.user = request.user
            log.save()
        else:
            msg = '保存成功!'
        return Response({
            'msg': msg,
            'status': 'success'
        },
                        status=status.HTTP_200_OK)

    # 添加运费
    @action(methods=['post'], detail=False, url_path='postage')
    def postage(self, request):
        ship_id = request.data['id']
        shipping_fee = request.data['shipping_fee']

        ship = Ship.objects.filter(id=ship_id).first()
        ship.shipping_fee = shipping_fee
        ship.save()

        all_fee = ship.shipping_fee + ship.extra_fee

        queryset = ShipDetail.objects.filter(ship=ship)
        total_weight = 0
        total_cbm = 0

        if ship.ship_type == '空运':
            for i in queryset:
                total_weight += (i.weight * i.qty)
            for i in queryset:
                if total_weight:
                    percent = (i.weight * i.qty) / total_weight
                else:
                    percent = 0
                avg_ship_fee = percent * all_fee / i.qty
                i.avg_ship_fee = avg_ship_fee
                i.save()

        if ship.ship_type == '海运':
            for i in queryset:
                cbm = i.length * i.width * i.heigth / 1000000
                total_cbm += (cbm * i.qty)
            for i in queryset:
                cbm = i.length * i.width * i.heigth / 1000000
                percent = cbm / total_cbm
                avg_ship_fee = percent * all_fee
                i.avg_ship_fee = avg_ship_fee
                i.save()

        # 创建操作日志
        log = MLOperateLog()
        log.op_module = 'SHIP'
        log.op_type = 'EDIT'
        log.target_type = 'SHIP'
        log.target_id = ship.id
        log.desc = '添加头程运费 {name}'.format(name=shipping_fee)
        log.user = request.user
        log.save()
        return Response({'msg': '操作成功!'}, status=status.HTTP_200_OK)

    # 添加杂费
    @action(methods=['post'], detail=False, url_path='extra_fee')
    def extra_fee(self, request):
        ship_id = request.data['id']
        extra_fee = request.data['extra_fee']

        ship = Ship.objects.filter(id=ship_id).first()
        ship.extra_fee = extra_fee
        ship.save()

        all_fee = ship.shipping_fee + ship.extra_fee

        queryset = ShipDetail.objects.filter(ship=ship)
        total_weight = 0
        total_cbm = 0

        if ship.ship_type == '空运':
            for i in queryset:
                total_weight += (i.weight * i.qty)
            for i in queryset:
                if total_weight:
                    percent = (i.weight * i.qty) / total_weight
                else:
                    percent = 0
                avg_ship_fee = percent * all_fee / i.qty
                i.avg_ship_fee = avg_ship_fee
                i.save()

        if ship.ship_type == '海运':
            for i in queryset:
                cbm = i.length * i.width * i.heigth / 1000000
                total_cbm += (cbm * i.qty)
            for i in queryset:
                cbm = i.length * i.width * i.heigth / 1000000
                percent = cbm / total_cbm
                avg_ship_fee = percent * all_fee
                i.avg_ship_fee = avg_ship_fee
                i.save()

        # 创建操作日志
        log = MLOperateLog()
        log.op_module = 'SHIP'
        log.op_type = 'EDIT'
        log.target_type = 'SHIP'
        log.target_id = ship.id
        log.desc = '添加杂费 {name}'.format(name=extra_fee)
        log.user = request.user
        log.save()
        return Response({'msg': '操作成功!'}, status=status.HTTP_200_OK)

    # 入仓
    @action(methods=['post'], detail=False, url_path='in_warehouse')
    def in_warehouse(self, request):
        ship_id = request.data['id']

        # 检查是否重复操作
        is_exist = Ship.objects.filter(id=ship_id, s_status='BOOKED').count()
        if not is_exist:
            return Response({
                'msg': '运单状态已变动，请检查!',
                'status': 'error'
            },
                            status=status.HTTP_202_ACCEPTED)

        ship = Ship.objects.filter(id=ship_id).first()

        if ship.target == 'FBM':
            ship_detail = ShipDetail.objects.filter(ship=ship)
            for i in ship_detail:
                shop_stock = ShopStock.objects.filter(
                    sku=i.sku, item_id=i.item_id).first()
                # 如果是补货产品
                if shop_stock:
                    shop_stock.qty += i.qty
                    shop_stock.onway_qty -= i.qty  # 减去在途库存

                    value1 = shop_stock.unit_cost * shop_stock.qty
                    value2 = i.unit_cost * i.qty
                    avg = (value1 + value2) / (shop_stock.qty + i.qty)
                    shop_stock.unit_cost = avg

                    ex_value1 = shop_stock.first_ship_cost * shop_stock.qty
                    ex_value2 = i.avg_ship_fee * i.qty
                    ex_avg = (ex_value1 + ex_value2) / (shop_stock.qty + i.qty)
                    shop_stock.first_ship_cost = ex_avg

                    shop_stock.weight = i.weight
                    shop_stock.length = i.length
                    shop_stock.width = i.width
                    shop_stock.heigth = i.heigth
                    shop_stock.save()

                    # 创建库存日志
                    stock_log = StockLog()
                    stock_log.shop_stock = shop_stock
                    stock_log.current_stock = shop_stock.qty
                    stock_log.qty = i.qty
                    stock_log.in_out = 'IN'
                    stock_log.action = 'INBOUND'
                    stock_log.desc = '补货入仓, 入仓批次: ' + ship.batch
                    stock_log.user_id = request.user.id
                    stock_log.save()
                else:
                    shop_stock = ShopStock()
                    shop = Shop.objects.filter(name=ship.shop).first()
                    url = ''
                    if shop.platform == 'MERCADO':
                        url = 'https://articulo.mercadolibre.com.mx/' + shop.site + '-' + i.item_id
                    if shop.platform == 'NOON':
                        url = 'https://www.noon.com/product/{item_id}/p/?o={item_id}-1'.format(
                            item_id=i.item_id)
                    if shop.platform == 'OZON':
                        url = 'https://www.ozon.ru/product/{item_id}'.format(
                            item_id=i.item_id)
                    if shop.platform == 'EMAG':
                        url = 'http://emag.ro/preview/pd/{item_id}/'.format(
                            item_id=i.item_id)
                    shop_stock.shop = shop
                    shop_stock.sku = i.sku
                    shop_stock.p_name = i.p_name
                    shop_stock.label_code = i.label_code
                    shop_stock.upc = i.upc
                    shop_stock.item_id = i.item_id
                    shop_stock.image = i.image
                    shop_stock.qty = i.qty
                    shop_stock.weight = i.weight
                    shop_stock.length = i.length
                    shop_stock.width = i.width
                    shop_stock.heigth = i.heigth
                    shop_stock.unit_cost = i.unit_cost
                    shop_stock.first_ship_cost = i.avg_ship_fee
                    shop_stock.sale_url = url
                    shop_stock.save()

                    # 创建库存日志
                    stock_log = StockLog()
                    stock_log.shop_stock = shop_stock
                    stock_log.current_stock = i.qty
                    stock_log.qty = i.qty
                    stock_log.in_out = 'IN'
                    stock_log.action = 'INBOUND'
                    stock_log.desc = '首次入仓, 入仓批次: ' + ship.batch
                    stock_log.user_id = request.user.id
                    stock_log.save()
        else:
            # 入仓中转仓
            ship_detail = ShipDetail.objects.filter(ship=ship)
            for i in ship_detail:
                trans_stock = TransStock()
                shop = Shop.objects.filter(name=ship.shop).first()
                trans_stock.shop = shop

                # 查该sku所属店铺下的用户id
                shop2 = Shop.objects.filter(name=i.target_FBM).first()
                if shop2:
                    if shop2.user:
                        trans_stock.user_id = shop2.user.id
                trans_stock.listing_shop = i.target_FBM
                trans_stock.sku = i.sku
                trans_stock.p_name = i.p_name
                trans_stock.label_code = i.label_code
                trans_stock.upc = i.upc
                trans_stock.item_id = i.item_id
                trans_stock.image = i.image
                trans_stock.qty = i.qty
                trans_stock.unit_cost = i.unit_cost
                trans_stock.first_ship_cost = i.avg_ship_fee
                trans_stock.s_number = ship.s_number
                trans_stock.batch = ship.batch
                trans_stock.box_number = i.ship.batch + '-' + i.box_number
                box = ShipBox.objects.filter(ship=ship,
                                             box_number=i.box_number).first()
                trans_stock.carrier_box_number = box.carrier_box_number
                trans_stock.box_weight = box.weight
                trans_stock.box_length = box.length
                trans_stock.box_width = box.width
                trans_stock.box_heigth = box.heigth
                trans_stock.box_cbm = box.cbm
                trans_stock.note = box.note
                trans_stock.arrived_date = ship.book_date
                trans_stock.save()

                # 增加fbm库存中转仓数量
                shop_stock = ShopStock.objects.filter(
                    sku=i.sku, item_id=i.item_id).first()
                if shop_stock:
                    shop_stock.trans_qty += i.qty
                    shop_stock.save()
                else:
                    shop_stock = ShopStock()
                    shop = Shop.objects.filter(name=i.target_FBM).first()
                    shop_stock.shop = shop
                    shop_stock.sku = i.sku
                    shop_stock.p_name = i.p_name
                    shop_stock.label_code = i.label_code
                    shop_stock.upc = i.upc
                    shop_stock.item_id = i.item_id
                    shop_stock.image = i.image
                    shop_stock.trans_qty = i.qty
                    shop_stock.weight = i.weight
                    shop_stock.length = i.length
                    shop_stock.width = i.width
                    shop_stock.heigth = i.heigth
                    shop_stock.unit_cost = i.unit_cost
                    shop_stock.first_ship_cost = i.avg_ship_fee
                    shop_stock.sale_url = 'https://articulo.mercadolibre.com.mx/' + shop.site + '-' + i.item_id
                    shop_stock.save()

        ship.s_status = 'FINISHED'
        ship.save()

        # 创建操作日志
        log = MLOperateLog()
        log.op_module = 'SHIP'
        log.op_type = 'EDIT'
        log.target_type = 'SHIP'
        log.target_id = ship.id
        log.desc = '操作入仓!'
        log.user = request.user
        log.save()

        return Response({'msg': '入仓成功!'}, status=status.HTTP_200_OK)

    # 转入异常
    @action(methods=['post'], detail=False, url_path='mark_error')
    def mark_error(self, request):
        ship_id = request.data['id']
        ship = Ship.objects.filter(id=ship_id).first()
        ship.s_status = 'ERROR'
        ship.save()

        # 创建操作日志
        log = MLOperateLog()
        log.op_module = 'SHIP'
        log.op_type = 'EDIT'
        log.target_type = 'SHIP'
        log.target_id = ship.id
        log.desc = '将运单转入异常状态'
        log.user = request.user
        log.save()
        return Response({'msg': '操作成功!'}, status=status.HTTP_200_OK)

    # 转出异常
    @action(methods=['post'], detail=False, url_path='mark_out_error')
    def mark_out_error(self, request):
        ship_id = request.data['id']
        ship = Ship.objects.filter(id=ship_id).first()
        ship.s_status = 'SHIPPED'
        ship.save()

        # 创建操作日志
        log = MLOperateLog()
        log.op_module = 'SHIP'
        log.op_type = 'EDIT'
        log.target_type = 'SHIP'
        log.target_id = ship.id
        log.desc = '将运单从异常状态转出'
        log.user = request.user
        log.save()
        return Response({'msg': '操作成功!'}, status=status.HTTP_200_OK)

    # 计算运单数量
    @action(methods=['get'], detail=False, url_path='calc_ships')
    def calc_ships(self, request):
        if request.user.is_superuser:
            pre_qty = Ship.objects.filter(s_status='PREPARING').count()
            shipped_qty = Ship.objects.filter(s_status='SHIPPED').count()
            booked_qty = Ship.objects.filter(s_status='BOOKED').count()
        else:
            pre_qty = Ship.objects.filter(s_status='PREPARING').filter(
                Q(user_id=request.user.id) | Q(user_id=0)).count()
            shipped_qty = Ship.objects.filter(s_status='SHIPPED').filter(
                Q(user_id=request.user.id) | Q(user_id=0)).count()
            booked_qty = Ship.objects.filter(s_status='BOOKED').filter(
                Q(user_id=request.user.id) | Q(user_id=0)).count()
        return Response(
            {
                'pre_qty': pre_qty,
                'shipped_qty': shipped_qty,
                'booked_qty': booked_qty
            },
            status=status.HTTP_200_OK)

    # 导出盛德物流申报单
    @action(methods=['post'], detail=False, url_path='export_logistic_decl')
    def export_logistic_decl(self, request):
        ship_id = request.data['id']
        logistic_name = request.data['name']
        ship = Ship.objects.filter(id=ship_id).first()
        from openpyxl.drawing.image import Image

        sd_set = ShipDetail.objects.filter(ship__id=ship_id)
        # 检查货品装箱情况
        for i in sd_set:
            if not i.box_number:
                return Response({
                    'msg': '有货品未装箱，请装箱后再导出!',
                    'status': 'error'
                },
                                status=status.HTTP_202_ACCEPTED)

        if logistic_name == 'SHENGDE':
            wb = openpyxl.Workbook()
            sh = wb.active
            sh.title = 'Sheet1'
            sh.column_dimensions['AA'].width = 15
            sh['A1'] = '货箱编号'
            sh['B1'] = '件数'
            sh['C1'] = '型号'
            sh['D1'] = '品牌'
            sh['E1'] = '英文品名'
            sh['F1'] = '中文品名'
            sh['G1'] = '单个产品申报价值(usd)'
            sh['H1'] = '单箱申报数量'
            sh['I1'] = '单箱申报总金额'
            sh['J1'] = '英文材质'
            sh['K1'] = '材质'
            sh['L1'] = '用途'
            sh['M1'] = '海关编码'
            sh['N1'] = '亚马逊内部编码ID'
            sh['O1'] = '销售网址'
            sh['P1'] = '销售价格'
            sh['Q1'] = '产品重量(kg)'
            sh['R1'] = '产品尺寸(长*宽*高cm)'
            sh['S1'] = '是否带电'
            sh['T1'] = '是否带磁'
            sh['U1'] = 'ASIN'
            sh['V1'] = 'FNSKU'
            sh['W1'] = '货箱重量(kg)'
            sh['X1'] = '货箱长度(cm)'
            sh['Y1'] = '货箱宽度(cm)'
            sh['Z1'] = '货箱高度(cm)'
            sh['AA1'] = '图片(不能超出单元格)'

            boxes = ShipBox.objects.filter(ship__id=ship_id)
            box_num = 0
            num = 0
            for b in boxes:
                ship_detail = ShipDetail.objects.filter(
                    ship__id=ship_id, box_number=b.box_number)

                box_tag = 1
                for i in ship_detail:
                    sh.row_dimensions[num + 2].height = 100
                    sh['A' +
                       str(num + 2)] = i.ship.batch + '/' + str(box_num + 1)
                    sh['B' + str(num + 2)] = box_tag
                    sh['C' + str(num + 2)] = i.sku
                    sh['D' + str(num + 2)] = i.brand
                    sh['E' + str(num + 2)] = i.en_name
                    sh['F' + str(num + 2)] = i.cn_name
                    sh['G' + str(num + 2)] = i.declared_value
                    sh['H' + str(num + 2)] = i.qty
                    sh['I' + str(num + 2)] = i.declared_value * i.qty
                    sh['J' + str(num + 2)] = i.en_material
                    sh['K' + str(num + 2)] = i.cn_material
                    sh['L' + str(num + 2)] = i.use
                    sh['M' + str(num + 2)] = i.custom_code
                    sh['N' + str(num + 2)] = ''
                    sh['O' + str(
                        num + 2
                    )] = 'https://articulo.mercadolibre.com.mx/MLM-' + i.item_id
                    sh['P' + str(num + 2)] = ''
                    sh['Q' + str(num + 2)] = ''
                    sh['R' + str(num + 2)] = ''
                    sh['S' + str(num + 2)] = '是' if i.is_elec else '否'
                    sh['T' + str(num + 2)] = '否'
                    sh['U' + str(num + 2)] = ''
                    sh['V' + str(num + 2)] = ''
                    sh['W' + str(num + 2)] = ''
                    sh['X' + str(num + 2)] = ''
                    sh['Y' + str(num + 2)] = ''
                    sh['Z' + str(num + 2)] = ''

                    img = Image('media/ml_product/' + i.sku + '_100x100.jpg')
                    img.width, img.height = 80, 80
                    sh.add_image(img, 'AA' + str(num + 2))

                    box_tag = 0
                    num += 1
                box_num += 1
            wb.save('media/export/盛德物流申报-' + ship.shop + '.xlsx')
            url = BASE_URL + '/media/export/盛德物流申报-' + ship.shop + '.xlsx'

            # 创建操作日志
            log = MLOperateLog()
            log.op_module = 'SHIP'
            log.op_type = 'CREATE'
            log.target_type = 'SHIP'
            log.target_id = ship.id
            log.desc = '导出盛德申报单'
            log.user = request.user
            log.save()
        if logistic_name == 'WEICAO':
            from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

            alignment = Alignment(horizontal='center', vertical='center')
            v_alignment = Alignment(vertical='center')
            title_font = Font(name='微软雅黑', sz=10, b=True)
            border = Border(left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'),
                            top=Side(style='thin', color='000000'),
                            bottom=Side(style='thin', color='000000'))
            wb = openpyxl.Workbook()
            boxes = ShipBox.objects.filter(ship__id=ship_id)
            n = 0
            for b in boxes:
                n += 1
                if n == 1:
                    sh = wb.active
                    sh.title = '第1箱'
                else:
                    sh = wb.create_sheet('第' + str(n) + '箱', n)

                sh.column_dimensions['A'].width = 30
                sh.column_dimensions['B'].width = 30
                sh.column_dimensions['F'].width = 30
                sh.column_dimensions['G'].width = 30
                sh.row_dimensions[1].height = 30
                sh.row_dimensions[2].height = 30
                sh.row_dimensions[3].height = 30
                sh.row_dimensions[4].height = 30
                sh.row_dimensions[5].height = 30
                sh['A1'] = '发货人公司名称（填写英文）：'
                sh['B1'] = 'Shenzhen Suke Technology Co., Ltd'
                sh['F1'] = '收货人名/送仓地址（英文）：'
                sh['A2'] = '地址（英文）：'
                sh['B2'] = 'Room 820, 8th Floor, Aihua Building, No. 2038, Shennan Middle Road, Fuqiang Community, Huaqiangbei Street, Futian District, Shenzhen'
                sh['F2'] = '地址（英文）'
                sh['A3'] = '联系电话：'
                sh['B3'] = '13823289200'
                sh['F3'] = '联系电话：'
                sh['A4'] = 'Amazon Reference ID:'
                sh['F4'] = 'shipment ID'
                sh['A5'] = '图片'
                sh['B5'] = '英文品名'
                sh['C5'] = '中文品名'
                sh['D5'] = '申报单价（USD）'
                sh['E5'] = '数量（个）'
                sh['F5'] = 'SKU'
                sh['G5'] = '申报总金额（USD）'

                sh.merge_cells('B1:E1')
                sh.merge_cells('B2:E2')
                sh.merge_cells('B3:E3')
                sh.merge_cells('B4:E4')

                area = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
                for i in area:
                    sh[i + '1'].border = border
                    sh[i + '2'].border = border
                    sh[i + '3'].border = border
                    sh[i + '4'].border = border

                    sh[i + '5'].font = title_font
                    sh[i + '5'].border = border
                    sh[i + '5'].alignment = alignment

                num = 0
                all_qty = 0
                ship_detail = ShipDetail.objects.filter(
                    ship__id=ship_id, box_number=b.box_number)
                for i in ship_detail:
                    sh.row_dimensions[num + 6].height = 100

                    img = Image('media/ml_product/' + i.sku + '_100x100.jpg')
                    img.width, img.height = 80, 80
                    sh.add_image(img, 'A' + str(num + 6))
                    sh['A' + str(num + 6)].border = border

                    sh['B' + str(num + 6)] = i.en_name
                    sh['B' + str(num + 6)].border = border
                    sh['B' + str(num + 6)].alignment = alignment
                    sh['C' + str(num + 6)] = i.cn_name
                    sh['C' + str(num + 6)].border = border
                    sh['C' + str(num + 6)].alignment = alignment
                    sh['D' + str(num + 6)] = i.declared_value
                    sh['D' + str(num + 6)].border = border
                    sh['D' + str(num + 6)].alignment = alignment
                    sh['E' + str(num + 6)] = i.qty
                    sh['E' + str(num + 6)].border = border
                    sh['E' + str(num + 6)].alignment = alignment
                    sh['F' + str(num + 6)] = i.sku
                    sh['F' + str(num + 6)].border = border
                    sh['F' + str(num + 6)].alignment = alignment
                    sh['G' + str(num + 6)] = i.declared_value * i.qty
                    sh['G' + str(num + 6)].border = border
                    sh['G' + str(num + 6)].alignment = alignment

                    num += 1
                    all_qty += i.qty
                sh['A' + str(num + 6)] = '总计'
                sh['A' + str(num + 6)].font = title_font
                sh['A' + str(num + 6)].border = border
                sh['A' + str(num + 6)].alignment = alignment
                sh['B' + str(num + 6)] = all_qty
                sh['B' + str(num + 6)].border = border
                sh['B' + str(num + 6)].alignment = alignment
                sh.row_dimensions[num + 6].height = 30
                sh.row_dimensions[num + 7].height = 30
                sh['A' + str(num + 7)] = '这里填这箱子尺寸及重量'
                sh['A' + str(num + 7)].font = title_font
                sh['A' + str(num + 7)].border = border
                sh['A' + str(num + 7)].alignment = alignment
                sh['B' + str(num + 7)] = '{w}kg / {l}x{h}x{k}cm'.format(
                    w=b.weight, l=b.length, h=b.heigth, k=b.width)
                sh['B' + str(num + 7)].border = border
                sh['B' + str(num + 7)].alignment = alignment

            wb.save('media/export/微草空运-' + ship.shop + '.xlsx')
            url = BASE_URL + '/media/export/微草空运-' + ship.shop + '.xlsx'

            # 创建操作日志
            log = MLOperateLog()
            log.op_module = 'SHIP'
            log.op_type = 'CREATE'
            log.target_type = 'SHIP'
            log.target_id = ship.id
            log.desc = '导出微草空运申报单'
            log.user = request.user
            log.save()

        return Response({
            'url': url,
            'status': 'success'
        },
                        status=status.HTTP_200_OK)

    # 导出采购单
    @action(methods=['post'], detail=False, url_path='export_purchase')
    def export_purchase(self, request):
        from openpyxl.drawing.image import Image

        ship_id = request.data['id']
        ship = Ship.objects.filter(id=ship_id).first()
        wb = openpyxl.Workbook()
        sh = wb.active
        sh.title = 'FBM发仓列表（' + ship.shop + '）'
        sh.column_dimensions['C'].width = 15
        sh.column_dimensions['D'].width = 15
        sh.column_dimensions['E'].width = 50
        sh.column_dimensions['F'].width = 15
        sh.column_dimensions['L'].width = 50
        sh['A1'] = '类型'
        sh['B1'] = 'SKU'
        sh['C1'] = 'UPC'
        sh['D1'] = 'MLM'
        sh['E1'] = '名称'
        sh['F1'] = '图片'
        sh['G1'] = '包材名称'
        sh['H1'] = '包材尺寸'
        sh['I1'] = '数量'
        sh['J1'] = '单重kg'
        sh['K1'] = '总重kg'
        sh['L1'] = '单价'
        sh['M1'] = '小计'
        sh['N1'] = '备注'

        ship_detail = ShipDetail.objects.filter(ship=ship)
        num = 0
        for i in ship_detail:
            sh['A' + str(num + 2)] = '新入仓' if i.s_type == 'NEW' else '补仓'
            sh['B' + str(num + 2)] = i.sku
            sh['C' + str(num + 2)] = i.upc
            sh['D' + str(num + 2)] = i.item_id
            sh['E' + str(num + 2)] = i.p_name

            sh.row_dimensions[num + 2].height = 100
            if i.image:
                # 获取完整文件名（含路径，如 "uploads/images/2025/07/01/test.jpg"）
                full_name = i.image.name
                # 提取仅文件名（不含路径和扩展名，如 "test"）
                file_base_name = os.path.basename(full_name)  # 结果："test.jpg"
                img_name = os.path.splitext(file_base_name)[0]  # 结果："test"
                img = Image('media/ml_product/' + img_name + '_100x100.jpg')
                img.width, img.height = 100, 100
                sh.add_image(img, 'F' + str(num + 2))

            sh['G' + str(num + 2)] = i.packing_name
            sh['H' + str(num + 2)] = i.packing_size
            sh['I' + str(num + 2)] = i.qty
            sh['J' + str(num + 2)] = i.weight
            sh['K' + str(num + 2)] = i.weight * i.qty
            sh['L' + str(num + 2)] = i.unit_cost
            sh['M' + str(num + 2)] = i.unit_cost * i.qty
            sh['N' + str(num + 2)] = i.note

            num += 1
        wb.save('media/export/美客多采购单-' + ship.shop + '.xlsx')
        url = BASE_URL + '/media/export/美客多采购单-' + ship.shop + '.xlsx'

        # 创建操作日志
        log = MLOperateLog()
        log.op_module = 'SHIP'
        log.op_type = 'CREATE'
        log.target_type = 'SHIP'
        log.target_id = ship.id
        log.desc = '导出采购单'
        log.user = request.user
        log.save()

        return Response({'url': url}, status=status.HTTP_200_OK)

    # 导出打包质检单
    @action(methods=['post'], detail=False, url_path='export_qc')
    def export_qc(self, request):
        from openpyxl.drawing.image import Image
        from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

        alignment = Alignment(horizontal='center', vertical='center')
        v_alignment = Alignment(vertical='center')
        title_font = Font(name='微软雅黑', sz=10, b=True)
        big_title_font = Font(name='微软雅黑', sz=15, b=True)
        border = Border(left=Side(border_style='thin', color='000000'),
                        right=Side(border_style='thin', color='000000'),
                        top=Side(style='thin', color='000000'),
                        bottom=Side(style='thin', color='000000'))

        ship_id = request.data['id']
        ship = Ship.objects.filter(id=ship_id).first()
        wb = openpyxl.Workbook()
        sh = wb.active
        sh.title = '美客多质检单（' + ship.shop + '）'
        sh.column_dimensions['C'].width = 50
        sh.column_dimensions['D'].width = 15
        sh.column_dimensions['E'].width = 15
        sh.column_dimensions['H'].width = 15
        sh.column_dimensions['I'].width = 15
        sh.column_dimensions['J'].width = 15
        sh['A1'] = '类型'
        sh['B1'] = 'SKU'
        sh['C1'] = '名称'
        sh['D1'] = '图片'
        sh['E1'] = '包材名称'
        sh['F1'] = '数量'
        sh['G1'] = '箱号'
        sh['H1'] = '配货抽检（签名）'
        sh['I1'] = '包货抽检（签名）'
        sh['J1'] = '装 箱（签名）'
        sh['K1'] = '备注'

        area = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
        for i in area:
            sh[i + '1'].alignment = alignment
            sh[i + '1'].font = title_font
            sh[i + '1'].border = border
            sh[i + '2'].border = border

        sh.merge_cells('A2:K2')
        sh['A2'] = ship.shop
        sh['A2'].alignment = alignment
        sh['A2'].font = big_title_font
        sh['A2'].fill = PatternFill(patternType='solid', fgColor='65d6b4')

        ship_detail = ShipDetail.objects.filter(ship=ship)
        num = 1
        for i in ship_detail:
            sh['A' + str(num + 2)] = '新入仓' if i.s_type == 'NEW' else '补仓'
            sh['A' + str(num + 2)].alignment = alignment
            sh['A' + str(num + 2)].border = border
            sh['B' + str(num + 2)] = i.sku
            sh['B' + str(num + 2)].alignment = alignment
            sh['B' + str(num + 2)].border = border
            sh['C' + str(num + 2)] = i.p_name
            sh['C' + str(num + 2)].alignment = v_alignment
            sh['C' + str(num + 2)].border = border

            sh.row_dimensions[num + 2].height = 100
            if i.image:
                # 获取完整文件名（含路径，如 "uploads/images/2025/07/01/test.jpg"）
                full_name = i.image.name
                # 提取仅文件名（不含路径和扩展名，如 "test"）
                file_base_name = os.path.basename(full_name)  # 结果："test.jpg"
                img_name = os.path.splitext(file_base_name)[0]  # 结果："test"
                img = Image('media/ml_product/' + img_name + '_100x100.jpg')
                img.width, img.height = 100, 100
                sh.add_image(img, 'D' + str(num + 2))
                sh['D' + str(num + 2)].alignment = alignment
                sh['D' + str(num + 2)].border = border

            sh['E' + str(num + 2)] = i.packing_name
            sh['E' + str(num + 2)].border = border
            sh['F' + str(num + 2)] = i.qty
            sh['F' + str(num + 2)].border = border
            sh['F' + str(num + 2)].alignment = alignment
            sh['G' + str(num + 2)] = i.box_number
            sh['G' + str(num + 2)].alignment = alignment
            sh['G' + str(num + 2)].border = border
            sh['H' + str(num + 2)].border = border
            sh['I' + str(num + 2)].border = border
            sh['J' + str(num + 2)].border = border
            sh['K' + str(num + 2)] = i.note
            sh['K' + str(num + 2)].border = border

            num += 1
        wb.save('media/export/美客多质检单-' + ship.shop + '.xlsx')
        url = BASE_URL + '/media/export/美客多质检单-' + ship.shop + '.xlsx'

        # 创建操作日志
        log = MLOperateLog()
        log.op_module = 'SHIP'
        log.op_type = 'CREATE'
        log.target_type = 'SHIP'
        log.target_id = ship.id
        log.desc = '导出质检单'
        log.user = request.user
        log.save()

        return Response({'url': url}, status=status.HTTP_200_OK)

    # 修改运费结算状态
    @action(methods=['post'], detail=False, url_path='change_logi_fee_status')
    def change_logi_fee_status(self, request):
        ship_id = request.data['id']
        logi_fee_clear = request.data['logi_fee_clear']

        ship = Ship.objects.filter(id=ship_id).first()
        ship.logi_fee_clear = logi_fee_clear
        ship.save()

        # 创建操作日志
        log = MLOperateLog()
        log.op_module = 'SHIP'
        log.op_type = 'EDIT'
        log.target_type = 'SHIP'
        log.target_id = ship.id
        log.desc = '标记物流费用已结算' if logi_fee_clear else '取消标记物流费用结算'
        log.user = request.user
        log.save()
        return Response({'msg': '操作成功'}, status=status.HTTP_200_OK)

    # 检查变动清单待处理产品
    @action(methods=['get'], detail=False, url_path='check_ship_change')
    def check_ship_change(self, request):
        # 变动清单待处理产品
        shops_set = Shop.objects.filter(user__id=request.user.id)
        remove_items_count = 0
        if request.user.is_superuser:
            remove_items_count = ShipItemRemove.objects.filter(
                handle=0).count()
        else:
            for i in shops_set:
                remove_items_count += ShipItemRemove.objects.filter(
                    handle=0).filter(belong_shop=i.name).count()

        return Response({'remove_items_count': remove_items_count},
                        status=status.HTTP_200_OK)

    # 检查运单是否已被编辑
    @action(methods=['post'], detail=False, url_path='check_if_ship_edit')
    def check_if_ship_edit(self, request):
        ship_id = request.data['ship_id']
        time_flag = request.data['time_flag']
        ship_changed = False

        # 查出最新修改的日志
        log = MLOperateLog.objects.filter(
            op_module='SHIP',
            target_id=ship_id).filter(Q(op_type='EDIT')
                                      | Q(op_type='DEL')).first()

        if not log:
            return Response({
                'time_flag': '',
                'ship_changed': ship_changed
            },
                            status=status.HTTP_200_OK)

        if log:
            latest_time = log.create_time.strftime("%Y-%m-%d %H:%M:%S")
            if not time_flag:
                time_flag = latest_time
            else:
                if time_flag != latest_time:
                    ship_changed = True

        return Response({
            'time_flag': time_flag,
            'ship_changed': ship_changed
        },
                        status=status.HTTP_200_OK)

    # 跟踪物流单号
    @action(methods=['post'], detail=False, url_path='ship_tracking')
    def ship_tracking(self, request):
        track_num = request.data['track_num']
        message = tasks.ship_tracking(track_num)

        # 创建操作日志
        log = MLOperateLog()
        log.op_module = 'SHIP'
        log.op_type = 'EDIT'
        log.target_type = 'SHIP'
        log.desc = '刷新物流跟踪-{message} 单号{track_num}'.format(track_num=track_num,
                                                           message=message)
        log.user = request.user
        log.save()

        if message != 'SUCCESS':
            return Response({
                'msg': message,
                'status': 'error'
            },
                            status=status.HTTP_202_ACCEPTED)
        return Response({'msg': '操作成功'}, status=status.HTTP_200_OK)

    # 批量更新物流跟踪信息
    @action(methods=['get'], detail=False, url_path='bulk_update_tracking')
    def bulk_update_tracking(self, request):
        tasks.bulk_ship_tracking()
        # 创建操作日志
        log = MLOperateLog()
        log.op_module = 'SHIP'
        log.op_type = 'EDIT'
        log.target_type = 'SHIP'
        log.desc = '手动批量刷新运单物流跟踪信息'
        log.user = request.user
        log.save()
        return Response({'msg': '操作成功'}, status=status.HTTP_200_OK)

    # 盛德标签
    @action(methods=['post'], detail=False, url_path='carrier_label')
    def carrier_label(self, request):
        label_type = request.data['label_type']  # 箱唛：BOX 交运单：RECEIPT
        s_number = request.data['s_number']

        c_path = os.path.join(BASE_DIR, "site_config.json")
        with open(c_path, "r", encoding="utf-8") as f:
            data = json.load(f)  # 加载配置数据
        header = {'Cookie': data['sd_cookies']}

        # 查询运单详情，获取operNo
        url_1 = 'http://client.sanstar.net.cn/console/customer_order/get_order_list'
        data_1 = {
            'thecompany': 'SO',
            'operNo': s_number,
            'op': 1,
        }
        resp = requests.post(url_1, data=data_1, headers=header)

        if 'success' in resp.json():
            # 查询运单异常情况
            return Response({
                'msg': resp.json()['msg'],
                'status': 'error'
            },
                            status=status.HTTP_202_ACCEPTED)

        # 查到运单（已受理）
        if resp.json()[0]['tb']:
            ship_order = resp.json()[0]['tb'][0]
            oper_no = ship_order['operNo']

            # 箱唛标签
            url_2 = ''
            data_2 = {}
            if label_type == 'BOX':
                url_2 = 'http://client.sanstar.net.cn/console/Report/shippingmark'
                data_2 = {
                    'param':
                    '[{"operNo":' + oper_no +
                    ',"dsid":"8CF7FA95-1F7B-4F03-BD27-F33E9EA9F6FC","type":1,"proc":"client_add_BigWaybill_warehousereceipt","TheCompany":"10571","country":"墨西哥","repottype":"pdf","EntrustType":"空运","serialno":0,"new_num":0}]'
                }

            # 交运单
            if label_type == 'RECEIPT':
                url_2 = 'http://client.sanstar.net.cn/console/Report/getwarehousereceipt'
                data_2 = {
                    'param':
                    '[{"type":1,"operNo":"' + oper_no +
                    '","dsid":"8CF7FA95-1F7B-4F03-BD27-F33E9EA9F6FC","proc":"cliect_交仓单","warehouseid":"0200","regionid":30,"repottype":"pdf"}]'
                }

            # 获取标签链接
            resp2 = requests.post(url_2, data=data_2, headers=header)
            if 'success' in resp2.json():
                if resp2.json()['success']:
                    # 创建操作日志
                    log = MLOperateLog()
                    log.op_module = 'SHIP'
                    log.op_type = 'EDIT'
                    log.target_type = 'SHIP'
                    log.desc = '打印物流标签-{message} 单号{track_num}'.format(
                        track_num=s_number,
                        message='箱唛' if label_type == 'BOX' else '交运单')
                    log.user = request.user
                    log.save()
                    return Response(
                        {
                            'link': resp2.json()['msg'],
                            'status': 'success'
                        },
                        status=status.HTTP_200_OK)

        else:
            return Response({
                'msg': '运单待受理，无法生成标签',
                'status': 'error'
            },
                            status=status.HTTP_202_ACCEPTED)

        return Response({
            'msg': '生成标签异常',
            'status': 'error'
        },
                        status=status.HTTP_202_ACCEPTED)

    # 盛德交运运单
    @action(methods=['post'], detail=False, url_path='carrier_place_order')
    def carrier_place_order(self, request):
        ship_id = request.data['ship_id']
        d_code = request.data['d_code']
        fbm = FBMWarehouse.objects.filter(w_code=d_code).first()
        if not fbm:
            return Response({
                'msg': 'fbm仓库不存在!',
                'status': 'error'
            },
                            status=status.HTTP_202_ACCEPTED)
        ship_data = {
            'd_code': d_code,
            'address1': fbm.address,
            'zip_code': fbm.zip if fbm.zip else ' ',
            'reserveid': request.data['reserveid'],
            'apptdate': request.data['apptdate'],
            'EntrustType': request.data['EntrustType'],
            'warehouseid': request.data['warehouseid'],
            'DeliveryTime': request.data['DeliveryTime'],
            'sellerid': request.data['sellerid'],
            'envio': request.data['envio'],
            'product': request.data['product'],
            'ProductNature': request.data['ProductNature'],
        }

        info = tasks.sd_place_order(ship_id, ship_data)
        if info['status'] == 'error':
            return Response({
                'msg': info['msg'],
                'status': 'error'
            },
                            status=status.HTTP_202_ACCEPTED)
        return Response({
            'msg': info['msg'],
            'status': 'success'
        },
                        status=status.HTTP_200_OK)

    # 盛德交运前检查
    @action(methods=['post'],
            detail=False,
            url_path='check_before_place_order')
    def check_before_place_order(self, request):
        ship_id = request.data['ship_id']
        ship = Ship.objects.filter(id=ship_id).first()
        sd_set = ShipDetail.objects.filter(ship=ship)
        is_packed = True  # 产品打包状态
        is_declare = True  # 产品申报状态
        is_file = False  # 箱唛附件
        all_status = False  # 总状态
        for i in sd_set:
            if not i.box_number:
                is_packed = False
            if not i.custom_code or not i.en_name or not i.brand or not i.declared_value or not i.cn_material:
                is_declare = False

        sa_set = ShipAttachment.objects.filter(ship=ship, a_type='BOX_LABEL')
        for i in sa_set:
            extension = i.name.split(".")[-1]
            if extension == 'pdf':
                is_file = True

        if is_packed and is_declare and is_file:
            all_status = True
        return Response(
            {
                'is_packed': is_packed,
                'is_declare': is_declare,
                'is_file': is_file,
                'all_status': all_status
            },
            status=status.HTTP_200_OK)

    # 查询盛德运单受理状态
    @action(methods=['get'], detail=False, url_path='check_sd_order_status')
    def check_sd_order_status(self, request):
        tasks.query_sd_order_status()
        return Response({
            'msg': '刷新成功！',
            'status': 'success'
        },
                        status=status.HTTP_200_OK)

    # 盛德对账查询
    @action(methods=['post'], detail=False, url_path='bill_check')
    def bill_check(self, request):
        ships = request.data['ships']
        checked_ships = []
        for i in ships:
            sp = Ship.objects.filter(
                s_number=i['in_s_number'],
                envio_number=i['in_envio_number']).first()
            in_weight = round(float(i['in_weight']), 2)
            in_price = round(float(i['in_price']), 2)
            if sp:
                checked_ships.append({
                    'in_s_number': i['in_s_number'],
                    'in_envio_number': i['in_envio_number'],
                    'in_total_box': i['in_total_box'],
                    'in_weight': in_weight,
                    'in_price': in_price,
                    'in_shipping_fee': i['in_shipping_fee'],
                    'id': sp.id,
                    'total_box': sp.total_box,
                    'weight': sp.weight,
                    'batch': sp.batch,
                    's_status': sp.s_status,
                    'shop': sp.shop,
                    'logi_fee_clear': sp.logi_fee_clear,
                    'note': sp.note,
                    'confirmed': False
                })
            else:
                checked_ships.append({
                    'in_s_number': i['in_s_number'],
                    'in_envio_number': i['in_envio_number'],
                    'in_total_box': i['in_total_box'],
                    'in_weight': in_weight,
                    'in_price': in_price,
                    'in_shipping_fee': i['in_shipping_fee'],
                    'id': 0,
                    'total_box': '',
                    'weight': '',
                    'batch': '',
                    's_status': '',
                    'shop': '',
                    'logi_fee_clear': '',
                    'note': '',
                    'confirmed': False
                })
        return Response({'ships': checked_ships}, status=status.HTTP_200_OK)

    # 盛德对账提交
    @action(methods=['post'], detail=False, url_path='bill_submit')
    def bill_submit(self, request):
        ships = request.data['ships']
        for i in ships:
            if i['confirmed'] and i['id']:
                sp = Ship.objects.filter(id=i['id']).first()
                sp.shipping_fee = i['in_shipping_fee']
                sp.logi_fee_clear = True
                sp.save()

                i['logi_fee_clear'] = True

                # 计算每个产品运费占比
                all_fee = sp.shipping_fee + sp.extra_fee
                queryset = ShipDetail.objects.filter(ship=sp)
                total_weight = 0
                total_cbm = 0

                if sp.ship_type == '空运':
                    for i in queryset:
                        total_weight += (i.weight * i.qty)
                    for i in queryset:
                        if total_weight:
                            percent = (i.weight * i.qty) / total_weight
                        else:
                            percent = 0
                        avg_ship_fee = percent * all_fee / i.qty
                        i.avg_ship_fee = avg_ship_fee
                        i.save()

                if sp.ship_type == '海运':
                    for i in queryset:
                        cbm = i.length * i.width * i.heigth / 1000000
                        total_cbm += (cbm * i.qty)
                    for i in queryset:
                        cbm = i.length * i.width * i.heigth / 1000000
                        percent = cbm / total_cbm
                        avg_ship_fee = percent * all_fee
                        i.avg_ship_fee = avg_ship_fee
                        i.save()

                # 创建操作日志
                log = MLOperateLog()
                log.op_module = 'SHIP'
                log.op_type = 'EDIT'
                log.target_type = 'SHIP'
                log.target_id = sp.id
                log.desc = '物流费用已结算(对账),结算运费 {shipping_fee}'.format(
                    shipping_fee=sp.shipping_fee)
                log.user = request.user
                log.save()
        return Response({
            'ships': ships,
            'msg': '操作成功！',
            'status': 'success'
        },
                        status=status.HTTP_200_OK)

    # 盛德cookies设置
    @action(methods=['post', 'get'],
            detail=False,
            url_path='sd_cookies_setting')
    def sd_cookies_setting(self, request):
        from pathlib import Path
        config_path = Path(__file__).parent.parent / "site_config.json"
        # 读取现有配置
        with open(config_path, 'r') as f:
            config = json.load(f)
        if request.method == 'GET':
            return Response({
                'sd_cookies': config['sd_cookies'],
            },
                            status=status.HTTP_200_OK)

        elif request.method == 'POST':
            cookies = request.data['sd_cookies']
            config['sd_cookies'] = cookies
            # 写回文件
            with open(config_path, 'w') as f:
                json.dump(config, f, indent=4)
            return Response({
                'msg': '操作成功！',
                'status': 'success'
            },
                            status=status.HTTP_200_OK)
        else:
            return Response({
                'msg': '请求方法错误!',
                'status': 'error'
            },
                            status=status.HTTP_202_ACCEPTED)

    # 导出OZON产品入仓单
    @action(methods=['post'],
            detail=False,
            url_path='export_ozon_product_import')
    def export_ozon_product_import(self, request):
        from openpyxl.styles import Font
        ship_id = request.data['id']
        ship = Ship.objects.filter(id=ship_id).first()
        wb = openpyxl.Workbook()
        sh = wb.active
        sh.title = 'Sheet1'
        title_font = Font(name='Arial')
        sh['A1'] = 'артикул'
        sh['A1'].font = title_font
        sh['B1'] = 'имя (необязательно)'
        sh['B1'].font = title_font
        sh['C1'] = 'количество'
        sh['C1'].font = title_font
        ship_detail = ShipDetail.objects.filter(ship__id=ship_id)
        num = 0
        for i in ship_detail:
            sh['A' + str(num + 2)] = i.sku
            sh['C' + str(num + 2)] = i.qty
            num += 1
        wb.save('media/export/OZON产品入仓单-' + ship.shop + '.xlsx')
        url = BASE_URL + '/media/export/OZON产品入仓单-' + ship.shop + '.xlsx'

        # 创建操作日志
        log = MLOperateLog()
        log.op_module = 'SHIP'
        log.op_type = 'CREATE'
        log.target_type = 'SHIP'
        log.target_id = ship.id
        log.desc = '导出OZON产品入仓单'
        log.user = request.user
        log.save()

        return Response({
            'url': url,
            'status': 'success'
        },
                        status=status.HTTP_200_OK)

    # 导出OZON产品装箱单
    @action(methods=['post'], detail=False, url_path='export_ozon_package')
    def export_ozon_package(self, request):
        from openpyxl.styles import Font
        ship_id = request.data['id']
        ship = Ship.objects.filter(id=ship_id).first()

        sd_set = ShipDetail.objects.filter(ship__id=ship_id)
        # 检查货品装箱情况
        for i in sd_set:
            if not i.box_number:
                return Response({
                    'msg': '有货品未装箱，请装箱后再导出!',
                    'status': 'error'
                },
                                status=status.HTTP_202_ACCEPTED)
        boxes = ShipBox.objects.filter(ship__id=ship_id)
        # 检查ozon箱唛号是否有填写
        for i in boxes:
            if not i.note:
                return Response({
                    'msg': '有平台箱唛未备注，请填写后再导出!',
                    'status': 'error'
                },
                                status=status.HTTP_202_ACCEPTED)

        wb = openpyxl.Workbook()
        sh = wb.active
        sh.title = 'Состав ГМ поставки'
        title_font = Font(name='Arial')
        sh['A1'] = 'ШК или артикул товара'
        sh['A1'].font = title_font
        sh['B1'] = 'Кол-во товаров'
        sh['B1'].font = title_font
        sh['C1'] = 'Зона размещения'
        sh['C1'].font = title_font
        sh['D1'] = 'ШК ГМ'
        sh['D1'].font = title_font
        sh['E1'] = 'Тип ГМ (не обязательно)'
        sh['E1'].font = title_font
        sh['F1'] = 'Срок годности ДО в формате YYYY-MM-DD (не более 1 СГ на 1 SKU в 1 ГМ)'
        sh['F1'].font = title_font
        ship_detail = ShipDetail.objects.filter(ship__id=ship_id)
        num = 0
        for i in ship_detail:
            box = ShipBox.objects.filter(ship__id=ship_id,
                                         box_number=i.box_number).first()
            sh['A' + str(num + 2)] = i.sku
            sh['B' + str(num + 2)] = i.qty
            sh['D' + str(num + 2)] = box.note
            sh['E' + str(num + 2)] = 'Коробка'
            sh['E' + str(num + 2)].font = title_font
            num += 1
        wb.save('media/export/OZON产品装箱单-' + ship.shop + '.xlsx')
        url = BASE_URL + '/media/export/OZON产品装箱单-' + ship.shop + '.xlsx'

        # 创建操作日志
        log = MLOperateLog()
        log.op_module = 'SHIP'
        log.op_type = 'CREATE'
        log.target_type = 'SHIP'
        log.target_id = ship.id
        log.desc = '导出OZON产品装箱单'
        log.user = request.user
        log.save()

        return Response({
            'url': url,
            'status': 'success'
        },
                        status=status.HTTP_200_OK)

    # 导出OZON产品装箱单-跨箱
    @action(methods=['post'],
            detail=False,
            url_path='export_ozon_package_in_muti_box')
    def export_ozon_package_in_muti_box(self, request):
        from openpyxl.styles import Font
        ship_id = request.data['id']
        ship = Ship.objects.filter(id=ship_id).first()

        p_list = request.data['p_list']
        sku_list = []  # 跨箱的sku
        for i in p_list:
            sku_list.append(i[0])
            # 检查格式
            if not ShipDetail.objects.filter(ship__id=ship_id,
                                             sku=i[0]).first():
                return Response({
                    'msg': '有平台箱唛备注有误，请检查!',
                    'status': 'error'
                },
                                status=status.HTTP_202_ACCEPTED)

        wb = openpyxl.Workbook()
        sh = wb.active
        sh.title = 'Состав ГМ поставки'
        title_font = Font(name='Arial')
        sh['A1'] = 'ШК или артикул товара'
        sh['A1'].font = title_font
        sh['B1'] = 'Кол-во товаров'
        sh['B1'].font = title_font
        sh['C1'] = 'Зона размещения'
        sh['C1'].font = title_font
        sh['D1'] = 'ШК ГМ'
        sh['D1'].font = title_font
        sh['E1'] = 'Тип ГМ (не обязательно)'
        sh['E1'].font = title_font
        sh['F1'] = 'Срок годности ДО в формате YYYY-MM-DD (не более 1 СГ на 1 SKU в 1 ГМ)'
        sh['F1'].font = title_font
        ship_detail = ShipDetail.objects.filter(ship__id=ship_id)
        # 不跨箱sku加入跨箱sku列表
        for i in ship_detail:
            if i.sku not in sku_list:
                p_list.append([i.sku, i.box_number, i.qty])

        num = 0
        for i in p_list:
            box = ShipBox.objects.filter(ship__id=ship_id,
                                         box_number=i[1]).first()
            sh['A' + str(num + 2)] = i[0]
            sh['B' + str(num + 2)] = int(i[2])
            sh['D' + str(num + 2)] = box.note
            sh['E' + str(num + 2)] = 'Коробка'
            sh['E' + str(num + 2)].font = title_font
            num += 1
        wb.save('media/export/OZON产品装箱单-' + ship.shop + '.xlsx')
        url = BASE_URL + '/media/export/OZON产品装箱单-' + ship.shop + '.xlsx'

        # 创建操作日志
        log = MLOperateLog()
        log.op_module = 'SHIP'
        log.op_type = 'CREATE'
        log.target_type = 'SHIP'
        log.target_id = ship.id
        log.desc = '导出OZON产品装箱单-跨箱'
        log.user = request.user
        log.save()

        return Response({
            'url': url,
            'status': 'success'
        },
                        status=status.HTTP_200_OK)

    # 导出OZON物流箱唛号映射表
    @action(methods=['post'],
            detail=False,
            url_path='export_ozon_package_match')
    def export_ozon_package_match(self, request):
        from openpyxl.styles import Font, Alignment
        alignment = Alignment(horizontal='center', vertical='center')
        title_font = Font(name='微软雅黑', sz=15, b=True)
        zi_font = Font(name='微软雅黑', sz=15)
        ship_id = request.data['id']
        ship = Ship.objects.filter(id=ship_id).first()

        boxes = ShipBox.objects.filter(ship__id=ship_id)
        # 检查ozon箱唛号是否有填写
        for i in boxes:
            if not i.note:
                return Response({
                    'msg': '有平台箱唛未备注，请填写后再导出!',
                    'status': 'error'
                },
                                status=status.HTTP_202_ACCEPTED)

        wb = openpyxl.Workbook()
        sh = wb.active
        sh.column_dimensions['A'].width = 40
        sh.row_dimensions[1].height = 30
        sh['A1'] = '箱唛号'
        sh['A1'].alignment = alignment
        sh['A1'].font = title_font
        sh['B1'] = '箱序号'
        sh['B1'].font = title_font
        sh['B1'].alignment = alignment

        num = 0
        for i in boxes:
            sh.row_dimensions[num + 2].height = 30
            sh['A' + str(num + 2)] = i.note
            sh['A' + str(num + 2)].alignment = alignment
            sh['A' + str(num + 2)].font = zi_font
            sh['B' + str(num + 2)] = i.box_number
            sh['B' + str(num + 2)].alignment = alignment
            sh['B' + str(num + 2)].font = zi_font
            num += 1
        wb.save('media/export/OZON物流箱唛号映射表-' + ship.shop + '.xlsx')
        url = BASE_URL + '/media/export/OZON物流箱唛号映射表-' + ship.shop + '.xlsx'

        # 创建操作日志
        log = MLOperateLog()
        log.op_module = 'SHIP'
        log.op_type = 'CREATE'
        log.target_type = 'SHIP'
        log.target_id = ship.id
        log.desc = '导出OZON物流箱唛号映射表'
        log.user = request.user
        log.save()

        return Response({
            'url': url,
            'status': 'success'
        },
                        status=status.HTTP_200_OK)

    # 重写
    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()

        # 创建操作日志
        log = MLOperateLog()
        log.op_module = 'SHIP'
        log.op_type = 'DEL'
        log.target_type = 'SHIP'
        log.target_id = instance.id
        log.desc = '删除运单 {batch}-{shop}'.format(batch=instance.batch,
                                                shop=instance.shop)
        log.user = request.user
        log.save()

        self.perform_destroy(instance)
        return Response(status=status.HTTP_204_NO_CONTENT)


class ShipDetailViewSet(mixins.ListModelMixin, mixins.CreateModelMixin,
                        mixins.UpdateModelMixin, mixins.DestroyModelMixin,
                        mixins.RetrieveModelMixin, viewsets.GenericViewSet):
    """
    list:
        运单详情列表,分页,过滤,搜索,排序
    create:
        运单详情新增
    retrieve:
        运单详情详情页
    update:
        运单详情修改
    destroy:
        运单详情删除
    """
    queryset = ShipDetail.objects.all()
    serializer_class = ShipDetailSerializer  # 序列化
    pagination_class = DefaultPagination  # 分页

    filter_backends = (DjangoFilterBackend, filters.SearchFilter,
                       filters.OrderingFilter)  # 过滤,搜索,排序
    filter_fields = ('ship', 'box_number', 's_type')  # 配置过滤字段
    search_fields = ('sku', 'p_name', 'label_code', 'upc', 'item_id')  # 配置搜索字段
    ordering_fields = ('create_time', 'qty')  # 配置排序字段


class ShipItemRemoveViewSet(mixins.ListModelMixin, mixins.CreateModelMixin,
                            mixins.UpdateModelMixin, mixins.DestroyModelMixin,
                            mixins.RetrieveModelMixin,
                            viewsets.GenericViewSet):
    """
    list:
        遗弃清单列表,分页,过滤,搜索,排序
    create:
        遗弃清单新增
    retrieve:
        遗弃清单详情页
    update:
        遗弃清单修改
    destroy:
        遗弃清单删除
    """
    queryset = ShipItemRemove.objects.all()
    serializer_class = ShipItemRemoveSerializer  # 序列化
    pagination_class = DefaultPagination  # 分页

    filter_backends = (DjangoFilterBackend, filters.SearchFilter,
                       filters.OrderingFilter)  # 过滤,搜索,排序
    # filter_fields = ('ship', 'item_type', 'handle', 'belong_shop')  # 配置过滤字段
    filterset_fields = {
        'ship': ['exact'],
        'item_type': ['exact'],
        'handle': ['exact'],
        'belong_shop': ['exact', 'in'],
    }
    search_fields = ('sku', 'p_name', 'item_id')  # 配置搜索字段
    ordering_fields = ('create_time', 'item_type', 'ship__id')  # 配置排序字段

    # 恢复遗弃项
    @action(methods=['post'], detail=False, url_path='restore_remove')
    def restore_remove(self, request):
        sir_id = request.data['id']

        sir = ShipItemRemove.objects.filter(id=sir_id).first()
        if sir:
            product = MLProduct.objects.filter(sku=sir.sku).first()
            sd = ShipDetail()
            sd.ship = sir.ship
            sd.s_type = 'REFILL'
            sd.qty = sir.plan_qty
            sd.plan_qty = sir.plan_qty
            sd.note = '恢复遗弃项'
            sd.sku = sir.sku
            sd.target_FBM = product.shop
            sd.p_name = product.p_name
            sd.label_code = product.label_code
            sd.upc = product.upc
            sd.item_id = product.item_id
            sd.custom_code = product.custom_code
            sd.cn_name = product.cn_name
            sd.en_name = product.en_name
            sd.brand = product.brand
            sd.declared_value = product.declared_value
            sd.cn_material = product.cn_material
            sd.en_material = product.en_material
            sd.use = product.use
            sd.is_elec = product.is_elec
            sd.is_magnet = product.is_magnet
            sd.is_water = product.is_water
            sd.image = product.image
            sd.unit_cost = product.unit_cost
            sd.weight = product.weight
            sd.length = product.length
            sd.width = product.width
            sd.heigth = product.heigth

            packing = Packing.objects.filter(id=product.packing_id).first()
            if packing:
                sd.packing_name = packing.name
                sd.packing_size = packing.size
            sd.save()

            sir.delete()

        return Response({'msg': '操作成功!'}, status=status.HTTP_200_OK)

    # 迁移遗弃项
    @action(methods=['post'], detail=False, url_path='move_items')
    def move_items(self, request):
        ship_id = request.data['ship_id']
        move_method = request.data['move_method']
        product_list = request.data['product_list']

        ship = Ship.objects.filter(id=ship_id).first()
        # 检查产品与目标店铺是否相符
        if ship.target == 'FBM':
            for i in product_list:
                if i['belong_shop'] != ship.shop:
                    return Response({
                        'msg': '产品与目标店铺不符，请核查',
                        'status': 'error'
                    },
                                    status=status.HTTP_202_ACCEPTED)

        sd_set = ShipDetail.objects.filter(ship=ship)
        for i in product_list:
            item_remove = ShipItemRemove.objects.filter(id=i['id']).first()
            p = sd_set.filter(sku=i['sku']).first()
            # 如果产品不存在，则直接新建添加
            if not p:
                product = MLProduct.objects.filter(sku=i['sku']).first()
                sd = ShipDetail()
                sd.ship = ship
                sd.s_type = 'REFILL'
                sd.qty = i['move_qty']
                sd.plan_qty = i['move_qty']
                sd.sku = i['sku']
                sd.target_FBM = product.shop
                sd.p_name = product.p_name
                sd.label_code = product.label_code
                sd.upc = product.upc
                sd.item_id = product.item_id
                sd.custom_code = product.custom_code
                sd.cn_name = product.cn_name
                sd.en_name = product.en_name
                sd.brand = product.brand
                sd.declared_value = product.declared_value
                sd.cn_material = product.cn_material
                sd.en_material = product.en_material
                sd.use = product.use
                sd.is_elec = product.is_elec
                sd.is_magnet = product.is_magnet
                sd.is_water = product.is_water
                sd.image = product.image
                sd.unit_cost = product.unit_cost
                sd.weight = product.weight
                sd.length = product.length
                sd.width = product.width
                sd.heigth = product.heigth

                # 如果店铺库存中没有，发货在途没有，自动标为新品
                is_exist = ShopStock.objects.filter(sku=sd.sku).count()
                is_ship = ShipDetail.objects.filter(sku=sd.sku).filter(
                    Q(ship__s_status='SHIPPED')
                    | Q(ship__s_status='BOOKED')).count()
                if not is_exist and not is_ship:
                    sd.s_type = 'NEW'

                sd.save()

                item_remove.handle = 1
                item_remove.handle_time = datetime.now()
                item_remove.save()

                # 创建操作日志
                log = MLOperateLog()
                log.op_module = 'SHIP'
                log.op_type = 'CREATE'
                log.target_type = 'SHIP'
                log.target_id = ship.id
                log.desc = '迁入产品 {sku} {p_name} {qty}个'.format(
                    sku=product.sku, p_name=product.p_name, qty=i['move_qty'])
                log.user = request.user
                log.save()
                continue

            if p and move_method == 'DEL':
                item_remove.handle = 2
                item_remove.handle_time = datetime.now()
                item_remove.save()
                # 创建操作日志
                log = MLOperateLog()
                log.op_module = 'SHIP'
                log.op_type = 'DEL'
                log.desc = '移除变动清单产品 {sku} {p_name} {qty}个'.format(
                    sku=p.sku, p_name=p.p_name, qty=i['move_qty'])
                log.user = request.user
                log.save()
                continue

            if p and move_method == 'ADD':
                p.qty += i['move_qty']
                p.plan_qty += i['move_qty']
                p.save()

                item_remove.handle = 1
                item_remove.handle_time = datetime.now()
                item_remove.save()
                # 创建操作日志
                log = MLOperateLog()
                log.op_module = 'SHIP'
                log.op_type = 'CREATE'
                log.target_type = 'SHIP'
                log.target_id = ship.id
                log.desc = '叠加迁入产品 {sku} {p_name} {qty}个'.format(
                    sku=p.sku, p_name=p.p_name, qty=i['move_qty'])
                log.user = request.user
                log.save()
                continue

        return Response({
            'msg': '操作成功!',
            'status': 'success'
        },
                        status=status.HTTP_200_OK)

    # 批量移除变动产品
    @action(methods=['post'], detail=False, url_path='del_items')
    def del_items(self, request):
        product_list = request.data
        for i in product_list:
            item_remove = ShipItemRemove.objects.filter(id=i['id']).first()
            item_remove.handle = 2
            item_remove.handle_time = datetime.now()
            item_remove.save()
            # 创建操作日志
            log = MLOperateLog()
            log.op_module = 'SHIP'
            log.op_type = 'DEL'
            log.desc = '移除变动清单产品 {sku} {p_name} {qty}个'.format(
                sku=item_remove.sku,
                p_name=item_remove.p_name,
                qty=item_remove.plan_qty - item_remove.send_qty)
            log.user = request.user
            log.save()
        return Response({
            'msg': '操作成功!',
            'status': 'success'
        },
                        status=status.HTTP_200_OK)

    # 批量保留变动产品
    @action(methods=['post'], detail=False, url_path='keep_items')
    def keep_items(self, request):
        product_list = request.data
        for i in product_list:
            item_remove = ShipItemRemove.objects.filter(id=i['id']).first()
            item_remove.handle = 3
            item_remove.handle_time = datetime.now()
            item_remove.save()
            # 创建操作日志
            log = MLOperateLog()
            log.op_module = 'SHIP'
            log.op_type = 'EDIT'
            log.desc = '保留变动清单产品 {sku} {p_name} {qty}个'.format(
                sku=item_remove.sku,
                p_name=item_remove.p_name,
                qty=item_remove.plan_qty - item_remove.send_qty)
            log.user = request.user
            log.save()
        return Response({
            'msg': '操作成功!',
            'status': 'success'
        },
                        status=status.HTTP_200_OK)


class ShipAttachmentViewSet(mixins.ListModelMixin, mixins.CreateModelMixin,
                            mixins.UpdateModelMixin, mixins.DestroyModelMixin,
                            mixins.RetrieveModelMixin,
                            viewsets.GenericViewSet):
    """
    list:
        运单附件 列表,分页,过滤,搜索,排序
    create:
        运单附件 新增
    retrieve:
        运单附件 详情页
    update:
        运单附件 修改
    destroy:
        运单附件 删除
    """
    queryset = ShipAttachment.objects.all()
    serializer_class = ShipAttachmentSerializer  # 序列化
    pagination_class = DefaultPagination  # 分页

    filter_backends = (DjangoFilterBackend, filters.SearchFilter,
                       filters.OrderingFilter)  # 过滤,搜索,排序
    filter_fields = ('ship', 'a_type')  # 配置过滤字段
    search_fields = ('name', )  # 配置搜索字段
    ordering_fields = ('create_time', 'a_type')  # 配置排序字段

    # 运单附件上传
    @action(methods=['post'], detail=False, url_path='bulk_upload')
    def bulk_upload(self, request):
        import os
        data = request.data
        path = 'media/ml_ships/'
        ship_id = data['id']
        a_type = data['a_type']
        ship = Ship.objects.filter(id=ship_id).first()

        head_path = '{path}{batch}_{id}'.format(path=path,
                                                batch=ship.batch,
                                                id=ship.id)
        # 判断是否存在文件夹,如果没有就创建文件路径
        if not os.path.exists(head_path):
            os.makedirs(head_path)

        file = request.FILES.get('file')
        file_name = file.name

        # 判断是否存在文件
        is_exist = ShipAttachment.objects.filter(ship=ship,
                                                 name=file_name).count()
        if is_exist:
            after_name = os.path.splitext(file_name)[-1]  # 获取扩展名
            first_name = os.path.splitext(file_name)[0]  # 获取文件名
            file_name = '{fn}({time_str})'.format(
                fn=first_name,
                time_str=time.strftime('%m%d%H%M%S')) + after_name

        head_path = head_path + '/' + file_name
        with open(head_path, 'wb') as f:
            for chunk in file.chunks():
                f.write(chunk)
        sa = ShipAttachment()
        sa.ship = ship
        sa.name = file_name
        sa.a_type = a_type
        sa.save()

        # 创建操作日志
        log = MLOperateLog()
        log.op_module = 'SHIP'
        log.op_type = 'CREATE'
        log.target_type = 'SHIP'
        log.target_id = ship_id
        log.desc = '上传附件：' + file_name
        log.user = request.user
        log.save()

        return Response({'msg': '成功上传'}, status=status.HTTP_200_OK)

    @action(methods=['post'], detail=False, url_path='sa_delete')
    def sa_delete(self, request):
        import os
        sa_id = request.data['id']
        sa = ShipAttachment.objects.filter(id=sa_id).first()

        # 创建操作日志
        log = MLOperateLog()
        log.op_module = 'SHIP'
        log.op_type = 'DEL'
        log.target_type = 'SHIP'
        log.target_id = sa.ship.id
        log.desc = '删除附件：' + sa.name
        log.user = request.user
        log.save()

        path = 'media/ml_ships/{batch}_{id}/{name}'.format(batch=sa.ship.batch,
                                                           id=sa.ship.id,
                                                           name=sa.name)
        # path = 'media/ml_ships/' + sa.ship.envio_number + '/' + sa.name
        os.remove(path)
        sa.delete()

        return Response({'msg': '成功删除'}, status=status.HTTP_200_OK)


class ShipBoxViewSet(mixins.ListModelMixin, mixins.CreateModelMixin,
                     mixins.UpdateModelMixin, mixins.DestroyModelMixin,
                     mixins.RetrieveModelMixin, viewsets.GenericViewSet):
    """
    list:
        包装箱列表,分页,过滤,搜索,排序
    create:
        包装箱新增
    retrieve:
        包装箱详情页
    update:
        包装箱修改
    destroy:
        包装箱删除
    """
    queryset = ShipBox.objects.all()
    serializer_class = ShipBoxSerializer  # 序列化
    pagination_class = DefaultPagination  # 分页

    filter_backends = (DjangoFilterBackend, filters.SearchFilter,
                       filters.OrderingFilter)  # 过滤,搜索,排序
    filter_fields = ('ship', )  # 配置过滤字段
    search_fields = ('box_number', 'carrier_box_number')  # 配置搜索字段
    ordering_fields = ('item_qty', 'box_number', 'id')  # 配置排序字段# 创建运单

    # add box
    @action(methods=['post'], detail=False, url_path='add_shipbox')
    def add_shipbox(self, request):
        ship_id = request.data['ship']
        box_number = request.data['box_number']
        length = float(request.data['length'])
        width = float(request.data['width'])
        heigth = float(request.data['heigth'])
        weight = request.data['weight']
        carrier_box_number = request.data['carrier_box_number']
        note = request.data['note']

        ship = Ship.objects.filter(id=ship_id).first()
        if ship:
            box = ShipBox()
            box.ship = ship
            box.box_number = box_number
            box.length = length
            box.width = width
            box.heigth = heigth
            box.weight = weight
            box.carrier_box_number = carrier_box_number
            box.note = note

            cbm = length * width * heigth / 1000000
            box.cbm = cbm
            size_weight = length * width * heigth / 6000
            box.size_weight = size_weight
            box.save()

        # 统计包装箱总重量
        sum_weight = ShipBox.objects.filter(ship=ship).aggregate(Sum('weight'))
        total_weight = sum_weight['weight__sum']
        ship.weight = total_weight

        # 统计包装箱总体积
        sum_cbm = ShipBox.objects.filter(ship=ship).aggregate(Sum('cbm'))
        total_cbm = sum_cbm['cbm__sum']
        ship.cbm = total_cbm
        ship.save()

        # 创建操作日志
        log = MLOperateLog()
        log.op_module = 'SHIP'
        log.op_type = 'CREATE'
        log.target_type = 'SHIP'
        log.target_id = ship.id
        log.desc = '新增包装箱 箱号{name}'.format(name=box.box_number)
        log.user = request.user
        log.save()

        return Response({'msg': '成功新增包装箱!'}, status=status.HTTP_200_OK)

    # edit box
    @action(methods=['put'], detail=False, url_path='update_shipbox')
    def update_shipbox(self, request):
        box_id = request.data['id']
        box_number = request.data['box_number']
        length = float(request.data['length'])
        width = float(request.data['width'])
        heigth = float(request.data['heigth'])
        weight = request.data['weight']
        carrier_box_number = request.data['carrier_box_number']
        note = request.data['note']

        box = ShipBox.objects.filter(id=box_id).first()
        box.box_number = box_number
        box.length = length
        box.width = width
        box.heigth = heigth
        box.weight = weight
        box.carrier_box_number = carrier_box_number
        box.note = note

        cbm = length * width * heigth / 1000000
        box.cbm = cbm
        size_weight = length * width * heigth / 6000
        box.size_weight = size_weight
        box.save()

        # 统计包装箱总重量
        sum_weight = ShipBox.objects.filter(ship=box.ship).aggregate(
            Sum('weight'))
        total_weight = sum_weight['weight__sum']
        box.ship.weight = total_weight

        # 统计包装箱总体积
        sum_cbm = ShipBox.objects.filter(ship=box.ship).aggregate(Sum('cbm'))
        total_cbm = sum_cbm['cbm__sum']
        box.ship.cbm = total_cbm
        box.ship.save()

        # 创建操作日志
        log = MLOperateLog()
        log.op_module = 'SHIP'
        log.op_type = 'EDIT'
        log.target_type = 'SHIP'
        log.target_id = box.ship.id
        log.desc = '修改包装箱 箱号{name}'.format(name=box.box_number)
        log.user = request.user
        log.save()

        return Response({'msg': '包装箱已更新!'}, status=status.HTTP_200_OK)

    # delete box
    @action(methods=['post'], detail=False, url_path='delete_shipbox')
    def delete_shipbox(self, request):
        box_id = request.data['id']
        ship_id = request.data['ship_id']
        box = ShipBox.objects.filter(id=box_id).first()
        box_number = box.box_number
        box.delete()

        ship = Ship.objects.filter(id=ship_id).first()
        # 统计包装箱总重量
        sum_weight = ShipBox.objects.filter(ship=ship).aggregate(Sum('weight'))
        total_weight = sum_weight['weight__sum']
        ship.weight = total_weight

        # 统计包装箱总体积
        sum_cbm = ShipBox.objects.filter(ship=ship).aggregate(Sum('cbm'))
        total_cbm = sum_cbm['cbm__sum']
        ship.cbm = total_cbm
        ship.save()

        # 创建操作日志
        log = MLOperateLog()
        log.op_module = 'SHIP'
        log.op_type = 'DEL'
        log.target_type = 'SHIP'
        log.target_id = ship.id
        log.desc = '删除包装箱 箱号{name}'.format(name=box_number)
        log.user = request.user
        log.save()
        return Response({'msg': '包装箱已删除!'}, status=status.HTTP_200_OK)


class CarrierViewSet(mixins.ListModelMixin, mixins.CreateModelMixin,
                     mixins.UpdateModelMixin, mixins.DestroyModelMixin,
                     mixins.RetrieveModelMixin, viewsets.GenericViewSet):
    """
    list:
        物流商列表,分页,过滤,搜索,排序
    create:
        物流商新增
    retrieve:
        物流商详情页
    update:
        物流商修改
    destroy:
        物流商删除
    """
    queryset = Carrier.objects.all()
    serializer_class = CarrierSerializer  # 序列化
    pagination_class = DefaultPagination  # 分页

    filter_backends = (DjangoFilterBackend, filters.SearchFilter,
                       filters.OrderingFilter)  # 过滤,搜索,排序
    filter_fields = ('name', )  # 配置过滤字段
    search_fields = ('name', )  # 配置搜索字段
    ordering_fields = ('od_num', 'id')  # 配置排序字段

    # 获取盛德预约入仓时间
    @action(methods=['post'], detail=False, url_path='get_sd_reserve')
    def get_sd_reserve(self, request):
        apptdate = request.data['apptdate']  # 预约日期
        cid = '0200'
        typeid = '612C36BE-4081-4F8D-9755-02C65B376B8D'

        payload = {
            'apptdate': apptdate,
            'cid': cid,
            'typeid': typeid,
            'somrequest': '',
        }
        url = 'http://client.sanstar.net.cn/ashx/reservationdeliverysys/handler/idata_context.ashx?action=get_reserveinto'
        c_path = os.path.join(BASE_DIR, "site_config.json")
        with open(c_path, "r", encoding="utf-8") as f:
            data = json.load(f)  # 加载配置数据
        header = {'Cookie': data['sd_cookies']}
        resp = requests.post(url, data=payload, headers=header)
        if 'success' in resp.json():
            if resp.json()['success']:
                return Response(
                    {
                        'status': 'success',
                        'data': resp.json()['data']
                    },
                    status=status.HTTP_200_OK)

        return Response({
            'msg': '查询异常',
            'status': 'error'
        },
                        status=status.HTTP_202_ACCEPTED)

    # 获取盛德fbm仓库信息
    @action(methods=['get'], detail=False, url_path='get_fbm_warehouse')
    def get_fbm_warehouse(self, request):
        payload = {
            'country': '墨西哥',
            'nid': '03E85014-136E-4E5B-B2F9-751E21FF5D88',
            'somrequest': '',
        }
        url = 'http://client.sanstar.net.cn/console/customer_order/fbawarehousecodedata'
        c_path = os.path.join(BASE_DIR, "site_config.json")
        with open(c_path, "r", encoding="utf-8") as f:
            data = json.load(f)  # 加载配置数据
        header = {'Cookie': data['sd_cookies']}
        resp = requests.post(url, data=payload, headers=header)
        if 'success' in resp.json():
            if resp.json()['success']:
                return Response(
                    {
                        'status': 'success',
                        'data': resp.json()['data']
                    },
                    status=status.HTTP_200_OK)

        return Response({
            'msg': '查询异常',
            'status': 'error'
        },
                        status=status.HTTP_202_ACCEPTED)


class TransStockViewSet(mixins.ListModelMixin, mixins.CreateModelMixin,
                        mixins.UpdateModelMixin, mixins.DestroyModelMixin,
                        mixins.RetrieveModelMixin, viewsets.GenericViewSet):
    """
    list:
        中转仓库存列表,分页,过滤,搜索,排序
    create:
        中转仓库存新增
    retrieve:
        中转仓库存详情页
    update:
        中转仓库存修改
    destroy:
        中转仓库存删除
    """
    queryset = TransStock.objects.all()
    serializer_class = TransStockSerializer  # 序列化
    pagination_class = DefaultPagination  # 分页

    filter_backends = (DjangoFilterBackend, filters.SearchFilter,
                       filters.OrderingFilter)  # 过滤,搜索,排序
    filter_fields = ('listing_shop', 'shop', 'is_out', 'user_id')  # 配置过滤字段
    search_fields = ('sku', 'p_name', 'label_code', 'upc', 'item_id',
                     's_number', 'batch', 'box_number', 'carrier_box_number',
                     'listing_shop')  # 配置搜索字段
    ordering_fields = ('sku', 'item_id', 'qty', 's_number', 'batch',
                       'arrived_date', 'stock_days', 'out_time')  # 配置排序字段

    # fbm发仓
    @action(methods=['post'], detail=False, url_path='send_fbm')
    def send_fbm(self, request):
        data = request.data
        shop = data[0]['listing_shop']

        batch = 'Z{time_str}'.format(time_str=time.strftime('%m%d'))

        # 检查拼箱产品是否全部选中
        box_number_set = []
        for i in data:
            box_number_set.append(i['box_number'])

        for i in data:
            # 检查是否拼箱
            box_count = TransStock.objects.filter(box_number=i['box_number'],
                                                  is_out=False).count()
            if box_count > 1:
                current_box_count = box_number_set.count(i['box_number'])
                if box_count != current_box_count:
                    return Response({
                        'msg': '拼箱货品需同时出库！',
                        'status': 'error'
                    },
                                    status=status.HTTP_202_ACCEPTED)

        # 检查产品是否已出库
        for i in data:
            ts_check = TransStock.objects.filter(id=i['id'],
                                                 is_out=False).first()
            if not ts_check:
                return Response({
                    'msg': '产品状态已变动，请检查！',
                    'status': 'error'
                },
                                status=status.HTTP_202_ACCEPTED)

        shop_obj = Shop.objects.filter(name=shop).first()
        platform = ''
        if shop_obj:
            platform = shop_obj.platform
        ship = Ship(s_status='SHIPPED',
                    send_from='LOCAL',
                    shop=shop,
                    platform=platform,
                    target='FBM',
                    batch=batch,
                    logi_fee_clear=True,
                    user_id=request.user.id,
                    sent_time=datetime.now())
        ship.save()

        # 创建操作日志
        log = MLOperateLog()
        log.op_module = 'SHIP'
        log.op_type = 'CREATE'
        log.target_type = 'SHIP'
        log.target_id = ship.id
        log.desc = '从中转仓创建运单'
        log.user = request.user
        log.save()

        total_box = 0  # 总箱数
        total_qty = 0  # 总数量
        total_weight = 0  # 总重量
        total_cbm = 0  # 总体积

        for i in data:
            # 创建产品
            product = MLProduct.objects.filter(sku=i['sku']).first()
            sd = ShipDetail()
            sd.ship = ship
            sd.s_type = 'REFILL'
            sd.qty = i['qty']
            sd.sku = i['sku']
            sd.target_FBM = product.shop
            sd.p_name = product.p_name
            sd.label_code = product.label_code
            sd.upc = product.upc
            sd.item_id = product.item_id
            sd.custom_code = product.custom_code
            sd.cn_name = product.cn_name
            sd.en_name = product.en_name
            sd.brand = product.brand
            sd.declared_value = product.declared_value
            sd.cn_material = product.cn_material
            sd.en_material = product.en_material
            sd.use = product.use
            sd.image = product.image
            sd.unit_cost = i['unit_cost']
            sd.avg_ship_fee = i['first_ship_cost']
            sd.box_number = i['box_number']
            sd.weight = product.weight
            sd.length = product.length
            sd.width = product.width
            sd.heigth = product.heigth
            sd.save()

            # 检查是否同箱号的拼箱
            is_exist = ShipBox.objects.filter(
                ship=ship, box_number=i['box_number']).count()
            if not is_exist:
                # 创建包装箱
                box = ShipBox()
                box.ship = ship
                box.box_number = i['box_number']
                box.length = i['box_length']
                box.width = i['box_width']
                box.heigth = i['box_heigth']
                box.weight = i['box_weight']
                box.carrier_box_number = i['carrier_box_number']
                box.note = i['s_number']
                cbm = i['box_cbm']
                box.cbm = cbm
                box.save()

                total_box += 1
                total_weight += i['box_weight']
                total_cbm += i['box_cbm']
            total_qty += i['qty']

            # 减去fbm库存 中转仓数量
            shop_stock = ShopStock.objects.filter(sku=sd.sku,
                                                  item_id=sd.item_id).first()
            if shop_stock:
                shop_stock.onway_qty += sd.qty  # 增加在途数量
                shop_stock.trans_qty -= sd.qty  # 减去中转仓数量
                shop_stock.save()

            ts = TransStock.objects.filter(id=i['id']).first()
            ts.is_out = True
            ts.out_time = datetime.now()
            ts.save()

        ship.total_box = total_box
        ship.total_qty = total_qty
        ship.weight = total_weight
        ship.cbm = total_cbm
        ship.save()

        # 创建操作日志
        log = MLOperateLog()
        log.op_module = 'TRANS'
        log.op_type = 'CREATE'
        log.target_type = 'TRANS'
        log.target_id = ship.id
        log.desc = '中转仓FBM发仓 目标店铺: {name}'.format(name=ship.shop)
        log.user = request.user
        log.save()

        return Response({'msg': '操作成功!'}, status=status.HTTP_200_OK)


class MLSiteViewSet(mixins.ListModelMixin, mixins.CreateModelMixin,
                    mixins.UpdateModelMixin, mixins.DestroyModelMixin,
                    mixins.RetrieveModelMixin, viewsets.GenericViewSet):
    """
    list:
        站点列表,分页,过滤,搜索,排序
    create:
        站点新增
    retrieve:
        站点详情页
    update:
        站点修改
    destroy:
        站点删除
    """
    queryset = MLSite.objects.all()
    serializer_class = MLSiteSerializer  # 序列化
    pagination_class = DefaultPagination  # 分页

    filter_backends = (DjangoFilterBackend, filters.SearchFilter,
                       filters.OrderingFilter)  # 过滤,搜索,排序
    filter_fields = ('name', )  # 配置过滤字段
    search_fields = ('name', )  # 配置搜索字段
    ordering_fields = ('od_num', 'id')  # 配置排序字段


class FBMWarehouseViewSet(mixins.ListModelMixin, mixins.CreateModelMixin,
                          mixins.UpdateModelMixin, mixins.DestroyModelMixin,
                          mixins.RetrieveModelMixin, viewsets.GenericViewSet):
    """
    list:
        FBM仓库列表,分页,过滤,搜索,排序
    create:
        FBM仓库新增
    retrieve:
        FBM仓库详情页
    update:
        FBM仓库修改
    destroy:
        FBM仓库删除
    """
    queryset = FBMWarehouse.objects.all()
    serializer_class = FBMWarehouseSerializer  # 序列化
    pagination_class = DefaultPagination  # 分页

    filter_backends = (DjangoFilterBackend, filters.SearchFilter,
                       filters.OrderingFilter)  # 过滤,搜索,排序
    filter_fields = ('country', 'is_active', 'platform')  # 配置过滤字段
    search_fields = ('w_code', 'name', 'address')  # 配置搜索字段
    ordering_fields = ('create_time', 'id')  # 配置排序字段


class FinanceViewSet(mixins.ListModelMixin, mixins.CreateModelMixin,
                     mixins.UpdateModelMixin, mixins.DestroyModelMixin,
                     mixins.RetrieveModelMixin, viewsets.GenericViewSet):
    """
    list:
        财务管理列表,分页,过滤,搜索,排序
    create:
        财务管理新增
    retrieve:
        财务管理详情页
    update:
        财务管理修改
    destroy:
        财务管理删除
    """
    queryset = Finance.objects.all()
    serializer_class = FinanceSerializer  # 序列化
    pagination_class = DefaultPagination  # 分页

    filter_backends = (DjangoFilterBackend, filters.SearchFilter,
                       filters.OrderingFilter)  # 过滤,搜索,排序
    # filter_fields = ('country', 'is_active')  # 配置过滤字段
    filterset_fields = {
        'income': ['gte', 'lte', 'exact', 'gt', 'lt'],
        'wd_date': ['gte', 'lte', 'exact', 'gt', 'lt'],
        'rec_date': ['gte', 'lte', 'exact', 'gt', 'lt'],
        'exchange': ['gte', 'lte', 'exact', 'gt', 'lt'],
        'income_rmb': ['gte', 'lte', 'exact', 'gt', 'lt'],
        'exc_date': ['gte', 'lte', 'exact', 'gt', 'lt'],
        'shop': ['exact'],
        'is_received': ['exact'],
        'f_type': ['exact'],
    }
    search_fields = ('income', )  # 配置搜索字段
    ordering_fields = ('wd_date', 'rec_date', 'exc_date', 'income_rmb',
                       'income', 'create_time')  # 配置排序字段

    # ML创建店铺提现
    @action(methods=['post'], detail=False, url_path='create_wd')
    def create_wd(self, request):
        data = request.data
        shop_id = data['shop']

        shop = Shop.objects.filter(id=shop_id).first()
        finance = Finance()
        finance.shop = shop
        finance.currency = shop.exc_currency
        finance.income = data['income']
        finance.wd_date = data['wd_date']
        finance.f_type = 'WD'
        finance.save()

        # 创建操作日志
        log = MLOperateLog()
        log.op_module = 'FINANCE'
        log.op_type = 'CREATE'
        log.target_type = 'FINANCE'
        log.desc = '新增店铺提现 店铺: {name}，提现资金: ${income}'.format(
            name=shop.name, income=finance.income)
        log.user = request.user
        log.save()

        return Response({'msg': '操作成功!'}, status=status.HTTP_200_OK)

    # ML创建店铺费用
    @action(methods=['post'], detail=False, url_path='create_fee')
    def create_fee(self, request):
        data = request.data
        shop_id = data['shop']

        shop = Shop.objects.filter(id=shop_id).first()
        finance = Finance()
        finance.shop = shop
        finance.note = data['note']
        finance.income = data['income']
        finance.wd_date = data['wd_date']
        finance.f_type = 'FEE'
        finance.save()

        # 创建操作日志
        log = MLOperateLog()
        log.op_module = 'FINANCE'
        log.op_type = 'CREATE'
        log.target_type = 'FINANCE'
        log.desc = '新增店铺费用 店铺: {name}，费用说明: ${currency}，费用金额: ${income}'.format(
            name=shop.name, currency=finance.note, income=finance.income)
        log.user = request.user
        log.save()

        return Response({'msg': '操作成功!'}, status=status.HTTP_200_OK)

    # ML创建结汇
    @action(methods=['post'], detail=False, url_path='create_exc')
    def create_exc(self, request):
        data = request.data
        shop_id = data['shop']

        # 店铺提现外汇
        sum_income_fund = Finance.objects.filter(shop__id=shop_id,
                                                 is_received=True,
                                                 f_type='WD').aggregate(
                                                     Sum('income'))
        income_fund = sum_income_fund['income__sum']
        if not income_fund:
            income_fund = 0
        if float(data['exchange']) > income_fund:
            return Response({'msg': '结汇金额超过账号资金'},
                            status=status.HTTP_202_ACCEPTED)

        shop = Shop.objects.filter(id=shop_id).first()
        finance = Finance()
        finance.shop = shop
        finance.currency = shop.currency
        finance.exchange = data['exchange']
        finance.income_rmb = data['income_rmb']
        finance.exc_date = data['exc_date']
        finance.f_type = 'EXC'
        finance.save()

        # 创建操作日志
        log = MLOperateLog()
        log.op_module = 'FINANCE'
        log.op_type = 'CREATE'
        log.target_type = 'FINANCE'
        log.desc = '新增店铺结汇 店铺: {name}，结汇资金: ${exchange}, 收入￥{income}'.format(
            name=shop.name,
            exchange=finance.exchange,
            income=finance.income_rmb)
        log.user = request.user
        log.save()

        return Response({'msg': '操作成功!'}, status=status.HTTP_200_OK)

    # 统计资金
    @action(methods=['post'], detail=False, url_path='calc_fund')
    def calc_fund(self, request):
        from decimal import Decimal
        data = request.data
        shop_id = data['shop']

        shop = Shop.objects.filter(id=shop_id).first()

        # 在途外汇
        sum_onway_fund = Finance.objects.filter(shop__id=shop_id,
                                                is_received=False,
                                                f_type='WD').aggregate(
                                                    Sum('income'))
        onway_fund = sum_onway_fund['income__sum']
        if not onway_fund:
            onway_fund = 0

        # 结汇资金
        sum_income_rmb = Finance.objects.filter(shop__id=shop_id,
                                                f_type='EXC').aggregate(
                                                    Sum('income_rmb'))
        income_rmb = sum_income_rmb['income_rmb__sum']
        if not income_rmb:
            income_rmb = 0

        # 结汇外汇
        sum_exchange = Finance.objects.filter(shop__id=shop_id,
                                              f_type='EXC').aggregate(
                                                  Sum('exchange'))
        exchange_fund = sum_exchange['exchange__sum']
        if not exchange_fund:
            exchange_fund = 0

        # 店铺提现外汇
        sum_income_fund = Finance.objects.filter(shop__id=shop_id,
                                                 is_received=True,
                                                 f_type='WD').aggregate(
                                                     Sum('income'))
        income_fund = sum_income_fund['income__sum']
        if not income_fund:
            income_fund = 0

        # rest_income = income_fund - exchange_fund
        rest_income = Decimal(income_fund).quantize(
            Decimal("0.00")) - Decimal(exchange_fund).quantize(Decimal("0.00"))

        return Response(
            {
                'onway_fund': onway_fund,
                'income_rmb': income_rmb,
                'rest_income': rest_income,
                'default_currency': shop.exc_currency
            },
            status=status.HTTP_200_OK)


class MLOrderViewSet(mixins.ListModelMixin, mixins.CreateModelMixin,
                     mixins.UpdateModelMixin, mixins.DestroyModelMixin,
                     mixins.RetrieveModelMixin, viewsets.GenericViewSet):
    """
    list:
        销售订单列表,分页,过滤,搜索,排序
    create:
        销售订单增
    retrieve:
        销售订单详情页
    update:
        销售订单改
    destroy:
        销售订单删除
    """
    queryset = MLOrder.objects.all()
    serializer_class = MLOrderSerializer  # 序列化
    pagination_class = DefaultPagination  # 分页

    filter_backends = (DjangoFilterBackend, filters.SearchFilter,
                       filters.OrderingFilter)  # 过滤,搜索,排序
    # filter_fields = ('shop', 'is_ad', 'order_status')  # 配置过滤字段
    filterset_fields = {
        'order_time': ['gte', 'lte', 'exact', 'gt', 'lt'],
        'order_time_bj': ['gte', 'lte', 'exact', 'gt', 'lt'],
        'shop': ['exact'],
        'is_ad': ['exact'],
        'order_status': ['exact'],
        'finance_check1': ['exact'],
    }
    search_fields = ('order_number', 'sku', 'p_name', 'item_id')  # 配置搜索字段
    ordering_fields = ('create_time', 'order_time', 'order_time_bj', 'price',
                       'profit')  # 配置排序字段

    # ML订单批量上传(旧)
    @action(methods=['post'], detail=False, url_path='bulk_upload')
    def bulk_upload(self, request):
        import warnings
        import re
        warnings.filterwarnings('ignore')

        data = request.data
        shop_id = data['id']
        wb = openpyxl.load_workbook(data['excel'])
        sheet = wb.active

        month_dict = {
            'enero': '01',
            'febrero': '02',
            'marzo': '03',
            'abril': '04',
            'mayo': '05',
            'junio': '06',
            'julio': '07',
            'agosto': '08',
            'septiembre': '09',
            'octubre': '10',
            'noviembre': '11',
            'diciembre': '12'
        }

        shop = Shop.objects.filter(id=shop_id).first()
        er = ExRate.objects.filter(currency=shop.currency).first()
        ex_rate = er.value

        add_list = []
        for cell_row in list(sheet)[3:]:
            qty = cell_row[5].value
            if not qty:
                continue
            sku = cell_row[13].value
            item_id = cell_row[14].value[3:]

            # 如果不在fmb库存中，或者所在店铺不对应，则跳出
            shop_stock = ShopStock.objects.filter(sku=sku,
                                                  item_id=item_id,
                                                  shop=shop).first()
            if not shop_stock:
                continue
            first_ship_cost = shop_stock.first_ship_cost
            if not first_ship_cost:
                first_ship_cost = 0

            order_number = cell_row[0].value

            t = cell_row[1].value
            de_locate = [m.start() for m in re.finditer('de', t)]
            day = t[:de_locate[0] - 1]
            if int(day) < 10:
                day = '0' + day
            month = t[de_locate[0] + 3:de_locate[1] - 1]
            year = t[de_locate[1] + 3:de_locate[1] + 7]
            hour = t[de_locate[1] + 8:de_locate[1] + 10]
            min = t[de_locate[1] + 11:de_locate[1] + 13]
            order_time = '%s-%s-%s %s:%s:00' % (year, month_dict[month], day,
                                                hour, min)

            bj = datetime.strptime(order_time,
                                   '%Y-%m-%d %H:%M:%S') + timedelta(hours=14)
            order_time_bj = bj.strftime('%Y-%m-%d %H:%M:%S')

            price = cell_row[17].value if cell_row[17].value else 0
            fees = cell_row[8].value if cell_row[8].value else 0
            postage = cell_row[9].value if cell_row[9].value else 0
            receive_fund = cell_row[11].value if cell_row[11].value else 0
            is_ad = True if cell_row[12].value == 'Sí' else False

            buyer_name = cell_row[24].value
            buyer_address = cell_row[26].value
            buyer_city = cell_row[27].value
            buyer_state = cell_row[28].value
            buyer_postcode = cell_row[29].value
            buyer_country = cell_row[30].value

            profit = (
                float(receive_fund) * 0.99
            ) * ex_rate - shop_stock.unit_cost * qty - shop_stock.first_ship_cost * qty
            profit_rate = profit / (price * ex_rate)
            if profit_rate < 0:
                profit_rate = 0

            order_status = 'FINISHED'
            if cell_row[2].value == 'Cancelada por el comprador':
                order_status = 'CANCEL'
            if cell_row[2].value == 'Paquete cancelado por Mercado Libre':
                order_status = 'CANCEL'
            if cell_row[2].value == 'Devolución en camino':
                order_status = 'RETURN'
            if cell_row[
                    2].value == 'Reclamo cerrado con reembolso al comprador':
                order_status = 'CASE'
            if cell_row[2].value[:8] == 'Devuelto':
                order_status = 'RETURN'

            # 检查同一店铺订单编号是否存在
            ml_order = MLOrder.objects.filter(order_number=order_number,
                                              shop=shop).first()
            if not ml_order:
                add_list.append(
                    MLOrder(
                        shop=shop,
                        platform='MERCADO',
                        order_number=order_number,
                        order_status=order_status,
                        order_time=order_time,
                        order_time_bj=order_time_bj,
                        qty=qty,
                        currency=shop.currency,
                        ex_rate=ex_rate,
                        price=price,
                        fees=fees,
                        postage=postage,
                        receive_fund=receive_fund,
                        is_ad=is_ad,
                        sku=sku,
                        p_name=shop_stock.p_name,
                        item_id=item_id,
                        image=shop_stock.image,
                        unit_cost=shop_stock.unit_cost * qty,
                        first_ship_cost=first_ship_cost * qty,
                        profit=profit,
                        profit_rate=profit_rate,
                        buyer_name=buyer_name,
                        buyer_address=buyer_address,
                        buyer_city=buyer_city,
                        buyer_state=buyer_state,
                        buyer_postcode=buyer_postcode,
                        buyer_country=buyer_country,
                    ))
                shop_stock.qty -= qty
                shop_stock.save()

                # 创建库存日志
                stock_log = StockLog()
                stock_log.shop_stock = shop_stock
                stock_log.current_stock = shop_stock.qty
                stock_log.qty = qty
                stock_log.in_out = 'OUT'
                stock_log.action = 'SALE'
                stock_log.desc = '销售订单号: ' + order_number
                stock_log.user_id = 0
                stock_log.save()
                stock_log.create_time = order_time  # order_time
                stock_log.save()
            else:
                if ml_order.order_status != order_status:
                    ml_order.order_status = order_status
                    ml_order.receive_fund = receive_fund
                    ml_order.profit = profit
                    ml_order.save()

                    # 如果订单是取消状态，库存增加回来
                    if order_status == 'CANCEL':
                        shop_stock.qty += qty
                        shop_stock.save()

                        # 创建库存日志
                        stock_log = StockLog()
                        stock_log.shop_stock = shop_stock
                        stock_log.current_stock = shop_stock.qty
                        stock_log.qty = qty
                        stock_log.in_out = 'IN'
                        stock_log.action = 'CANCEL'
                        stock_log.desc = '取消订单入库, 订单号: ' + order_number
                        stock_log.user_id = 0
                        stock_log.save()
        if len(add_list):
            MLOrder.objects.bulk_create(add_list)

        # 计算产品销量
        tasks.calc_product_sales.delay()
        # 统计过去30天每天销量
        tasks.calc_shop_sale.delay()

        # 创建操作日志
        log = MLOperateLog()
        log.op_module = 'ORDER'
        log.op_type = 'CREATE'
        log.target_type = 'ORDER'
        log.desc = '销售订单导入 店铺: {name}'.format(name=shop.name)
        log.user = request.user
        log.save()

        return Response({'msg': '成功上传'}, status=status.HTTP_200_OK)

    # 订单上传接口2
    @action(methods=['post'], detail=False, url_path='bulk_upload2')
    def bulk_upload2(self, request):

        data = request.data
        shop_id = data['id']
        mel_row = data['mel_row']
        shop = Shop.objects.filter(id=shop_id).first()

        if not shop:
            return Response({
                'msg': '店铺状态异常',
                'status': 'error'
            },
                            status=status.HTTP_202_ACCEPTED)

        # 保存上传excel文件
        path = 'media/upload_file/order_excel_' + shop_id + '.xlsx'
        excel = data['excel']
        content = excel.chunks()
        with open(path, 'wb') as f:
            for i in content:
                f.write(i)

        # 创建上传通知
        file_upload = FileUploadNotify()
        file_upload.shop = shop
        file_upload.user_id = request.user.id
        file_upload.upload_type = 'ORDER'
        file_upload.upload_status = 'LOADING'
        file_upload.desc = '订单正在上传中...'
        file_upload.save()

        if shop.platform == 'MERCADO':
            tasks.upload_mercado_order.delay(shop_id, file_upload.id, mel_row)
        if shop.platform == 'NOON':
            tasks.upload_noon_order.delay(shop_id, file_upload.id)
        if shop.platform == 'OZON':
            tasks.upload_ozon_order.delay(shop_id, file_upload.id)

        # 计算产品销量
        tasks.calc_product_sales.delay()

        # 统计过去30天每天销量
        tasks.calc_shop_sale.delay()

        # 创建操作日志
        log = MLOperateLog()
        log.op_module = 'ORDER'
        log.op_type = 'CREATE'
        log.target_type = 'ORDER'
        log.desc = '销售订单导入 店铺: {name}'.format(name=shop.name)
        log.user = request.user
        log.save()
        return Response({
            'msg': '文件已上传，后台处理中，稍微刷新查看结果',
            'status': 'success'
        },
                        status=status.HTTP_200_OK)


class MLOperateLogViewSet(mixins.ListModelMixin, mixins.CreateModelMixin,
                          mixins.UpdateModelMixin, mixins.DestroyModelMixin,
                          mixins.RetrieveModelMixin, viewsets.GenericViewSet):
    """
    list:
        操作日志列表,分页,过滤,搜索,排序
    create:
        操作日志新增
    retrieve:
        操作日志详情页
    update:
        操作日志修改
    destroy:
        操作日志删除
    """
    queryset = MLOperateLog.objects.all()
    serializer_class = MLOperateLogSerializer  # 序列化
    pagination_class = DefaultPagination  # 分页

    filter_backends = (DjangoFilterBackend, filters.SearchFilter,
                       filters.OrderingFilter)  # 过滤,搜索,排序
    filter_fields = ('op_module', 'op_type', 'target_id', 'target_type', 'user'
                     )  # 配置过滤字段
    filterset_fields = {
        'create_time': ['gte', 'lte', 'exact', 'gt', 'lt'],
        'op_module': ['exact'],
        'op_type': ['exact'],
        'target_id': ['exact'],
        'target_type': ['exact'],
        'user': ['exact'],
    }
    search_fields = ('desc', )  # 配置搜索字段
    ordering_fields = ('create_time', 'id')  # 配置排序字段


class ShopReportViewSet(mixins.ListModelMixin, viewsets.GenericViewSet):
    """
    list:
        统计店铺每天销量
    """
    queryset = ShopReport.objects.all()
    serializer_class = ShopReportSerializer  # 序列化

    filter_backends = (DjangoFilterBackend, )  # 过滤
    filterset_fields = {
        'calc_date': ['gte', 'lte', 'exact'],
        'shop': ['exact'],
    }

    def get_queryset(self):
        user = self.request.user
        # 返回指定用户数据
        if user.is_superuser:
            queryset = ShopReport.objects.all()
        else:
            queryset = ShopReport.objects.filter(shop__user=user)
        return queryset

    # 统计30天店铺每天销量
    @action(methods=['get'], detail=False, url_path='calc_shop_sale')
    def calc_shop_sale(self, request):
        tasks.calc_shop_sale()
        return Response({'msg': '更新成功'}, status=status.HTTP_200_OK)


class PurchaseManageViewSet(mixins.ListModelMixin, mixins.CreateModelMixin,
                            mixins.UpdateModelMixin, mixins.DestroyModelMixin,
                            mixins.RetrieveModelMixin,
                            viewsets.GenericViewSet):
    """
    list:
        采购管理列表,分页,过滤,搜索,排序
    create:
        采购管理新增
    retrieve:
        采购管理详情页
    update:
        采购管理修改
    destroy:
        采购管理删除
    """
    queryset = PurchaseManage.objects.all()
    serializer_class = PurchaseManageSerializer  # 序列化
    pagination_class = DefaultPagination  # 分页

    filter_backends = (DjangoFilterBackend, filters.SearchFilter,
                       filters.OrderingFilter)  # 过滤,搜索,排序
    # filter_fields = ('p_status', 'shop', 's_type', 'create_type', 'is_urgent')  # 配置过滤字段
    filterset_fields = {
        'buy_qty': ['gte', 'lte', 'exact', 'gt', 'lt'],
        'platform': ['exact'],
        'p_status': ['exact'],
        'shop': ['exact'],
        's_type': ['exact'],
        'create_type': ['exact'],
        'is_urgent': ['exact'],
    }
    search_fields = ('sku', 'p_name', 'item_id')  # 配置搜索字段
    ordering_fields = ('create_time', 'shop', 'item_id', 'buy_time',
                       'rec_time', 'pack_time', 'used_time', 'p_name'
                       )  # 配置排序字段

    # 拉取运单备货产品
    @action(methods=['get'], detail=False, url_path='pull_purchase')
    def pull_purchase(self, request):
        pm_queryset = PurchaseManage.objects.filter(p_status='WAITBUY')
        for i in pm_queryset:
            sku_count = ShipDetail.objects.filter(ship__s_status='PREPARING',
                                                  sku=i.sku).count()
            if sku_count == 0:
                if i.create_type == 'SYS':
                    i.delete()
            # 计算该sku已采购数量
            aready_buy_set = PurchaseManage.objects.filter(sku=i.sku).filter(
                Q(p_status='PURCHASED') | Q(p_status='RECEIVED')
                | Q(p_status='PACKED'))
            total_buy = 0
            for item in aready_buy_set:
                if item.p_status == 'PURCHASED':
                    total_buy += item.buy_qty
                if item.p_status == 'RECEIVED':
                    total_buy += item.rec_qty
                if item.p_status == 'PACKED':
                    total_buy += item.pack_qty
            if sku_count == 1:
                sd = ShipDetail.objects.filter(ship__s_status='PREPARING',
                                               sku=i.sku).first()
                i.buy_qty = sd.qty - total_buy if sd.qty - total_buy > 0 else 0
                i.create_time = datetime.now()
                i.save()
            if sku_count > 1:
                sd_set = ShipDetail.objects.filter(ship__s_status='PREPARING',
                                                   sku=i.sku)
                total_ship_qty = 0
                for item in sd_set:
                    total_ship_qty += item.qty
                i.buy_qty = total_ship_qty - total_buy if total_ship_qty - total_buy > 0 else 0
                i.create_time = datetime.now()
                i.save()

        ship_queryset = ShipDetail.objects.filter(ship__s_status='PREPARING')
        for i in ship_queryset:
            wait_buy_sku_count = PurchaseManage.objects.filter(
                sku=i.sku, p_status='WAITBUY').count()
            if not wait_buy_sku_count:
                aready_buy_set = PurchaseManage.objects.filter(
                    sku=i.sku).filter(
                        Q(p_status='PURCHASED') | Q(p_status='RECEIVED')
                        | Q(p_status='PACKED'))
                total_buy = 0
                for item in aready_buy_set:
                    if item.p_status == 'PURCHASED':
                        total_buy += item.buy_qty
                    if item.p_status == 'RECEIVED':
                        total_buy += item.rec_qty
                    if item.p_status == 'PACKED':
                        total_buy += item.pack_qty
                sd_set = ShipDetail.objects.filter(ship__s_status='PREPARING',
                                                   sku=i.sku)
                total_ship_qty = 0
                for item in sd_set:
                    total_ship_qty += item.qty

                shop = Shop.objects.filter(name=i.target_FBM).first()
                purchase_manage = PurchaseManage(p_status='WAITBUY',
                                                 s_type=i.s_type,
                                                 create_type='SYS',
                                                 from_batch=i.ship.batch,
                                                 platform=i.ship.platform,
                                                 sku=i.sku,
                                                 p_name=i.p_name,
                                                 item_id=i.item_id,
                                                 label_code=i.label_code,
                                                 image=i.image,
                                                 unit_cost=i.unit_cost,
                                                 weight=i.weight,
                                                 length=i.length,
                                                 width=i.width,
                                                 heigth=i.heigth,
                                                 buy_qty=total_ship_qty -
                                                 total_buy if total_ship_qty -
                                                 total_buy > 0 else 0,
                                                 note=i.note,
                                                 shop=shop.name,
                                                 shop_color=shop.name_color,
                                                 packing_size=i.packing_size,
                                                 packing_name=i.packing_name,
                                                 create_time=datetime.now())
                purchase_manage.save()

        # 创建操作日志
        log = MLOperateLog()
        log.op_module = 'PURCHASE'
        log.op_type = 'CREATE'
        log.target_type = 'PURCHASE'
        log.desc = '拉取采购需求'
        log.user = request.user
        log.save()
        return Response({'msg': '操作成功!'}, status=status.HTTP_200_OK)

    # 下单采购
    @action(methods=['post'], detail=False, url_path='place_buy')
    def place_buy(self, request):
        data = request.data
        products = data['products']
        is_change = data['is_change']
        # 检查是否有产品重复下单采购
        for i in products:
            is_exist = PurchaseManage.objects.filter(
                id=i['id'], p_status='PURCHASED').count()
            if is_exist:
                return Response({
                    'msg': '产品重复下单采购，请刷新页面检查！',
                    'status': 'error'
                },
                                status=status.HTTP_202_ACCEPTED)

        for i in products:
            purchase = PurchaseManage.objects.filter(id=i['id']).first()
            purchase.p_status = 'PURCHASED'
            purchase.buy_qty = i['buy_qty']
            purchase.rec_qty = i['buy_qty']
            purchase.buy_time = datetime.now()
            if is_change:
                purchase.unit_cost = i['unit_cost']
            purchase.save()

            # 是否需要更新产品价格
            if is_change:
                p = MLProduct.objects.filter(sku=purchase.sku).first()
                if p.unit_cost == purchase.unit_cost:
                    continue
                p.unit_cost = purchase.unit_cost
                p.save()

            # 创建操作日志
            log = MLOperateLog()
            log.op_module = 'PURCHASE'
            log.op_type = 'CREATE'
            log.target_type = 'PURCHASE'
            log.desc = '下单采购 {sku} {p_name} {qty}个'.format(sku=i['sku'],
                                                           p_name=i['p_name'],
                                                           qty=i['buy_qty'])
            log.user = request.user
            log.save()
        return Response({
            'msg': '操作成功!',
            'status': 'success'
        },
                        status=status.HTTP_200_OK)

    # 确认收货
    @action(methods=['post'], detail=False, url_path='rec_buy')
    def rec_buy(self, request):
        data = request.data

        # 检查是否重复操作
        for i in data:
            pm = PurchaseManage.objects.filter(id=i['id'],
                                               p_status='PURCHASED').first()
            if not pm:
                return Response({
                    'msg': '产品重复收货，请刷新页面检查！',
                    'status': 'error'
                },
                                status=status.HTTP_202_ACCEPTED)
            if i['rec_qty'] > pm.buy_qty:
                return Response({
                    'msg': '产品收货数量错误，请刷新页面检查！',
                    'status': 'error'
                },
                                status=status.HTTP_202_ACCEPTED)

        for i in data:
            buy_pm = PurchaseManage.objects.filter(id=i['id']).first()
            purchase_manage = PurchaseManage(p_status='RECEIVED',
                                             platform=buy_pm.platform,
                                             s_type=buy_pm.s_type,
                                             create_type=buy_pm.create_type,
                                             sku=buy_pm.sku,
                                             p_name=buy_pm.p_name,
                                             item_id=buy_pm.item_id,
                                             label_code=buy_pm.label_code,
                                             image=buy_pm.image,
                                             unit_cost=buy_pm.unit_cost,
                                             weight=buy_pm.weight,
                                             length=buy_pm.length,
                                             width=buy_pm.width,
                                             heigth=buy_pm.heigth,
                                             rec_qty=i['rec_qty'],
                                             pack_qty=i['rec_qty'],
                                             note=buy_pm.note,
                                             shop=buy_pm.shop,
                                             shop_color=buy_pm.shop_color,
                                             packing_size=buy_pm.packing_size,
                                             packing_name=buy_pm.packing_name,
                                             rec_time=datetime.now())
            purchase_manage.save()

            # 创建操作日志
            log = MLOperateLog()
            log.op_module = 'PURCHASE'
            log.op_type = 'CREATE'
            log.target_type = 'PURCHASE'
            log.desc = '确认收货 {sku} {p_name} {qty}个'.format(sku=i['sku'],
                                                           p_name=i['p_name'],
                                                           qty=i['rec_qty'])
            log.user = request.user
            log.save()

            if i['rec_qty'] >= i['buy_qty']:
                buy_pm.delete()
            else:
                buy_pm.buy_qty = i['buy_qty'] - i['rec_qty']
                buy_pm.rec_qty = i['buy_qty'] - i['rec_qty']
                buy_pm.save()

        return Response({
            'msg': '操作成功!',
            'status': 'success'
        },
                        status=status.HTTP_200_OK)

    # 确认打包
    @action(methods=['post'], detail=False, url_path='pack_buy')
    def pack_buy(self, request):
        data = request.data

        # 检查是否重复操作
        for i in data:
            pm = PurchaseManage.objects.filter(id=i['id'],
                                               p_status='RECEIVED').first()
            if not pm:
                return Response({
                    'msg': '产品重复打包，请刷新页面检查！',
                    'status': 'error'
                },
                                status=status.HTTP_202_ACCEPTED)
            if i['pack_qty'] > pm.rec_qty:
                return Response({
                    'msg': '产品打包数量错误，请刷新页面检查！',
                    'status': 'error'
                },
                                status=status.HTTP_202_ACCEPTED)

        for i in data:
            rec_pm = PurchaseManage.objects.filter(id=i['id']).first()
            pack_sku = PurchaseManage.objects.filter(
                sku=i['sku'], p_status='PACKED').first()
            if not pack_sku:
                purchase_manage = PurchaseManage(
                    p_status='PACKED',
                    platform=rec_pm.platform,
                    s_type=rec_pm.s_type,
                    create_type=rec_pm.create_type,
                    sku=rec_pm.sku,
                    p_name=rec_pm.p_name,
                    item_id=rec_pm.item_id,
                    label_code=rec_pm.label_code,
                    image=rec_pm.image,
                    unit_cost=rec_pm.unit_cost,
                    weight=rec_pm.weight,
                    length=rec_pm.length,
                    width=rec_pm.width,
                    heigth=rec_pm.heigth,
                    pack_qty=i['pack_qty'],
                    note=rec_pm.note,
                    shop=rec_pm.shop,
                    shop_color=rec_pm.shop_color,
                    packing_size=rec_pm.packing_size,
                    packing_name=rec_pm.packing_name,
                    pack_time=datetime.now())
                purchase_manage.save()
            else:
                pack_sku.pack_qty += i['pack_qty']
                pack_sku.save()

            if i['pack_qty'] >= i['rec_qty']:
                rec_pm.delete()
            else:
                rec_pm.rec_qty = i['rec_qty'] - i['pack_qty']
                rec_pm.pack_qty = i['rec_qty'] - i['pack_qty']
                rec_pm.save()

            # 创建操作日志
            log = MLOperateLog()
            log.op_module = 'PURCHASE'
            log.op_type = 'CREATE'
            log.target_type = 'PURCHASE'
            log.desc = '确认打包 {sku} {p_name} {qty}个'.format(sku=i['sku'],
                                                           p_name=i['p_name'],
                                                           qty=i['pack_qty'])
            log.user = request.user
            log.save()

        return Response({
            'msg': '操作成功!',
            'status': 'success'
        },
                        status=status.HTTP_200_OK)

    # 确认质检
    @action(methods=['post'], detail=False, url_path='product_qc')
    def product_qc(self, request):
        data = request.data
        for i in data:
            pm = PurchaseManage.objects.filter(id=i['id']).first()
            if pm:
                pm.is_qc = True
                pm.save()

            # 创建操作日志
            log = MLOperateLog()
            log.op_module = 'PURCHASE'
            log.op_type = 'EDIT'
            log.target_type = 'PURCHASE'
            log.desc = '完成质检 {sku} {p_name}'.format(sku=i['sku'],
                                                    p_name=i['p_name'])
            log.user = request.user
            log.save()
        return Response({'msg': '操作成功!'}, status=status.HTTP_200_OK)

    # 计算采购单数量
    @action(methods=['get'], detail=False, url_path='calc_purchase')
    def calc_purchase(self, request):
        wait_buy_num = PurchaseManage.objects.filter(
            p_status='WAITBUY').count()
        purchased_num = PurchaseManage.objects.filter(
            p_status='PURCHASED').count()
        rec_num = PurchaseManage.objects.filter(p_status='RECEIVED').count()
        pack_num = PurchaseManage.objects.filter(p_status='PACKED').count()

        return Response(
            {
                'wait_buy_num': wait_buy_num,
                'purchased_num': purchased_num,
                'rec_num': rec_num,
                'pack_num': pack_num
            },
            status=status.HTTP_200_OK)

    # 手动新建采购
    @action(methods=['post'], detail=False, url_path='manuel_create_buy')
    def manuel_create_buy(self, request):
        data = request.data
        for i in data:
            product = MLProduct.objects.filter(id=i['id']).first()
            shop = Shop.objects.filter(name=product.shop).first()
            packing = Packing.objects.filter(id=product.packing_id).first()
            purchase_manage = PurchaseManage(
                p_status='WAITBUY',
                s_type='REFILL',
                create_type='MANUAL',
                platform=product.platform,
                sku=product.sku,
                p_name=product.p_name,
                item_id=product.item_id,
                label_code=product.label_code,
                image=product.image,
                unit_cost=product.unit_cost,
                weight=product.weight,
                length=product.length,
                width=product.width,
                heigth=product.heigth,
                buy_qty=1,
                shop=shop.name,
                shop_color=shop.name_color,
                packing_size=packing.size if packing else '',
                packing_name=packing.name if packing else '',
                create_time=datetime.now())
            purchase_manage.save()
        # 创建操作日志
        log = MLOperateLog()
        log.op_module = 'PURCHASE'
        log.op_type = 'CREATE'
        log.target_type = 'PURCHASE'
        log.desc = '手动添加{name}个采购产品'.format(name=len(data))
        log.user = request.user
        log.save()
        return Response({'msg': '操作成功!'}, status=status.HTTP_200_OK)

    # 产品数据核查
    @action(methods=['post'], detail=False, url_path='check_product')
    def check_product(self, request):
        data = request.data
        packing = Packing.objects.filter(id=data['packing_id']).first()
        pm = PurchaseManage.objects.filter(id=data['pm_id']).first()
        pm.packing_name = packing.name
        pm.packing_size = packing.size
        pm.length = data['length']
        pm.unit_cost = data['unit_cost']
        pm.width = data['width']
        pm.heigth = data['heigth']
        pm.weight = data['weight']
        pm.save()
        product = MLProduct.objects.filter(sku=pm.sku).first()
        product.packing_id = data['packing_id']
        product.unit_cost = data['unit_cost']
        product.length = data['length']
        product.width = data['width']
        product.heigth = data['heigth']
        product.weight = data['weight']
        product.is_checked = data['is_checked']
        product.save()
        return Response({'msg': '操作成功!'}, status=status.HTTP_200_OK)

    # 迁移sku查询
    @action(methods=['post'], detail=False, url_path='move_check')
    def move_check(self, request):
        sku = request.data['sku']
        product = MLProduct.objects.filter(sku=sku).first()
        if not product:
            return Response({
                'msg': '产品不存在，请检查sku是否正确！',
                'status': 'error'
            },
                            status=status.HTTP_202_ACCEPTED)
        url = '{base_url}/media/ml_product/{sku}.jpg'.format(base_url=BASE_URL,
                                                             sku=product.sku)
        p = {
            'sku': product.sku,
            'p_name': product.p_name,
            'image': url if product.image else '',
            'shop': product.shop,
        }

        return Response({'product': p}, status=status.HTTP_200_OK)

    # 迁移sku
    @action(methods=['post'], detail=False, url_path='move_sku')
    def move_sku(self, request):
        data = request.data
        product = MLProduct.objects.filter(sku=data['target_sku']).first()
        shop = Shop.objects.filter(name=product.shop).first()
        packing = Packing.objects.filter(id=product.packing_id).first()
        purchase_manage = PurchaseManage(
            p_status='RECEIVED',
            s_type='REFILL',
            create_type='MOVE',
            platform=product.platform,
            sku=product.sku,
            p_name=product.p_name,
            item_id=product.item_id,
            label_code=product.label_code,
            image=product.image,
            unit_cost=product.unit_cost,
            weight=product.weight,
            length=product.length,
            width=product.width,
            heigth=product.heigth,
            rec_qty=data['move_qty'],
            pack_qty=data['move_qty'],
            shop=shop.name,
            shop_color=shop.name_color,
            packing_size=packing.size if packing else '',
            packing_name=packing.name if packing else '',
            rec_time=datetime.now(),
            create_time=datetime.now())
        purchase_manage.save()

        pm = PurchaseManage.objects.filter(id=data['from_id']).first()
        if data['move_qty'] < data['from_qty']:
            pm.rec_qty -= data['move_qty']
            pm.pack_qty -= data['move_qty']
            pm.save()
        else:
            pm.delete()

        # 创建操作日志
        log = MLOperateLog()
        log.op_module = 'PURCHASE'
        log.op_type = 'CREATE'
        log.target_type = 'PURCHASE'
        log.desc = '迁移 {from_sku} 数量{move_qty}个至 {target_sku}'.format(
            from_sku=data['from_sku'],
            move_qty=data['move_qty'],
            target_sku=data['target_sku'])
        log.user = request.user
        log.save()
        return Response({
            'msg': '操作成功',
            'code': 'success'
        },
                        status=status.HTTP_200_OK)

    # 重写
    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()

        # 创建操作日志
        log = MLOperateLog()
        log.op_module = 'PURCHASE'
        log.op_type = 'DEL'
        log.target_type = 'PURCHASE'
        log.target_id = instance.id
        log.desc = '删除产品 {sku} {name}'.format(sku=instance.sku,
                                              name=instance.p_name)
        log.user = request.user
        log.save()

        self.perform_destroy(instance)
        return Response(status=status.HTTP_204_NO_CONTENT)


class UPCViewSet(mixins.ListModelMixin, mixins.CreateModelMixin,
                 mixins.UpdateModelMixin, mixins.DestroyModelMixin,
                 mixins.RetrieveModelMixin, viewsets.GenericViewSet):
    """
    list:
        UPC号码池列表,分页,过滤,搜索,排序
    create:
        UPC号码池新增
    retrieve:
        UPC号码池详情页
    update:
        UPC号码池修改
    destroy:
        UPC号码池删除
    """
    queryset = UPC.objects.all()
    serializer_class = UPCSerializer  # 序列化
    pagination_class = DefaultPagination  # 分页

    filter_backends = (DjangoFilterBackend, filters.SearchFilter,
                       filters.OrderingFilter)  # 过滤,搜索,排序
    filter_fields = ('is_used', 'user', 'use_time')  # 配置过滤字段
    filterset_fields = {
        'use_time': ['gte', 'lte', 'exact', 'gt', 'lt'],
        'is_used': ['exact'],
        'user': ['exact'],
    }
    search_fields = ('number', )  # 配置搜索字段
    ordering_fields = ('create_time', 'use_time')  # 配置排序字段

    # upc号码上传
    @action(methods=['post'], detail=False, url_path='bulk_upload')
    def bulk_upload(self, request):
        data = request.data
        wb = openpyxl.load_workbook(data['excel'])
        sheet = wb.active

        add_list = []
        for cell_row in list(sheet):
            number = cell_row[0].value
            if not number:
                continue
            is_exist = UPC.objects.filter(number=number).count()
            if not is_exist:
                add_list.append(UPC(number=number))
        UPC.objects.bulk_create(add_list)
        msg = '成功上传 {s} 条'.format(s=len(add_list))

        return Response({'msg': msg}, status=status.HTTP_200_OK)

    # 申请upc号码
    @action(methods=['post'], detail=False, url_path='request_upc')
    def request_upc(self, request):
        qty = request.data['qty']
        available = UPC.objects.filter(is_used=False).count()
        if qty > available:
            return Response({'msg': 'UPC号码数量不足，请联系管理员'},
                            status=status.HTTP_202_ACCEPTED)
        upcs = UPC.objects.filter(is_used=False)[:qty]
        for i in upcs:
            i.is_used = True
            i.user = request.user
            i.use_time = datetime.now()
            i.save()
        return Response({'msg': '申请成功'}, status=status.HTTP_200_OK)


class RefillRecommendViewSet(mixins.ListModelMixin, mixins.CreateModelMixin,
                             mixins.UpdateModelMixin, mixins.DestroyModelMixin,
                             mixins.RetrieveModelMixin,
                             viewsets.GenericViewSet):
    """
    list:
        补货推荐列表,分页,过滤,搜索,排序
    create:
        补货推荐新增
    retrieve:
        补货推荐详情页
    update:
        补货推荐修改
    destroy:
        补货推荐删除
    """
    queryset = RefillRecommend.objects.all()
    serializer_class = RefillRecommendSerializer  # 序列化
    pagination_class = DefaultPagination  # 分页

    filter_backends = (DjangoFilterBackend, filters.SearchFilter,
                       filters.OrderingFilter)  # 过滤,搜索,排序
    filter_fields = ('is_new', 'shop', 'trend')  # 配置过滤字段
    search_fields = ('sku', 'p_name', 'item_id')  # 配置搜索字段
    ordering_fields = ('create_time', 'item_id', 'all_sold', 'days30_sold',
                       'days15_sold', 'days7_sold', 'min_send', 'full_send'
                       )  # 配置排序字段

    # 创建补货推荐列表
    @action(methods=['post'], detail=False, url_path='create_refill')
    def create_refill(self, request):
        ship_type = request.data['ship_type']
        end_date = request.data['end_date']
        shop_id = request.data['shop_id']
        shop = Shop.objects.filter(id=shop_id).first()

        # 删除旧数据
        RefillRecommend.objects.filter(shop=shop).delete()

        # 生成批次号
        b_num = end_date.replace('-', '')[2:]
        batch = 'P{time_str}'.format(time_str=b_num)

        refill_setting = RefillSettings.objects.filter(
            platform=shop.platform).first()
        if not refill_setting:
            return Response({
                'msg': '补货参数未初始化!',
                'status': 'error'
            },
                            status=status.HTTP_202_ACCEPTED)
        # 获取设置参数
        ship_days = refill_setting.fly_days  # 截单日期到上架所需天数
        batch_period = refill_setting.fly_batch_period  # 批次周期（发货批次间隔天数）
        is_include_trans = refill_setting.is_include_trans
        if ship_type == 'SEA':
            ship_days = refill_setting.sea_days
            batch_period = refill_setting.sea_batch_period

        # 计算FBM已入仓产品
        queryset = ShopStock.objects.filter(shop__id=shop_id)
        add_list = []
        for i in queryset:
            # 停售和清仓产品不计算
            if i.p_status == 'OFFLINE' or i.p_status == 'CLEAN':
                continue
            # 该批次已备货的产品不计算
            is_exist = ShipDetail.objects.filter(ship__s_status='PREPARING',
                                                 ship__batch=batch,
                                                 sku=i.sku).count()
            if is_exist:
                continue

            # 获取首次入仓的天数
            sd = ShipDetail.objects.filter(
                ship__s_status='FINISHED',
                sku=i.sku).order_by('-create_time').first()
            is_new = False  # 是否新品
            first_list_days = 0  # 首次上架天数
            avg_sale = 0  # 日均销量
            trend = 'UP'  # 15天销量趋势
            if sd:
                first_list_date = str(sd.create_time.date())
                dd = datetime.strptime(first_list_date, '%Y-%m-%d')
                delta = datetime.now() - dd
                is_new = True if delta.days < 15 else False
                first_list_days = delta.days

            # 入仓时间小于15天为新品，新品日均销量为7天日均，非新品为15天日均
            if is_new:
                avg_sale = i.day7_sold / 7
            else:
                avg_sale = i.day15_sold / 15

            # 15天销量趋势
            avg15 = i.day15_sold / 15
            avg30 = i.day30_sold / 30
            if avg15 < avg30:
                trend = 'DOWN'

            trans_qty = 0  # 中转仓数量
            trans_onway_qty = 0  # 中转在途数量
            all_stock = i.qty + i.onway_qty  # 总库存
            # 中转在仓数量
            ts = TransStock.objects.filter(sku=i.sku, is_out=False)
            for t in ts:
                trans_qty += t.qty
            # 中转在途数量
            ssd = ShipDetail.objects.filter(
                sku=i.sku, ship__target='TRANSIT').filter(
                    Q(ship__s_status='SHIPPED') | Q(ship__s_status='BOOKED'))
            for s in ssd:
                trans_onway_qty += s.qty
            # 如果计算包含中转仓数量
            if is_include_trans:
                all_stock = i.qty + i.onway_qty + trans_qty + trans_onway_qty

            # 库存维持天数,日均销量为0，默认维持天数为100天
            keep_days = int(all_stock / avg_sale) if avg_sale else 100

            dd = datetime.strptime(end_date, '%Y-%m-%d')
            delta = dd - datetime.now()
            current_days = delta.days + 1

            # 如果维持天数大于 从现在起到下个周期发货所需天数，说明库存充足， 无需补货
            if keep_days > current_days + ship_days + batch_period and i.p_status != 'HOT_SALE':
                continue

            min_send = 0  # 最低发货数量
            full_send = 0  # 完整周期发货数量
            advice = ''

            # 维持天数小于货运时间 + 当前时间,将会断货
            if keep_days < current_days + ship_days:
                # 补货天数：周期 + 货运时间
                min_send = int(avg_sale * (batch_period + ship_days))
                full_send = int(avg_sale *
                                (current_days + batch_period + ship_days))
                advice = '即将缺货'
            # 维持天数大于货运时间 + 当前时间,不断货
            elif keep_days < current_days + ship_days + batch_period:
                min_send = int(avg_sale * batch_period)
                full_send = int(avg_sale * ship_days)
                advice = '本批次不补货将可能断货'
            # 维持天数大于从现在起到下个周期发货所需天数,库存充足
            else:
                min_send = 0
                full_send = 0
                advice = '库存充足'

            # 备货中数量
            prepare_qty = 0
            sd = ShipDetail.objects.filter(ship__s_status='PREPARING',
                                           sku=i.sku)
            for s in sd:
                prepare_qty += s.qty
            own_qty = 0
            # 已采购现货数量
            pm = PurchaseManage.objects.filter(sku=i.sku).filter(
                Q(p_status='RECEIVED') | Q(p_status='PACKED'))
            for p in pm:
                if p.p_status == 'RECEIVED':
                    own_qty += p.rec_qty
                if p.p_status == 'PACKED':
                    own_qty += p.pack_qty

            # fbm在途库存数量
            fbm_onway_qty = 0
            sd_set = ShipDetail.objects.filter(
                sku=i.sku, ship__target='TRANSIT').filter(
                    Q(ship__s_status='SHIPPED') | Q(ship__s_status='BOOKED'))
            for sst in sd_set:
                fbm_onway_qty += sst.qty

            add_list.append(
                RefillRecommend(
                    shop=shop,
                    sku=i.sku,
                    p_name=i.p_name,
                    image=i.image,
                    item_id=i.item_id,
                    is_new=is_new,
                    first_list_days=first_list_days,
                    trend=trend,
                    all_sold=i.total_sold,
                    days30_sold=i.day30_sold,
                    days15_sold=i.day15_sold,
                    days7_sold=i.day7_sold,
                    fbm_qty=i.qty,
                    onway_qty=fbm_onway_qty,
                    trans_qty=trans_qty,
                    trans_onway_qty=trans_onway_qty,
                    prepare_qty=prepare_qty,
                    own_qty=own_qty,
                    avg_sale=avg_sale,
                    keep_days=keep_days,
                    min_send=min_send,
                    full_send=full_send,
                    advice=advice,
                ))
        if len(add_list) > 0:
            RefillRecommend.objects.bulk_create(add_list)
        msg = '成功推荐{n}个产品'.format(n=len(add_list))

        return Response({'msg': msg}, status=status.HTTP_200_OK)


class RefillSettingsViewSet(mixins.ListModelMixin, mixins.CreateModelMixin,
                            mixins.UpdateModelMixin, mixins.DestroyModelMixin,
                            mixins.RetrieveModelMixin,
                            viewsets.GenericViewSet):
    """
    list:
        补货推荐设置列表,分页,过滤,搜索,排序
    create:
        补货推荐设置新增
    retrieve:
        补货推荐设置详情页
    update:
        补货推荐设置修改
    destroy:
        补货推荐设置删除
    """
    queryset = RefillSettings.objects.all()
    serializer_class = RefillSettingsSerializer  # 序列化
    pagination_class = DefaultPagination  # 分页

    filter_backends = (DjangoFilterBackend, filters.SearchFilter,
                       filters.OrderingFilter)  # 过滤,搜索,排序
    filter_fields = ('is_include_trans', )  # 配置过滤字段
    search_fields = ('fly_days', )  # 配置搜索字段
    ordering_fields = ('fly_days', )  # 配置排序字段


class CarrierTrackViewSet(mixins.ListModelMixin, mixins.CreateModelMixin,
                          mixins.UpdateModelMixin, mixins.DestroyModelMixin,
                          mixins.RetrieveModelMixin, viewsets.GenericViewSet):
    """
    list:
        运单物流跟踪列表,分页,过滤,搜索,排序
    create:
        运单物流跟踪新增
    retrieve:
        运单物流跟踪详情页
    update:
        运单物流跟踪修改
    destroy:
        运单物流跟踪删除
    """
    queryset = CarrierTrack.objects.all()
    serializer_class = CarrierTrackSerializer  # 序列化
    pagination_class = DefaultPagination  # 分页

    filter_backends = (DjangoFilterBackend, filters.SearchFilter,
                       filters.OrderingFilter)  # 过滤,搜索,排序
    filter_fields = ('carrier_name', 'carrier_number')  # 配置过滤字段
    search_fields = ('context', )  # 配置搜索字段
    ordering_fields = ('create_time', 'time', 'optime')  # 配置排序字段


class MLStockLogViewSet(mixins.ListModelMixin, mixins.CreateModelMixin,
                        mixins.UpdateModelMixin, mixins.DestroyModelMixin,
                        mixins.RetrieveModelMixin, viewsets.GenericViewSet):
    """
    list:
        库存日志列表,分页,过滤,搜索,排序
    create:
        库存日志新增
    retrieve:
        库存日志详情页
    update:
        库存日志修改
    destroy:
        库存日志删除
    """
    queryset = StockLog.objects.all()
    serializer_class = StockLogSerializer  # 序列化
    pagination_class = DefaultPagination  # 分页

    filter_backends = (DjangoFilterBackend, filters.SearchFilter,
                       filters.OrderingFilter)  # 过滤,搜索,排序
    filter_fields = ('shop_stock', 'in_out', 'action', 'user_id')  # 配置过滤字段
    search_fields = ('desc', )  # 配置搜索字段
    ordering_fields = ('create_time', )  # 配置排序字段


class FileUploadNotifyViewSet(mixins.ListModelMixin, mixins.CreateModelMixin,
                              mixins.UpdateModelMixin,
                              mixins.DestroyModelMixin,
                              mixins.RetrieveModelMixin,
                              viewsets.GenericViewSet):
    """
    list:
        文件上传通知列表,分页,过滤,搜索,排序
    create:
        文件上传通知新增
    retrieve:
        文件上传通知详情页
    update:
        文件上传通知修改
    destroy:
        文件上传通知删除
    """
    queryset = FileUploadNotify.objects.all()
    serializer_class = FileUploadNotifySerializer  # 序列化
    pagination_class = DefaultPagination  # 分页

    filter_backends = (DjangoFilterBackend, filters.SearchFilter,
                       filters.OrderingFilter)  # 过滤,搜索,排序
    filter_fields = ('shop', 'upload_status', 'upload_type', 'user_id'
                     )  # 配置过滤字段
    search_fields = ('desc', )  # 配置搜索字段
    ordering_fields = ('create_time', )  # 配置排序字段

    # 获取订单上传记录
    @action(methods=['post'], detail=False, url_path='get_upload_result')
    def get_upload_result(self, request):
        data = request.data
        shop_id = data['id']
        shop = Shop.objects.filter(id=shop_id).first()
        # 开发产品订单记录
        if shop_id == 0:
            shop = None

        res = FileUploadNotify.objects.filter(shop=shop).first()
        if not res:
            return Response({'status': '无记录！'}, status=status.HTTP_200_OK)

        return Response(
            {
                'upload_status': res.upload_status,
                'create_time': res.create_time,
                'desc': res.desc,
                'status': 'success'
            },
            status=status.HTTP_200_OK)


class PlatformCategoryRateViewSet(mixins.ListModelMixin,
                                  mixins.CreateModelMixin,
                                  mixins.UpdateModelMixin,
                                  mixins.DestroyModelMixin,
                                  mixins.RetrieveModelMixin,
                                  viewsets.GenericViewSet):
    """
    list:
        平台类目佣金费率列表,分页,过滤,搜索,排序
    create:
        平台类目佣金费率新增
    retrieve:
        平台类目佣金费率详情页
    update:
        平台类目佣金费率修改
    destroy:
        平台类目佣金费率删除
    """
    queryset = PlatformCategoryRate.objects.all()
    serializer_class = PlatformCategoryRateSerializer  # 序列化
    pagination_class = DefaultPagination  # 分页

    filter_backends = (DjangoFilterBackend, filters.SearchFilter,
                       filters.OrderingFilter)  # 过滤,搜索,排序
    filter_fields = ('platform', 'site', 'first_category')  # 配置过滤字段
    search_fields = ('en_name', 'cn_name', 'first_category')  # 配置搜索字段

    # 上传平台类目费率表
    @action(methods=['get'], detail=False, url_path='upload_category')
    def upload_category(self, request):
        import warnings
        import openpyxl
        warnings.filterwarnings('ignore')

        data = MEDIA_ROOT + '/upload_file/ebay澳洲类目.xlsx'
        wb = openpyxl.load_workbook(data)
        sheet = wb.active

        for cell_row in list(sheet)[2:]:
            pc = PlatformCategoryRate()
            pc.platform = 'ebay'
            pc.site = cell_row[0].value
            pc.category_id = cell_row[1].value
            pc.en_name = cell_row[2].value
            pc.cn_name = cell_row[3].value
            pc.first_category = cell_row[4].value
            pc.fixed_fee0 = cell_row[5].value
            pc.fixed_fee1 = cell_row[6].value
            pc.fee0 = cell_row[7].value
            pc.fee1 = cell_row[8].value
            pc.fee2 = cell_row[9].value
            pc.fee3 = cell_row[10].value
            pc.save()

        return Response({'message': 'ok'}, status=status.HTTP_200_OK)

    # 上传虚拟仓物流价格表
    @action(methods=['get'], detail=False, url_path='upload_shipping')
    def upload_shipping(self, request):
        import warnings
        import openpyxl
        warnings.filterwarnings('ignore')

        data = MEDIA_ROOT + '/upload_file/物流价格表.xlsx'
        wb = openpyxl.load_workbook(data)
        sheet = wb.active

        for cell_row in list(sheet)[2:]:
            pc = ShippingPrice()
            pc.country = cell_row[0].value
            pc.carrier_name = cell_row[1].value
            pc.carrier_code = cell_row[2].value
            pc.area = cell_row[3].value
            pc.min_weight = cell_row[4].value
            pc.max_weight = cell_row[5].value
            pc.basic_price = cell_row[6].value
            pc.calc_price = cell_row[7].value
            pc.volume_ratio = cell_row[8].value
            pc.is_elec = True if cell_row[9].value else False
            pc.save()

        return Response({'message': 'ok'}, status=status.HTTP_200_OK)

    # 上传虚拟仓物流分区表
    @action(methods=['get'],
            detail=False,
            url_path='upload_shipping_area_code')
    def upload_shipping_area_code(self, request):
        import warnings
        import openpyxl
        warnings.filterwarnings('ignore')
        ShippingAreaCode.objects.all().delete()

        data = MEDIA_ROOT + '/upload_file/物流分区表.xlsx'
        wb = openpyxl.load_workbook(data)
        sheet = wb.active

        for cell_row in list(sheet)[1:]:
            pc = ShippingAreaCode()
            pc.country = cell_row[0].value
            pc.carrier_name = cell_row[1].value
            pc.carrier_code = cell_row[2].value
            pc.postcode = cell_row[3].value
            pc.area = cell_row[4].value
            pc.is_avaiable = False if cell_row[
                5].value == 'out of network' else True
            pc.save()

        return Response({'message': 'ok'}, status=status.HTTP_200_OK)

    # 获取计算物流运费和利润
    @action(methods=['post'], detail=False, url_path='get_shipping_price')
    def get_shipping_price(self, request):
        data = request.data
        site = data['site']  # 站点、国家
        p_width = data['p_width']  # 宽
        p_height = data['p_height']  # 高
        p_length = data['p_length']  # 长
        p_weight = data['p_weight']  # 重量g
        p_cost = data['p_cost']  # 产品成本价
        carrier_code = data['carrier_code']  # 物流渠道
        rec_rmb = data['rec_rmb']  # rmb收款
        price = data['price']  # 售价
        ex_rate = data['ex_rate']  # 汇率

        if carrier_code == 'ALL':
            sp = ShippingPrice.objects.filter(country=site,
                                              min_weight__lte=p_weight,
                                              max_weight__gte=p_weight)
        else:
            sp = ShippingPrice.objects.filter(country=site,
                                              carrier_code=carrier_code,
                                              min_weight__lte=p_weight,
                                              max_weight__gte=p_weight)

        p_list = []
        for i in sp:
            # 计算计费重量
            calc_weight = 0
            # 体积重
            v_weight = p_width * p_height * p_length / i.volume_ratio
            # ubi计费方式
            if i.carrier_code in [
                    'UBI_AUS_POST', 'UBI_AUS_POST_E', 'UBI_ARAMEX',
                    'UBI_ARAMEX_E'
            ]:
                # 体积重量与实际重量比低于2:1的，按照实际重量收费。超过2:1的，按照体积重量收取
                if p_weight * 2 < v_weight:
                    calc_weight = v_weight
                else:
                    calc_weight = p_weight
            # 其它物流计费方式
            else:
                if v_weight > p_weight:
                    calc_weight = v_weight
                else:
                    calc_weight = p_weight

            # 运费计算
            cost = calc_weight * i.calc_price / 1000 + i.basic_price

            # 利润
            profit = rec_rmb - cost - p_cost

            # 毛利率
            gross_margin = profit / (price * ex_rate)

            p_list.append({
                'name': i.carrier_name,
                'area': i.area,
                'calc_weight': calc_weight,
                'cost': cost,
                'profit': profit,
                'gross_margin': gross_margin,
            })

        return Response(p_list, status=status.HTTP_200_OK)

    # 获取汇率
    @action(methods=['post'], detail=False, url_path='get_ex_rate')
    def get_ex_rate(self, request):
        currency = request.data['currency']
        rate = ExRate.objects.filter(currency=currency).first()

        value = 0
        update = ''
        if rate:
            value = rate.value
            update = rate.update_time.strftime('%Y%m%d')
        return Response({
            'value': value,
            'update': update
        },
                        status=status.HTTP_200_OK)

    # 查询邮编分区
    @action(methods=['post'], detail=False, url_path='get_postcode_area')
    def get_postcode_area(self, request):
        postcode = request.data['postcode'].strip()
        postage = request.data['postage']
        weight = request.data['weight']
        cost = 0

        sac_list = ShippingAreaCode.objects.filter(postcode=postcode)
        add_list = []
        for i in sac_list:
            # 计算运费
            if i.is_avaiable and postage:
                sp = ShippingPrice.objects.filter(country=i.country,
                                                  carrier_code=i.carrier_code,
                                                  area=i.area).first()
                if sp:
                    cost = weight * sp.calc_price / 1000 + sp.basic_price
            add_list.append({
                'id': i.id,
                'country': i.country,
                'carrier_name': i.carrier_name,
                'area': i.area,
                'postcode': i.postcode,
                'is_avaiable': i.is_avaiable,
                'note': i.note,
                'update_time': i.update_time,
                'cost': cost,
            })

        if not len(add_list):
            return Response({
                'msg': '找不到邮编分区!',
                'status': 'error'
            },
                            status=status.HTTP_202_ACCEPTED)
        return Response(add_list, status=status.HTTP_200_OK)

    # 修改邮编服务
    @action(methods=['post'], detail=False, url_path='change_postcode_tag')
    def change_postcode_tag(self, request):
        id = request.data['id']
        is_avaiable = request.data['is_avaiable']

        sac = ShippingAreaCode.objects.filter(id=id).first()
        sac.is_avaiable = not is_avaiable
        sac.note = '{name}修改了服务标记'.format(name=request.user.first_name)
        sac.update_time = datetime.now()
        sac.save()

        return Response({
            'msg': '成功修改!',
            'status': 'success'
        },
                        status=status.HTTP_200_OK)

    # 批量查询邮编分区
    @action(methods=['post'], detail=False, url_path='bulk_get_postcode_area')
    def bulk_get_postcode_area(self, request):
        import warnings
        warnings.filterwarnings('ignore')

        data = request.data
        carrier_code = request.data['carrier_code']
        is_remote = request.data['is_remote']
        wb = openpyxl.load_workbook(data['excel'])
        # sheet = wb['Sheet1']
        sheet = wb.active

        if sheet.max_row <= 1:
            return Response({
                'msg': '表格不能为空',
                'status': 'error'
            },
                            status=status.HTTP_202_ACCEPTED)
        if sheet['A1'].value != '订单编号' or sheet['B1'].value != '邮政编码':
            return Response({
                'msg': '表格有误，请下载最新模板',
                'status': 'error'
            },
                            status=status.HTTP_202_ACCEPTED)

        new_wb = openpyxl.Workbook()
        sh = new_wb.active
        sh.title = 'Sheet1'
        sh['A1'] = '订单编号'
        sh['B1'] = '邮政编码'
        sh['C1'] = '分区'
        num = 0
        for cell_row in list(sheet)[1:]:
            order_number = cell_row[0].value.strip()
            postcode = cell_row[1].value.strip()
            sac = ShippingAreaCode.objects.filter(
                postcode=postcode, carrier_code=carrier_code).first()
            area = '分区不存在'
            if sac:
                area = sac.area

            # 仅筛选偏远分区
            if is_remote == 'true':
                if area == '1区' or area == '2区':
                    continue

            sh['A' + str(num + 2)] = order_number
            sh['B' + str(num + 2)] = postcode
            sh['C' + str(num + 2)] = area
            num += 1

        new_wb.save('media/export/批量邮编查询结果-' + carrier_code + '.xlsx')
        url = BASE_URL + '/media/export/批量邮编查询结果-' + carrier_code + '.xlsx'
        print(is_remote)

        return Response({
            'msg': '操作成功!',
            'url': url,
            'status': 'success'
        },
                        status=status.HTTP_200_OK)
