from django.db.models import Sum
from rest_framework import mixins, status
from rest_framework import viewsets
from rest_framework.response import Response
from rest_framework import filters
from rest_framework.pagination import PageNumberPagination
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.decorators import action
import openpyxl
import barcode
from barcode.writer import ImageWriter
from fpdf import FPDF
from product import tasks

from casedance.settings import BASE_URL
from purchase.models import PurchaseDetail, PurchaseOrder
from sale.models import OrderDetail
from setting.models import OperateLog, Tag
from store.models import Store, Stock, StockInOutDetail
from .models import Product, ProductExtraInfo, DeviceModel, CompatibleModel, ProductTag, Supplier, DeviceBrand
from .serializers import ProductSerializer, ProductExtraInfoSerializer, DeviceModelSerializer, \
    CompatibleModelSerializer, ProductTagSerializer, SupplierSerializer, SimpleProductSerializer, DeviceBrandSerializer


# Create your views here.
class DefaultPagination(PageNumberPagination):
    """
    分页设置
    """
    page_size = 20
    page_size_query_param = 'page_size'
    page_query_param = 'page'
    max_page_size = 10000


class SimpleProductViewSet(mixins.ListModelMixin,
                           viewsets.GenericViewSet):
    """
        list:
            简易产品列表

        """
    queryset = Product.objects.all()
    serializer_class = SimpleProductSerializer  # 序列化


class ProductViewSet(mixins.ListModelMixin,
                     mixins.CreateModelMixin,
                     mixins.UpdateModelMixin,
                     mixins.DestroyModelMixin,
                     mixins.RetrieveModelMixin,
                     viewsets.GenericViewSet):
    """
    list:
        产品列表,分页,过滤,搜索,排序
    create:
        产品新增
    retrieve:
        产品详情页
    update:
        产品修改
    destroy:
        产品删除
    """
    queryset = Product.objects.all()
    serializer_class = ProductSerializer  # 序列化
    pagination_class = DefaultPagination  # 分页

    filter_backends = (DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter)  # 过滤,搜索,排序
    filter_fields = ('status', 'brand', 'series', 'p_type', 'is_auto_promote', 'stock_strategy',
                     'product_p_tag__tag__tag_name', 'product_stock__store', 'product_stock__product')  # 配置过滤字段
    search_fields = ('sku', 'p_name', 'product_p_tag__tag__tag_name')  # 配置搜索字段
    ordering_fields = ('create_time', 'sku', 'p_name')  # 配置排序字段

    #  重写产品删除
    def destroy(self, request, *args, **kwargs):
        product = self.get_object()
        # 检查是否有手机兼容产品存在
        is_comp_product = CompatibleModel.objects.filter(product=product).count()
        # 检查是否有产品标签存在
        is_product_tag = ProductTag.objects.filter(product=product).count()
        # 检查是否有采购单存在
        is_purchase_order = PurchaseDetail.objects.filter(product=product).count()
        # 检查是否有销售单存在
        is_order = OrderDetail.objects.filter(product=product).count()
        # 检查是否有手工出入库单存在
        is_stock_inout = StockInOutDetail.objects.filter(product=product).count()
        # 检查是否有库存
        stock_num = Stock.objects.filter(product=product).aggregate(Sum('qty'))
        is_stock = stock_num['qty__sum']

        if is_comp_product or is_product_tag or is_purchase_order or is_order or is_stock_inout or is_stock:
            return Response({'msg': 'sku有关联数据，无法删除'}, status=status.HTTP_406_NOT_ACCEPTABLE)

        # 如果满足条件，则可以删除产品
        product.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    # 产品批量修改
    @action(methods=['post'], detail=False, url_path='bulk_edit')
    def bulk_edit(self, request):
        if not 'ids' in request.data:
            return Response({'msg': '非法操作'}, status=status.HTTP_400_BAD_REQUEST)
        change_item = request.data['changeItems']
        for id in request.data['ids']:
            if Product.objects.filter(id=id).count():
                product = Product.objects.get(id=id)
                if 'image' in change_item:
                    product.image = change_item['image']
                if 'status' in change_item:
                    product.status = change_item['status']
                if 'series' in change_item:
                    product.series = change_item['series']
                if 'p_type' in change_item:
                    product.p_type = change_item['p_type']
                if 'unit_cost' in change_item:
                    product.unit_cost = change_item['unit_cost']
                if 'sale_price' in change_item:
                    product.sale_price = change_item['sale_price']
                if 'length' in change_item:
                    product.length = change_item['length']
                if 'width' in change_item:
                    product.width = change_item['width']
                if 'heigth' in change_item:
                    product.heigth = change_item['heigth']
                if 'weight' in change_item:
                    product.weight = change_item['weight']
                if 'is_auto_promote' in change_item:
                    product.is_auto_promote = change_item['is_auto_promote']
                if 'stock_strategy' in change_item:
                    product.stock_strategy = change_item['stock_strategy']
                if 'stock_days' in change_item:
                    product.stock_days = change_item['stock_days']
                if 'alert_qty' in change_item:
                    product.alert_qty = change_item['alert_qty']
                if 'alert_days' in change_item:
                    product.alert_days = change_item['alert_days']
                if 'mini_pq' in change_item:
                    product.mini_pq = change_item['mini_pq']
                if 'max_pq' in change_item:
                    product.max_pq = change_item['max_pq']
                if 'note' in change_item:
                    product.note = change_item['note']
                product.save()
                if 'tag' in change_item:
                    tag_id = change_item['tag']
                    tag = Tag.objects.get(id=tag_id)
                    # 产品标签在这个产品中不存在才添加
                    if not ProductTag.objects.filter(tag=tag, product=product).count():
                        product_tag = ProductTag()
                        product_tag.product = product
                        product_tag.tag = tag
                        product_tag.save()

        return Response({'msg': '修改成功'}, status=status.HTTP_200_OK)

    # 产品excel批量上传
    @action(methods=['post'], detail=False, url_path='bulk_upload')
    def bulk_upload(self, request):
        data = request.data
        wb = openpyxl.load_workbook(data['excel'])
        sheet = wb['sku上传模板']

        # 取出从第二行开始的所有数据, 检查必填项列是否为空
        # 1 SKU
        # 2 产品名称
        # 4 成本价
        # 5 售价
        # 6 产品系列
        err_list = []  # 错误列表
        add_list = []  # 批量新增sku
        if sheet.max_row <= 1:
            err_list.append({'msg': '表格不能为空'})
            return Response(err_list, status=status.HTTP_202_ACCEPTED)
        for cell_row in list(sheet)[1:]:
            err_item = {}
            row_status = cell_row[0].value and cell_row[1].value and cell_row[3].value and cell_row[4].value and \
                         cell_row[5].value

            if not row_status:
                err_item.update({'row': '第 %s 行' % cell_row[0].row})
                err_item.update({'msg': '价格有误'})
                err_list.append(err_item)
                continue
            if not type(cell_row[3].value) in [float, int]:
                err_item.update({'row': '第 %s 行' % cell_row[0].row})
                err_item.update({'msg': '价格有误'})
                err_list.append(err_item)
                continue
            if not type(cell_row[4].value) in [float, int]:
                err_item.update({'row': '第 %s 行' % cell_row[0].row})
                err_item.update({'msg': '价格有误'})
                err_list.append(err_item)
                continue
            if cell_row[6].value and not cell_row[6].value.strip() in ['ON_SALE', 'OFFLINE', 'CLEAN', 'UN_LISTED',
                                                                       'PRIVATE']:
                err_item.update({'row': '第 %s 行' % cell_row[0].row})
                err_item.update({'msg': '产品状态有误'})
                err_list.append(err_item)
                continue

            # 检查sku是否已存在
            sku_is_exist = Product.objects.filter(sku=cell_row[0].value.strip()).count()
            if sku_is_exist:
                err_item = {}
                err_item.update({'row': '第 %s 行' % cell_row[0].row})
                err_item.update({'msg': 'sku: %s 已存在' % cell_row[0].value.strip()})
                err_list.append(err_item)
                continue

            # 检查产品系列是否存在
            series_is_exist = ProductExtraInfo.objects.filter(type='SERIES', name=cell_row[5].value.strip()).count()
            if not series_is_exist:
                err_item = {}
                err_item.update({'row': '第 %s 行' % cell_row[0].row})
                err_item.update({'msg': '产品系列: %s 不存在' % cell_row[5].value.strip()})
                err_list.append(err_item)
                continue

            # 创建产品对象
            sku = cell_row[0].value.strip()
            p_name = cell_row[1].value.strip()
            label_name = cell_row[2].value.strip() if cell_row[2].value else ''
            unit_cost = cell_row[3].value
            sale_price = cell_row[4].value
            series = cell_row[5].value.strip()
            p_status = cell_row[6].value.strip() if cell_row[6].value else 'UN_LISTED'
            image = cell_row[7].value.strip() if cell_row[7].value else None
            brand = cell_row[8].value.strip() if cell_row[8].value else ''
            p_type = cell_row[9].value.strip() if cell_row[9].value else ''
            length = cell_row[10].value if cell_row[10].value else None
            width = cell_row[11].value if cell_row[11].value else None
            heigth = cell_row[12].value if cell_row[12].value else None
            weight = cell_row[13].value if cell_row[13].value else None
            stock_strategy = cell_row[14].value.strip() if cell_row[14].value else 'STANDARD'
            stock_days = cell_row[15].value if cell_row[15].value else None
            alert_qty = cell_row[16].value if cell_row[16].value else None
            alert_days = cell_row[17].value if cell_row[17].value else None
            mini_pq = cell_row[18].value if cell_row[18].value else None
            max_pq = cell_row[19].value if cell_row[19].value else None
            note = cell_row[20].value.strip() if cell_row[20].value else ''
            add_list.append(Product(
                sku=sku,
                p_name=p_name,
                label_name=label_name,
                unit_cost=unit_cost,
                sale_price=sale_price,
                series=series,
                status=p_status,
                image=image,
                brand=brand,
                p_type=p_type,
                length=length,
                width=width,
                heigth=heigth,
                weight=weight,
                stock_strategy=stock_strategy,
                stock_days=stock_days,
                alert_qty=alert_qty,
                alert_days=alert_days,
                mini_pq=mini_pq,
                max_pq=max_pq,
                note=note
            ))
        Product.objects.bulk_create(add_list)

        # bulk_create 无法发送信号，所以手动为所有门店创建一份产品库存记录
        queryset = Store.objects.all()
        if queryset:
            store_add_list = []
            for p in add_list:
                # 根据sku查询拿出批量创建的产品
                product = Product.objects.all().get(sku=p.sku)
                # 记录创建产品日志
                op = OperateLog()
                op.user = request.user
                op.op_log = '通过批量导入创建了产品'
                op.op_type = 'PRODUCT'
                op.target_id = product.id
                op.save()
                for store in queryset:
                    store_add_list.append(Stock(product=product, store=store))
            Stock.objects.bulk_create(store_add_list)

        success_count = len(add_list)
        fail_count = sheet.max_row - 1 - success_count
        all_data = {}
        all_data.update({'err_list': err_list})
        all_data.update({'fail_count': fail_count})
        all_data.update({'success_count': success_count})
        return Response(all_data, status=status.HTTP_201_CREATED)

    # 产品标签生成
    @action(methods=['post'], detail=False, url_path='create_label')
    def create_label(self, request):
        data = request.data['products']

        path = 'media/label/'
        # pdf文件信息
        pdf = FPDF('P', 'mm', (40, 30))
        pdf.add_font('fireflysung', fname='media/sys/fireflysung.ttf')
        pdf.set_font('fireflysung', size=8)
        pdf.set_margin(0.5)

        # 条码图片信息
        options = {'module_height': 4,  # 默认值15.0，条码高度，单位为毫米
                   'module_width': 0.13,  # 默认值0.2，每个条码宽度（？），单位为毫米
                   'quiet_zone': 2,  # 默认值6.5，两端空白宽度，单位为毫米
                   'font_size': 3,  # 默认值10，文本字体大小，单位为磅
                   'text_distance': 1,  # 默认值5.0，文本和条码之间的距离，单位为毫米
                   'dpi': 600,  # 图片分辨率
                   }
        if data:
            file_name = 'product_label'
            for item in data:
                sku = item['sku']
                qty = item['qty']
                product = Product.objects.filter(sku=sku).first()
                if len(data) == 1:
                    file_name = sku

                # 生成条码图片png
                barcode.generate('code128', product.sku,
                                 writer=barcode.writer.ImageWriter(),
                                 output=path + product.sku,
                                 writer_options=options, )
                for i in range(qty):
                    pdf.add_page()
                    pdf.image("media/sys/logo.png", x=8, y=2, w=25)
                    pdf.cell(0, 20, align='C', txt=product.label_name)
                    pdf.image(path + product.sku + '.png', x=-1, y=12, w=44)
            output_name = path + file_name + '.pdf'
            pdf.output(output_name)
            url = BASE_URL + '/' + output_name
        return Response({'url': url}, status=status.HTTP_200_OK)


class ProductExtraInfoViewSet(mixins.ListModelMixin,
                              mixins.CreateModelMixin,
                              mixins.UpdateModelMixin,
                              mixins.DestroyModelMixin,
                              mixins.RetrieveModelMixin,
                              viewsets.GenericViewSet):
    """
    list:
        产品附属信息列表,分页,过滤,搜索,排序
    create:
        产品附属信息新增
    retrieve:
        产品附属信息详情页
    update:
        产品附属信息修改
    destroy:
        产品附属信息删除
    """
    queryset = ProductExtraInfo.objects.all()
    serializer_class = ProductExtraInfoSerializer  # 序列化
    pagination_class = DefaultPagination  # 分页

    filter_backends = (DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter)  # 过滤,搜索,排序
    filter_fields = ('type',)  # 配置过滤字段
    search_fields = ('name',)  # 配置搜索字段


class DeviceModelViewSet(mixins.ListModelMixin,
                         mixins.CreateModelMixin,
                         mixins.UpdateModelMixin,
                         mixins.DestroyModelMixin,
                         mixins.RetrieveModelMixin,
                         viewsets.GenericViewSet):
    """
    list:
        市面手机型号表,分页,过滤,搜索,排序
    create:
        市面手机型号表新增
    retrieve:
        市面手机型号表详情页
    update:
        市面手机型号表修改
    destroy:
        市面手机型号表删除
    """
    queryset = DeviceModel.objects.all()
    serializer_class = DeviceModelSerializer  # 序列化
    pagination_class = DefaultPagination  # 分页

    filter_backends = (DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter)  # 过滤,搜索,排序
    filter_fields = ('brand', 'type',)  # 配置过滤字段
    search_fields = ('model', 'brand', 'note', 'detail_model')  # 配置搜索字段

    # 市面手机型号excel批量上传
    @action(methods=['post'], detail=False, url_path='bulk_upload')
    def bulk_upload(self, request):
        data = request.data
        wb = openpyxl.load_workbook(data['excel'])
        sheet = wb['型号上传模板']

        add_list = []
        if sheet.max_row <= 1:
            return Response({'msg': '表格不能为空'}, status=status.HTTP_202_ACCEPTED)

        for cell_row in list(sheet)[1:]:
            row_status = cell_row[0].value and cell_row[1].value and cell_row[2].value
            if not row_status:
                return Response({'msg': '必填项不能为空'}, status=status.HTTP_202_ACCEPTED)

            # 检查型号是否已存在
            is_exist = DeviceModel.objects.filter(model=cell_row[0].value.strip()).count()
            if is_exist:
                continue

            model = cell_row[0].value.strip()
            brand = cell_row[1].value.strip()
            m_type = cell_row[2].value.strip()
            note = cell_row[3].value

            add_list.append(DeviceModel(
                model=model,
                brand=brand,
                type=m_type,
                note=note
            ))
        DeviceModel.objects.bulk_create(add_list)

        return Response({'msg': '成功上传'}, status=status.HTTP_200_OK)

    # 型号绑定
    @action(methods=['post'], detail=False, url_path='cp_band')
    def cp_band(self, request):
        ids = request.data['ids']

        exist_cp = []
        for i in ids:
            dm = DeviceModel.objects.get(id=i)
            if dm.cp_id:
                # 如果关联id不在列表中就加进去
                if dm.cp_id not in exist_cp:
                    exist_cp.append(dm.cp_id)
        if len(exist_cp) > 1:
            return Response({'msg': '兼容型号绑定冲突，请检查'}, status=status.HTTP_202_ACCEPTED)

        if len(exist_cp) == 1:
            cp_id = exist_cp[0]
            for i in ids:
                dm = DeviceModel.objects.get(id=i)
                dm.cp_id = cp_id
                dm.save()

        if len(exist_cp) == 0:
            cp_id = ids[0]
            for i in ids:
                dm = DeviceModel.objects.get(id=i)
                dm.cp_id = cp_id
                dm.save()

        return Response({'msg': '成功绑定'}, status=status.HTTP_200_OK)

    # 型号解绑
    @action(methods=['post'], detail=False, url_path='cp_unband')
    def cp_unband(self, request):
        m_id = request.data['id']
        dm = DeviceModel.objects.get(id=m_id)
        delete_cp_id = dm.cp_id
        dm.cp_id = None
        dm.save()

        rest_count = DeviceModel.objects.filter(cp_id=delete_cp_id).count()
        # 如果只有剩下1
        if rest_count == 1:
            dm1 = DeviceModel.objects.get(cp_id=delete_cp_id)
            dm1.cp_id = None
            dm.save()

        if rest_count > 1 and DeviceModel.objects.filter(cp_id=m_id).count():
            new_cp = DeviceModel.objects.filter(cp_id=m_id).first()
            new_cp_id = new_cp.id
            qs = DeviceModel.objects.filter(cp_id=m_id)
            for i in qs:
                i.cp_id = new_cp_id
                i.save()

        return Response({'msg': '成功解绑'}, status=status.HTTP_200_OK)

    # 爬取品牌列表
    @action(methods=['get'], detail=False, url_path='get_brands')
    def get_brands(self, request):
        tasks.get_brands.delay()
        return Response({'msg': 'OK'}, status=status.HTTP_200_OK)

    # 爬取型号信息
    @action(methods=['get'], detail=False, url_path='get_device_models')
    def get_device_models(self, request):
        tasks.get_device_models()
        return Response({'msg': 'OK'}, status=status.HTTP_200_OK)

    # 爬取型号详细信息
    @action(methods=['get'], detail=True, url_path='get_model_info')
    def get_model_info(self, request, pk):
        tasks.get_models_info(pk)
        return Response({'msg': '更新成功'}, status=status.HTTP_200_OK)

    # 获取最近更新的手机型号
    @action(methods=['get'], detail=False, url_path='check_new_models')
    def check_new_models(self, request):
        tasks.check_new_models()
        return Response({'msg': '更新成功'}, status=status.HTTP_200_OK)

    # 获取手机型号参数
    @action(methods=['get'], detail=False, url_path='update_spec')
    def update_spec(self, request):
        tasks.update_spec()
        return Response({'msg': '更新成功'}, status=status.HTTP_200_OK)


class DeviceBrandViewSet(mixins.ListModelMixin,
                         viewsets.GenericViewSet):
    """
        list:
            市面手机品牌

        """
    queryset = DeviceBrand.objects.all()
    serializer_class = DeviceBrandSerializer  # 序列化


class CompatibleModelViewSet(mixins.ListModelMixin,
                             mixins.CreateModelMixin,
                             mixins.UpdateModelMixin,
                             mixins.DestroyModelMixin,
                             mixins.RetrieveModelMixin,
                             viewsets.GenericViewSet):
    """
    list:
        产品兼容手机型号,分页,过滤,搜索,排序
    create:
        产品兼容手机型号新增
    retrieve:
        产品兼容手机型号详情页
    update:
        产品兼容手机型号修改
    destroy:
        产品兼容手机型号删除
    """
    queryset = CompatibleModel.objects.all()
    serializer_class = CompatibleModelSerializer  # 序列化
    pagination_class = DefaultPagination  # 分页

    filter_backends = (DjangoFilterBackend, filters.SearchFilter)  # 过滤,搜索,排序
    filter_fields = ('product', 'phone_model',)  # 配置过滤字段
    search_fields = ('phone_model',)  # 配置搜索字段


class ProductTagViewSet(mixins.ListModelMixin,
                        mixins.CreateModelMixin,
                        mixins.UpdateModelMixin,
                        mixins.DestroyModelMixin,
                        mixins.RetrieveModelMixin,
                        viewsets.GenericViewSet):
    """
    list:
        产品标签,分页,过滤,搜索,排序
    create:
        产品标签新增
    retrieve:
        产品标签详情页
    update:
        产品标签修改
    destroy:
        产品标签删除
    """
    queryset = ProductTag.objects.all()
    serializer_class = ProductTagSerializer  # 序列化

    filter_backends = (DjangoFilterBackend,)  # 过滤
    filter_fields = ('product', 'tag',)  # 配置过滤字段


class SupplierViewSet(mixins.ListModelMixin,
                      mixins.CreateModelMixin,
                      mixins.UpdateModelMixin,
                      mixins.DestroyModelMixin,
                      mixins.RetrieveModelMixin,
                      viewsets.GenericViewSet):
    """
    list:
        供应商列表,分页,过滤,搜索,排序
    create:
        供应商新增
    retrieve:
        供应商详情页
    update:
        供应商修改
    destroy:
        供应商删除
    """
    queryset = Supplier.objects.all()
    serializer_class = SupplierSerializer  # 序列化
    pagination_class = DefaultPagination  # 分页

    filter_backends = (DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter)  # 过滤,搜索,排序
    filter_fields = ('buy_way', 'is_active')  # 配置过滤字段
    search_fields = ('supplier_name', 'contact_name', 'phone', 'email')  # 配置搜索字段
    ordering_fields = ('create_time',)  # 配置排序字段

    #  重写供应商删除
    def destroy(self, request, *args, **kwargs):
        supplier = self.get_object()
        # 检查是否有采购单存在
        is_exist = PurchaseOrder.objects.filter(supplier=supplier).count()

        if is_exist:
            return Response({'msg': '该供应商有关联数据，无法删除'}, status=status.HTTP_406_NOT_ACCEPTABLE)
        # 如果满足条件，则可以删除该供应商
        supplier.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)
