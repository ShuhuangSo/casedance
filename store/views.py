import time
from random import Random

from rest_framework import mixins, status
from rest_framework import viewsets
from rest_framework.response import Response
from rest_framework import filters
from rest_framework.pagination import PageNumberPagination
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.decorators import action
import openpyxl

from product.models import Product
from .models import Store, StockInOut, StockInOutDetail, Stock, StockLog
from .serializers import StoreSerializer, StockInOutSerializer, StockLogSerializer, Stock2Serializer


# Create your views here.


class DefaultPagination(PageNumberPagination):
    """
    分页设置
    """
    page_size = 20
    page_size_query_param = 'page_size'
    page_query_param = 'page'
    max_page_size = 100


class StoreViewSet(mixins.ListModelMixin,
                   mixins.CreateModelMixin,
                   mixins.UpdateModelMixin,
                   mixins.DestroyModelMixin,
                   mixins.RetrieveModelMixin,
                   viewsets.GenericViewSet):
    """
    list:
        仓库、销售门店列表,分页,过滤,搜索,排序
    create:
        仓库、销售门店新增
    retrieve:
        仓库、销售门店详情页
    update:
        仓库、销售门店修改
    destroy:
        仓库、销售门店删除
    """
    queryset = Store.objects.all()
    serializer_class = StoreSerializer  # 序列化
    pagination_class = DefaultPagination  # 分页

    filter_backends = (DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter)  # 过滤,搜索,排序
    filter_fields = ('type', 'is_active')  # 配置过滤字段
    search_fields = ('store_name', 'contact_name', 'phone')  # 配置搜索字段
    ordering_fields = ('create_time',)  # 配置排序字段

    # 重写create
    def create(self, request, *args, **kwargs):
        store_name = request.data['store_name']
        s_type = request.data['type']
        address = request.data['address']
        contact_name = request.data['contact_name']
        phone = request.data['phone']
        qq = request.data['qq']
        wechat = request.data['wechat']
        website = request.data['website']
        note = request.data['note']

        store = Store()
        store.store_name = store_name
        store.type = s_type
        store.address = address
        store.contact_name = contact_name
        store.phone = phone
        store.qq = qq
        store.wechat = wechat
        store.website = website
        store.note = note
        store.save()

        queryset = Product.objects.all()
        add_list = []
        for i in queryset:
            add_list.append(
                Stock(
                    product=i,
                    store=store
                )
            )
        Stock.objects.bulk_create(add_list)

        return Response({'msg': '操作成功！'}, status=status.HTTP_200_OK)


class StockInOutViewSet(mixins.ListModelMixin,
                        mixins.CreateModelMixin,
                        mixins.UpdateModelMixin,
                        mixins.DestroyModelMixin,
                        mixins.RetrieveModelMixin,
                        viewsets.GenericViewSet):
    """
    list:
        手工出入库/调拨列表,分页,过滤,搜索,排序
    create:
        手工出入库/调拨新增
    retrieve:
        手工出入库/调拨详情页
    update:
        手工出入库/调拨修改
    destroy:
        手工出入库/调拨删除
    """
    queryset = StockInOut.objects.all()
    serializer_class = StockInOutSerializer  # 序列化
    pagination_class = DefaultPagination  # 分页

    filter_backends = (DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter)  # 过滤,搜索,排序
    filter_fields = (
        'origin_store', 'target_store', 'user', 'type', 'reason_in', 'reason_out', 'reason_move', 'is_active')  # 配置过滤字段
    search_fields = ('batch_number', 'inout_detail__product__sku')  # 配置搜索字段
    ordering_fields = ('create_time',)  # 配置排序字段

    # 重写create
    def create(self, request, *args, **kwargs):

        # 获取不重复批次单号(18位：M+17位数字)
        random_ins = Random()
        batch_number = 'M{time_str}{userid}{ranstr}'.format(time_str=time.strftime('%Y%m%d%H%M%S'),
                                                            userid=request.user.id,
                                                            ranstr=random_ins.randint(10, 99))
        # 创建出入库单
        origin_store = None
        target_store = None
        if Store.objects.filter(id=request.data['target_store']):
            target_store = Store.objects.all().get(id=request.data['target_store'])
        if request.data['origin_store']:
            if Store.objects.filter(id=request.data['origin_store']):
                origin_store = Store.objects.all().get(id=request.data['origin_store'])
        else:
            origin_store = target_store

        stock_inout = StockInOut()
        stock_inout.batch_number = batch_number
        stock_inout.origin_store = origin_store
        stock_inout.target_store = target_store
        stock_inout.user = request.user
        stock_inout.type = request.data['type']
        if request.data['reason_in']:
            stock_inout.reason_in = request.data['reason_in']
        if request.data['reason_out']:
            stock_inout.reason_out = request.data['reason_out']
        if request.data['reason_move']:
            stock_inout.reason_move = request.data['reason_move']
        stock_inout.note = request.data['note']
        stock_inout.save()

        # 创建出入库单产品详情
        inout_detail = request.data['inout_detail']
        if inout_detail:
            add_list = []
            for i in inout_detail:
                product = Product.objects.all().get(id=i['id'])
                stock = Stock.objects.filter(store=target_store).get(product=product)
                add_list.append(
                    StockInOutDetail(
                        stock_in_out=stock_inout,
                        product=product,
                        qty=i['qty'],
                        stock_before=stock.qty
                    )
                )
            StockInOutDetail.objects.bulk_create(add_list)

        # 产品入库操作
        if stock_inout.type == 'IN' and target_store:
            for i in add_list:
                stock = Stock.objects.filter(store=target_store).get(product=i.product)
                stock.qty += i.qty  # 现有库存加上入库库存
                stock.save()

                #  产品手工入库日志记录保存
                stock_log = StockLog()
                stock_log.qty = i.qty
                stock_log.product = i.product
                stock_log.store = target_store
                stock_log.user = request.user
                stock_log.op_type = 'M_IN'
                stock_log.op_origin_id = stock_inout.id
                stock_log.save()

        #  产品出库操作
        if stock_inout.type == 'OUT' and target_store:
            for i in add_list:
                stock = Stock.objects.filter(store=target_store).get(product=i.product)
                stock.qty -= i.qty  # 现有库存减去出库库存
                stock.save()

                #  产品手工出库日志记录保存
                stock_log = StockLog()
                stock_log.qty = i.qty
                stock_log.product = i.product
                stock_log.store = target_store
                stock_log.user = request.user
                stock_log.op_type = 'M_OUT'
                stock_log.op_origin_id = stock_inout.id
                stock_log.save()

        #  库存调拨操作
        if stock_inout.type == 'MOVE' and target_store and origin_store:
            for i in add_list:
                ta_stock = Stock.objects.filter(store=target_store).get(product=i.product)
                or_stock = Stock.objects.filter(store=origin_store).get(product=i.product)
                or_stock.qty -= i.qty  # 源仓库减去出库库存
                or_stock.save()
                ta_stock.qty += i.qty  # 目标仓库加上入库库存
                ta_stock.save()

                #  源仓库产品手工出库日志记录保存
                or_stock_log = StockLog()
                or_stock_log.qty = i.qty
                or_stock_log.product = i.product
                or_stock_log.store = origin_store
                or_stock_log.user = request.user
                or_stock_log.op_type = 'M_OUT'
                or_stock_log.op_origin_id = stock_inout.id
                or_stock_log.save()

                #  目标仓库产品手工入库日志记录保存
                ta_stock_log = StockLog()
                ta_stock_log.qty = i.qty
                ta_stock_log.product = i.product
                ta_stock_log.store = target_store
                ta_stock_log.user = request.user
                ta_stock_log.op_type = 'M_IN'
                ta_stock_log.op_origin_id = stock_inout.id
                ta_stock_log.save()

        return Response({'msg': '操作成功！'}, status=status.HTTP_200_OK)

    # excel批量上传库存盘点
    @action(methods=['post'], detail=False, url_path='stock_taking')
    def stock_taking(self, request):
        data = request.data
        target_store = data['target_store']
        wb = openpyxl.load_workbook(data['excel'])
        sheet = wb['盘点上传模板']

        err_list = []  # 错误列表
        add_list = []  # 盘点sku
        if sheet.max_row <= 1:
            err_list.append({'msg': '表格不能为空'})
            return Response(err_list, status=status.HTTP_202_ACCEPTED)

        # 获取不重复批次单号(18位：M+17位数字)
        random_ins = Random()
        batch_number = 'M{time_str}{userid}{ranstr}'.format(time_str=time.strftime('%Y%m%d%H%M%S'),
                                                            userid=request.user.id,
                                                            ranstr=random_ins.randint(10, 99))
        store = Store.objects.filter(id=target_store).first()
        stock_inout = StockInOut()
        stock_inout.batch_number = batch_number
        stock_inout.origin_store = store
        stock_inout.target_store = store
        stock_inout.user = request.user
        stock_inout.type = 'TAKING'
        stock_inout.save()

        for cell_row in list(sheet)[1:]:
            err_item = {}
            row_status = cell_row[0].value and cell_row[2].value

            if not row_status:
                err_item.update({'row': '第 %s 行' % cell_row[0].row})
                err_item.update({'msg': '数据格式错误'})
                err_list.append(err_item)
                continue
            if not type(cell_row[2].value) in [int]:
                err_item.update({'row': '第 %s 行' % cell_row[0].row})
                err_item.update({'msg': '盘点数量格式不正确'})
                err_list.append(err_item)
                continue
            # 检查sku是否存在
            sku_is_exist = Product.objects.filter(sku=cell_row[0].value.strip()).count()
            if not sku_is_exist:
                err_item = {}
                err_item.update({'row': '第 %s 行' % cell_row[0].row})
                err_item.update({'msg': 'sku: %s 不存在' % cell_row[0].value.strip()})
                err_list.append(err_item)
                continue

            # 创建出入口单详情对象
            sku = cell_row[0].value.strip()
            qty = cell_row[2].value
            product = Product.objects.filter(sku=sku).first()
            stock = Stock.objects.filter(store=store).get(product=product)
            add_list.append(
                StockInOutDetail(
                    stock_in_out=stock_inout,
                    product=product,
                    qty=qty,
                    stock_before=stock.qty
                )
            )
        StockInOutDetail.objects.bulk_create(add_list)

        # 库存盘点变动操作
        for i in add_list:
            stock = Stock.objects.filter(store=target_store).get(product=i.product)
            stock.qty = i.qty
            stock.save()

            #  产品盘点日志记录保存
            stock_log = StockLog()
            stock_log.qty = i.qty - i.stock_before
            stock_log.product = i.product
            stock_log.store = store
            stock_log.user = request.user
            stock_log.op_type = 'TAKING'
            stock_log.op_origin_id = stock_inout.id
            stock_log.save()

        success_count = len(add_list)
        if not success_count:
            stock_inout.delete()
        fail_count = sheet.max_row - 1 - success_count
        all_data = {}
        all_data.update({'err_list': err_list})
        all_data.update({'fail_count': fail_count})
        all_data.update({'success_count': success_count})

        return Response(all_data, status=status.HTTP_200_OK)


class StockLogViewSet(mixins.ListModelMixin,
                      mixins.CreateModelMixin,
                      mixins.UpdateModelMixin,
                      mixins.DestroyModelMixin,
                      mixins.RetrieveModelMixin,
                      viewsets.GenericViewSet):
    """
    list:
        库存出入日志列表,分页,过滤,搜索,排序
    create:
        库存出入日志新增
    retrieve:
        库存出入日志详情页
    update:
        库存出入日志修改
    destroy:
        库存出入日志删除
    """
    queryset = StockLog.objects.all()
    serializer_class = StockLogSerializer  # 序列化
    pagination_class = DefaultPagination  # 分页

    filter_backends = (DjangoFilterBackend, filters.OrderingFilter)  # 过滤,搜索,排序
    filter_fields = ('op_type', 'user', 'store', 'product')  # 配置过滤字段
    ordering_fields = ('create_time',)  # 配置排序字段


class StockViewSet(mixins.ListModelMixin,
                   mixins.CreateModelMixin,
                   mixins.UpdateModelMixin,
                   mixins.DestroyModelMixin,
                   mixins.RetrieveModelMixin,
                   viewsets.GenericViewSet):
    """
    list:
        库存列表,分页,过滤,搜索,排序
    create:
        库存新增
    retrieve:
        库存详情页
    update:
        库存修改
    destroy:
        库存删除
    """
    queryset = Stock.objects.all()
    serializer_class = Stock2Serializer  # 序列化
    pagination_class = DefaultPagination  # 分页

    filter_backends = (DjangoFilterBackend, filters.OrderingFilter)  # 过滤,搜索,排序
    filter_fields = ('product__series', 'store')  # 配置过滤字段
    ordering_fields = ('qty',)  # 配置排序字段
